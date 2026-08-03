"""
OCR Router: tries PaddleOCR first (free, local), escalates to Groq (Fast Vision),
Gemini (Reliable Vision), and finally Azure (Enterprise Fallback).

Units: Quintal (Qtl) throughout — NOT metric tons.
1 Quintal = 100 kg. Bills in this business are always in Qtl.
"""
import re, json, base64, logging, os, asyncio, time, tempfile
import sys
from pathlib import Path
from typing import Optional, List, Tuple
from dotenv import load_dotenv

log = logging.getLogger(__name__)

_OPENROUTER_FREE_BLOCKED_UNTIL_TS = 0.0
_OPENROUTER_FREE_VISION_MODELS_CACHE: List[str] = []
_OPENROUTER_FREE_VISION_MODELS_CACHE_TS = 0.0

KNOWN_MATERIALS = {"maize", "dorb", "domc", "rice ddgs", "ddgs"}
VEHICLE_RE = re.compile(r'[A-Z]{2}\s*\d{2}\s*[A-Z]{1,2}\s*\d{4}', re.IGNORECASE)


def _normalize_vehicle_number(value: Optional[str]) -> Optional[str]:
    raw = str(value or "").strip().upper()
    if not raw:
        return None

    # Keep only alphanumerics, removing spaces and separators.
    compact = re.sub(r"[^A-Z0-9]", "", raw)
    if not compact:
        return None

    # Prefer canonical vehicle pattern when detectable from noisy text.
    m = VEHICLE_RE.search(raw)
    if m:
        return re.sub(r"\s+", "", m.group(0)).upper()
    return compact


def _normalize_bill_number(value: Optional[str]) -> Optional[str]:
    s = str(value or "").strip()
    if not s:
        return None

    # Remove common label prefixes.
    s = re.sub(r"(?i)^\s*(invoice|inv|bill)\s*(no|number)?\s*[:#\-]*\s*", "", s).strip()

    # Ignore fixed path-like prefixes and keep the last segment.
    parts = [p.strip() for p in re.split(r"[/\\]+", s) if p.strip()]
    if parts:
        s = parts[-1]

    # Business rule: keep the trailing numeric identifier when present.
    nums = re.findall(r"\d+", s)
    if nums:
        return nums[-1]

    # Fallback to cleaned alphanumeric token.
    cleaned = re.sub(r"\s+", "", s)
    return cleaned or None


def _llm_sleep_seconds() -> float:
    load_dotenv()
    raw = (os.getenv("LLM_REQUEST_SLEEP_SEC") or "2.5").strip()
    try:
        sec = float(raw)
    except Exception:
        sec = 2.5
    return max(0.0, min(sec, 10.0))


async def _sleep_before_llm_request_async() -> None:
    sec = _llm_sleep_seconds()
    if sec > 0:
        await asyncio.sleep(sec)


def _sleep_before_llm_request_sync() -> None:
    sec = _llm_sleep_seconds()
    if sec > 0:
        time.sleep(sec)


def _openrouter_free_only_enabled() -> bool:
    load_dotenv()
    return (os.getenv("OPENROUTER_FREE_ONLY", "true").strip().lower() in {"1", "true", "yes", "on"})


def _openrouter_api_base_url() -> str:
    load_dotenv()
    return (os.getenv("OPENROUTER_API_BASE_URL") or "https://openrouter.ai/api/v1").strip().rstrip("/")


def _openrouter_model_blocklist() -> set:
    load_dotenv()
    raw_env = os.getenv("OPENROUTER_MODEL_BLOCKLIST")
    raw = ("nvidia/nemotron-nano-12b-v2-vl:free" if raw_env is None else raw_env).strip()
    out = set()
    for token in re.split(r"[,\s]+", raw):
        s = str(token or "").strip()
        if s:
            out.add(s)
    return out


def _openrouter_doc_preferred_models() -> List[str]:
    """Preferred free OCR candidates inspired by free-llm-image-to-text docs."""
    return [
        "amazon/nova-2-lite-v1:free",
        "google/gemma-4-31b-it:free",
        "google/gemma-4-26b-a4b-it:free",
        "google/gemma-3-27b-it:free",
        "google/gemma-3-12b-it:free",
        "google/gemma-3-4b-it:free",
        "openai/gpt-4o-mini:free",
    ]


def _openrouter_discover_free_vision_models(api_key: str) -> List[str]:
    global _OPENROUTER_FREE_VISION_MODELS_CACHE
    global _OPENROUTER_FREE_VISION_MODELS_CACHE_TS

    load_dotenv()
    now = time.time()
    ttl_raw = (os.getenv("OPENROUTER_MODEL_DISCOVERY_TTL_SEC") or "900").strip()
    try:
        ttl_sec = max(60, min(3600, int(ttl_raw)))
    except Exception:
        ttl_sec = 900

    if _OPENROUTER_FREE_VISION_MODELS_CACHE and (now - _OPENROUTER_FREE_VISION_MODELS_CACHE_TS) < ttl_sec:
        return list(_OPENROUTER_FREE_VISION_MODELS_CACHE)

    try:
        from openai import OpenAI

        client = OpenAI(base_url=_openrouter_api_base_url(), api_key=api_key, max_retries=0)
        data = client.models.list()
        models = getattr(data, "data", []) or []

        found: List[str] = []
        for item in models:
            model_id = str(getattr(item, "id", "") or "").strip()
            if not model_id or not model_id.endswith(":free"):
                continue

            try:
                meta = item.model_dump()  # pydantic model
            except Exception:
                meta = {}

            arch = meta.get("architecture") if isinstance(meta, dict) else {}
            if not isinstance(arch, dict):
                arch = {}

            input_modalities = [str(x).lower() for x in (arch.get("input_modalities") or [])]
            modality = str(arch.get("modality") or "").lower()
            if ("image" in input_modalities) or ("image" in modality):
                found.append(model_id)

        deduped = [m for i, m in enumerate(found) if m and m not in found[:i]]
        _OPENROUTER_FREE_VISION_MODELS_CACHE = deduped
        _OPENROUTER_FREE_VISION_MODELS_CACHE_TS = now
        return list(deduped)
    except Exception as e:
        # Keep flow resilient: fall back to static candidates when discovery endpoint fails.
        log.warning("OpenRouter model discovery failed, using static fallback: %s", str(e)[:220])
        return list(_OPENROUTER_FREE_VISION_MODELS_CACHE)


def _openrouter_resolve_model_candidates(
    api_key: str,
    *,
    configured_items: Optional[List[str]] = None,
    include_paid: bool = False,
) -> List[str]:
    configured_items = [str(x or "").strip() for x in (configured_items or []) if str(x or "").strip()]
    preferred = [*configured_items, *_openrouter_doc_preferred_models()]
    preferred = [m for i, m in enumerate(preferred) if m and m not in preferred[:i]]

    blocklist = _openrouter_model_blocklist()
    available_free = _openrouter_discover_free_vision_models(api_key)

    out: List[str] = []

    if available_free:
        for model_id in preferred:
            if model_id.endswith(":free") and model_id in available_free and model_id not in blocklist and model_id not in out:
                out.append(model_id)
        for model_id in available_free:
            if model_id not in blocklist and model_id not in out:
                out.append(model_id)
    else:
        for model_id in preferred:
            if model_id.endswith(":free") and model_id not in blocklist and model_id not in out:
                out.append(model_id)

    if include_paid:
        paid_fallbacks = [
            "qwen/qwen2.5-vl-72b-instruct",
            "qwen/qwen2.5-vl-32b-instruct",
            "qwen/qwen-vl-max",
            "qwen/qwen-vl-plus",
        ]
        for model_id in paid_fallbacks:
            if model_id not in out:
                out.append(model_id)

    return out


def _is_openrouter_free_quota_error(err_text: str) -> bool:
    e = str(err_text or "").lower()
    if "free-models-per-day" in e:
        return True
    if "rate limit exceeded" in e and "openrouter" in e:
        return True
    if "too many requests" in e:
        return True
    if "temporarily rate-limited upstream" in e:
        return True
    if "'code': 429" in e or '"code": 429' in e:
        return True
    if "error code: 429" in e and ("x-ratelimit-remaining" in e or "rate limit" in e):
        return True
    return False


def _mark_openrouter_free_quota_exhausted(err_text: str) -> None:
    global _OPENROUTER_FREE_BLOCKED_UNTIL_TS

    now = time.time()
    reset_ts = now + 1800  # default 30 min cooldown when reset header is unavailable
    text = str(err_text or "")

    m = re.search(r"X-RateLimit-Reset['\"]?\s*[:=]\s*['\"]?(\d{10,13})", text, flags=re.IGNORECASE)
    if m:
        try:
            raw = int(m.group(1))
            parsed = (raw / 1000.0) if raw > 10_000_000_000 else float(raw)
            if parsed > now:
                reset_ts = parsed
        except Exception:
            pass

    _OPENROUTER_FREE_BLOCKED_UNTIL_TS = max(_OPENROUTER_FREE_BLOCKED_UNTIL_TS, reset_ts)
    wait_sec = int(max(1, _OPENROUTER_FREE_BLOCKED_UNTIL_TS - now))
    log.warning("OpenRouter free tier quota exhausted; skipping OpenRouter for ~%ss", wait_sec)


def _openrouter_free_temporarily_blocked() -> bool:
    return _openrouter_free_only_enabled() and (time.time() < _OPENROUTER_FREE_BLOCKED_UNTIL_TS)


def _openrouter_retry_after_sec() -> int:
    if _OPENROUTER_FREE_BLOCKED_UNTIL_TS <= 0:
        return 0
    return int(max(0, _OPENROUTER_FREE_BLOCKED_UNTIL_TS - time.time()))


def _mistral_enabled() -> bool:
    load_dotenv()
    return (os.getenv("MISTRAL_ENABLED", "true").strip().lower() in {"1", "true", "yes", "on"})


def _mistral_api_base_url() -> str:
    load_dotenv()
    return (os.getenv("MISTRAL_BASE_URL") or "https://api.mistral.ai/v1").strip().rstrip("/")


def _mistral_resolve_model_candidates(*, configured_items: Optional[List[str]] = None) -> List[str]:
    configured_items = [str(x or "").strip() for x in (configured_items or []) if str(x or "").strip()]
    defaults = [
        "ministral-3-14b-2512",
        "ministral-3-8b-2512",
        "pixtral-large-2411",
        "pixtral-12b-2409",
    ]
    out = [*configured_items, *defaults]
    return [m for i, m in enumerate(out) if m and m not in out[:i]]


def _is_mistral_rate_limit_error(err_text: str) -> bool:
    e = str(err_text or "").lower()
    if "error code: 429" in e:
        return True
    if "\"code\": 429" in e or "'code': 429" in e:
        return True
    if "too many requests" in e:
        return True
    if "rate limit" in e:
        return True
    if "quota" in e and "mistral" in e:
        return True
    return False

# ── PaddleOCR ─────────────────────────────────────────────────────────────

_paddle = None

def get_paddle():
    global _paddle
    if _paddle is None:
        try:
            from paddleocr import PaddleOCR
            _paddle = PaddleOCR(use_angle_cls=True, lang='en', show_log=False)
            log.info("PaddleOCR loaded")
        except Exception as e:
            log.warning(f"PaddleOCR not available: {e}")
    return _paddle

def extract_with_paddle(image_path: str) -> dict:
    paddle = get_paddle()
    if paddle is None:
        return {"high_confidence": False, "fields_found": 0, "source": "paddle", "error": "PaddleOCR not installed"}

    try:
        result = paddle.ocr(image_path, cls=True)
        lines, confidences = [], []
        for block in (result or []):
            for line in (block or []):
                if line and len(line) > 1 and line[1]:
                    text, conf = line[1][0], line[1][1]
                    lines.append(text.strip())
                    confidences.append(conf)

        full_text = "\n".join(lines)
        avg_conf = sum(confidences) / len(confidences) if confidences else 0

        extracted = _parse_bill_text(full_text)
        extracted.update({
            "raw_text":       full_text,
            "ocr_confidence": round(avg_conf, 3),
            "source":         "paddle",
        })
        return extracted
    except Exception as e:
        log.error(f"PaddleOCR error: {e}")
        return {"high_confidence": False, "fields_found": 0, "source": "paddle", "error": str(e)}


def _parse_bill_text(text: str) -> dict:
    r = {}

    # Vehicle number (Rajasthan pattern)
    v = VEHICLE_RE.search(text)
    r["vehicle_number"] = _normalize_vehicle_number(v.group(0) if v else None)

    # Quantity — keep in Quintal, also handle MT/KG and convert
    qt_qtl = re.search(r'(\d[\d,]*\.?\d*)\s*(?:Qtl|Quintal|QTL|qtl)', text, re.IGNORECASE)
    qt_mt  = re.search(r'(\d[\d,]*\.?\d*)\s*(?:MT|M\.T\.?|metric\s*ton)', text, re.IGNORECASE)
    qt_kg  = re.search(r'(\d[\d,]*\.?\d*)\s*(?:KGS?|K\.G\.?)', text, re.IGNORECASE)
    if qt_qtl:
        r["quantity_qtl"] = float(qt_qtl.group(1).replace(",", ""))
        r["quantity_unit"] = "qtl"
    elif qt_mt:
        r["quantity_qtl"] = round(float(qt_mt.group(1).replace(",", "")) * 10, 3)
        r["quantity_unit"] = "mt"
    elif qt_kg:
        r["quantity_qtl"] = round(float(qt_kg.group(1).replace(",", "")) / 100, 3)
        r["quantity_unit"] = "kg"
    else:
        r["quantity_qtl"] = None
        r["quantity_unit"] = None

    # Rate per Quintal
    rate_unit_match = re.search(
        r'(?:rate|price|@)?\s*per\s*(qtl|quintal|mt|m\.t\.?|metric\s*ton|kg|k\.g\.?)\s*[:\-]?\s*(?:Rs\.?|₹)?\s*([\d,]+\.?\d*)',
        text,
        re.IGNORECASE,
    )
    if rate_unit_match:
        unit_raw = (rate_unit_match.group(1) or "").lower()
        rate_val = float(rate_unit_match.group(2).replace(",", ""))
        if "kg" in unit_raw:
            r["rate_per_qtl"] = round(rate_val * 100, 3)
            r["rate_unit"] = "kg"
        elif "mt" in unit_raw or "ton" in unit_raw:
            r["rate_per_qtl"] = round(rate_val / 10, 3)
            r["rate_unit"] = "mt"
        else:
            r["rate_per_qtl"] = rate_val
            r["rate_unit"] = "qtl"
    else:
        rate = re.search(r'(?:rate|price|@)\s*[:\-]?\s*(?:Rs\.?|₹)?\s*([\d,]+\.?\d*)', text, re.IGNORECASE)
        r["rate_per_qtl"] = float(rate.group(1).replace(",", "")) if rate else None
        r["rate_unit"] = None

    # Total amount
    total = re.search(r'(?:grand\s*total|total|amount|net)\s*[:\-]?\s*(?:Rs\.?|₹)\s*([\d,]+\.?\d*)', text, re.IGNORECASE)
    r["total_amount"] = float(total.group(1).replace(",", "")) if total else None

    # Vendor name — look for company name patterns
    company = re.search(r'(?:M/S\.?|M/s\.?|Messrs\.?)?\s*([A-Z][A-Z\s&\.]+(?:TRADING|INDUSTRIES|ENTERPRISES|COMPANY|SUPPLIERS|AGRO|FOODS|TRADERS))', text)
    r["vendor_name"] = company.group(0).strip() if company else None

    # Material
    tl = text.lower()
    mat = None
    for m in ["Rice DDGS", "Maize", "Dorb", "Domc"]:
        if m.lower() in tl or (m == "Rice DDGS" and "ddgs" in tl):
            mat = m
            break
    r["material_type"] = mat

    # Bill number
    bn = re.search(r'(?:invoice\s*no|bill\s*no|inv\.?\s*no|invoice)\s*[:\-#]?\s*([A-Z0-9\-/]+)', text, re.IGNORECASE)
    r["bill_number"] = _normalize_bill_number(bn.group(1) if bn else None)

    # Bill date
    dt = re.search(r'(?:dated?|date)\s*[:\-]?\s*(\d{1,2})[/\-\.](\d{1,2})[/\-\.](\d{2,4})', text, re.IGNORECASE)
    if not dt:
        dt = re.search(r'(\d{1,2})[/\-\.](\d{1,2})[/\-\.](\d{2,4})', text)
    if dt:
        d, mo, y = dt.group(1), dt.group(2), dt.group(3)
        y = f"20{y}" if len(y) == 2 else y
        r["bill_date"] = f"{y}-{mo.zfill(2)}-{d.zfill(2)}"
    else:
        r["bill_date"] = None

    # Destination plant
    plant_found = None
    delivery = re.search(r'(?:delivery\s*at|deliver\s*to)\s*[:\-]?\s*(.+)', text, re.IGNORECASE)
    if delivery:
        dlv = delivery.group(1).strip()
        for p in ["Ajmer", "Jodhpur", "Kaladers", "Kaladera", "Nadbai", "Bikaner", "Pali", "Lambiyan"]:
            if p.lower() in dlv.lower():
                plant_found = "Kaladers" if p == "Kaladera" else p
                break
        if not plant_found:
            plant_found = dlv[:50]
    else:
        for p in ["Ajmer", "Jodhpur", "Kaladers", "Kaladera", "Nadbai", "Bikaner", "Pali", "Lambiyan"]:
            if p.lower() in tl:
                plant_found = "Kaladers" if p == "Kaladera" else p
                break
    r["destination_plant"] = plant_found

    key_fields = ["vehicle_number", "quantity_qtl", "rate_per_qtl", "total_amount", "material_type"]
    filled = sum(1 for f in key_fields if r.get(f) is not None)
    r["fields_found"]    = filled
    r["high_confidence"] = filled >= 4
    return r


# ── Gemini prompt ──────────────────────────────────────────────────────────

GEMINI_PROMPT = """
You are reading an Indian cattle feed raw material tax invoice / purchase bill photo.
The bill may be handwritten or printed in English or Hindi.
This is for a cattle feed plant in Rajasthan, India.

Extract ALL fields below. Return ONLY a valid JSON object — no markdown, no explanation, no preamble.

{
  "vendor_name": "The SELLER company name at the TOP of the bill (e.g. M/S. SALASAR TRADING COMPANY). This is the company ISSUING the bill, not the buyer. Look for large bold text at the top.",
  "vehicle_number": "truck/transport vehicle registration number e.g. RJ04GB5606 — NEVER guess, return null if not clearly visible",
    "material_type": "one of exactly: Maize, Dorb, Domc, Rice DDGS — match to description of goods if you see musturd its domc as domc is de-oiled musturd",
  "quantity_qtl": its the number under/near qty/quantity weight of material<quantity in QUINTAL (Qtl) as a number if no unit then keep it as it is. If bill shows Qtl/QTL keep as-is. If MT multiply by 10. e.g. 420.700>, dont take number of bags in this field, only the weight. If both number of bags and weight are visible, use weight for quantity and ignore number of bags.
  "rate_per_qtl": <price per Quintal in INR as a number e.g. 1925.00> its mostly under "Rate" column or near "@" symbol.
  "total_amount": <Grand Total in INR as a number. Indian format: 8,09,848 = 809848>,
  "bill_date": "Date of bill if visible, else null. Look for any date format or similar patterns of date",
  "bill_number": "Invoice number e.g. 184",
  "gstin_vendor": "GSTIN of the seller if visible",
  "broker_name": "Broker name if mentioned on bill e.g. YESH DALAL",
  "destination_plant": "Delivery location / plant name if mentioned e.g. Kaladera, Bikaner,jodhpur, Ajmer,nadbai,pali, lambiyan ,Kaladera, bhilwara. Look for 'Delivery at' or 'Shipped to' CATTLE FEED PLANT NADBAI means plant name is Nadbai, if you see only NADBAI then also its plant name, if you see KALADERA or KALADERS then plant name is Kaladers",
  "transport_company": "Transport/logistics company name if visible",
  "is_handwritten": <true if handwritten, false if printed>,
  "confidence": <0.0 to 1.0 — your overall extraction confidence>,
  "unclear_fields": ["list field names you were uncertain about"]
}

Critical rules:
- vendor_name = the SELLER at top of bill (large heading), NOT the buyer
- quantity_qtl must always be in Quintal — convert if needed,its the number under/near QTY./quantity weight of material, dont take number of bags in this field or near witg kgs as unit
- Indian lakh format: 8,09,848 = 809848 (not 80984.8)
- Never invent vehicle numbers — return null if unclear
- bill_number is just the number e.g. "184" not "Invoice No. 184"
- For destination_plant: "KALADERA" or "KALADERS" = "Kaladers"  Kaladera, Bikaner,jodhpur, Ajmer,nadbai,pali, lambiyan ,Kaladera, bhilwara. Look for 'Delivery at' or 'Shipped to' CATTLE FEED PLANT NADBAI means plant name is Nadbai, if you see only NADBAI then also its plant name, if you see KALADERA or KALADERS then plant name is Kaladers",
"""

TENDER_NOTICE_PROMPT = """
You are reading a tender/NIT (Notice Inviting Tender) document for cattle feed raw material supply setting up requirements.
Return ONLY one valid JSON object.

{
    "document_type": "tender_notice",
    "tender_rm_number": "Tender/NIT/RM number (often starts with RM- or e-NIT No., e.g., RM-828) or null",
    "notice_date": "YYYY-MM-DD or null",
    "plant_name": "Plant name where supply is required (e.g. Kaladera, Bikaner, Jodhpur) or null",
    "supply_period_start": "YYYY-MM-DD (e.g. from 01-04-2026) or null",
    "supply_period_end": "YYYY-MM-DD (e.g. to 30-04-2026) or null",
    "items": [
        {
            "material_type": "Material name (e.g., Maize, Dorb, Rice DDGS)",
            "quantity_qtl": 0.0,
            "quantity_unit": "Qtl/MT/KG or null"
        }
    ],
    "confidence": 0.0,
    "unclear_fields": ["field names"]
}

Rules:
- This is NOT a purchase order. It does not have a winner yet.
- Do not invent values. Keep dates in YYYY-MM-DD.
- quantity_qtl must be numeric; if source uses MT convert to Qtl (x10), if KG convert to Qtl (/100).
- Extract overall plant name from header/subject, and find supply period dates (often written 'supply period from X to Y').
"""

PURCHASE_ORDER_PROMPT = """
You are reading a Purchase Order / Work Order / Allotment Letter for cattle feed raw material supply.
Return ONLY one valid JSON object (no markdown, no explanation).

Important business filter:
We only care about rows where winner/approved party is one of these companies:
1) Shree Nath Industries
2) Shree Vinayak Trading Company
3) Shree Ganpati Enterpriese
IGNORE ALL OTHER COMPANIES AND THERE DETAILS 
Treat small spelling variations/case differences as same company.

{
    "document_type": "purchase_order",
    "po_number": "PO/Order number or null",
    "tender_rm_number": "RM/Tender reference (e.g. RM-828) or null",
    "main_tender_name": "same as tender_rm_number when available, else null",
    "po_date": "YYYY-MM-DD or null",
    "supply_period_start": "YYYY-MM-DD or null",
    "supply_period_end": "YYYY-MM-DD or null",
    "plant_name": "Primary destination plant/location or null",
    "winner_party_name": "Winner party for this PO if clearly stated, else null",
    "winner_party_email": "Winner email if visible",
    "our_company_winner_name": "One of: Shree Nath Industries / Shree Vinayak Trading Company / Shree Ganpati Enterpriese, else null",
    "has_our_company_winner": true,
    "items": [
        {
            "material_type": "Material name (Maize/Dorb/Domc/Rice DDGS or null)",
            "approved_quantity_qtl": 0.0,
            "approved_rate_per_qtl": 0.0,
            "approved_party_name": "Winning party for this line item",
            "plant_name": "Plant for this item if shown, else parent plant_name",
            "line_amount": 0.0,
            "quantity_unit": "Qtl/MT/KG or null"
        }
    ],
    "sub_tenders": [
        {
            "sub_tender_name": "{tender_rm_number}+{plant_name}",
            "plant_name": "Plant name",
            "material_type": "Material won",
            "tender_quantity_qtl": 0.0,
            "week1_target_qty_qtl": 0.0,
            "week1_deadline_date": "YYYY-MM-DD (supply_period_start + 7 days)",
            "week2_deadline_date": "YYYY-MM-DD (supply_period_end)",
            "cycle": "Any"
        }
    ],
    "total_amount": 0.0,
    "confidence": 0.0,
    "unclear_fields": ["field names"]
}

Rules:
- IGNORE ALL OTHER COMPANIES AND THERE DETAILS 
- Extract plant_name, supply_period_start, supply_period_end carefully.
- approved_quantity_qtl is mandatory for item rows; convert MT to Qtl (x10), KG to Qtl (/100).
- approved_rate_per_qtl should be per Qtl value.
- Keep ONLY those item rows where approved_party_name belongs to one of the 3 companies listed above.
- If no row belongs to those 3 companies: set has_our_company_winner=false, our_company_winner_name=null, items=[].
- Set main_tender_name = tender_rm_number.
- For each returned item, build one sub_tenders row:
  sub_tender_name = tender_rm_number + plant_name,
  tender_quantity_qtl = approved_quantity_qtl,
  week1_target_qty_qtl = approved_quantity_qtl / 2,
  week1_deadline_date = supply_period_start + 7 days,
  week2_deadline_date = supply_period_end,
  cycle = "Any" (if not specified).
- Do not invent PO number, tender_rm_number, dates, quantities, rates, or winner names.
"""

REJECTION_NOTICE_PROMPT = """
You are reading a plant rejection notice/letter for a received truck of raw material that failed quality checks.
Return ONLY one valid JSON object.

{
    "document_type": "rejection_notice",
    "tender_rm_number": "RM/Tender reference if visible (e.g., against RM-828)",
    "vendor_name": "Supplier/party name whose truck is being rejected",
    "truck_number": "Truck/vehicle registration number (e.g., RJ14GB1234) - DO NOT include extra words, max 20 chars",
    "material_type": "Material in the truck (Maize/Dorb/Domc/Rice DDGS) or null",
    "rejection_date": "Date of rejection or notice date in YYYY-MM-DD or null",
    "plant_name": "Plant/location issuing the rejection (e.g. Kaladera, Bikaner, Jodhpur) or null",
    "rejection_type": "'partial' if accepting with deduction/shortage, 'complete' if fully returning truck, or null",
    "rejected_qty_qtl": 0.0,
    "reason": "Text describing rejection reason (e.g., 'wet material', 'moisture high', 'fungus') or null",
    "confidence": 0.0,
    "unclear_fields": ["field names"]
}

Rules:
- If quantity unit is MT convert to Qtl (x10). If KG convert to Qtl (/100).
- If rejected quantity is not clearly readable or it's a complete rejection without a specific weight, set it to null.
- reason should capture specifically why it was rejected.
- truck_number is extremely critical, extract exact alphanumeric (e.g., RJ20GB6238).
"""

SUPPORTED_DOCUMENT_TYPES = {
        "purchase_bill",
        "tender_notice",
        "purchase_order",
        "rejection_notice",
        "plant_unloading",
    "not_classified",
}

# ── Groq Vision Logic ───────────────────────────────────────────────────

async def extract_with_groq(image_path: str) -> dict:
    from groq import Groq
    load_dotenv()
    api_key = os.getenv("GROQ_API_KEY")

    if not api_key:
        return {"source": "groq", "error": "GROQ_API_KEY not set", "high_confidence": False}

    groq_input_path: Optional[str] = None
    cleanup_temp = False
    compact_temp_path: Optional[str] = None
    try:
        src_path = Path(image_path)
        if src_path.suffix.lower() == ".pdf":
            converted = _convert_pdf_first_page_to_image(image_path)
            if not converted:
                return {
                    "source": "groq",
                    "error": "PDF->image conversion failed for Groq",
                    "high_confidence": False,
                }
            groq_input_path = converted
            cleanup_temp = True
        else:
            groq_input_path = image_path

        client = Groq(api_key=api_key)
        data_url, _mime, compact_temp_path = _build_data_url_for_llm(groq_input_path)

        # Try fast vision model first, fall back to slower model on failure
        vision_models = [
            "llama-3.2-90b-vision-preview",  # High quality
            "llama-3.2-11b-vision-preview",  # Fallback
        ]
        last_error = None
        for model_id in vision_models:
            try:
                await _sleep_before_llm_request_async()
                response = client.chat.completions.create(
                    messages=[
                        {
                            "role": "user",
                            "content": [
                                {"type": "text", "text": GEMINI_PROMPT},
                                {
                                    "type": "image_url",
                                    "image_url": {"url": data_url}
                                }
                            ]
                        }
                    ],
                    model=model_id,
                    temperature=0.1,
                    response_format={"type": "json_object"}
                )
                raw_content = response.choices[0].message.content or "{}"
                result = _load_json_object_loose(raw_content)
                result = _normalize_model_bill_payload(result)
                result["source"] = "groq"
                result["groq_model"] = model_id
                fields_found = _count_bill_fields(result)
                result["fields_found"] = fields_found
                conf = float(result.get("confidence") or 0)
                result["high_confidence"] = (fields_found >= 5) or (fields_found >= 4 and conf >= 0.65)

                if "quantity_qtl" not in result and "quantity_mt" in result:
                    try:
                        result["quantity_qtl"] = round(float(result["quantity_mt"]) * 10, 3)
                    except Exception:
                        pass
                log.info("Groq Vision succeeded: model=%s fields=%s high_conf=%s", model_id, fields_found, result["high_confidence"])
                return result
            except Exception as me:
                last_error = me
                log.warning("Groq Vision attempt with %s failed: %s", model_id, me)
                continue

        log.error("Groq Vision: all vision models failed: %s", last_error)
        return {"source": "groq", "error": str(last_error), "high_confidence": False}
    except Exception as e:
        log.error(f"Groq Vision error: {e}")
        return {"source": "groq", "error": str(e), "high_confidence": False}
    finally:
        if compact_temp_path:
            try:
                Path(compact_temp_path).unlink(missing_ok=True)
            except Exception:
                pass
        if cleanup_temp and groq_input_path:
            try:
                Path(groq_input_path).unlink(missing_ok=True)
            except Exception:
                pass


async def extract_with_mistral(image_path: str) -> dict:
    try:
        from openai import OpenAI
    except Exception as e:
        log.error("Mistral unavailable: cannot import openai on %s: %s", sys.executable, e)
        return {
            "source": "mistral",
            "error": f"openai import failed: {e}",
            "high_confidence": False,
        }

    load_dotenv()
    if not _mistral_enabled():
        return {"source": "mistral", "error": "MISTRAL_ENABLED is false", "high_confidence": False}

    api_key = (os.getenv("MISTRAL_API_KEY") or "").strip()
    if not api_key:
        return {"source": "mistral", "error": "MISTRAL_API_KEY not set", "high_confidence": False}

    mistral_input_path: Optional[str] = None
    cleanup_temp = False
    compact_temp_path: Optional[str] = None

    configured_raw = (os.getenv("MISTRAL_MODEL") or "").strip()
    configured_items = [s.strip() for s in re.split(r"[,\s]+", configured_raw) if s.strip()]
    model_candidates = _mistral_resolve_model_candidates(configured_items=configured_items)

    try:
        src_path = Path(image_path)
        if src_path.suffix.lower() == ".pdf":
            converted = _convert_pdf_first_page_to_image(image_path)
            if not converted:
                return {
                    "source": "mistral",
                    "error": "PDF->image conversion failed for Mistral",
                    "high_confidence": False,
                }
            mistral_input_path = converted
            cleanup_temp = True
        else:
            mistral_input_path = image_path

        data_url, _mime, compact_temp_path = _build_data_url_for_llm(mistral_input_path)

        client = OpenAI(
            base_url=_mistral_api_base_url(),
            api_key=api_key,
            max_retries=0,
        )

        last_err = None
        best_result = None
        best_fields = -1
        non_rate_failures: List[str] = []
        for model_id in model_candidates:
            try:
                request_kwargs = {
                    "model": model_id,
                    "messages": [
                        {
                            "role": "user",
                            "content": [
                                {"type": "text", "text": GEMINI_PROMPT},
                                {
                                    "type": "image_url",
                                    "image_url": {"url": data_url}
                                }
                            ]
                        }
                    ],
                    "temperature": 0.1,
                    "max_tokens": 1600,
                    "response_format": {"type": "json_object"},
                }

                try:
                    await _sleep_before_llm_request_async()
                    response = client.chat.completions.create(**request_kwargs)
                except Exception as e:
                    em = str(e).lower()
                    if "json mode is not enabled" in em or "response_format" in em:
                        request_kwargs.pop("response_format", None)
                        await _sleep_before_llm_request_async()
                        response = client.chat.completions.create(**request_kwargs)
                    else:
                        raise

                raw = (response.choices[0].message.content or "").strip()
                result = _load_json_object_loose(raw)
                result = _normalize_model_bill_payload(result)
                result["source"] = "mistral"
                result["mistral_model"] = model_id

                if "quantity_qtl" not in result and "quantity_mt" in result:
                    try:
                        result["quantity_qtl"] = round(float(result["quantity_mt"]) * 10, 3)
                    except Exception:
                        pass

                fields_found = _count_bill_fields(result)
                result["fields_found"] = fields_found
                conf = float(result.get("confidence") or 0)
                result["high_confidence"] = (fields_found >= 5) or (fields_found >= 4 and conf >= 0.65)

                if fields_found > best_fields:
                    best_fields = fields_found
                    best_result = result

                if result["high_confidence"]:
                    return result
                continue
            except Exception as e:
                last_err = e
                if _is_mistral_rate_limit_error(str(e)):
                    log.warning("Mistral rate-limited on model %s: %s", model_id, str(e)[:220])
                else:
                    non_rate_failures.append(f"{model_id}: {str(e)[:180]}")
                continue

        if isinstance(best_result, dict):
            return best_result

        if non_rate_failures:
            log.debug("Mistral non-rate-limit failures: %s", " | ".join(non_rate_failures[:3]))

        return {
            "source": "mistral",
            "error": f"All Mistral model candidates failed: {str(last_err)[:260]}",
            "high_confidence": False,
        }
    except Exception as e:
        log.error("Mistral Vision error: %s", e)
        return {"source": "mistral", "error": str(e), "high_confidence": False}
    finally:
        if compact_temp_path:
            try:
                Path(compact_temp_path).unlink(missing_ok=True)
            except Exception:
                pass
        if cleanup_temp and mistral_input_path:
            try:
                Path(mistral_input_path).unlink(missing_ok=True)
            except Exception:
                pass


# ── Gemini Logic ──────────────────────────────────────────────────────────

async def extract_with_gemini(image_path: str) -> dict:
    from database import get_settings
    settings = get_settings()

    if not settings.gemini_api_key:
        return {"source": "gemini", "error": "GEMINI_API_KEY not set", "high_confidence": False}

    gemini_input_path: Optional[str] = None
    cleanup_temp = False
    try:
        # Handle PDF: convert first page to image
        src_path = Path(image_path)
        if src_path.suffix.lower() == ".pdf":
            converted = _convert_pdf_first_page_to_image(image_path)
            if not converted:
                log.warning("Gemini: PDF->image conversion failed, attempting to read PDF bytes directly")
                gemini_input_path = image_path
            else:
                gemini_input_path = converted
                cleanup_temp = True
        else:
            gemini_input_path = image_path

        suffix = Path(gemini_input_path).suffix.lower().lstrip(".")
        mime = {"jpg": "image/jpeg", "jpeg": "image/jpeg", "png": "image/png",
                "webp": "image/webp", "pdf": "application/pdf"}.get(suffix, "image/jpeg")
        img_bytes = Path(gemini_input_path).read_bytes()

        text = _gemini_generate_text(
            api_key=settings.gemini_api_key,
            model="gemini-2.5-flash",
            prompt_text=GEMINI_PROMPT,
            blob=img_bytes,
            mime=mime,
        )
        text = re.sub(r"^```json\s*", "", text)
        text = re.sub(r"\s*```$", "", text)
        result = _load_json_object_loose(text)
        result = _normalize_model_bill_payload(result)
        result["source"] = "gemini"
        fields_found = _count_bill_fields(result)
        result["fields_found"] = fields_found
        conf = float(result.get("confidence") or 0)
        result["high_confidence"] = (fields_found >= 5) or (fields_found >= 4 and conf >= 0.65)

        if "quantity_qtl" not in result and "quantity_mt" in result:
            try:
                result["quantity_qtl"] = round(float(result["quantity_mt"]) * 10, 3)
            except Exception:
                pass
        log.info("Gemini OCR succeeded: fields=%s high_conf=%s", fields_found, result["high_confidence"])
        return result

    except Exception as e:
        log.error(f"Gemini OCR error: {e}")
        return {"source": "gemini", "error": str(e), "high_confidence": False}
    finally:
        if cleanup_temp and gemini_input_path:
            try:
                Path(gemini_input_path).unlink(missing_ok=True)
            except Exception:
                pass


# ── Azure Logic ──────────────────────────────────────────────────────────

async def extract_with_azure(image_path: str) -> dict:
    load_dotenv()
    endpoint = os.getenv("AZURE_OCR_ENDPOINT")
    key = os.getenv("AZURE_OCR_KEY")

    if not endpoint or not key:
        return {"source": "azure", "error": "Azure credentials missing", "high_confidence": False}

    try:
        from azure.core.credentials import AzureKeyCredential
        from azure.ai.formrecognizer import DocumentAnalysisClient

        def _field_value(field):
            if field is None:
                return None
            val = getattr(field, "value", None)
            if val is not None:
                return val
            return getattr(field, "content", None)

        def _num_from_any(v):
            if v is None:
                return None
            if hasattr(v, "amount") and getattr(v, "amount", None) is not None:
                try:
                    return float(v.amount)
                except Exception:
                    return None
            if hasattr(v, "value"):
                return _num_from_any(getattr(v, "value", None))
            if isinstance(v, (int, float)):
                return float(v)
            s = str(v).strip().replace(",", "")
            m = re.search(r"-?\d+(?:\.\d+)?", s)
            if m:
                try:
                    return float(m.group(0))
                except Exception:
                    return None
            return None

        def _str_from_any(v):
            if v is None:
                return None
            if hasattr(v, "value") and getattr(v, "value", None) not in (None, ""):
                return str(v.value).strip()
            if hasattr(v, "content") and getattr(v, "content", None) not in (None, ""):
                return str(v.content).strip()
            s = str(v).strip()
            return s if s else None

        def _normalize_material(name: Optional[str]) -> Optional[str]:
            if not name:
                return None
            s = str(name).lower()
            if "ddgs" in s:
                return "Rice DDGS"
            if "maize" in s or "makka" in s:
                return "Maize"
            if "dorb" in s:
                return "Dorb"
            if "domc" in s or "doms" in s:
                return "Domc"
            return None

        def _qty_to_qtl(qty: Optional[float], unit_text: str) -> Optional[float]:
            if qty is None:
                return None
            u = (unit_text or "").lower()
            if "kg" in u:
                return round(qty / 100, 3)
            if "mt" in u or "ton" in u:
                return round(qty * 10, 3)
            return float(qty)

        def _rate_to_qtl(rate: Optional[float], unit_text: str) -> Optional[float]:
            if rate is None:
                return None
            u = (unit_text or "").lower()
            if "kg" in u:
                return round(rate * 100, 3)
            if "mt" in u or "ton" in u:
                return round(rate / 10, 3)
            return float(rate)

        client = DocumentAnalysisClient(endpoint=endpoint, credential=AzureKeyCredential(key))
        with open(image_path, "rb") as f:
            poller = client.begin_analyze_document("prebuilt-invoice", document=f)
            result = poller.result()

        extracted = _parse_bill_text(result.content)
        if result.documents:
            inv = result.documents[0].fields
            if not extracted.get("vendor_name") and inv.get("VendorName"):
                extracted["vendor_name"] = inv.get("VendorName").value
            if not extracted.get("total_amount") and inv.get("InvoiceTotal"):
                extracted["total_amount"] = inv.get("InvoiceTotal").value.amount
            if not extracted.get("bill_date") and inv.get("InvoiceDate"):
                extracted["bill_date"] = str(inv.get("InvoiceDate").value)
            if not extracted.get("bill_number") and inv.get("InvoiceId"):
                extracted["bill_number"] = inv.get("InvoiceId").value

            # Pull qty/rate/material from invoice line items when regex text parsing misses them.
            items = _field_value(inv.get("Items"))
            if isinstance(items, list):
                for item in items:
                    item_map = _field_value(item)
                    if not isinstance(item_map, dict):
                        continue

                    unit_text = _str_from_any(_field_value(item_map.get("Unit"))) or _str_from_any(_field_value(item_map.get("UnitOfMeasure"))) or ""
                    qty_raw = _num_from_any(_field_value(item_map.get("Quantity")))
                    rate_raw = _num_from_any(_field_value(item_map.get("UnitPrice")))
                    amt_raw = _num_from_any(_field_value(item_map.get("Amount")))
                    desc = _str_from_any(_field_value(item_map.get("Description")))

                    qty_qtl = _qty_to_qtl(qty_raw, unit_text)
                    rate_qtl = _rate_to_qtl(rate_raw, unit_text)

                    if extracted.get("quantity_qtl") is None and qty_qtl is not None:
                        extracted["quantity_qtl"] = qty_qtl
                    if extracted.get("rate_per_qtl") is None and rate_qtl is not None:
                        extracted["rate_per_qtl"] = rate_qtl
                    if extracted.get("total_amount") is None and amt_raw is not None:
                        extracted["total_amount"] = amt_raw
                    if extracted.get("material_type") is None:
                        extracted["material_type"] = _normalize_material(desc)

                    if extracted.get("quantity_qtl") is not None and extracted.get("rate_per_qtl") is not None:
                        break

        if extracted.get("fields_found", 0) < 4:
            with open(image_path, "rb") as f:
                poller_read = client.begin_analyze_document("prebuilt-read", document=f)
                result_read = poller_read.result()
            
            read_ext = _parse_bill_text(result_read.content)
            for k, v in read_ext.items():
                if extracted.get(k) is None: extracted[k] = v
            extracted["raw_text"] = result_read.content
        else:
            extracted["raw_text"] = result.content

        key_fields = ["vehicle_number", "quantity_qtl", "rate_per_qtl", "total_amount", "material_type"]
        extracted["fields_found"] = sum(1 for f in key_fields if extracted.get(f) is not None)
        extracted.update({"source": "azure", "high_confidence": extracted.get("fields_found", 0) >= 4})
        return extracted
    except Exception as e:
        log.error(f"Azure OCR error: {e}")
        return {"source": "azure", "error": str(e), "high_confidence": False}


# ── Main router ───────────────────────────────────────────────────────────

async def extract_bill(image_path: str) -> dict:
    """
    Priority: 1. Paddle (Local), 2. Groq, 3. Mistral, 4. SambaNova, 5. Gemini, 6. GitHub Models, 7. Azure
    """
    log.info(f"Extracting bill: {image_path}")

    paddle_res = {}
    mistral_res: dict = {}
    sambanova_res: dict = {}
    gemini_res: dict = {}
    github_res: dict = {}
    azure_res: dict = {}

    # 1. Paddle
    # paddle_res = extract_with_paddle(image_path)
    # if paddle_res.get("high_confidence"):
    #     log.info("Paddle succeeded")
    #     return _validate_bill(paddle_res)

    # 2. Groq
    log.info("Paddle low confidence — escalating to Groq")
    groq_res = await extract_with_groq(image_path)
    if groq_res.get("high_confidence"):
        log.info("Groq succeeded")
        return _merge_and_validate(groq_res, paddle_res)
    
    # 3. Mistral
    log.info("Groq low confidence/failed — escalating to Mistral")
    mistral_res = await extract_with_mistral(image_path)
    if not isinstance(mistral_res, dict):
        mistral_res = {}
    log.info(
        "Mistral result: model='%s' fields=%s high_conf=%s vendor='%s' vehicle='%s' bill='%s'",
        mistral_res.get("mistral_model"),
        mistral_res.get("fields_found"),
        mistral_res.get("high_confidence"),
        mistral_res.get("vendor_name") or mistral_res.get("broker_name"),
        mistral_res.get("vehicle_number"),
        mistral_res.get("bill_number"),
    )
    if mistral_res.get("high_confidence"):
        return _merge_and_validate(mistral_res, paddle_res)

    # 4. SambaNova
    log.info("Mistral low confidence/failed — escalating to SambaNova")
    sambanova_res = await extract_with_sambanova(image_path)
    if not isinstance(sambanova_res, dict):
        sambanova_res = {}
    log.info(
        "SambaNova result: model='%s' fields=%s high_conf=%s vendor='%s' vehicle='%s' bill='%s'",
        sambanova_res.get("sambanova_model"),
        sambanova_res.get("fields_found"),
        sambanova_res.get("high_confidence"),
        sambanova_res.get("vendor_name") or sambanova_res.get("broker_name"),
        sambanova_res.get("vehicle_number"),
        sambanova_res.get("bill_number"),
    )
    if sambanova_res.get("high_confidence"):
        return _merge_and_validate(sambanova_res, paddle_res)

    # 5. Gemini
    log.info("SambaNova low confidence/failed — escalating to Gemini")
    gemini_res = await extract_with_gemini(image_path)
    if not isinstance(gemini_res, dict):
        gemini_res = {}
    if gemini_res.get("fields_found") is None:
        try:
            gemini_res["fields_found"] = _count_bill_fields(gemini_res)
        except Exception:
            pass
    log.info(
        "Gemini result: fields=%s high_conf=%s vendor='%s' vehicle='%s' bill='%s'",
        gemini_res.get("fields_found"),
        gemini_res.get("high_confidence"),
        gemini_res.get("vendor_name") or gemini_res.get("broker_name"),
        gemini_res.get("vehicle_number"),
        gemini_res.get("bill_number"),
    )
    if gemini_res.get("high_confidence"):
        return _merge_and_validate(gemini_res, paddle_res)

    # 6. GitHub
    log.info("Gemini low confidence/failed — escalating to GitHub Models (GPT-4o-mini)")
    github_res = await extract_with_github_models(image_path)
    if not isinstance(github_res, dict):
        github_res = {}
    log.info(
        "GitHub Models result: model='%s' fields=%s high_conf=%s vendor='%s' vehicle='%s' bill='%s'",
        github_res.get("github_model"),
        github_res.get("fields_found"),
        github_res.get("high_confidence"),
        github_res.get("vendor_name") or github_res.get("broker_name"),
        github_res.get("vehicle_number"),
        github_res.get("bill_number"),
    )
    if github_res.get("high_confidence"):
        return _merge_and_validate(github_res, paddle_res)

    # 7. Azure
    log.info("GitHub low confidence/failed — escalating to Azure")
    azure_res = await extract_with_azure(image_path)
    if not isinstance(azure_res, dict):
        azure_res = {}
    if azure_res.get("high_confidence"):
        return _merge_and_validate(azure_res, paddle_res)
    
    # Keep the best low-signal cloud result if none reached confidence.
    cloud_candidates = [mistral_res, sambanova_res, gemini_res, github_res, azure_res]
    cloud_best = max(
        cloud_candidates,
        key=lambda r: int((r or {}).get("fields_found") or _count_bill_fields(r or {})),
    )
    return _merge_and_validate(cloud_best if isinstance(cloud_best, dict) else {}, paddle_res)


async def extract_with_github_models(image_path: str) -> dict:
    try:
        from openai import OpenAI
    except Exception as e:
        log.error("GitHub Models unavailable: cannot import openai on %s: %s", sys.executable, e)
        return {
            "source": "github_models",
            "error": f"openai import failed: {e}",
            "high_confidence": False,
        }

    load_dotenv()
    enabled = (os.getenv("GITHUB_MODELS_ENABLED", "true").strip().lower() in {"1", "true", "yes", "on"})
    if not enabled:
        return {
            "source": "github_models",
            "error": "GITHUB_MODELS_ENABLED is false",
            "high_confidence": False,
        }

    api_key = (os.getenv("GITHUB_TOKEN") or os.getenv("GITHUB_MODELS_API_KEY") or "").strip()
    if not api_key:
        return {
            "source": "github_models",
            "error": "GITHUB token not set (GITHUB_TOKEN or GITHUB_MODELS_API_KEY)",
            "high_confidence": False,
        }

    configured_model = (os.getenv("GITHUB_MODEL") or "openai/gpt-4o-mini").strip()
    configured_base = (os.getenv("GITHUB_MODELS_BASE_URL") or "").strip().rstrip("/")

    model_candidates = [configured_model, "openai/gpt-4o-mini", "gpt-4o-mini"]
    model_candidates = [m for i, m in enumerate(model_candidates) if m and m not in model_candidates[:i]]

    base_candidates = [
        configured_base,
        "https://models.github.ai/inference",
        "https://models.inference.ai.azure.com",
    ]
    base_candidates = [b for i, b in enumerate(base_candidates) if b and b not in base_candidates[:i]]

    github_input_path: Optional[str] = None
    cleanup_temp = False
    compact_temp_path: Optional[str] = None
    try:
        src_path = Path(image_path)
        if src_path.suffix.lower() == ".pdf":
            converted = _convert_pdf_first_page_to_image(image_path)
            if not converted:
                return {
                    "source": "github_models",
                    "error": "PDF->image conversion failed for GitHub Models",
                    "high_confidence": False,
                }
            github_input_path = converted
            cleanup_temp = True
        else:
            github_input_path = image_path

        data_url, _mime, compact_temp_path = _build_data_url_for_llm(github_input_path)

        best_result = None
        best_fields = -1
        last_err = None

        for base_url in base_candidates:
            client = OpenAI(base_url=base_url, api_key=api_key, max_retries=0)
            for model_id in model_candidates:
                try:
                    request_kwargs = {
                        "model": model_id,
                        "messages": [
                            {
                                "role": "user",
                                "content": [
                                    {"type": "text", "text": f"{GEMINI_PROMPT}\\n\\nReturn only one valid JSON object with no extra text."},
                                    {
                                        "type": "image_url",
                                        "image_url": {"url": data_url}
                                    },
                                ],
                            }
                        ],
                        "temperature": 0.1,
                        "max_tokens": 1600,
                        "response_format": {"type": "json_object"},
                    }

                    try:
                        await _sleep_before_llm_request_async()
                        response = client.chat.completions.create(**request_kwargs)
                    except Exception as e:
                        em = str(e).lower()
                        if "json mode is not enabled" in em or "response_format" in em:
                            request_kwargs.pop("response_format", None)
                            await _sleep_before_llm_request_async()
                            response = client.chat.completions.create(**request_kwargs)
                        else:
                            raise

                    raw = (response.choices[0].message.content or "").strip()
                    result = _load_json_object_loose(raw)
                    result = _normalize_model_bill_payload(result)
                    result["source"] = "github_models"
                    result["github_model"] = model_id
                    result["github_base_url"] = base_url

                    if "quantity_qtl" not in result and "quantity_mt" in result:
                        try:
                            result["quantity_qtl"] = round(float(result["quantity_mt"]) * 10, 3)
                        except Exception:
                            pass

                    fields_found = _count_bill_fields(result)
                    result["fields_found"] = fields_found
                    conf = float(result.get("confidence") or 0)
                    result["high_confidence"] = (fields_found >= 5) or (fields_found >= 4 and conf >= 0.65)

                    if fields_found > best_fields:
                        best_fields = fields_found
                        best_result = result

                    if result["high_confidence"]:
                        return result
                except Exception as e:
                    last_err = e
                    log.warning("GitHub Models failed (%s @ %s): %s", model_id, base_url, e)
                    continue

        if isinstance(best_result, dict):
            return best_result

        return {
            "source": "github_models",
            "error": f"All GitHub Models candidates failed: {last_err}",
            "high_confidence": False,
        }
    except Exception as e:
        return {
            "source": "github_models",
            "error": str(e),
            "high_confidence": False,
        }
    finally:
        if compact_temp_path:
            try:
                Path(compact_temp_path).unlink(missing_ok=True)
            except Exception:
                pass
        if cleanup_temp and github_input_path:
            try:
                Path(github_input_path).unlink(missing_ok=True)
            except Exception:
                pass


async def extract_with_sambanova(image_path: str) -> dict:
    try:
        from openai import OpenAI
    except Exception as e:
        log.error("SambaNova unavailable: cannot import openai on %s: %s", sys.executable, e)
        return {
            "source": "sambanova",
            "error": f"openai import failed: {e}",
            "high_confidence": False,
        }

    load_dotenv()
    enabled = (os.getenv("SAMBANOVA_ENABLED", "true").strip().lower() in {"1", "true", "yes", "on"})
    if not enabled:
        return {
            "source": "sambanova",
            "error": "SAMBANOVA_ENABLED is false",
            "high_confidence": False,
        }

    api_key = os.getenv("SAMBANOVA_API_KEY")
    if not api_key:
        return {
            "source": "sambanova",
            "error": "SAMBANOVA_API_KEY not set",
            "high_confidence": False,
        }

    model_id = (os.getenv("SAMBANOVA_MODEL") or "Llama-4-Maverick-17B-128E-Instruct").strip()
    sambanova_input_path: Optional[str] = None
    cleanup_temp = False
    compact_temp_path: Optional[str] = None

    try:
        src_path = Path(image_path)
        if src_path.suffix.lower() == ".pdf":
            converted = _convert_pdf_first_page_to_image(image_path)
            if not converted:
                return {
                    "source": "sambanova",
                    "error": "PDF->image conversion failed for SambaNova",
                    "high_confidence": False,
                }
            sambanova_input_path = converted
            cleanup_temp = True
        else:
            sambanova_input_path = image_path

        data_url, _mime, compact_temp_path = _build_data_url_for_llm(sambanova_input_path)
        client = OpenAI(
            base_url="https://api.sambanova.ai/v1",
            api_key=api_key,
            max_retries=0,
        )

        await _sleep_before_llm_request_async()
        response = client.chat.completions.create(
            model=model_id,
            messages=[
                {
                    "role": "user",
                    "content": [
                        {"type": "text", "text": f"{GEMINI_PROMPT}\n\nReturn only one valid JSON object with no extra text."},
                        {
                            "type": "image_url",
                            "image_url": {"url": data_url}
                        },
                    ],
                }
            ],
            temperature=0.1,
            max_tokens=1600,
        )

        raw = (response.choices[0].message.content or "").strip()
        result = _load_json_object_loose(raw)
        result = _normalize_model_bill_payload(result)
        result["source"] = "sambanova"
        result["sambanova_model"] = model_id

        if "quantity_qtl" not in result and "quantity_mt" in result:
            try:
                result["quantity_qtl"] = round(float(result["quantity_mt"]) * 10, 3)
            except Exception:
                pass

        fields_found = _count_bill_fields(result)
        result["fields_found"] = fields_found
        conf = float(result.get("confidence") or 0)
        result["high_confidence"] = (fields_found >= 5) or (fields_found >= 4 and conf >= 0.65)
        return result
    except Exception as e:
        err_text = str(e)
        err_l = err_text.lower()
        payload = {
            "source": "sambanova",
            "error": err_text,
            "high_confidence": False,
            "sambanova_model": model_id,
        }
        if "rate limit" in err_l or "429" in err_l:
            payload["rate_limited"] = True
        return payload
    finally:
        if compact_temp_path:
            try:
                Path(compact_temp_path).unlink(missing_ok=True)
            except Exception:
                pass
        if cleanup_temp and sambanova_input_path:
            try:
                Path(sambanova_input_path).unlink(missing_ok=True)
            except Exception:
                pass


def _merge_and_validate(primary: dict, fallback: dict) -> dict:
    merged = dict(primary)
    fields = ["vehicle_number", "quantity_qtl", "rate_per_qtl", "total_amount",
              "material_type", "bill_date", "bill_number", "vendor_name"]
    for f in fields:
        if merged.get(f) is None and fallback.get(f) is not None:
            merged[f] = fallback[f]
    merged["raw_text"] = merged.get("raw_text", fallback.get("raw_text", ""))
    return _validate_bill(merged)


def _as_num(v) -> Optional[float]:
    if v is None:
        return None
    try:
        return float(str(v).replace(",", "").strip())
    except Exception:
        return None


def _unit_kind(v: Optional[str]) -> Optional[str]:
    s = str(v or "").strip().lower()
    if not s:
        return None
    if "kg" in s:
        return "kg"
    if "mt" in s or "ton" in s:
        return "mt"
    if "qtl" in s or "quintal" in s:
        return "qtl"
    return None


def _apply_qty_rate_unit(qty: Optional[float], rate: Optional[float], unit_kind: Optional[str]) -> tuple[Optional[float], Optional[float]]:
    if unit_kind == "kg":
        qty = (qty / 100.0) if qty is not None else None
        rate = (rate * 100.0) if rate is not None else None
    elif unit_kind == "mt":
        qty = (qty * 10.0) if qty is not None else None
        rate = (rate / 10.0) if rate is not None else None
    return qty, rate


def _reconcile_bill_qty_rate_units(data: dict) -> dict:
    if not isinstance(data, dict):
        return {}

    out = dict(data)

    qty = _as_num(out.get("quantity_qtl"))
    rate = _as_num(out.get("rate_per_qtl"))
    total = _as_num(out.get("total_amount"))

    qty_mt = _as_num(out.get("quantity_mt"))
    rate_mt = _as_num(out.get("rate_per_mt"))

    if qty is None and qty_mt is not None:
        qty = qty_mt * 10.0
    if rate is None and rate_mt is not None:
        rate = rate_mt / 10.0

    unit_hint = out.get("quantity_unit") or out.get("rate_unit") or out.get("unit")
    hinted_kind = _unit_kind(unit_hint)
    if hinted_kind in {"kg", "mt"}:
        qty, rate = _apply_qty_rate_unit(qty, rate, hinted_kind)

    # If total is present, pick the most consistent scale variant for qty/rate.
    if qty is not None and rate is not None and total is not None and total > 0:
        def _err(q: float, r: float) -> float:
            computed = q * r
            return abs(computed - total) / max(total, 1.0)

        candidates = [
            ("as_is", qty, rate),
            ("rate_div_10", qty, rate / 10.0),
            ("qty_mul_10", qty * 10.0, rate),
            ("qty_div_10", qty / 10.0, rate),
            ("both_mt", qty * 10.0, rate / 10.0),
            ("both_kg", qty / 100.0, rate * 100.0),
            ("rate_mul_10", qty, rate * 10.0),
        ]

        base_err = _err(qty, rate)
        best_name, best_q, best_r = min(candidates, key=lambda c: _err(c[1], c[2]))
        best_err = _err(best_q, best_r)

        # Only auto-adjust when clearly better to avoid overfitting noise.
        if best_name != "as_is" and best_err <= 0.25 and (base_err - best_err) >= 0.10:
            qty, rate = best_q, best_r

    out["quantity_qtl"] = round(qty, 3) if qty is not None else None
    out["rate_per_qtl"] = round(rate, 3) if rate is not None else None
    return out


def _validate_bill(data: dict) -> dict:
    data = _reconcile_bill_qty_rate_units(data)
    errors = []
    qty = data.get("quantity_qtl")
    rate = data.get("rate_per_qtl")
    total = data.get("total_amount")
    
    if qty and rate and total:
        computed = float(qty) * float(rate)
        data["validation_amount"] = abs(computed - float(total)) / float(total) <= 0.05
    else:
        data["validation_amount"] = False

    veh = _normalize_vehicle_number(data.get("vehicle_number"))
    data["vehicle_number"] = veh
    data["bill_number"] = _normalize_bill_number(data.get("bill_number"))
    data["validation_vehicle"] = bool(re.match(r'^[A-Z]{2}\d{2}[A-Z]{1,2}\d{4}$', veh)) if veh else False
    
    mat = data.get("material_type")
    data["validation_material"] = mat in ["Maize", "Dorb", "Domc", "Rice DDGS"]

    data["needs_review"] = not (data["validation_amount"] and data["validation_vehicle"] and data["validation_material"])
    return data


def _normalize_model_bill_payload(result: dict) -> dict:
    if not isinstance(result, dict):
        return {}

    out = dict(result)
    aliases = {
        "vendor_name": ["seller_name", "vendor", "supplier_name", "company_name", "party_name"],
        "vehicle_number": ["truck_number", "truck_no", "vehicle_no", "vehicle"],
        "material_type": ["material", "item_name", "item"],
        "quantity_qtl": ["qty_qtl", "quantity", "qty", "net_qty_qtl"],
        "rate_per_qtl": ["rate", "unit_rate", "rate_per_unit"],
        "total_amount": ["amount", "grand_total", "total", "invoice_total"],
        "bill_number": ["invoice_number", "invoice_no", "bill_no", "invoice"],
        "bill_date": ["invoice_date", "date"],
    }

    for target, alias_keys in aliases.items():
        if out.get(target) in (None, "", "null", "None"):
            for key in alias_keys:
                val = out.get(key)
                if val not in (None, "", "null", "None"):
                    out[target] = val
                    break

    for key in ["quantity_qtl", "rate_per_qtl", "total_amount", "confidence"]:
        val = out.get(key)
        if isinstance(val, str):
            s = val.strip().replace(",", "")
            try:
                out[key] = float(s)
            except Exception:
                pass

    for key in ["quantity_mt", "rate_per_mt"]:
        val = out.get(key)
        if isinstance(val, str):
            s = val.strip().replace(",", "")
            try:
                out[key] = float(s)
            except Exception:
                pass

    if out.get("bill_date"):
        bd = str(out.get("bill_date")).strip()
        m = re.search(r"(\d{1,2})[/\-.](\d{1,2})[/\-.](\d{2,4})", bd)
        if m:
            d, mo, y = m.group(1), m.group(2), m.group(3)
            if len(y) == 2:
                y = f"20{y}"
            out["bill_date"] = f"{y}-{mo.zfill(2)}-{d.zfill(2)}"

    out["vehicle_number"] = _normalize_vehicle_number(out.get("vehicle_number"))
    out["bill_number"] = _normalize_bill_number(out.get("bill_number"))

    return _reconcile_bill_qty_rate_units(out)


def _count_bill_fields(result: dict) -> int:
    keys = [
        "vendor_name",
        "vehicle_number",
        "material_type",
        "quantity_qtl",
        "rate_per_qtl",
        "total_amount",
        "bill_number",
    ]
    return sum(1 for k in keys if result.get(k) not in (None, "", "null", "None"))


def _load_json_object_loose(raw: str) -> dict:
    text = (raw or "").strip()
    if not text:
        raise ValueError("empty model content")

    # Remove markdown fences when providers return ```json ... ```.
    text = re.sub(r"^```(?:json)?\s*", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\s*```$", "", text).strip()

    try:
        return json.loads(text)
    except Exception:
        start = text.find("{")
        end = text.rfind("}")
        if start != -1 and end != -1 and end > start:
            return json.loads(text[start:end + 1])
        raise


def _gemini_generate_text(api_key: str, model: str, prompt_text: str, blob: bytes, mime: str) -> str:
    """Generate Gemini text with either google.genai (new SDK) or google-generativeai (legacy SDK)."""
    err_new = None
    try:
        from google import genai as google_genai
        from google.genai import types

        client = google_genai.Client(api_key=api_key)
        _sleep_before_llm_request_sync()
        response = client.models.generate_content(
            model=model,
            contents=[prompt_text, types.Part.from_bytes(data=blob, mime_type=mime)],
        )
        return (response.text or "").strip()
    except Exception as e:
        err_new = e

    try:
        import google.generativeai as legacy_genai

        legacy_genai.configure(api_key=api_key)
        model_obj = legacy_genai.GenerativeModel(model)
        _sleep_before_llm_request_sync()
        response = model_obj.generate_content([
            prompt_text,
            {"mime_type": mime, "data": blob},
        ])

        text = (getattr(response, "text", None) or "").strip()
        if text:
            return text

        parts_text: List[str] = []
        for cand in (getattr(response, "candidates", None) or []):
            content = getattr(cand, "content", None)
            for part in (getattr(content, "parts", None) or []):
                ptxt = getattr(part, "text", None)
                if ptxt:
                    parts_text.append(str(ptxt))
        return "".join(parts_text).strip()
    except Exception as err_legacy:
        raise RuntimeError(f"Gemini SDK unavailable: google.genai={err_new}; google.generativeai={err_legacy}")


PLANT_UNLOADING_PROMPT = """
You are reading an Indian plant unloading register page (raw material control register).
The sheet may be handwritten or printed and may be rotated.
This is for a cattle feed plant workflow in Rajasthan, India.

Extract ALL fields below. Return ONLY a valid JSON object - no markdown, no explanation, no preamble.

{
    "plant_name": "Plant/location from header (e.g. Kaladera, Bikaner, Jodhpur, Ajmer, Nadbai, Pali, Lambiyan, Bhilwara) or null. It may appear as CATTLE FEED PLANT {location} or CFP {location}",
    "item_name": "Material/item name (e.g. Maize, Dorb, Domc, Rice DDGS) or null",
    "party_name": "Supplier/party/vendor name from register header (e.g. M/S SALASAR TRADING COMPANY) or null",
    "rm_number": "RM register/reference number (e.g. RM-2026-184) or null",
    "po_number": "PO/Purchase Order number (e.g. PO-5562) or null",
    "sheet_date": "Main sheet date in YYYY-MM-DD format (e.g. 2026-03-20) or null",
    "rows": [
        {
            "sno": "Serial number as visible (e.g. 1, 2, 3) or null",
            "ws_no": "WS/weighment slip number (e.g. WS-1245) or null",
            "date": "Row date in YYYY-MM-DD format (e.g. 2026-03-20) or null",
            "truck_number": "Truck/vehicle registration (e.g. RJ20GB6238) - NEVER guess",
            "no_of_bags": "Number of bags as numeric value (e.g. 420) or null",
            "received_qty_mt": "Received quantity in QUINTAL (Qtl) as number (legacy key name with _mt) (e.g. 427.000) or null",
            "net_qty_mt": "Net quantity in QUINTAL (Qtl) as number (legacy key name with _mt) (e.g. 419.800) or null",
            "total_qty_mt": "Running/total quantity in QUINTAL (Qtl) as number (legacy key name with _mt) (e.g. 4205.600) or null"
        }
    ],
    "confidence": "0.0 to 1.0 overall extraction confidence"
}

Critical rules:
- Extract one JSON row per unloaded truck entry line.
- Business unit is QUINTAL (Qtl) everywhere.
- Keep output quantity values in Qtl for all rows.
- If source shows MT, convert to Qtl by multiplying by 10.
- If source shows KG, convert to Qtl by dividing by 100.
- Handwritten compact weight notation may omit decimal point in quantity columns. For quantity fields only, interpret values like 34822 as 348.22 Qtl, 30768 as 307.68 Qtl, 65590 as 655.90 Qtl.
- If a row has only one clear weight value in the right-side quantity column, map that value to net_qty_mt for that row.
- If both Received and Net columns are present, net_qty_mt must come from the Net Quantity column, not from Received or Total.
- Keep legacy JSON keys (received_qty_mt/net_qty_mt/total_qty_mt) unchanged for API compatibility, but values must be in Qtl.
- Do not invent truck numbers, row dates, WS numbers, or quantities.
- Plant name normalization: if text is like "CATTLE FEED PLANT KALADERA" or "CFP KALADERA", extract location as "Kaladera".
- Accept plant/location variants from this set when clearly present: Kaladera, Bikaner, Jodhpur, Ajmer, Nadbai, Pali, Lambiyan, Bhilwara.
- If a field is unclear or missing, set it to null.
- Keep row order exactly as the sheet order.
- If table borders are faint, infer columns from alignment but do not hallucinate values.
- If there are no valid rows, return rows as an empty list.
"""


PLANT_UNLOADING_FALLBACK_PROMPT = """
Extract ONLY unloading table rows from this document.
Return ONLY valid JSON in this exact structure:
{"rows":[{"sno":null,"ws_no":null,"date":null,"truck_number":null,"no_of_bags":null,"received_qty_mt":null,"net_qty_mt":null,"total_qty_mt":null}],"confidence":0.0}

Field expectations:
- truck_number example: RJ20GB6238 (never guess)
- date format: YYYY-MM-DD when derivable, otherwise null
- quantities are numeric QUINTAL (Qtl) values, not strings
- no_of_bags is numeric count

Rules:
- Keep one object per table row.
- If source table values are in MT, convert to Qtl (x10) before returning.
- If source table values are in KG, convert to Qtl (/100) before returning.
- For quantity fields, when handwritten compact values omit decimal (e.g., 34822, 30768, 65590), decode as two-decimal Qtl values (348.22, 307.68, 655.90).
- If only one per-row weight is visible, place it in net_qty_mt.
- If both Received and Net values are visible for a row, always use Net for net_qty_mt.
- Keep legacy keys (received_qty_mt/net_qty_mt/total_qty_mt) but store Qtl values in them.
- Do not return prose.
- Do not include keys outside the defined schema.
- If no usable row exists, return {"rows":[],"confidence":0.0}.
"""


PLANT_UNLOADING_TEXT_PARSE_PROMPT = """
You are given OCR text extracted from a plant unloading register (raw material control register).
Parse it and return ONLY valid JSON with this schema:
{
    "plant_name": "header plant/location (e.g. Kaladera, Bikaner, Jodhpur, Ajmer, Nadbai, Pali, Lambiyan, Bhilwara) or null; may be written as CATTLE FEED PLANT {location} or CFP {location}",
    "item_name": "material/item (e.g. Maize, Dorb, Domc, Rice DDGS) or null",
    "party_name": "supplier/party name (e.g. M/S SALASAR TRADING COMPANY) or null",
    "rm_number": "RM reference number (e.g. RM-2026-184) or null",
    "po_number": "PO number (e.g. PO-5562) or null",
    "sheet_date": "YYYY-MM-DD or null",
    "rows": [
        {
            "sno": "string/number serial or null",
            "ws_no": "weighment slip number or null",
            "date": "YYYY-MM-DD or null",
            "truck_number": "registration like RJ20GB6238 or null",
            "no_of_bags": number or null,
            "received_qty_mt": "QUINTAL (Qtl) value in legacy _mt key" or null,
            "net_qty_mt": "QUINTAL (Qtl) value in legacy _mt key" or null,
            "total_qty_mt": "QUINTAL (Qtl) value in legacy _mt key" or null
        }
    ],
    "confidence": 0.0
}

Rules:
- Do not hallucinate values not present in OCR text.
- For plant_name, normalize "CATTLE FEED PLANT {location}" and "CFP {location}" to just the location name.
- Use Quintal (Qtl) as output unit for all quantities.
- If OCR text shows MT, convert to Qtl (x10). If it shows KG, convert to Qtl (/100).
- Quantity columns may use compact handwritten format without decimal, such as 34822 or 30768. In quantity fields, decode these as 348.22 and 307.68 Qtl respectively.
- If a row has a single recognized quantity value and no explicit received/net label, assign it to net_qty_mt.
- If both Received Quantity and Net Quantity are present, set net_qty_mt from Net Quantity only.
- Keep legacy _mt key names for compatibility, but place Qtl values in those keys.
- Preserve table row order.
- If uncertain, use null values but keep valid rows.
- Return only JSON, no explanations.
"""


def _normalize_date_text(val: Optional[str]) -> Optional[str]:
    if val is None:
        return None
    s = str(val).strip()
    if not s or s.lower() in ("none", "null", "n/a", "na", "-"):
        return None
    from datetime import datetime as _dt
    fmts = ["%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%d.%m.%Y", "%d.%m.%y", "%d/%m/%y"]
    for fmt in fmts:
        try:
            return _dt.strptime(s, fmt).date().isoformat()
        except Exception:
            continue
    return s


def _convert_pdf_first_page_to_image(file_path: str) -> Optional[str]:
    """Convert first PDF page to PNG using available local render engines."""
    src = Path(file_path)
    if src.suffix.lower() != ".pdf":
        return None

    # Use a short temp filename in the same directory to avoid Windows long-path issues.
    try:
        fd, tmp_name = tempfile.mkstemp(prefix="ocrp1_", suffix=".png", dir=str(src.parent))
        os.close(fd)
        out = Path(tmp_name)
        try:
            out.unlink(missing_ok=True)
        except Exception:
            pass
    except Exception:
        out = src.with_suffix(".page1.png")
        if out.exists():
            try:
                out.unlink()
            except Exception:
                pass

    errors: List[str] = []

    # Engine 1: PyMuPDF (fitz)
    try:
        import fitz  # PyMuPDF
        doc = fitz.open(file_path)
        try:
            if doc.page_count < 1:
                return None
            page = doc.load_page(0)
            pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))
            pix.save(str(out))
            if out.exists():
                return str(out)
        finally:
            doc.close()
    except Exception as e:
        errors.append(f"pymupdf: {e}")

    # Engine 2: pypdfium2 (often works where fitz wheels are unavailable)
    try:
        import pypdfium2 as pdfium
        pdf = pdfium.PdfDocument(file_path)
        try:
            if len(pdf) < 1:
                return None
            page = pdf[0]
            bitmap = page.render(scale=2)
            image = bitmap.to_pil()
            image.save(str(out), format="PNG")
            if out.exists():
                return str(out)
        finally:
            pdf.close()
    except Exception as e:
        errors.append(f"pypdfium2: {e}")

    if errors:
        log.warning(f"PDF conversion failed with all engines: {' | '.join(errors)}")
    return None


def _extract_pdf_text_with_azure_read(file_path: str) -> Optional[str]:
    """Fallback text extraction for PDFs when local conversion libs are missing."""
    load_dotenv()
    endpoint = os.getenv("AZURE_OCR_ENDPOINT")
    key = os.getenv("AZURE_OCR_KEY")
    if not endpoint or not key:
        return None
    try:
        from azure.core.credentials import AzureKeyCredential
        from azure.ai.formrecognizer import DocumentAnalysisClient

        client = DocumentAnalysisClient(endpoint=endpoint, credential=AzureKeyCredential(key))
        with open(file_path, "rb") as f:
            poller = client.begin_analyze_document("prebuilt-read", document=f)
            result = poller.result()
        return result.content or None
    except Exception as e:
        log.warning(f"Azure PDF read fallback failed: {e}")
        return None


def _extract_unloading_from_text_with_groq(raw_text: str, api_key: str, source_label: str = "azure+groq") -> Optional[dict]:
    if not raw_text:
        return None
    try:
        from groq import Groq
        client = Groq(api_key=api_key)
        _sleep_before_llm_request_sync()
        response = client.chat.completions.create(
            messages=[
                {
                    "role": "user",
                    "content": f"{PLANT_UNLOADING_TEXT_PARSE_PROMPT}\n\nOCR_TEXT:\n{raw_text[:45000]}",
                }
            ],
            model="llama-3.3-70b-versatile",
            temperature=0.1,
            response_format={"type": "json_object"},
        )
        parsed = json.loads(response.choices[0].message.content)
        rows = _normalize_unloading_rows(parsed)
        return {
            "source": source_label,
            "plant_name": parsed.get("plant_name"),
            "item_name": parsed.get("item_name"),
            "party_name": parsed.get("party_name"),
            "rm_number": parsed.get("rm_number"),
            "po_number": parsed.get("po_number"),
            "sheet_date": _normalize_date_text(parsed.get("sheet_date")),
            "rows": rows,
            "confidence": parsed.get("confidence", 0.0),
            "high_confidence": len(rows) > 0,
            "raw_text": raw_text,
        }
    except Exception as e:
        log.warning(f"Text->rows fallback via Groq failed: {e}")
        return None


def _excel_cell_text_for_unloading_prompt(value) -> str:
    if value is None:
        return ""
    if hasattr(value, "isoformat") and not isinstance(value, str):
        try:
            return str(value.isoformat())
        except Exception:
            pass
    txt = str(value).strip()
    if not txt:
        return ""
    return re.sub(r"\s+", " ", txt)


def _build_excel_text_for_unloading_prompt(file_path: str, max_rows: int = 1500, max_cols: int = 24) -> Optional[str]:
    wb = None
    try:
        from openpyxl import load_workbook

        wb = load_workbook(file_path, read_only=True, data_only=True)
        if not wb.sheetnames:
            return None

        ws = wb[wb.sheetnames[0]]
        lines: List[str] = [f"SHEET: {ws.title}"]
        char_count = len(lines[0]) + 1
        blank_streak = 0
        max_scan_rows = min(int(ws.max_row or 0), int(max_rows))

        for ridx, row_vals in enumerate(ws.iter_rows(min_row=1, max_row=max_scan_rows, values_only=True), start=1):
            vals = [_excel_cell_text_for_unloading_prompt(v) for v in list(row_vals)[:max_cols]]
            while vals and not vals[-1]:
                vals.pop()

            if not vals:
                blank_streak += 1
                if blank_streak >= 80 and len(lines) > 20:
                    break
                continue

            blank_streak = 0
            line = f"R{ridx}: {' | '.join(vals)}"
            if len(line) > 800:
                line = line[:800]
            lines.append(line)
            char_count += len(line) + 1
            if char_count >= 65000:
                break

        text_payload = "\n".join(lines).strip()
        return text_payload or None
    except Exception as e:
        log.warning(f"Excel->text preparation for unloading prompt failed: {e}")
        return None
    finally:
        try:
            if wb is not None:
                wb.close()
        except Exception:
            pass


def _pick_first(d: dict, keys: List[str]):
    for k in keys:
        if k in d and d.get(k) not in (None, "", "null", "None"):
            return d.get(k)
    return None


def _to_float(v):
    if v is None:
        return None
    s = str(v).strip().replace(',', '')
    if not s or s.lower() in ("none", "null", "na", "n/a", "-"):
        return None
    try:
        return float(s)
    except Exception:
        m = re.search(r"-?\d+(?:\.\d+)?", s)
        if not m:
            return None
        try:
            return float(m.group(0))
        except Exception:
            return None


def _to_unloading_qty_qtl(v):
    if v is None:
        return None

    raw = str(v).strip().replace(',', '')
    if not raw or raw.lower() in ("none", "null", "na", "n/a", "-"):
        return None

    num = _to_float(raw)
    if num is None:
        return None

    # Register pages often write quantity as compact integer with implied 2 decimals.
    # Example: 34822 => 348.22 Qtl. Keep this only for quantity fields.
    digits_only = re.sub(r"\D", "", raw)
    if "." not in raw and 5 <= len(digits_only) <= 6:
        try:
            compact = int(digits_only)
            scaled = compact / 100.0
            if 50.0 <= scaled <= 5000.0:
                return round(scaled, 3)
        except Exception:
            pass

    return num


def _extract_compact_qty_candidate_qtl(source_row: Optional[object], bag_count_hint: Optional[float] = None) -> Optional[float]:
    if source_row is None:
        return None

    texts: List[str] = []
    bag_count: Optional[float] = _to_float(bag_count_hint)
    if isinstance(source_row, dict):
        if bag_count is None:
            bag_count = _to_float(_pick_first(source_row, ["no_of_bags", "bags", "bag_count"]))
        for k, v in source_row.items():
            if v is None:
                continue
            key = str(k or "").strip().lower()
            if any(tok in key for tok in [
                "truck", "vehicle", "date", "ws", "bag", "sno", "serial",
                "party", "item", "plant", "rm", "po", "total", "cumulative",
                "running", "rate", "amount", "remark",
            ]):
                continue
            texts.append(str(v))
    elif isinstance(source_row, (list, tuple)):
        texts = [str(v) for v in source_row if v is not None]
    else:
        texts = [str(source_row)]

    candidates: List[float] = []
    for raw in texts:
        s = str(raw or "")
        if not s:
            continue

        # Remove common non-quantity patterns before scanning compact tokens.
        s = re.sub(r"[A-Z]{2}\s*\d{2}\s*[A-Z]{1,3}\s*\d{4,5}", " ", s.upper())
        s = re.sub(r"\b\d{1,2}[./-]\d{1,2}(?:[./-]\d{2,4})?\b", " ", s)

        for m in re.finditer(r"(?<!\d)(\d{5,6})(?!\d)", s):
            try:
                token = int(m.group(1))
            except Exception:
                continue

            # Typical handwritten compact qty tokens like 34822 => 348.22 Qtl.
            if token < 20000:
                continue
            qtl = token / 100.0
            if 100.0 <= qtl <= 1000.0:
                candidates.append(round(qtl, 3))

    if not candidates:
        return None

    # Prefer candidate with realistic Qtl-per-bag ratio when bag count is available.
    if bag_count is not None and bag_count > 0:
        scored: List[Tuple[float, float]] = []
        for q in candidates:
            ratio = q / bag_count
            penalty = 0.0
            # Typical 50kg bag implies around 0.5 Qtl per bag. Keep this band broad for noisy OCR.
            if ratio < 0.30:
                penalty += (0.30 - ratio) * 10.0
            elif ratio > 0.80:
                penalty += (ratio - 0.80) * 10.0
            penalty += abs(ratio - 0.50)
            scored.append((penalty, q))
        scored.sort(key=lambda x: (x[0], x[1]))
        if scored:
            return round(float(scored[0][1]), 3)

    # Without bag signal, compact per-row qty is usually smaller than running-total token.
    return round(float(min(candidates)), 3)


def _coerce_unloading_row_quantities(row: dict, source_row: Optional[object] = None) -> dict:
    out = dict(row or {})

    # Try additional model key aliases when net quantity is missing.
    if out.get("net_qty_mt") is None and isinstance(source_row, dict):
        alt = _to_unloading_qty_qtl(_pick_first(source_row, [
            "net_weight_qtl",
            "weight_qtl",
            "quantity_qtl",
            "qty_qtl",
            "unload_qty_qtl",
            "actual_weight",
            "weight",
            "qty",
            "quantity",
            "wt",
        ]))

        if alt is None:
            for key, val in source_row.items():
                k = str(key or "").strip().lower()
                if not k:
                    continue
                if any(tok in k for tok in [
                    "bag", "ws", "date", "truck", "vehicle", "sno", "serial", "party", "item",
                    "plant", "rm", "po", "total", "cumulative", "running", "gross", "received",
                    "rate", "amount",
                ]):
                    continue
                if ("qty" in k) or ("weight" in k) or k.endswith("_wt"):
                    alt = _to_unloading_qty_qtl(val)
                    if alt is not None:
                        break

        if alt is not None:
            out["net_qty_mt"] = alt

    compact_qtl = _extract_compact_qty_candidate_qtl(source_row, out.get("no_of_bags"))
    out["_compact_qty_candidate"] = compact_qtl
    if out.get("net_qty_mt") is None and compact_qtl is not None:
        out["net_qty_mt"] = compact_qtl

    # If net appears copied from running-total column, prefer compact per-row quantity.
    if (
        compact_qtl is not None
        and out.get("net_qty_mt") is not None
        and out.get("total_qty_mt") is not None
    ):
        try:
            net_v = float(out.get("net_qty_mt"))
            tot_v = float(out.get("total_qty_mt"))
            if abs(net_v - tot_v) <= 1e-6 and abs(compact_qtl - tot_v) >= 1.0:
                out["net_qty_mt"] = compact_qtl
        except Exception:
            pass

    # Common handwritten register case: single quantity column is net quantity.
    # Prefer received/per-row quantity over running total even when total is present.
    if out.get("net_qty_mt") is None and out.get("received_qty_mt") is not None:
        out["net_qty_mt"] = out.get("received_qty_mt")

    if out.get("received_qty_mt") is None and out.get("net_qty_mt") is not None:
        out["received_qty_mt"] = out.get("net_qty_mt")

    return out


def _is_likely_cumulative_series(values: List[Optional[float]]) -> bool:
    seq = [float(v) for v in values if v is not None and float(v) > 0]
    if len(seq) < 3:
        return False

    non_decreasing = sum(1 for i in range(1, len(seq)) if seq[i] >= (seq[i - 1] - 1e-6))
    ratio = non_decreasing / max(1, len(seq) - 1)
    return ratio >= 0.85 and (seq[-1] - seq[0]) > 1.0


def _repair_unloading_net_quantities(rows: List[dict]) -> List[dict]:
    if not isinstance(rows, list) or not rows:
        return rows

    def _strip_internal_keys(row_list: List[dict]) -> List[dict]:
        cleaned: List[dict] = []
        for row in row_list:
            if not isinstance(row, dict):
                continue
            cleaned.append({k: v for k, v in row.items() if not str(k).startswith("_")})
        return cleaned

    out_rows = [dict(r or {}) for r in rows]
    rec_vals = [_to_float((r or {}).get("received_qty_mt")) for r in out_rows]
    net_vals = [_to_float((r or {}).get("net_qty_mt")) for r in out_rows]
    tot_vals = [_to_float((r or {}).get("total_qty_mt")) for r in out_rows]

    rec_count = sum(1 for v in rec_vals if v is not None)
    net_count = sum(1 for v in net_vals if v is not None)
    tot_count = sum(1 for v in tot_vals if v is not None)

    rec_cum = _is_likely_cumulative_series(rec_vals)
    net_cum = _is_likely_cumulative_series(net_vals)
    tot_cum = _is_likely_cumulative_series(tot_vals)

    # If only running total exists, derive per-row by differencing.
    if rec_count == 0 and net_count == 0 and tot_count >= 2 and tot_cum:
        prev_total = None
        diffs: List[Tuple[int, float]] = []
        for idx, row in enumerate(out_rows):
            t = _to_float(row.get("total_qty_mt"))
            if t is None:
                continue
            if prev_total is None:
                diff = t
            else:
                diff = (t - prev_total) if t >= prev_total else t
            prev_total = t
            if diff is not None and diff > 0:
                diffs.append((idx, float(diff)))

        if diffs:
            diff_vals = [d for _, d in diffs]
            mt_like = sum(1 for d in diff_vals if d < 100.0) >= max(1, int(len(diff_vals) * 0.7))
            for idx, diff in diffs:
                q = (diff * 10.0) if mt_like else diff
                out_rows[idx]["net_qty_mt"] = round(float(q), 3)
                if out_rows[idx].get("received_qty_mt") is None:
                    out_rows[idx]["received_qty_mt"] = out_rows[idx].get("net_qty_mt")

        return _strip_internal_keys(out_rows)

    def _normalize_qty_candidate_qtl(value, bags: Optional[float], prefer_exact: bool = False) -> Tuple[Optional[float], float]:
        base = _to_float(value)
        if base is None:
            return None, 9999.0

        if prefer_exact:
            cands = [float(base)]
        else:
            cands = [float(base), float(base) * 10.0, float(base) / 10.0]

        if bags is None or bags <= 0:
            # Without bag context, keep original unless clearly too small.
            chosen = cands[0]
            if not prefer_exact and chosen < 80.0 and (chosen * 10.0) <= 2000.0:
                chosen = chosen * 10.0
            return chosen, 0.8

        best_q = None
        best_pen = 1e9
        for q in cands:
            if q <= 0:
                continue
            ratio = q / bags
            pen = abs(ratio - 0.50)
            if ratio < 0.25:
                pen += (0.25 - ratio) * 8.0
            elif ratio > 0.95:
                pen += (ratio - 0.95) * 8.0
            if q < 80.0:
                pen += 1.0
            if q > 1500.0:
                pen += (q - 1500.0) / 300.0
            if pen < best_pen:
                best_pen = pen
                best_q = q

        if best_q is None:
            return None, 9999.0
        return float(best_q), float(best_pen)

    for row in out_rows:
        bags = _to_float(row.get("no_of_bags"))
        compact_raw = _to_float(row.get("_compact_qty_candidate"))

        rec_q, rec_pen = _normalize_qty_candidate_qtl(row.get("received_qty_mt"), bags)
        net_q, net_pen = _normalize_qty_candidate_qtl(row.get("net_qty_mt"), bags)
        tot_q, tot_pen = _normalize_qty_candidate_qtl(row.get("total_qty_mt"), bags)
        cmp_q, cmp_pen = _normalize_qty_candidate_qtl(compact_raw, bags, prefer_exact=True)

        # Persist normalized side columns when available.
        if rec_q is not None:
            row["received_qty_mt"] = round(rec_q, 3)
        if tot_q is not None:
            row["total_qty_mt"] = round(tot_q, 3)

        candidates: List[Tuple[float, float]] = []
        if net_q is not None:
            candidates.append((net_pen + (1.2 if net_cum else 0.0), net_q))
        if rec_q is not None:
            candidates.append((rec_pen + (0.15 if not rec_cum else 0.8), rec_q))
        if cmp_q is not None:
            candidates.append((cmp_pen - 0.1, cmp_q))
        if tot_q is not None:
            candidates.append((tot_pen + (2.5 if tot_cum else 1.3), tot_q))

        if candidates:
            best = min(candidates, key=lambda x: x[0])
            row["net_qty_mt"] = round(float(best[1]), 3)
        elif row.get("net_qty_mt") is None and row.get("received_qty_mt") is not None:
            row["net_qty_mt"] = row.get("received_qty_mt")

        if row.get("received_qty_mt") is None and row.get("net_qty_mt") is not None:
            row["received_qty_mt"] = row.get("net_qty_mt")

    return _strip_internal_keys(out_rows)


def _unloading_rows_quality(rows: Optional[List[dict]]) -> dict:
    row_list = rows if isinstance(rows, list) else []
    rows_count = len(row_list)
    rows_with_any_qty = 0
    rows_with_net = 0

    for row in row_list:
        if not isinstance(row, dict):
            continue
        qty_vals = [
            row.get("received_qty_mt"),
            row.get("net_qty_mt"),
            row.get("total_qty_mt"),
        ]
        has_any = any(v is not None for v in qty_vals)
        if has_any:
            rows_with_any_qty += 1
        if row.get("net_qty_mt") is not None:
            rows_with_net += 1

    score = (rows_with_net * 1000) + (rows_with_any_qty * 10) + rows_count
    return {
        "rows": rows_count,
        "rows_with_any_qty": rows_with_any_qty,
        "rows_with_net": rows_with_net,
        "score": score,
    }


def _normalize_unloading_rows(data: dict) -> List[dict]:
    candidates = []
    for k in ["rows", "entries", "trucks", "unloading_rows", "table_rows", "details"]:
        v = data.get(k)
        if isinstance(v, list):
            candidates.extend(v)

    out_rows: List[dict] = []
    for r in candidates:
        if isinstance(r, dict):
            row = {
                "sno": _pick_first(r, ["sno", "sr_no", "serial_no"]),
                "ws_no": _pick_first(r, ["ws_no", "ws_number", "ws", "w_s_no"]),
                "date": _normalize_date_text(_pick_first(r, ["date", "entry_date", "truck_date"])),
                "truck_number": _pick_first(r, ["truck_number", "truck_no", "vehicle_number", "vehicle_no", "truck"]),
                "no_of_bags": _to_float(_pick_first(r, ["no_of_bags", "bags", "bag_count"])),
                "received_qty_mt": _to_unloading_qty_qtl(_pick_first(r, ["received_qty_mt", "received_quantity_mt", "received_qty", "gross_qty_mt", "received"])),
                "net_qty_mt": _to_unloading_qty_qtl(_pick_first(r, [
                    "net_qty_mt", "net_quantity_mt", "net_weight", "net_qty", "net",
                    "net_weight_qtl", "weight_qtl", "quantity_qtl", "qty_qtl", "actual_weight",
                ])),
                "total_qty_mt": _to_unloading_qty_qtl(_pick_first(r, ["total_qty_mt", "total_qty", "cumulative_qty_mt", "running_total"])),
                "item_name": _pick_first(r, ["item_name", "item"]) or data.get("item_name"),
                "party_name": _pick_first(r, ["party_name", "vendor_name", "party"]) or data.get("party_name"),
                "rm_number": _pick_first(r, ["rm_number", "rm_no", "rm"]) or data.get("rm_number"),
                "po_number": _pick_first(r, ["po_number", "po_no", "purchase_order_no"]) or data.get("po_number"),
                "plant_name": _pick_first(r, ["plant_name", "plant"]) or data.get("plant_name"),
            }
            row = _coerce_unloading_row_quantities(row, r)
            out_rows.append(row)
        elif isinstance(r, (list, tuple)):
            vals = list(r)
            row = {
                "sno": vals[0] if len(vals) > 0 else None,
                "ws_no": vals[1] if len(vals) > 1 else None,
                "date": _normalize_date_text(vals[2] if len(vals) > 2 else None),
                "truck_number": vals[3] if len(vals) > 3 else None,
                "no_of_bags": _to_float(vals[4] if len(vals) > 4 else None),
                "received_qty_mt": _to_unloading_qty_qtl(vals[5] if len(vals) > 5 else None),
                "net_qty_mt": _to_unloading_qty_qtl(vals[6] if len(vals) > 6 else None),
                "total_qty_mt": _to_unloading_qty_qtl(vals[7] if len(vals) > 7 else None),
                "item_name": data.get("item_name"),
                "party_name": data.get("party_name"),
                "rm_number": data.get("rm_number"),
                "po_number": data.get("po_number"),
                "plant_name": data.get("plant_name"),
            }
            row = _coerce_unloading_row_quantities(row, vals)
            out_rows.append(row)

    # Keep rows with at least one strong signal (truck/date/net/ws)
    filtered = []
    for row in out_rows:
        if row.get("truck_number") or row.get("date") or row.get("net_qty_mt") is not None or row.get("ws_no"):
            filtered.append(row)

    return _repair_unloading_net_quantities(filtered)


def _build_unloading_result_from_payload(payload: Optional[dict], fallback_source: str) -> Optional[dict]:
    if not isinstance(payload, dict):
        return None

    rows = _normalize_unloading_rows(payload)
    if not rows:
        return None

    raw_text = payload.get("raw_text")
    if raw_text in (None, ""):
        try:
            raw_text = json.dumps(payload, ensure_ascii=False)
        except Exception:
            raw_text = ""

    return {
        "source": str(payload.get("source") or fallback_source),
        "plant_name": payload.get("plant_name"),
        "item_name": payload.get("item_name"),
        "party_name": payload.get("party_name"),
        "rm_number": payload.get("rm_number"),
        "po_number": payload.get("po_number"),
        "sheet_date": _normalize_date_text(payload.get("sheet_date")),
        "rows": rows,
        "confidence": payload.get("confidence", 0.0),
        "high_confidence": len(rows) > 0,
        "raw_text": raw_text,
    }


def _mime_type_for_path(file_path: str) -> str:
    ext = Path(file_path).suffix.lower()
    return {
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".png": "image/png",
        ".webp": "image/webp",
        ".pdf": "application/pdf",
    }.get(ext, "application/octet-stream")


def _env_int(name: str, default: int, low: int, high: int) -> int:
    load_dotenv()
    raw = (os.getenv(name) or "").strip()
    try:
        val = int(raw) if raw else int(default)
    except Exception:
        val = int(default)
    return max(int(low), min(int(high), int(val)))


def _compact_image_for_llm(file_path: str) -> Tuple[str, bool]:
    """Downscale/re-encode image payload before base64 upload to avoid 413 errors."""
    src = Path(file_path)
    if src.suffix.lower() not in {".jpg", ".jpeg", ".png", ".webp", ".bmp", ".tif", ".tiff", ".gif"}:
        return file_path, False

    # Keep default compact enough for strict provider gateways after base64 expansion.
    max_bytes = _env_int("OCR_LLM_IMAGE_MAX_BYTES", 450000, 120000, 4000000)
    max_side = _env_int("OCR_LLM_IMAGE_MAX_SIDE", 1280, 512, 4096)
    min_side = _env_int("OCR_LLM_IMAGE_MIN_SIDE", 640, 320, 2048)
    jpeg_quality = _env_int("OCR_LLM_IMAGE_JPEG_QUALITY", 68, 35, 95)

    try:
        from PIL import Image, ImageOps

        with Image.open(src) as im:
            im = ImageOps.exif_transpose(im)
            width, height = im.size
            size_bytes = src.stat().st_size if src.exists() else 0
            if (
                size_bytes > 0
                and size_bytes <= max_bytes
                and max(width, height) <= max_side
                and src.suffix.lower() in {".jpg", ".jpeg"}
            ):
                return file_path, False

            if im.mode not in {"RGB", "L"}:
                im = im.convert("RGB")
            elif im.mode == "L":
                im = im.convert("RGB")

            if max(width, height) > max_side:
                scale = max_side / float(max(width, height))
                nw = max(1, int(width * scale))
                nh = max(1, int(height * scale))
                im = im.resize((nw, nh), Image.Resampling.LANCZOS)

            out_path = str(src.with_suffix("")) + ".llm.jpg"
            work = im
            quality = jpeg_quality
            attempts = 0
            while attempts < 18:
                work.save(out_path, format="JPEG", quality=quality, optimize=True, progressive=True)
                out_size = Path(out_path).stat().st_size if Path(out_path).exists() else 0
                if out_size <= max_bytes:
                    return out_path, True

                if quality > 42:
                    quality = max(42, quality - 7)
                else:
                    if max(work.size) <= min_side:
                        break
                    shrink = 0.82 if max(work.size) > 900 else 0.75
                    ww = max(1, int(work.size[0] * shrink))
                    hh = max(1, int(work.size[1] * shrink))
                    work = work.resize((ww, hh), Image.Resampling.LANCZOS)
                attempts += 1

            # Last-resort tiny JPEG so request does not fail entire sync on 413.
            if max(work.size) > min_side:
                tiny_scale = min_side / float(max(work.size))
                tw = max(1, int(work.size[0] * tiny_scale))
                th = max(1, int(work.size[1] * tiny_scale))
                work = work.resize((tw, th), Image.Resampling.LANCZOS)
            quality = min(quality, 40)
            work.save(out_path, format="JPEG", quality=quality, optimize=True, progressive=True)
            out_size = Path(out_path).stat().st_size if Path(out_path).exists() else 0
            if out_size > max_bytes:
                log.warning(
                    "LLM compact image still above target (%s bytes > %s) for %s",
                    out_size,
                    max_bytes,
                    file_path,
                )

            return out_path, True
    except Exception as e:
        log.warning("LLM image compact skipped for %s: %s", file_path, e)
        return file_path, False


def _build_data_url_for_llm(file_path: str) -> Tuple[str, str, Optional[str]]:
    compact_path, compact_temp = _compact_image_for_llm(file_path)
    mime = _mime_type_for_path(compact_path)
    with open(compact_path, "rb") as f:
        blob_b64 = base64.b64encode(f.read()).decode("utf-8")
    return f"data:{mime};base64,{blob_b64}", mime, (compact_path if compact_temp else None)


def _extract_unloading_with_groq_doc(file_path: str, api_key: str) -> Optional[dict]:
    """Layer 1: send uploaded document directly to Groq with unloading prompt."""
    groq_input_path: Optional[str] = None
    cleanup_temp = False
    compact_temp_path: Optional[str] = None
    try:
        from groq import Groq

        src_path = Path(file_path)
        if src_path.suffix.lower() == ".pdf":
            converted = _convert_pdf_first_page_to_image(file_path)
            if not converted:
                log.warning("Groq direct-doc layer skipped: PDF->PNG conversion unavailable")
                return None
            groq_input_path = converted
            cleanup_temp = True
        else:
            groq_input_path = file_path

        client = Groq(api_key=api_key)
        data_url, _mime, compact_temp_path = _build_data_url_for_llm(groq_input_path)

        _sleep_before_llm_request_sync()
        response = client.chat.completions.create(
            messages=[
                {
                    "role": "user",
                    "content": [
                        {"type": "text", "text": PLANT_UNLOADING_PROMPT},
                        {"type": "image_url", "image_url": {"url": data_url}},
                    ],
                }
            ],
            model="llama-3.3-70b-versatile",
            temperature=0.1,
            response_format={"type": "json_object"},
        )

        raw = response.choices[0].message.content
        data = json.loads(raw)
        rows = _normalize_unloading_rows(data)

        if not rows:
            _sleep_before_llm_request_sync()
            retry = client.chat.completions.create(
                messages=[
                    {
                        "role": "user",
                        "content": [
                            {"type": "text", "text": PLANT_UNLOADING_FALLBACK_PROMPT},
                            {"type": "image_url", "image_url": {"url": data_url}},
                        ],
                    }
                ],
                model="llama-3.3-70b-versatile",
                temperature=0.1,
                response_format={"type": "json_object"},
            )
            retry_data = json.loads(retry.choices[0].message.content)
            retry_rows = _normalize_unloading_rows(retry_data)
            if retry_rows:
                rows = retry_rows
                raw = json.dumps({"first": data, "fallback": retry_data})

        return {
            "source": "groq",
            "plant_name": data.get("plant_name"),
            "item_name": data.get("item_name"),
            "party_name": data.get("party_name"),
            "rm_number": data.get("rm_number"),
            "po_number": data.get("po_number"),
            "sheet_date": _normalize_date_text(data.get("sheet_date")),
            "rows": rows,
            "confidence": data.get("confidence", 0.0),
            "high_confidence": len(rows) > 0,
            "raw_text": raw,
        }
    except Exception as e:
        log.warning(f"Groq direct-doc layer failed: {e}")
        return None
    finally:
        if cleanup_temp and groq_input_path:
            try:
                Path(groq_input_path).unlink()
            except Exception:
                pass


def _extract_unloading_with_gemini_doc(file_path: str) -> Optional[dict]:
    """Layer 2: send uploaded document directly to Gemini and normalize rows."""
    try:
        from database import get_settings
        settings = get_settings()
        if not settings.gemini_api_key:
            return None

        mime = _mime_type_for_path(file_path)
        blob = Path(file_path).read_bytes()
        text = _gemini_generate_text(
            api_key=settings.gemini_api_key,
            model="gemini-2.5-flash",
            prompt_text=PLANT_UNLOADING_PROMPT,
            blob=blob,
            mime=mime,
        )
        text = re.sub(r"^```json\s*", "", text)
        text = re.sub(r"\s*```$", "", text)
        data = json.loads(text)
        rows = _normalize_unloading_rows(data)

        if not rows:
            text2 = _gemini_generate_text(
                api_key=settings.gemini_api_key,
                model="gemini-2.5-flash",
                prompt_text=PLANT_UNLOADING_FALLBACK_PROMPT,
                blob=blob,
                mime=mime,
            )
            text2 = re.sub(r"^```json\s*", "", text2)
            text2 = re.sub(r"\s*```$", "", text2)
            data2 = json.loads(text2)
            rows2 = _normalize_unloading_rows(data2)
            if rows2:
                rows = rows2
                text = json.dumps({"first": data, "fallback": data2})

        return {
            "source": "gemini",
            "plant_name": data.get("plant_name"),
            "item_name": data.get("item_name"),
            "party_name": data.get("party_name"),
            "rm_number": data.get("rm_number"),
            "po_number": data.get("po_number"),
            "sheet_date": _normalize_date_text(data.get("sheet_date")),
            "rows": rows,
            "confidence": data.get("confidence", 0.0),
            "high_confidence": len(rows) > 0,
            "raw_text": text,
        }
    except Exception as e:
        log.warning(f"Gemini direct-doc layer failed: {e}")
        return None


async def extract_plant_unloading_sheet(file_path: str, preferred_provider: Optional[str] = None) -> dict:
    """Extract unloading rows with preferred sequence: GitHub -> Mistral -> Groq -> Gemini -> Azure+Groq."""
    ext = Path(file_path).suffix.lower()
    if ext in {".xlsx", ".xls"}:
        try:
            # Reuse the local spreadsheet parser already used by email ingest.
            from services.email_sync_service import _extract_plant_unloading_from_excel

            excel_res = _extract_plant_unloading_from_excel(file_path, subject="", file_name=Path(file_path).name)
            q = _unloading_rows_quality((excel_res or {}).get("rows"))
            if isinstance(excel_res, dict):
                attempts = [{
                    "provider": "excel_local",
                    "source": "excel_local",
                    "status": "accepted" if q["rows"] > 0 else "no_rows",
                    "rows": int(q.get("rows") or 0),
                    "rows_with_any_qty": int(q.get("rows_with_any_qty") or 0),
                    "rows_with_net": int(q.get("rows_with_net") or 0),
                    "confidence": float(excel_res.get("confidence") or 0),
                    "error": str(excel_res.get("error") or "") or None,
                }]

                # If Excel header mapping found only partial row signals (e.g., truck+bags)
                # run a text-based LLM parser over sheet text to recover qty/header fields.
                weak_excel_parse = (
                    int(q.get("rows") or 0) == 0
                    or int(q.get("rows_with_any_qty") or 0) == 0
                    or not any(excel_res.get(k) for k in ["item_name", "party_name", "rm_number", "po_number", "plant_name"])
                )

                selected = excel_res
                selected_q = q
                if weak_excel_parse:
                    load_dotenv()
                    groq_key = (os.getenv("GROQ_API_KEY") or "").strip()
                    excel_text = _build_excel_text_for_unloading_prompt(file_path)
                    llm_res = None
                    llm_q = {"rows": 0, "rows_with_any_qty": 0, "rows_with_net": 0, "score": -1}
                    llm_error = None

                    if not excel_text:
                        llm_error = "excel_text_unavailable"
                    elif not groq_key:
                        llm_error = "GROQ_API_KEY not set for excel-text fallback"
                    else:
                        llm_res = _extract_unloading_from_text_with_groq(
                            excel_text,
                            groq_key,
                            source_label="excel_text+groq",
                        )
                        if isinstance(llm_res, dict):
                            llm_q = _unloading_rows_quality(llm_res.get("rows"))
                        else:
                            llm_error = "excel_text_llm_parse_failed"

                    attempts.append({
                        "provider": "excel_text_groq",
                        "source": "excel_text+groq",
                        "status": "accepted" if int(llm_q.get("rows") or 0) > 0 else ("error" if llm_error else "no_rows"),
                        "rows": int(llm_q.get("rows") or 0),
                        "rows_with_any_qty": int(llm_q.get("rows_with_any_qty") or 0),
                        "rows_with_net": int(llm_q.get("rows_with_net") or 0),
                        "confidence": float((llm_res or {}).get("confidence") or 0) if isinstance(llm_res, dict) else 0,
                        "error": llm_error,
                    })

                    if isinstance(llm_res, dict):
                        if int(llm_q.get("score") or -1) > int(selected_q.get("score") or -1):
                            selected = llm_res
                            selected_q = llm_q

                        # Preserve stronger local parse header fields/rows and fill missing metadata from LLM.
                        for key in ["plant_name", "item_name", "party_name", "rm_number", "po_number", "sheet_date"]:
                            if not selected.get(key) and llm_res.get(key):
                                selected[key] = llm_res.get(key)

                selected["provider_attempts"] = attempts
                selected.setdefault("high_confidence", bool(int(selected_q.get("rows") or 0) > 0))
                selected.setdefault("source", "excel_local")
                return selected
        except Exception as e:
            return {
                "source": "excel_local",
                "rows": [],
                "error": f"excel_local_parser_failed: {str(e)[:220]}",
                "provider_attempts": [{
                    "provider": "excel_local",
                    "source": "excel_local",
                    "status": "error",
                    "rows": 0,
                    "rows_with_any_qty": 0,
                    "rows_with_net": 0,
                    "confidence": 0,
                    "error": f"excel_local_parser_failed: {str(e)[:220]}",
                }],
                "high_confidence": False,
            }

    load_dotenv()
    api_key = os.getenv("GROQ_API_KEY")
    provider_attempts: List[dict] = []
    best_result: Optional[dict] = None
    best_score = -1
    provider_key = str(preferred_provider or "").strip().lower()
    if provider_key in {"", "auto"}:
        provider_key = ""
    supported_keys = {"github", "mistral", "groq", "gemini", "azure"}
    if provider_key and provider_key not in supported_keys:
        return {
            "source": "plant-unloading-router",
            "rows": [],
            "error": f"Unsupported ocr_engine '{preferred_provider}' for plant_unloading",
            "provider_attempts": [],
            "high_confidence": False,
        }

    def _enabled(key: str) -> bool:
        if not provider_key:
            return True
        if provider_key == "azure":
            return key == "azure"
        return key == provider_key

    def _consider_best(result: Optional[dict]) -> None:
        nonlocal best_result, best_score
        if not isinstance(result, dict):
            return
        q = _unloading_rows_quality(result.get("rows"))
        if q["score"] > best_score:
            best_score = q["score"]
            best_result = result

    def _record_attempt(
        provider: str,
        payload: Optional[dict],
        quality: Optional[dict] = None,
        error: Optional[str] = None,
        accepted: bool = False,
    ) -> None:
        src = provider
        conf: Optional[float] = None
        err = error
        q = quality or {"rows": 0, "rows_with_any_qty": 0, "rows_with_net": 0}
        if isinstance(payload, dict):
            src = str(payload.get("source") or provider)
            if err is None and payload.get("error"):
                err = str(payload.get("error"))
            try:
                if payload.get("confidence") is not None:
                    conf = float(payload.get("confidence"))
            except Exception:
                conf = None

        rows_count = int(q.get("rows") or 0)
        if accepted:
            status = "accepted"
        elif rows_count > 0:
            status = "partial"
        else:
            status = "error" if err else "no_rows"

        provider_attempts.append({
            "provider": provider,
            "source": src,
            "status": status,
            "rows": rows_count,
            "rows_with_any_qty": int(q.get("rows_with_any_qty") or 0),
            "rows_with_net": int(q.get("rows_with_net") or 0),
            "confidence": conf,
            "error": err,
        })

    # Layer 1: GitHub direct document
    if _enabled("github"):
        github_payload = await _extract_with_github_prompt(file_path, PLANT_UNLOADING_PROMPT, "github_unloading")
        github_res = _build_unloading_result_from_payload(github_payload, "github")
        _consider_best(github_res)
        github_q = _unloading_rows_quality((github_res or {}).get("rows"))
        github_accept = github_q["rows_with_net"] > 0
        _record_attempt("github", github_payload, github_q, accepted=github_accept)
        if github_accept:
            log.info("Plant unloading: layer1 GitHub succeeded")
            github_res["provider_attempts"] = provider_attempts
            return github_res

    # Layer 2: Mistral direct document
    if _enabled("mistral"):
        mistral_payload = await _extract_with_mistral_prompt(file_path, PLANT_UNLOADING_PROMPT, "mistral_unloading")
        mistral_res = _build_unloading_result_from_payload(mistral_payload, "mistral")
        _consider_best(mistral_res)
        mistral_q = _unloading_rows_quality((mistral_res or {}).get("rows"))
        mistral_accept = mistral_q["rows_with_net"] > 0
        _record_attempt("mistral", mistral_payload, mistral_q, accepted=mistral_accept)
        if mistral_accept:
            log.info("Plant unloading: layer2 Mistral succeeded")
            mistral_res["provider_attempts"] = provider_attempts
            return mistral_res

    # Layer 3: Groq direct document
    if _enabled("groq"):
        if api_key:
            groq_res = _extract_unloading_with_groq_doc(file_path, api_key)
            _consider_best(groq_res)
            groq_q = _unloading_rows_quality((groq_res or {}).get("rows"))
            groq_accept = groq_q["rows_with_net"] > 0
            _record_attempt("groq", groq_res, groq_q, accepted=groq_accept)
            if groq_accept:
                log.info("Plant unloading: layer3 Groq succeeded")
                groq_res["provider_attempts"] = provider_attempts
                return groq_res
        else:
            _record_attempt("groq", None, {"rows": 0, "rows_with_any_qty": 0, "rows_with_net": 0}, "GROQ_API_KEY not set")

    # Layer 4: Gemini direct document
    if _enabled("gemini"):
        gem_res = _extract_unloading_with_gemini_doc(file_path)
        _consider_best(gem_res)
        gem_q = _unloading_rows_quality((gem_res or {}).get("rows"))
        gem_accept = gem_q["rows_with_net"] > 0
        _record_attempt("gemini", gem_res, gem_q, accepted=gem_accept)
        if gem_accept:
            log.info("Plant unloading: layer4 Gemini succeeded")
            gem_res["provider_attempts"] = provider_attempts
            return gem_res

    # Layer 5: Azure read + Groq text parsing fallback
    if _enabled("azure"):
        if api_key:
            txt = _extract_pdf_text_with_azure_read(file_path)
            if txt:
                parsed = _extract_unloading_from_text_with_groq(txt, api_key)
                _consider_best(parsed)
                parsed_q = _unloading_rows_quality((parsed or {}).get("rows"))
                parsed_accept = parsed_q["rows_with_net"] > 0
                _record_attempt("azure+groq", parsed, parsed_q, accepted=parsed_accept)
                if parsed_accept:
                    log.info("Plant unloading: layer5 Azure+Groq succeeded")
                    parsed["provider_attempts"] = provider_attempts
                    return parsed
            else:
                _record_attempt("azure+groq", None, {"rows": 0, "rows_with_any_qty": 0, "rows_with_net": 0}, "Azure read text unavailable")
        else:
            _record_attempt("azure+groq", None, {"rows": 0, "rows_with_any_qty": 0, "rows_with_net": 0}, "GROQ_API_KEY not set")

    if isinstance(best_result, dict) and ((best_result.get("rows") or []) != []):
        best_result["provider_attempts"] = provider_attempts
        q = _unloading_rows_quality(best_result.get("rows"))
        best_result["high_confidence"] = q["rows_with_net"] > 0
        best_result.setdefault(
            "error",
            "No provider extracted net_qty_mt confidently; returning best available unloading rows",
        )
        return best_result

    return {
        "source": "plant-unloading-router",
        "rows": [],
        "error": "All extraction layers failed (GitHub direct, Mistral direct, Groq direct, Gemini direct, Azure+Groq fallback)",
        "provider_attempts": provider_attempts,
        "high_confidence": False,
    }

async def extract_with_openrouter(image_path: str) -> dict:
    try:
        from openai import OpenAI
    except Exception as e:
        log.error("OpenRouter unavailable: cannot import openai on %s: %s", sys.executable, e)
        return {
            "source": "openrouter",
            "error": f"openai import failed: {e}",
            "high_confidence": False,
        }
    load_dotenv()
    api_key = os.getenv("OPENROUTER_API_KEY")
    
    if not api_key:
        log.warning("OpenRouter skipped: OPENROUTER_API_KEY not set")
        return {"source": "openrouter", "error": "OPENROUTER_API_KEY not set", "high_confidence": False}

    openrouter_input_path: Optional[str] = None
    cleanup_temp = False
    compact_temp_path: Optional[str] = None
    configured_raw = (os.getenv("OPENROUTER_MODEL") or "").strip()
    configured_items = [s.strip() for s in re.split(r"[,\s]+", configured_raw) if s.strip()]
    free_only = _openrouter_free_only_enabled()
    if free_only and _openrouter_free_temporarily_blocked():
        retry_after = _openrouter_retry_after_sec()
        return {
            "source": "openrouter",
            "error": f"OpenRouter free quota cooldown active; retry in ~{retry_after}s",
            "high_confidence": False,
            "rate_limited": True,
            "retry_after_sec": retry_after,
        }
    try:
        src_path = Path(image_path)
        if src_path.suffix.lower() == ".pdf":
            converted = _convert_pdf_first_page_to_image(image_path)
            if not converted:
                return {
                    "source": "openrouter",
                    "error": "PDF->image conversion failed for OpenRouter",
                    "high_confidence": False,
                }
            openrouter_input_path = converted
            cleanup_temp = True
        else:
            openrouter_input_path = image_path

        data_url, _mime, compact_temp_path = _build_data_url_for_llm(openrouter_input_path)

        # OpenRouter uses the exact same structure as OpenAI
        client = OpenAI(
            base_url=_openrouter_api_base_url(),
            api_key=api_key,
            max_retries=0,
        )

        # Keep candidates aligned with discovered free vision models for this key.
        model_candidates = _openrouter_resolve_model_candidates(
            api_key,
            configured_items=configured_items,
            include_paid=(not free_only),
        )

        if free_only:
            bad_config = [m for m in configured_items if not m.endswith(":free")]
            if bad_config:
                log.warning(
                    "OPENROUTER_MODEL contains non-free entries ignored by free-only mode: %s",
                    bad_config,
                )
            model_candidates = [m for m in model_candidates if m.endswith(":free")]

        if not model_candidates:
            return {
                "source": "openrouter",
                "error": "No OpenRouter free vision model available for this key",
                "high_confidence": False,
            }

        last_err = None
        best_result = None
        best_fields = -1
        non_rate_failures: List[str] = []
        for model_id in model_candidates:
            try:
                request_kwargs = {
                    "model": model_id,
                    "messages": [
                        {
                            "role": "user",
                            "content": [
                                {"type": "text", "text": GEMINI_PROMPT},
                                {
                                    "type": "image_url",
                                    "image_url": {"url": data_url}
                                }
                            ]
                        }
                    ],
                    "temperature": 0.1,
                    "max_tokens": 1600,
                    "response_format": {"type": "json_object"},
                }

                try:
                    await _sleep_before_llm_request_async()
                    response = client.chat.completions.create(**request_kwargs)
                except Exception as e:
                    em = str(e).lower()
                    if "json mode is not enabled" in em or "response_format" in em:
                        log.info("Model %s does not support response_format JSON mode; retrying without it", model_id)
                        request_kwargs.pop("response_format", None)
                        await _sleep_before_llm_request_async()
                        response = client.chat.completions.create(**request_kwargs)
                    else:
                        raise

                raw = (response.choices[0].message.content or "").strip()
                try:
                    result = _load_json_object_loose(raw)
                except Exception as parse_err:
                    preview = (raw or "")[:220].replace("\n", " ")
                    raise ValueError(f"Invalid JSON payload: {parse_err}; preview='{preview}'")

                result = _normalize_model_bill_payload(result)
                result["source"] = "openrouter"
                result["openrouter_model"] = model_id

                if "quantity_qtl" not in result and "quantity_mt" in result:
                    try:
                        result["quantity_qtl"] = round(float(result["quantity_mt"]) * 10, 3)
                    except Exception:
                        pass

                fields_found = _count_bill_fields(result)
                result["fields_found"] = fields_found
                conf = float(result.get("confidence") or 0)
                result["high_confidence"] = (fields_found >= 5) or (fields_found >= 4 and conf >= 0.65)

                if fields_found > best_fields:
                    best_fields = fields_found
                    best_result = result

                if result["high_confidence"]:
                    return result
                continue
            except Exception as e:
                last_err = e
                err_text = str(e)
                if _is_openrouter_free_quota_error(err_text):
                    _mark_openrouter_free_quota_exhausted(err_text)
                    log.warning("OpenRouter free quota exhausted; stopping fallback chain: %s", e)
                    return {
                        "source": "openrouter",
                        "error": err_text,
                        "high_confidence": False,
                        "rate_limited": True,
                        "openrouter_model": model_id,
                    }
                non_rate_failures.append(f"{model_id}: {str(e)[:180]}")
                continue

        if isinstance(best_result, dict):
            return best_result

        if non_rate_failures:
            log.debug("OpenRouter non-rate-limit failures: %s", " | ".join(non_rate_failures[:3]))

        return {
            "source": "openrouter",
            "error": f"All OpenRouter model candidates failed: {str(last_err)[:260]}",
            "high_confidence": False,
        }
    except Exception as e:
        log.error(f"OpenRouter Vision error: {e}")
        return {"source": "openrouter", "error": str(e), "high_confidence": False}
    finally:
        if compact_temp_path:
            try:
                Path(compact_temp_path).unlink(missing_ok=True)
            except Exception:
                pass
        if cleanup_temp and openrouter_input_path:
            try:
                Path(openrouter_input_path).unlink(missing_ok=True)
            except Exception:
                pass


def classify_document_type(file_path: str, file_name: Optional[str] = None) -> Tuple[str, float, List[dict]]:
    """LLM-first document classifier with ordered fallback: Groq -> Mistral -> OpenRouter -> Gemini -> GitHub.
    Directly sends document/image to Vision LLMs without upfront text extraction.

    Returns: (document_type, confidence, candidates)
    """

    def _normalize_doc_type(raw: Optional[str]) -> Optional[str]:
        val = str(raw or "").strip().lower()
        alias = {
            "bill": "purchase_bill",
            "invoice": "purchase_bill",
            "nit": "tender_notice",
            "tender": "tender_notice",
            "po": "purchase_order",
            "order": "purchase_order",
            "rejection": "rejection_notice",
            "reject": "rejection_notice",
            "unloading": "plant_unloading",
            "unknown": "not_classified",
            "other": "not_classified",
            "other_document": "not_classified",
            "other_doc": "not_classified",
            "quality_report": "not_classified",
            "lab_report": "not_classified",
            "coa": "not_classified",
            "certificate_of_analysis": "not_classified",
        }
        val = alias.get(val, val)
        if val in SUPPORTED_DOCUMENT_TYPES:
            return val
        return None

    def _vision_classifier_prompt(name_text: str) -> str:
        return f"""
Classify this document visually into EXACTLY ONE type from this list only:
- purchase_bill : Tax invoices, bills of supply for raw materials. Look for "Invoice", "Bill of Supply", total amount, GST numbers.
- tender_notice : NIT (Notice Inviting Tender), asking for bids, listing materials and EMD but no winner yet. Look for "NIT", "Notice Inviting".
- purchase_order : PO, Work Order, or Allotment Letter approving a specific party/winner to supply materials at a specific rate.
- rejection_notice : Document rejecting a truck/material due to quality/moisture, mentioning deduction or total rejection. Look for "Reject", "Quality Issue".
- plant_unloading : Raw material control register, WS (weighment slip) log, tabular list of trucks unloaded. Often handwritten or printed table of truck numbers and weights.
- not_classified : Any non-target document such as quality/lab/analysis reports, correspondence, or unclear files that do not fit the above.

Filename/context for hints: {name_text}

Return only strict JSON:
{{"document_type":"one_of_the_types","confidence":0.0,"reason":"short reason based on visual evidence"}}
""".strip()

    def _groq_vision_classify(prompt: str, image_path: str, mime: str) -> Optional[dict]:
        load_dotenv()
        api_key = (os.getenv("GROQ_API_KEY") or "").strip()
        if not api_key:
            return None
        compact_temp_path: Optional[str] = None
        try:
            from groq import Groq

            data_url, _mime, compact_temp_path = _build_data_url_for_llm(image_path)

            configured = (os.getenv("DOC_CLASSIFIER_MODEL") or "").strip()
            # Fallbacks for vision capable models
            model_candidates = [
                configured,
                "llama-3.3-70b-versatile",
                "llama-3.1-8b-instant",
            ]
            model_candidates = [m for i, m in enumerate(model_candidates) if m and m not in model_candidates[:i]]
            
            client = Groq(api_key=api_key)
            for model in model_candidates:
                try:
                    _sleep_before_llm_request_sync()
                    resp = client.chat.completions.create(
                        messages=[
                            {
                                "role": "user",
                                "content": [
                                    {"type": "text", "text": prompt},
                                    {"type": "image_url", "image_url": {"url": data_url}},
                                ],
                            }
                        ],
                        model=model,
                        temperature=0,
                        response_format={"type": "json_object"},
                    )
                    payload = _load_json_object_loose((resp.choices[0].message.content or "").strip())
                    doc_type = _normalize_doc_type(payload.get("document_type"))
                    if not doc_type:
                        continue
                    try:
                        conf = float(payload.get("confidence") or 0)
                    except Exception:
                        conf = 0.0
                    return {
                        "provider": "groq",
                        "model": model,
                        "document_type": doc_type,
                        "confidence": max(0.0, min(conf, 0.99)),
                        "reason": str(payload.get("reason") or "").strip(),
                    }
                except Exception as e:
                    log.warning("Groq classifier failed (%s): %s", model, e)
            return None
        except Exception as e:
            log.warning("Groq classifier init failed: %s", e)
            return None
        finally:
            if compact_temp_path:
                try:
                    Path(compact_temp_path).unlink(missing_ok=True)
                except Exception:
                    pass

    def _mistral_vision_classify(prompt: str, image_path: str, mime: str) -> Optional[dict]:
        load_dotenv()
        if not _mistral_enabled():
            return None
        api_key = (os.getenv("MISTRAL_API_KEY") or "").strip()
        if not api_key:
            return None

        compact_temp_path: Optional[str] = None
        try:
            from openai import OpenAI

            data_url, _mime, compact_temp_path = _build_data_url_for_llm(image_path)

            configured_raw = [
                (os.getenv("MISTRAL_CLASSIFIER_MODEL") or "").strip(),
                (os.getenv("MISTRAL_MODEL") or "").strip(),
            ]
            configured_items: List[str] = []
            for raw in configured_raw:
                configured_items.extend([s.strip() for s in re.split(r"[,\s]+", raw) if s.strip()])
            model_candidates = _mistral_resolve_model_candidates(configured_items=configured_items)
            if not model_candidates:
                return None

            client = OpenAI(base_url=_mistral_api_base_url(), api_key=api_key, max_retries=0)
            non_rate_failures: List[str] = []
            for model in model_candidates:
                try:
                    kwargs = {
                        "model": model,
                        "messages": [
                            {
                                "role": "user",
                                "content": [
                                    {"type": "text", "text": prompt},
                                    {"type": "image_url", "image_url": {"url": data_url}},
                                ],
                            }
                        ],
                        "temperature": 0,
                        "max_tokens": 300,
                        "response_format": {"type": "json_object"},
                    }

                    try:
                        _sleep_before_llm_request_sync()
                        resp = client.chat.completions.create(**kwargs)
                    except Exception as e:
                        em = str(e).lower()
                        if "json mode is not enabled" in em or "response_format" in em:
                            kwargs.pop("response_format", None)
                            _sleep_before_llm_request_sync()
                            resp = client.chat.completions.create(**kwargs)
                        else:
                            raise

                    choices = getattr(resp, "choices", None) or []
                    message = choices[0].message if choices else None
                    content = (getattr(message, "content", None) or "").strip()
                    if not content:
                        raise ValueError("empty model content")

                    payload = _load_json_object_loose(content)
                    doc_type = _normalize_doc_type(payload.get("document_type"))
                    if not doc_type:
                        continue
                    try:
                        conf = float(payload.get("confidence") or 0)
                    except Exception:
                        conf = 0.0
                    return {
                        "provider": "mistral",
                        "model": model,
                        "document_type": doc_type,
                        "confidence": max(0.0, min(conf, 0.99)),
                        "reason": str(payload.get("reason") or "").strip(),
                    }
                except Exception as e:
                    err_text = str(e)
                    if _is_mistral_rate_limit_error(err_text):
                        log.warning("Mistral classifier rate-limited (%s): %s", model, err_text[:220])
                    else:
                        non_rate_failures.append(f"{model}: {str(e)[:160]}")
            if non_rate_failures:
                log.debug("Mistral classifier non-rate-limit failures: %s", " | ".join(non_rate_failures[:2]))
            return None
        except Exception as e:
            log.warning("Mistral classifier init failed: %s", e)
            return None
        finally:
            if compact_temp_path:
                try:
                    Path(compact_temp_path).unlink(missing_ok=True)
                except Exception:
                    pass

    def _openrouter_vision_classify(prompt: str, image_path: str, mime: str) -> Optional[dict]:
        load_dotenv()
        api_key = (os.getenv("OPENROUTER_API_KEY") or "").strip()
        if not api_key:
            return None
        free_only = _openrouter_free_only_enabled()
        if free_only and _openrouter_free_temporarily_blocked():
            return None
        compact_temp_path: Optional[str] = None
        try:
            from openai import OpenAI
            data_url, _mime, compact_temp_path = _build_data_url_for_llm(image_path)

            configured = (os.getenv("OPENROUTER_CLASSIFIER_MODEL") or "").strip()
            configured_items = [s for s in [configured, (os.getenv("OPENROUTER_MODEL") or "").strip()] if s]
            model_candidates = _openrouter_resolve_model_candidates(
                api_key,
                configured_items=configured_items,
                include_paid=(not free_only),
            )
            if free_only:
                model_candidates = [m for m in model_candidates if m.endswith(":free")]
            if not model_candidates:
                return None

            client = OpenAI(base_url=_openrouter_api_base_url(), api_key=api_key, max_retries=0)
            non_rate_failures: List[str] = []
            for model in model_candidates:
                try:
                    kwargs = {
                        "model": model,
                        "messages": [
                            {
                                "role": "user",
                                "content": [
                                    {"type": "text", "text": prompt},
                                    {"type": "image_url", "image_url": {"url": data_url}},
                                ],
                            }
                        ],
                        "temperature": 0,
                        "max_tokens": 300,
                        "response_format": {"type": "json_object"},
                    }
                    try:
                        _sleep_before_llm_request_sync()
                        resp = client.chat.completions.create(**kwargs)
                    except Exception as e:
                        em = str(e).lower()
                        if "json mode is not enabled" in em or "response_format" in em:
                            kwargs.pop("response_format", None)
                            _sleep_before_llm_request_sync()
                            resp = client.chat.completions.create(**kwargs)
                        else:
                            raise

                    choices = getattr(resp, "choices", None) or []
                    message = choices[0].message if choices else None
                    content = (getattr(message, "content", None) or "").strip()
                    if not content:
                        raise ValueError("empty model content")

                    payload = _load_json_object_loose(content)
                    doc_type = _normalize_doc_type(payload.get("document_type"))
                    if not doc_type:
                        continue
                    try:
                        conf = float(payload.get("confidence") or 0)
                    except Exception:
                        conf = 0.0
                    return {
                        "provider": "openrouter",
                        "model": model,
                        "document_type": doc_type,
                        "confidence": max(0.0, min(conf, 0.99)),
                        "reason": str(payload.get("reason") or "").strip(),
                    }
                except Exception as e:
                    err_text = str(e)
                    if _is_openrouter_free_quota_error(err_text):
                        _mark_openrouter_free_quota_exhausted(err_text)
                        return None
                    non_rate_failures.append(f"{model}: {str(e)[:160]}")
            if non_rate_failures:
                log.debug("OpenRouter classifier non-rate-limit failures: %s", " | ".join(non_rate_failures[:2]))
            return None
        except Exception as e:
            log.warning("OpenRouter classifier init failed: %s", e)
            return None
        finally:
            if compact_temp_path:
                try:
                    Path(compact_temp_path).unlink(missing_ok=True)
                except Exception:
                    pass

    def _gemini_vision_classify(prompt: str, file_to_use: str, mime: str) -> Optional[dict]:
        try:
            from database import get_settings
            settings = get_settings()
            api_key = (settings.gemini_api_key or "").strip()
            if not api_key:
                return None

            model = "gemini-2.5-flash"
            src = Path(file_to_use)
            if not src.exists():
                log.warning("Gemini classifier skipped missing file: %s", file_to_use)
                return None
            blob = src.read_bytes()
            text_out = ""
            
            try:
                from google import genai as google_genai
                from google.genai import types
                client = google_genai.Client(api_key=api_key)
                _sleep_before_llm_request_sync()
                resp = client.models.generate_content(
                    model=model,
                    contents=[prompt, types.Part.from_bytes(data=blob, mime_type=mime)]
                )
                text_out = (getattr(resp, "text", None) or "").strip()
            except Exception:
                try:
                    import google.generativeai as legacy_genai
                    legacy_genai.configure(api_key=api_key)
                    model_obj = legacy_genai.GenerativeModel(model)
                    _sleep_before_llm_request_sync()
                    resp = model_obj.generate_content([prompt, {"mime_type": mime, "data": blob}])
                    text_out = (getattr(resp, "text", None) or "").strip()
                    if not text_out:
                        parts_text: List[str] = []
                        for cand in (getattr(resp, "candidates", None) or []):
                            content = getattr(cand, "content", None)
                            for part in (getattr(content, "parts", None) or []):
                                ptxt = getattr(part, "text", None)
                                if ptxt:
                                    parts_text.append(str(ptxt))
                        text_out = "".join(parts_text).strip()
                except Exception as e2:
                    log.warning("Gemini classifier failed: %s", e2)
                    return None

            if not text_out:
                return None

            payload = _load_json_object_loose(text_out)
            doc_type = _normalize_doc_type(payload.get("document_type"))
            if not doc_type:
                return None
            try:
                conf = float(payload.get("confidence") or 0)
            except Exception:
                conf = 0.0
            return {
                "provider": "gemini",
                "model": model,
                "document_type": doc_type,
                "confidence": max(0.0, min(conf, 0.99)),
                "reason": str(payload.get("reason") or "").strip(),
            }
        except Exception as e:
            log.warning("Gemini classifier init failed: %s", e)
            return None

    def _github_vision_classify(prompt: str, image_path: str, mime: str) -> Optional[dict]:
        load_dotenv()
        api_key = (os.getenv("GITHUB_TOKEN") or os.getenv("GITHUB_MODELS_API_KEY") or "").strip()
        if not api_key:
            return None
        compact_temp_path: Optional[str] = None
        try:
            from openai import OpenAI

            data_url, _mime, compact_temp_path = _build_data_url_for_llm(image_path)

            base = (os.getenv("GITHUB_MODELS_BASE_URL") or "https://models.github.ai/inference").rstrip("/")
            configured = (os.getenv("GITHUB_CLASSIFIER_MODEL") or os.getenv("GITHUB_MODEL") or "").strip()
            model_candidates = [
                configured,
                "openai/gpt-4o-mini",
            ]
            model_candidates = [m for i, m in enumerate(model_candidates) if m and m not in model_candidates[:i]]

            client = OpenAI(base_url=base, api_key=api_key, max_retries=0)
            for model in model_candidates:
                try:
                    kwargs = {
                        "model": model,
                        "messages": [
                            {
                                "role": "user",
                                "content": [
                                    {"type": "text", "text": prompt},
                                    {"type": "image_url", "image_url": {"url": data_url}},
                                ],
                            }
                        ],
                        "temperature": 0,
                        "max_tokens": 300,
                        "response_format": {"type": "json_object"},
                    }
                    try:
                        _sleep_before_llm_request_sync()
                        resp = client.chat.completions.create(**kwargs)
                    except Exception as e:
                        em = str(e).lower()
                        if "json mode is not enabled" in em or "response_format" in em:
                            kwargs.pop("response_format", None)
                            _sleep_before_llm_request_sync()
                            resp = client.chat.completions.create(**kwargs)
                        else:
                            raise

                    choices = getattr(resp, "choices", None) or []
                    message = choices[0].message if choices else None
                    content = (getattr(message, "content", None) or "").strip()
                    if not content:
                        raise ValueError("empty model content")

                    payload = _load_json_object_loose(content)
                    doc_type = _normalize_doc_type(payload.get("document_type"))
                    if not doc_type:
                        continue
                    try:
                        conf = float(payload.get("confidence") or 0)
                    except Exception:
                        conf = 0.0
                    return {
                        "provider": "github",
                        "model": model,
                        "document_type": doc_type,
                        "confidence": max(0.0, min(conf, 0.99)),
                        "reason": str(payload.get("reason") or "").strip(),
                    }
                except Exception as e:
                    log.warning("GitHub classifier failed (%s): %s", model, e)
            return None
        except Exception as e:
            log.warning("GitHub classifier init failed: %s", e)
            return None
        finally:
            if compact_temp_path:
                try:
                    Path(compact_temp_path).unlink(missing_ok=True)
                except Exception:
                    pass

    name = (file_name or Path(file_path).name or "").strip()
    name_text = f"{name} {str(file_path)}"
    prompt = _vision_classifier_prompt(name_text)

    # Convert PDF to image if necessary (Vision models usually expect images)
    vision_file_path = file_path
    cleanup_temp = False
    
    if Path(file_path).suffix.lower() == ".pdf":
        converted = _convert_pdf_first_page_to_image(file_path)
        if converted:
            vision_file_path = converted
            cleanup_temp = True

    try:
        mime = _mime_type_for_path(vision_file_path)
        llm_candidates: List[dict] = []
        
        min_accept_conf = 0.62

        # 1. Groq Vision
        groq_res = _groq_vision_classify(prompt, vision_file_path, mime)
        if groq_res:
            llm_candidates.append(groq_res)
            conf = float(groq_res.get("confidence") or 0)
            if groq_res.get("document_type") == "not_classified" and conf >= 0.45:
                return "not_classified", round(conf, 3), llm_candidates
            if conf >= min_accept_conf:
                return groq_res["document_type"], round(float(groq_res.get("confidence")), 3), llm_candidates

        # 2. Mistral Vision
        mistral_res = _mistral_vision_classify(prompt, vision_file_path, mime)
        if mistral_res:
            llm_candidates.append(mistral_res)
            conf = float(mistral_res.get("confidence") or 0)
            if mistral_res.get("document_type") == "not_classified" and conf >= 0.45:
                return "not_classified", round(conf, 3), llm_candidates
            if conf >= min_accept_conf:
                return mistral_res["document_type"], round(float(mistral_res.get("confidence")), 3), llm_candidates

        # 3. OpenRouter Vision
        or_res = _openrouter_vision_classify(prompt, vision_file_path, mime)
        if or_res:
            llm_candidates.append(or_res)
            conf = float(or_res.get("confidence") or 0)
            if or_res.get("document_type") == "not_classified" and conf >= 0.45:
                return "not_classified", round(conf, 3), llm_candidates
            if conf >= min_accept_conf:
                return or_res["document_type"], round(float(or_res.get("confidence")), 3), llm_candidates

        # 4. Gemini Vision
        gemini_res = _gemini_vision_classify(prompt, vision_file_path, mime)
        if gemini_res:
            llm_candidates.append(gemini_res)
            if float(gemini_res.get("confidence") or 0) >= 0.55:
                return gemini_res["document_type"], round(float(gemini_res.get("confidence")), 3), llm_candidates

        # 5. GitHub Vision
        github_res = _github_vision_classify(prompt, vision_file_path, mime)
        if github_res:
            llm_candidates.append(github_res)
            if float(github_res.get("confidence") or 0) >= 0.55:
                return github_res["document_type"], round(float(github_res.get("confidence")), 3), llm_candidates

        if llm_candidates:
            best = max(llm_candidates, key=lambda c: float(c.get("confidence") or 0))
            best_conf = max(0.0, min(float(best.get("confidence") or 0), 0.99))
            best_type = str(best.get("document_type") or "").strip().lower()
            if best_type == "not_classified":
                return "not_classified", round(max(best_conf, 0.35), 3), llm_candidates
            if best_conf >= min_accept_conf:
                return str(best_type or "purchase_bill"), round(best_conf, 3), llm_candidates
            # Keep low-confidence classifications out of operational queue.
            return "not_classified", round(max(best_conf, 0.35), 3), llm_candidates

        # Final fallback - keep non-target/unclear docs as not_classified.
        return "not_classified", 0.0, [{"provider": "fallback", "document_type": "not_classified", "confidence": 0.0, "reason": "All vision classifiers failed"}]
    finally:
        if cleanup_temp and vision_file_path:
            try:
                Path(vision_file_path).unlink(missing_ok=True)
            except Exception:
                pass


def _to_qtl(value, unit_hint: Optional[str] = None) -> Optional[float]:
    if value is None:
        return None
    try:
        val = float(str(value).replace(",", "").strip())
    except Exception:
        return None
    u = (unit_hint or "").lower()
    if "kg" in u:
        return round(val / 100, 3)
    if "mt" in u or "ton" in u:
        return round(val * 10, 3)
    return round(val, 3)


async def _extract_with_groq_prompt(file_path: str, prompt_text: str, source_tag: str) -> dict:
    load_dotenv()
    api_key = os.getenv("GROQ_API_KEY")
    if not api_key:
        return {"source": source_tag, "error": "GROQ_API_KEY not set", "high_confidence": False}

    groq_input_path: Optional[str] = None
    cleanup_temp = False
    compact_temp_path: Optional[str] = None
    try:
        from groq import Groq

        src = Path(file_path)
        if src.suffix.lower() == ".pdf":
            converted = _convert_pdf_first_page_to_image(file_path)
            if not converted:
                return {
                    "source": source_tag,
                    "error": "PDF->image conversion failed",
                    "high_confidence": False,
                }
            groq_input_path = converted
            cleanup_temp = True
        else:
            groq_input_path = file_path

        data_url, _mime, compact_temp_path = _build_data_url_for_llm(groq_input_path)

        client = Groq(api_key=api_key)
        await _sleep_before_llm_request_async()
        response = client.chat.completions.create(
            messages=[
                {
                    "role": "user",
                    "content": [
                        {"type": "text", "text": f"{prompt_text}\n\nReturn only one valid JSON object with no extra text."},
                        {"type": "image_url", "image_url": {"url": data_url}},
                    ],
                }
            ],
            model="llama-3.3-70b-versatile",
            temperature=0.1,
            response_format={"type": "json_object"},
        )

        raw = (response.choices[0].message.content or "").strip()
        payload = _load_json_object_loose(raw)
        if not isinstance(payload, dict):
            payload = {}
        payload["source"] = source_tag
        conf = float(payload.get("confidence") or 0)
        payload["high_confidence"] = conf >= 0.7
        return payload
    except Exception as e:
        log.warning("Groq extraction failed for %s: %s", source_tag, e)
        return {"source": source_tag, "error": str(e), "high_confidence": False}
    finally:
        if compact_temp_path:
            try:
                Path(compact_temp_path).unlink(missing_ok=True)
            except Exception:
                pass
        if cleanup_temp and groq_input_path:
            try:
                Path(groq_input_path).unlink(missing_ok=True)
            except Exception:
                pass


async def _extract_with_mistral_prompt(file_path: str, prompt_text: str, source_tag: str) -> dict:
    try:
        from openai import OpenAI
    except Exception as e:
        return {"source": source_tag, "error": f"openai import failed: {e}", "high_confidence": False}

    load_dotenv()
    if not _mistral_enabled():
        return {"source": source_tag, "error": "MISTRAL_ENABLED is false", "high_confidence": False}
    api_key = (os.getenv("MISTRAL_API_KEY") or "").strip()
    if not api_key:
        return {"source": source_tag, "error": "MISTRAL_API_KEY not set", "high_confidence": False}

    mistral_input_path: Optional[str] = None
    cleanup_temp = False
    compact_temp_path: Optional[str] = None
    try:
        src = Path(file_path)
        if src.suffix.lower() == ".pdf":
            converted = _convert_pdf_first_page_to_image(file_path)
            if not converted:
                return {"source": source_tag, "error": "PDF->image conversion failed", "high_confidence": False}
            mistral_input_path = converted
            cleanup_temp = True
        else:
            mistral_input_path = file_path

        data_url, _mime, compact_temp_path = _build_data_url_for_llm(mistral_input_path)

        configured_raw = (os.getenv("MISTRAL_MODEL") or "").strip()
        configured_items = [s.strip() for s in re.split(r"[,\s]+", configured_raw) if s.strip()]
        model_candidates = _mistral_resolve_model_candidates(configured_items=configured_items)
        client = OpenAI(base_url=_mistral_api_base_url(), api_key=api_key, max_retries=0)

        last_err = None
        for model in model_candidates:
            try:
                request_kwargs = {
                    "model": model,
                    "messages": [
                        {
                            "role": "user",
                            "content": [
                                {"type": "text", "text": f"{prompt_text}\n\nReturn only one valid JSON object with no extra text."},
                                {"type": "image_url", "image_url": {"url": data_url}},
                            ],
                        }
                    ],
                    "temperature": 0.1,
                    "max_tokens": 2000,
                    "response_format": {"type": "json_object"},
                }

                try:
                    await _sleep_before_llm_request_async()
                    response = client.chat.completions.create(**request_kwargs)
                except Exception as e:
                    em = str(e).lower()
                    if "json mode is not enabled" in em or "response_format" in em:
                        request_kwargs.pop("response_format", None)
                        await _sleep_before_llm_request_async()
                        response = client.chat.completions.create(**request_kwargs)
                    else:
                        raise

                raw = (response.choices[0].message.content or "").strip()
                payload = _load_json_object_loose(raw)
                if not isinstance(payload, dict):
                    payload = {}
                payload["source"] = f"{source_tag}:{model}"
                c = float(payload.get("confidence") or 0)
                payload["high_confidence"] = c >= 0.7
                return payload
            except Exception as e:
                last_err = e
                if _is_mistral_rate_limit_error(str(e)):
                    log.warning("Mistral prompt-chain rate-limited (%s): %s", model, str(e)[:220])
                continue
        return {"source": source_tag, "error": str(last_err), "high_confidence": False}
    except Exception as e:
        return {"source": source_tag, "error": str(e), "high_confidence": False}
    finally:
        if compact_temp_path:
            try:
                Path(compact_temp_path).unlink(missing_ok=True)
            except Exception:
                pass
        if cleanup_temp and mistral_input_path:
            try:
                Path(mistral_input_path).unlink(missing_ok=True)
            except Exception:
                pass


def _signal_score_for_doc(doc_type: str, payload: dict) -> float:
    if not isinstance(payload, dict):
        return 0.0
    conf = 0.0
    try:
        conf = float(payload.get("confidence") or 0)
    except Exception:
        conf = 0.0

    if doc_type == "purchase_bill":
        return float(_count_bill_fields(payload)) + conf
    if doc_type == "purchase_order":
        score = 0.0
        if payload.get("po_number"):
            score += 2.0
        if isinstance(payload.get("items"), list) and payload.get("items"):
            score += 2.0
        if payload.get("winner_party_name") or payload.get("buyer_name"):
            score += 1.0
        return score + conf
    if doc_type == "tender_notice":
        score = 0.0
        if payload.get("tender_rm_number") or payload.get("tender_number"):
            score += 2.0
        if payload.get("plant_name"):
            score += 1.0
        items = payload.get("items") if isinstance(payload.get("items"), list) else []
        if items:
            score += 1.0
        return score + conf
    if doc_type == "rejection_notice":
        score = 0.0
        if payload.get("vehicle_number") or payload.get("truck_number"):
            score += 2.0
        if payload.get("reason"):
            score += 1.0
        if payload.get("rejected_qty_qtl") not in (None, "", "null", "None"):
            score += 1.0
        return score + conf
    if doc_type == "plant_unloading":
        rows = payload.get("rows") if isinstance(payload.get("rows"), list) else []
        return min(len(rows), 5) + conf
    return conf


def _doc_has_signal(doc_type: str, payload: dict) -> bool:
    return _signal_score_for_doc(doc_type, payload) >= 2.0


def _looks_handwritten_hint(file_path: str) -> bool:
    load_dotenv()
    force_hw = (os.getenv("DOC_FORCE_HANDWRITTEN", "").strip().lower() in {"1", "true", "yes", "on"})
    force_non_hw = (os.getenv("DOC_FORCE_NON_HANDWRITTEN", "").strip().lower() in {"1", "true", "yes", "on"})
    if force_hw:
        return True
    if force_non_hw:
        return False
    name = str(Path(file_path).name or "").lower()
    return any(token in name for token in ["hand", "handwritten", "manual"])


async def _extract_with_openrouter_prompt(file_path: str, prompt_text: str, source_tag: str) -> dict:
    try:
        from openai import OpenAI
    except Exception as e:
        return {"source": source_tag, "error": f"openai import failed: {e}", "high_confidence": False}

    load_dotenv()
    api_key = (os.getenv("OPENROUTER_API_KEY") or "").strip()
    if not api_key:
        return {"source": source_tag, "error": "OPENROUTER_API_KEY not set", "high_confidence": False}

    free_only = _openrouter_free_only_enabled()
    if free_only and _openrouter_free_temporarily_blocked():
        retry_after = _openrouter_retry_after_sec()
        return {
            "source": source_tag,
            "error": f"OpenRouter free quota cooldown active; retry in ~{retry_after}s",
            "high_confidence": False,
            "rate_limited": True,
            "retry_after_sec": retry_after,
        }

    openrouter_input_path: Optional[str] = None
    cleanup_temp = False
    compact_temp_path: Optional[str] = None
    try:
        src = Path(file_path)
        if src.suffix.lower() == ".pdf":
            converted = _convert_pdf_first_page_to_image(file_path)
            if not converted:
                return {"source": source_tag, "error": "PDF->image conversion failed", "high_confidence": False}
            openrouter_input_path = converted
            cleanup_temp = True
        else:
            openrouter_input_path = file_path

        data_url, _mime, compact_temp_path = _build_data_url_for_llm(openrouter_input_path)

        client = OpenAI(base_url=_openrouter_api_base_url(), api_key=api_key, max_retries=0)
        configured_items = [os.getenv("OPENROUTER_MODEL", "").strip()]
        model_candidates = _openrouter_resolve_model_candidates(
            api_key,
            configured_items=configured_items,
            include_paid=(not free_only),
        )
        if free_only:
            model_candidates = [m for m in model_candidates if m.endswith(":free")]
        if not model_candidates:
            return {"source": source_tag, "error": "No OpenRouter free model candidate configured", "high_confidence": False}

        last_err = None
        non_rate_failures: List[str] = []
        for model in model_candidates:
            try:
                request_kwargs = {
                    "model": model,
                    "messages": [
                        {
                            "role": "user",
                            "content": [
                                {"type": "text", "text": f"{prompt_text}\n\nReturn only one valid JSON object with no extra text."},
                                {"type": "image_url", "image_url": {"url": data_url}},
                            ],
                        }
                    ],
                    "temperature": 0.1,
                    "max_tokens": 2000,
                    "response_format": {"type": "json_object"},
                }

                try:
                    await _sleep_before_llm_request_async()
                    response = client.chat.completions.create(**request_kwargs)
                except Exception as e:
                    em = str(e).lower()
                    if "json mode is not enabled" in em or "response_format" in em:
                        request_kwargs.pop("response_format", None)
                        await _sleep_before_llm_request_async()
                        response = client.chat.completions.create(**request_kwargs)
                    else:
                        raise

                choices = getattr(response, "choices", None) or []
                message = choices[0].message if choices else None
                raw = (getattr(message, "content", None) or "").strip()
                if not raw:
                    raise ValueError("empty model content")
                payload = _load_json_object_loose(raw)
                if not isinstance(payload, dict):
                    payload = {}
                payload["source"] = f"{source_tag}:{model}"
                c = float(payload.get("confidence") or 0)
                payload["high_confidence"] = c >= 0.7
                return payload
            except Exception as e:
                err_text = str(e)
                if _is_openrouter_free_quota_error(err_text):
                    _mark_openrouter_free_quota_exhausted(err_text)
                    return {
                        "source": source_tag,
                        "error": err_text,
                        "high_confidence": False,
                        "rate_limited": True,
                        "openrouter_model": model,
                    }
                non_rate_failures.append(f"{model}: {str(e)[:180]}")
                last_err = e
                continue
        if non_rate_failures:
            log.debug("OpenRouter prompt-chain non-rate-limit failures: %s", " | ".join(non_rate_failures[:2]))
        return {"source": source_tag, "error": str(last_err), "high_confidence": False}
    except Exception as e:
        return {"source": source_tag, "error": str(e), "high_confidence": False}
    finally:
        if compact_temp_path:
            try:
                Path(compact_temp_path).unlink(missing_ok=True)
            except Exception:
                pass
        if cleanup_temp and openrouter_input_path:
            try:
                Path(openrouter_input_path).unlink(missing_ok=True)
            except Exception:
                pass


async def _extract_with_gemini_prompt(file_path: str, prompt_text: str, source_tag: str) -> dict:
    from database import get_settings
    settings = get_settings()
    if not settings.gemini_api_key:
        return {"source": source_tag, "error": "GEMINI_API_KEY not set", "high_confidence": False}

    gemini_input_path = file_path
    cleanup_temp = False
    try:
        src = Path(file_path)
        if src.suffix.lower() == ".pdf":
            converted = _convert_pdf_first_page_to_image(file_path)
            if converted:
                gemini_input_path = converted
                cleanup_temp = True

        src_gemini = Path(gemini_input_path)
        if not src_gemini.exists():
            return {"source": source_tag, "error": f"input file not found: {gemini_input_path}", "high_confidence": False}

        mime = _mime_type_for_path(gemini_input_path)
        blob = src_gemini.read_bytes()
        await _sleep_before_llm_request_async()
        text = _gemini_generate_text(
            api_key=settings.gemini_api_key,
            model="gemini-2.5-flash",
            prompt_text=f"{prompt_text}\n\nReturn only one valid JSON object with no extra text.",
            blob=blob,
            mime=mime,
        )
        text = re.sub(r"^```json\s*", "", text)
        text = re.sub(r"\s*```$", "", text)
        payload = _load_json_object_loose(text)
        if not isinstance(payload, dict):
            payload = {}
        payload["source"] = source_tag
        c = float(payload.get("confidence") or 0)
        payload["high_confidence"] = c >= 0.7
        return payload
    except Exception as e:
        return {"source": source_tag, "error": str(e), "high_confidence": False}
    finally:
        if cleanup_temp and gemini_input_path != file_path:
            try:
                Path(gemini_input_path).unlink(missing_ok=True)
            except Exception:
                pass


async def _extract_with_github_prompt(file_path: str, prompt_text: str, source_tag: str) -> dict:
    try:
        from openai import OpenAI
    except Exception as e:
        return {"source": source_tag, "error": f"openai import failed: {e}", "high_confidence": False}

    load_dotenv()
    api_key = (os.getenv("GITHUB_TOKEN") or os.getenv("GITHUB_MODELS_API_KEY") or "").strip()
    if not api_key:
        return {"source": source_tag, "error": "GITHUB token not set", "high_confidence": False}

    gh_input_path: Optional[str] = None
    cleanup_temp = False
    compact_temp_path: Optional[str] = None
    try:
        src = Path(file_path)
        if src.suffix.lower() == ".pdf":
            converted = _convert_pdf_first_page_to_image(file_path)
            if not converted:
                return {"source": source_tag, "error": "PDF->image conversion failed", "high_confidence": False}
            gh_input_path = converted
            cleanup_temp = True
        else:
            gh_input_path = file_path

        data_url, _mime, compact_temp_path = _build_data_url_for_llm(gh_input_path)

        base = (os.getenv("GITHUB_MODELS_BASE_URL") or "https://models.github.ai/inference").rstrip("/")
        model_candidates = [
            (os.getenv("GITHUB_MODEL") or "openai/gpt-4o-mini").strip(),
            "openai/gpt-4o-mini",
        ]
        model_candidates = [m for i, m in enumerate(model_candidates) if m and m not in model_candidates[:i]]
        client = OpenAI(base_url=base, api_key=api_key, max_retries=0)

        last_err = None
        for model in model_candidates:
            try:
                await _sleep_before_llm_request_async()
                response = client.chat.completions.create(
                    model=model,
                    messages=[
                        {
                            "role": "user",
                            "content": [
                                {"type": "text", "text": f"{prompt_text}\n\nReturn only one valid JSON object with no extra text."},
                                {"type": "image_url", "image_url": {"url": data_url}},
                            ],
                        }
                    ],
                    temperature=0.1,
                    max_tokens=2000,
                    response_format={"type": "json_object"},
                )
                raw = (response.choices[0].message.content or "").strip()
                payload = _load_json_object_loose(raw)
                if not isinstance(payload, dict):
                    payload = {}
                payload["source"] = f"{source_tag}:{model}"
                c = float(payload.get("confidence") or 0)
                payload["high_confidence"] = c >= 0.7
                return payload
            except Exception as e:
                last_err = e
                continue
        return {"source": source_tag, "error": str(last_err), "high_confidence": False}
    except Exception as e:
        return {"source": source_tag, "error": str(e), "high_confidence": False}
    finally:
        if compact_temp_path:
            try:
                Path(compact_temp_path).unlink(missing_ok=True)
            except Exception:
                pass
        if cleanup_temp and gh_input_path:
            try:
                Path(gh_input_path).unlink(missing_ok=True)
            except Exception:
                pass


def _extract_from_text_simple(doc_type: str, text: str, source_tag: str) -> dict:
    txt = str(text or "")
    out = {"source": source_tag, "high_confidence": False}
    if doc_type == "purchase_bill":
        parsed = _parse_bill_text(txt)
        parsed["source"] = source_tag
        parsed["raw_text"] = txt
        parsed["high_confidence"] = parsed.get("fields_found", 0) >= 4
        return parsed
    if doc_type == "tender_notice":
        nit = re.search(r"(?:NIT|RM)[-\s]*([A-Z0-9\-/]+)", txt, re.IGNORECASE)
        mats = []
        for m in ["Maize", "Dorb", "Domc", "Rice DDGS"]:
            if m.lower() in txt.lower() or (m == "Rice DDGS" and "ddgs" in txt.lower()):
                mats.append({"material_type": m, "quantity_qtl": None})
        out.update({
            "document_type": "tender_notice",
            "tender_rm_number": nit.group(0).strip() if nit else None,
            "tender_number": nit.group(0).strip() if nit else None,
            "notice_date": None,
            "supply_period_start": None,
            "supply_period_end": None,
            "items": mats,
            "raw_text": txt,
        })
        out["high_confidence"] = bool(out.get("tender_rm_number"))
        return out
    if doc_type == "purchase_order":
        po = re.search(r"(?:PO\s*(?:No\.?|Number)?|Purchase\s*Order)\s*[:#\- ]*([A-Z0-9\-/]+)", txt, re.IGNORECASE)
        rm = re.search(r"(?:RM\s*(?:No\.?|Number)?|Tender\s*(?:No\.?|Number)?)\s*[:#\- ]*([A-Z0-9\-/]+)", txt, re.IGNORECASE)
        out.update({
            "document_type": "purchase_order",
            "po_number": po.group(1).strip() if po else None,
            "tender_rm_number": rm.group(1).strip() if rm else None,
            "po_date": None,
            "supply_period_start": None,
            "supply_period_end": None,
            "winner_party_name": None,
            "buyer_name": None,
            "seller_name": None,
            "items": [],
            "raw_text": txt,
        })
        out["high_confidence"] = bool(out.get("po_number"))
        return out
    if doc_type == "rejection_notice":
        veh = VEHICLE_RE.search(txt)
        out.update({
            "document_type": "rejection_notice",
            "truck_number": veh.group(0).replace(" ", "").upper() if veh else None,
            "vehicle_number": veh.group(0).replace(" ", "").upper() if veh else None,
            "rejection_date": None,
            "rejection_type": "partial" if re.search(r"partial", txt, re.IGNORECASE) else ("complete" if re.search(r"complete", txt, re.IGNORECASE) else None),
            "reason": "Rejection" if re.search(r"reject", txt, re.IGNORECASE) else None,
            "rejected_qty_qtl": None,
            "raw_text": txt,
        })
        out["high_confidence"] = bool(out.get("truck_number") or out.get("reason"))
        return out
    return out


def _normalize_tender_notice_result(result: dict) -> dict:
    if not isinstance(result, dict):
        return {}
    out = dict(result)
    out["tender_rm_number"] = _pick_first(out, ["tender_rm_number", "tender_number", "rm_number", "nit_number"])
    if out.get("tender_number") in (None, "", "null", "None") and out.get("tender_rm_number"):
        out["tender_number"] = out.get("tender_rm_number")

    out["supply_period_start"] = _normalize_date_text(_pick_first(out, ["supply_period_start", "start_date", "from_date", "supply_from"]))
    out["supply_period_end"] = _normalize_date_text(_pick_first(out, ["supply_period_end", "end_date", "to_date", "supply_to", "deadline_date"]))

    items = out.get("items")
    if not isinstance(items, list) or not items:
        mat_list = out.get("materials") if isinstance(out.get("materials"), list) else []
        items = [{"material_type": m, "quantity_qtl": None} for m in mat_list if str(m or "").strip()]

    normalized_items = []
    for item in items:
        if not isinstance(item, dict):
            continue
        qty_raw = _pick_first(item, ["quantity_qtl", "quantity", "qty", "approved_quantity_qtl"])
        qty_unit = _pick_first(item, ["quantity_unit", "unit"]) or ""
        normalized_items.append({
            "material_type": _pick_first(item, ["material_type", "material", "item"]),
            "quantity_qtl": _to_qtl(qty_raw, str(qty_unit)),
        })
    out["items"] = normalized_items
    return out


def _normalize_party_key(name: Optional[str]) -> str:
    s = re.sub(r"[^a-z0-9]+", " ", str(name or "").lower())
    return re.sub(r"\s+", " ", s).strip()


def _canonical_our_company_name(name: Optional[str]) -> Optional[str]:
    key = _normalize_party_key(name)
    if not key:
        return None

    # Accept common spacing/spelling variations, including 'shri' vs 'shree'.
    if "vinayak" in key and ("shree" in key or "shri" in key):
        return "Shree Vinayak Trading Company"
    if "nath" in key and ("shree" in key or "shri" in key):
        return "Shree Nath Industries"
    if "ganpati" in key and ("shree" in key or "shri" in key):
        return "Shree Ganpati Enterpriese"
    return None


def _build_po_sub_tenders(out: dict, normalized_items: List[dict]) -> List[dict]:
    from datetime import date as _date, timedelta as _td

    tender_rm = str(out.get("tender_rm_number") or "").strip()
    base_plant = str(out.get("plant_name") or "").strip()

    start_raw = out.get("supply_period_start")
    end_raw = out.get("supply_period_end")

    start_dt: Optional[_date] = None
    end_dt: Optional[_date] = None
    try:
        if start_raw:
            start_dt = _date.fromisoformat(str(start_raw))
    except Exception:
        start_dt = None
    try:
        if end_raw:
            end_dt = _date.fromisoformat(str(end_raw))
    except Exception:
        end_dt = None

    rows: List[dict] = []
    for item in normalized_items:
        qty = _to_float(item.get("approved_quantity_qtl"))
        if qty is None or qty <= 0:
            continue

        plant = str(item.get("plant_name") or base_plant).strip()
        material = _pick_first(item, ["material_type", "material", "item"])
        sub_name = f"{tender_rm}{plant}" if tender_rm and plant else (tender_rm or plant or None)

        week1_deadline = None
        if start_dt:
            d = start_dt + _td(days=7)
            if end_dt and d > end_dt:
                d = end_dt
            week1_deadline = d.isoformat()

        rows.append({
            "sub_tender_name": sub_name,
            "plant_name": plant or None,
            "material_type": material,
            "tender_quantity_qtl": qty,
            "week1_target_qty_qtl": round(qty / 2.0, 3),
            "week1_deadline_date": week1_deadline,
            "week2_deadline_date": end_dt.isoformat() if end_dt else None,
            "cycle": "Any",
        })

    return rows


def _normalize_tender_rm_text(value: Optional[str]) -> Optional[str]:
    raw = str(value or "").strip()
    if not raw:
        return None
    m = re.search(r"\bRM\s*[- ]*([A-Z0-9]+(?:-[A-Z0-9]+)*)\b", raw, re.IGNORECASE)
    if m:
        return f"RM-{m.group(1).upper()}"
    m2 = re.search(r"\b([0-9]{2,4}(?:-[A-Z0-9]+)+)\b", raw, re.IGNORECASE)
    if m2:
        return f"RM-{m2.group(1).upper()}"
    return raw.upper()


def _normalize_material_name_text(value: Optional[str]) -> Optional[str]:
    v = _normalize_party_key(value).upper()
    if not v:
        return None
    if "RICE" in v and "DDGS" in v:
        return "Rice DDGS"
    if "RICE" in v and "BRAN" in v:
        return "Rice Bran"
    if "MAIZE" in v:
        return "Maize"
    if "DORB" in v:
        return "Dorb"
    if "DOMC" in v:
        return "Domc"
    if "GWAR" in v and "KORMA" in v:
        return "Gwar Korma"
    if "BY" in v and "PASS" in v and "FAT" in v:
        return "By Pass Fat"
    return str(value).strip() if value else None


def _extract_po_from_raw_text(raw_text: str) -> dict:
    txt = str(raw_text or "")
    flat = re.sub(r"\s+", " ", txt).strip()

    tender_rm = None
    m_rm = re.search(r"\bRM\s*[- ]*([A-Z0-9]+(?:-[A-Z0-9]+)*)\b", flat, re.IGNORECASE)
    if m_rm:
        tender_rm = f"RM-{m_rm.group(1).upper()}"

    po_date = None
    m_dated = re.search(r"\bDated\s*[:\-]?\s*(\d{1,2}[./-]\d{1,2}[./-]\d{2,4})", flat, re.IGNORECASE)
    if m_dated:
        po_date = _normalize_date_text(m_dated.group(1))

    supply_start = None
    supply_end = None
    m_period = re.search(
        r"supply\s*period[^.]{0,120}?from\s*(\d{1,2}[./-]\d{1,2}[./-]\d{2,4})[^0-9]{1,20}(?:to|upto|up\s*to)\s*(\d{1,2}[./-]\d{1,2}[./-]\d{2,4})",
        flat,
        re.IGNORECASE,
    )
    if m_period:
        supply_start = _normalize_date_text(m_period.group(1))
        supply_end = _normalize_date_text(m_period.group(2))

    plant_name = None
    m_plant = re.search(r"CATTLE\s+FEED\s+PLANT\s*,?\s*([A-Z][A-Z ]{2,30})", txt, re.IGNORECASE)
    if m_plant:
        plant_name = str(m_plant.group(1)).strip().title()

    material_matches = list(
        re.finditer(
            r"\b(DORB|DOMC|MAIZE|RICE\s*BRAN|RICE\s*DDGS|GWAR\s*KORMA|BY\s*PASS\s*FAT)\b",
            txt,
            re.IGNORECASE,
        )
    )
    qty_values: List[float] = []
    for qm in re.finditer(r"\b(\d{3,6}(?:\.\d{1,3})?)\b", txt):
        token = qm.group(1)
        if "." not in token:
            continue
        left = txt[max(0, qm.start() - 32): qm.start()].lower()
        # Skip supplier-rate numbers and obvious rate columns.
        if re.search(r"m/?s\.?", left):
            continue
        if re.search(r"\brate\b|rs\.?/?qtl|inclusive", left):
            continue
        try:
            val = float(token)
        except Exception:
            continue
        if 200 <= val <= 50000:
            qty_values.append(val)

    material_rows: List[dict] = []
    for idx, mm in enumerate(material_matches):
        material_rows.append(
            {
                "material_type": _normalize_material_name_text(mm.group(1)),
                "approved_quantity_qtl": qty_values[idx] if idx < len(qty_values) else None,
            }
        )

    party_matches = list(
        re.finditer(
            r"\bM/?s\.?\s*([A-Za-z][A-Za-z .&]{2,90}?)\s+(\d{3,6}(?:\.\d{1,3})?)\b",
            txt,
            re.IGNORECASE,
        )
    )

    items: List[dict] = []
    rows_count = len(material_rows)
    parties_count = len(party_matches)
    offset = max(0, rows_count - parties_count)

    for pidx, pm in enumerate(party_matches):
        party_raw = pm.group(1)
        canonical_party = _canonical_our_company_name(party_raw)
        if not canonical_party:
            continue

        try:
            rate_val = float(pm.group(2))
        except Exception:
            rate_val = None

        row_idx = min(max(0, pidx + offset), max(0, rows_count - 1)) if rows_count else None
        base_row = material_rows[row_idx] if row_idx is not None and rows_count else {}

        items.append(
            {
                "material_type": base_row.get("material_type"),
                "approved_quantity_qtl": _to_qtl(base_row.get("approved_quantity_qtl"), "Qtl"),
                "approved_rate_per_qtl": rate_val,
                "approved_party_name": canonical_party,
                "line_amount": None,
                "plant_name": plant_name,
            }
        )

    return {
        "tender_rm_number": tender_rm,
        "main_tender_name": tender_rm,
        "po_date": po_date,
        "supply_period_start": supply_start,
        "supply_period_end": supply_end,
        "plant_name": plant_name,
        "items": items,
    }


def _normalize_purchase_order_result(result: dict) -> dict:
    if not isinstance(result, dict):
        return {}
    out = dict(result)

    out["tender_rm_number"] = _normalize_tender_rm_text(_pick_first(out, ["tender_rm_number", "rm_number", "tender_number", "rm_ref"]))
    out["po_number"] = _pick_first(out, ["po_number", "po_no", "purchase_order_no"])
    if out.get("po_number") and not re.search(r"\d", str(out.get("po_number"))):
        out["po_number"] = None
    out["po_date"] = _normalize_date_text(_pick_first(out, ["po_date", "date", "order_date"]))
    out["supply_period_start"] = _normalize_date_text(_pick_first(out, ["supply_period_start", "supply_start", "from_date", "start_date", "period_start"]))
    out["supply_period_end"] = _normalize_date_text(_pick_first(out, ["supply_period_end", "supply_end", "to_date", "end_date", "period_end"]))
    out["main_tender_name"] = out.get("main_tender_name") or out.get("tender_rm_number")

    winner_name = _pick_first(out, ["winner_party_name", "approved_party_name", "awarded_to", "buyer_name", "company_name"])
    out["winner_party_name"] = winner_name
    out["winner_party_email"] = _pick_first(out, ["winner_party_email", "winner_email", "buyer_email", "email"])
    winner_canonical = _canonical_our_company_name(winner_name)

    items = out.get("items") if isinstance(out.get("items"), list) else []
    normalized_items: List[dict] = []
    for item in items:
        if not isinstance(item, dict):
            continue
        qty_raw = _pick_first(item, ["approved_quantity_qtl", "quantity_qtl", "quantity", "qty"])
        unit_hint = _pick_first(item, ["quantity_unit", "unit"]) or ""
        line_party = _pick_first(item, ["approved_party_name", "winner_party_name", "party_name"]) or winner_name
        line_party_canonical = _canonical_our_company_name(line_party)
        if not line_party_canonical:
            continue
        normalized_items.append({
            "material_type": _pick_first(item, ["material_type", "material", "item"]),
            "approved_quantity_qtl": _to_qtl(qty_raw, str(unit_hint)),
            "approved_rate_per_qtl": _to_float(_pick_first(item, ["approved_rate_per_qtl", "rate_per_qtl", "rate"])),
            "approved_party_name": line_party_canonical,
            "line_amount": _to_float(_pick_first(item, ["line_amount", "amount"])),
            "plant_name": _pick_first(item, ["plant_name", "destination_plant", "plant"]) or out.get("plant_name"),
        })

    if not normalized_items and out.get("raw_text"):
        parsed = _extract_po_from_raw_text(str(out.get("raw_text") or ""))
        if not out.get("tender_rm_number"):
            out["tender_rm_number"] = _normalize_tender_rm_text(parsed.get("tender_rm_number"))
        if not out.get("main_tender_name"):
            out["main_tender_name"] = parsed.get("main_tender_name")
        if not out.get("po_date"):
            out["po_date"] = parsed.get("po_date")
        if not out.get("supply_period_start"):
            out["supply_period_start"] = parsed.get("supply_period_start")
        if not out.get("supply_period_end"):
            out["supply_period_end"] = parsed.get("supply_period_end")
        if not out.get("plant_name"):
            out["plant_name"] = parsed.get("plant_name")
        parsed_items = parsed.get("items") if isinstance(parsed.get("items"), list) else []
        for item in parsed_items:
            if not isinstance(item, dict):
                continue
            canonical_party = _canonical_our_company_name(item.get("approved_party_name"))
            if not canonical_party:
                continue
            normalized_items.append({
                "material_type": _normalize_material_name_text(item.get("material_type")),
                "approved_quantity_qtl": _to_qtl(item.get("approved_quantity_qtl"), "Qtl"),
                "approved_rate_per_qtl": _to_float(item.get("approved_rate_per_qtl")),
                "approved_party_name": canonical_party,
                "line_amount": _to_float(item.get("line_amount")),
                "plant_name": item.get("plant_name") or out.get("plant_name"),
            })

    out["items"] = normalized_items
    out["our_company_winner_name"] = winner_canonical or (normalized_items[0]["approved_party_name"] if normalized_items else None)
    out["has_our_company_winner"] = bool(out.get("our_company_winner_name") or normalized_items)
    out["sub_tenders"] = _build_po_sub_tenders(out, normalized_items)
    return out


def _normalize_rejection_notice_result(result: dict) -> dict:
    if not isinstance(result, dict):
        return {}
    out = dict(result)
    truck = _pick_first(out, ["truck_number", "vehicle_number", "truck_no", "vehicle_no"])
    if isinstance(truck, str):
        v = VEHICLE_RE.search(truck)
        if v:
            truck = v.group().replace(" ", "").upper()
    out["truck_number"] = truck
    out["vehicle_number"] = truck

    out["rejection_date"] = _normalize_date_text(_pick_first(out, ["rejection_date", "notice_date", "bill_date", "date"]))
    out["rejection_type"] = _pick_first(out, ["rejection_type", "rejection_status", "status"])
    out["tender_rm_number"] = _pick_first(out, ["tender_rm_number", "rm_number", "tender_number", "rm_ref"])

    qty_raw = _pick_first(out, ["rejected_qty_qtl", "quantity_qtl", "weight_qtl", "weight"])
    unit_hint = _pick_first(out, ["quantity_unit", "weight_unit", "unit"]) or ""
    out["rejected_qty_qtl"] = _to_qtl(qty_raw, str(unit_hint))
    return out


async def _extract_with_azure_prompt(file_path: str, doc_type: str, source_tag: str) -> dict:
    txt = _extract_pdf_text_with_azure_read(file_path) or ""
    if not txt:
        return {"source": source_tag, "error": "Azure read text unavailable", "high_confidence": False}
    return _extract_from_text_simple(doc_type, txt, source_tag)


async def _extract_with_prompt_chain(file_path: str, prompt_text: str, doc_type: str, preferred_provider: Optional[str] = None) -> dict:
    handwritten = _looks_handwritten_hint(file_path)
    if handwritten:
        providers = [
            ("gemini", lambda: _extract_with_gemini_prompt(file_path, prompt_text, "gemini_prompt")),
            ("mistral", lambda: _extract_with_mistral_prompt(file_path, prompt_text, "mistral_prompt")),
            ("github", lambda: _extract_with_github_prompt(file_path, prompt_text, "github_prompt")),
        ]
    else:
        providers = [
            ("groq", lambda: _extract_with_groq_prompt(file_path, prompt_text, "groq_prompt")),
            ("mistral", lambda: _extract_with_mistral_prompt(file_path, prompt_text, "mistral_prompt")),
            ("openrouter", lambda: _extract_with_openrouter_prompt(file_path, prompt_text, "openrouter_prompt")),
            ("gemini", lambda: _extract_with_gemini_prompt(file_path, prompt_text, "gemini_prompt")),
            ("azure", lambda: _extract_with_azure_prompt(file_path, doc_type, "azure_text_prompt")),
            ("github", lambda: _extract_with_github_prompt(file_path, prompt_text, "github_prompt")),
        ]

    provider_key = str(preferred_provider or "").strip().lower()
    if provider_key in {"", "auto"}:
        provider_key = ""
    if provider_key:
        forced = [pair for pair in providers if pair[0] == provider_key]
        if not forced:
            return {
                "source": "prompt_chain",
                "error": f"Unsupported ocr_engine '{preferred_provider}' for {doc_type}",
                "provider_chain": [p for p, _ in providers],
                "provider_attempts": [],
                "high_confidence": False,
            }
        providers = forced

    def _payload_preview(payload: dict, limit: int = 300) -> str:
        if not isinstance(payload, dict):
            return str(payload)

        slim = {}
        for key in [
            "source",
            "document_type",
            "high_confidence",
            "confidence",
            "fields_found",
            "vendor_name",
            "vehicle_number",
            "bill_number",
            "po_number",
            "tender_rm_number",
            "truck_number",
            "material_type",
            "error",
        ]:
            val = payload.get(key)
            if val not in (None, "", [], {}):
                slim[key] = val

        if not slim:
            for key, val in payload.items():
                if key in {
                    "raw_text",
                    "ocr_raw_text",
                    "provider_errors",
                    "provider_attempts",
                    "rows",
                    "items",
                    "sub_tenders",
                    "classifier_candidates",
                }:
                    continue
                if val in (None, "", [], {}):
                    continue
                slim[key] = val
                if len(slim) >= 8:
                    break

        try:
            out = json.dumps(slim, ensure_ascii=False)
        except Exception:
            out = str(slim)

        if len(out) > max(40, int(limit)):
            out = out[: max(37, int(limit) - 3)] + "..."
        return out

    best = None
    best_score = -1.0
    errors: List[str] = []
    attempts: List[dict] = []

    for provider_name, fn in providers:
        try:
            payload = await fn()
        except Exception as e:
            err_text = str(e)
            errors.append(f"{provider_name}: {err_text}")
            attempts.append({
                "provider": provider_name,
                "source": provider_name,
                "status": "exception",
                "high_confidence": False,
                "signal_score": 0.0,
                "fields_found": None,
                "error": err_text[:320],
                "response_preview": "",
            })
            continue

        if not isinstance(payload, dict):
            payload = {"error": "invalid response", "source": provider_name, "high_confidence": False}
        payload.setdefault("source", provider_name)

        score = _signal_score_for_doc(doc_type, payload)
        is_high_conf = bool(payload.get("high_confidence"))
        has_signal = _doc_has_signal(doc_type, payload)
        error_text = str(payload.get("error") or "").strip()
        fields_raw = payload.get("fields_found")
        fields_found = None
        if fields_raw is not None:
            try:
                fields_found = int(fields_raw)
            except Exception:
                fields_found = None

        status = "accepted" if (is_high_conf and has_signal) else "candidate"
        if error_text:
            status = "error"
            errors.append(f"{provider_name}: {error_text}")

        attempts.append({
            "provider": provider_name,
            "source": str(payload.get("source") or provider_name),
            "status": status,
            "high_confidence": is_high_conf,
            "signal_score": round(float(score), 3),
            "fields_found": fields_found,
            "error": (error_text[:320] if error_text else None),
            "response_preview": _payload_preview(payload),
        })

        if score > best_score:
            best = payload
            best_score = score

        if is_high_conf and has_signal:
            payload.setdefault("provider_chain", [p for p, _ in providers])
            if errors:
                payload.setdefault("provider_errors", errors)
            payload.setdefault("provider_attempts", attempts)
            return payload

    if isinstance(best, dict):
        best.setdefault("provider_chain", [p for p, _ in providers])
        if errors:
            best.setdefault("provider_errors", errors)
        best.setdefault("provider_attempts", attempts)
        return best
    return {
        "source": "prompt_chain",
        "error": "all providers failed",
        "provider_chain": [p for p, _ in providers],
        "provider_errors": errors,
        "provider_attempts": attempts,
        "high_confidence": False,
    }


async def extract_tender_notice(file_path: str, preferred_provider: Optional[str] = None) -> dict:
    result = await _extract_with_prompt_chain(file_path, TENDER_NOTICE_PROMPT, "tender_notice", preferred_provider=preferred_provider)
    result = _normalize_tender_notice_result(result)
    result["document_type"] = "tender_notice"
    return result


async def extract_purchase_order_doc(file_path: str, preferred_provider: Optional[str] = None) -> dict:
    result = await _extract_with_prompt_chain(file_path, PURCHASE_ORDER_PROMPT, "purchase_order", preferred_provider=preferred_provider)
    result = _normalize_purchase_order_result(result)
    result["document_type"] = "purchase_order"
    return result


async def extract_rejection_notice(file_path: str, preferred_provider: Optional[str] = None) -> dict:
    result = await _extract_with_prompt_chain(file_path, REJECTION_NOTICE_PROMPT, "rejection_notice", preferred_provider=preferred_provider)
    result = _normalize_rejection_notice_result(result)
    result["document_type"] = "rejection_notice"
    return result


async def extract_document_by_type(file_path: str, document_type: str, prefer_ocr_engine: Optional[str] = None) -> dict:
    doc_type = str(document_type or "").strip().lower()
    if doc_type not in SUPPORTED_DOCUMENT_TYPES:
        raise ValueError(f"Unsupported document_type: {document_type}")

    provider_key = str(prefer_ocr_engine or "").strip().lower()
    if provider_key in {"", "auto"}:
        provider_key = None

    if doc_type == "purchase_bill":
        res = await _extract_with_prompt_chain(file_path, GEMINI_PROMPT, "purchase_bill", preferred_provider=provider_key)
        if isinstance(res, dict):
            res = _normalize_model_bill_payload(res)
            if "quantity_qtl" not in res and "quantity_mt" in res:
                try:
                    res["quantity_qtl"] = round(float(res["quantity_mt"]) * 10, 3)
                except Exception:
                    pass
            res["fields_found"] = _count_bill_fields(res)
            conf = float(res.get("confidence") or 0)
            res["high_confidence"] = (res.get("fields_found", 0) >= 5) or (res.get("fields_found", 0) >= 4 and conf >= 0.65)
        # Fallback to full router if Groq bill path fails hard.
        if (not provider_key) and (not isinstance(res, dict) or (str(res.get("error") or "") and (res.get("fields_found") or 0) == 0)):
            res = await extract_bill(file_path)
    elif doc_type == "plant_unloading":
        res = await extract_plant_unloading_sheet(file_path, preferred_provider=provider_key)
    elif doc_type == "tender_notice":
        res = await extract_tender_notice(file_path, preferred_provider=provider_key)
    elif doc_type == "purchase_order":
        res = await extract_purchase_order_doc(file_path, preferred_provider=provider_key)
    elif doc_type == "rejection_notice":
        res = await extract_rejection_notice(file_path, preferred_provider=provider_key)
    elif doc_type == "not_classified":
        res = {
            "document_type": "not_classified",
            "source": "classifier",
            "high_confidence": False,
            "manual_required": True,
            "note": "Document was not classified into predefined types; manual review required.",
        }
    else:
        res = {"error": "unknown_document_type", "high_confidence": False}

    if not isinstance(res, dict):
        return {"document_type": doc_type, "error": "invalid extractor response", "high_confidence": False}
    res.setdefault("document_type", doc_type)
    return res