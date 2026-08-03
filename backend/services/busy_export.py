"""
Busy accounting software export.
Generates purchase bills, sales bills, and payment vouchers
in Busy-compatible CSV format for direct import.
"""
import csv, io
from io import BytesIO
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook


BUSY_XLSX_HEADERS = [
    "VCH_SERIES",
    "VCH/BILL_DATE",
    "VCH/BILL_NO",
    "SALE/PURC_TYPE",
    "PARTY_NAME",
    "MC_NAME",
    "ITEM_NAME",
    "QUANTITY",
    "UNIT",
    "PRICE",
    "ITC_ELIGIBILITY_TYPE",
    "NARRATION",
]

def _fmt_date(d) -> str:
    if isinstance(d, (date, datetime)):
        return d.strftime("%d/%m/%Y")
    return str(d) if d else ""

def _fmt_amt(v) -> str:
    return f"{float(v):.2f}" if v else "0.00"


def export_purchase_bills(purchase_bills: list, settings) -> str:
    """
    Export purchase bills to Busy-compatible CSV.
    Columns match Busy's voucher import format.
    """
    output = io.StringIO()
    writer = csv.writer(output)

    # Busy purchase voucher header
    writer.writerow([
        "VoucherType", "VoucherDate", "VoucherNo",
        "PartyName",   "PartyGSTIN",
        "ItemName",    "Quantity",    "Unit",
        "Rate",        "Amount",
        "TaxRate",     "TaxAmount",  "TotalAmount",
        "Narration"
    ])

    for pb in purchase_bills:
        qty   = float(pb.qty_mt or 0)
        rate  = float(pb.rate_per_mt or 0)
        total = float(pb.total_amount or qty * rate)
        broker_name = pb.broker.name if pb.broker else "Unknown Broker"

        writer.writerow([
            "Purchase",
            _fmt_date(pb.bill_date),
            pb.pb_number or f"PB-{pb.id}",
            broker_name,
            pb.broker.gstin if pb.broker else "",
            "Cattle Feed Raw Material",
            f"{qty:.3f}",
            "MT",
            _fmt_amt(rate),
            _fmt_amt(total),
            "0",     # GST — adjust if applicable
            "0.00",
            _fmt_amt(total),
            f"Purchase against deal - Bill ID {pb.bill_id}"
        ])

    return output.getvalue()


def export_sales_bills(sales_bills: list, settings) -> str:
    """Export sales bills (to RCDF plants) for Busy import."""
    output = io.StringIO()
    writer = csv.writer(output)

    writer.writerow([
        "VoucherType", "VoucherDate", "VoucherNo",
        "PartyName",
        "ItemName",  "Quantity", "Unit",
        "Rate",      "Amount",   "TotalAmount",
        "Narration"
    ])

    for sb in sales_bills:
        qty   = float(sb.qty_mt or 0)
        rate  = float(sb.rate_per_mt or 0)
        total = float(sb.total_amount or qty * rate)
        plant = sb.plant.name if sb.plant else "Unknown Plant"

        writer.writerow([
            "Sales",
            _fmt_date(sb.bill_date),
            sb.sb_number or f"SB-{sb.id}",
            f"RCDF - {plant}",
            "Cattle Feed Raw Material",
            f"{qty:.3f}",
            "MT",
            _fmt_amt(rate),
            _fmt_amt(total),
            _fmt_amt(total),
            f"Supply to {plant} - Tender {sb.tender_id}"
        ])

    return output.getvalue()


def export_payment_vouchers(payments: list, settings) -> str:
    """Export broker payment vouchers for Busy import."""
    output = io.StringIO()
    writer = csv.writer(output)

    writer.writerow([
        "VoucherType", "VoucherDate", "VoucherNo",
        "PartyName",   "PartyGSTIN",
        "Amount",      "PaymentMode", "RefNo",
        "Narration"
    ])

    for pay in payments:
        broker_name = pay.broker.name if pay.broker else "Unknown Broker"
        writer.writerow([
            "Payment",
            _fmt_date(pay.payment_date or date.today()),
            pay.voucher_number or f"PMT-{pay.id}",
            broker_name,
            pay.broker.gstin if pay.broker else "",
            _fmt_amt(pay.amount),
            pay.payment_mode.upper() if pay.payment_mode else "NEFT",
            pay.reference_no or "",
            f"Payment for purchase bill {pay.purchase_bill_id}"
        ])

    return output.getvalue()


def save_export(content: str, export_type: str, settings) -> str:
    """Save export to file and return the path."""
    ts     = datetime.now().strftime("%Y%m%d_%H%M%S")
    fname  = f"{export_type}_{ts}.csv"
    fpath  = Path(settings.busy_export_dir) / fname
    fpath.write_text(content, encoding="utf-8-sig")  # BOM for Excel compatibility
    return str(fpath)


def save_export_bytes(content: bytes, export_type: str, settings, ext: str = "xlsx") -> str:
    """Save binary export (for XLSX) and return the path."""
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    fname = f"{export_type}_{ts}.{ext}"
    fpath = Path(settings.busy_export_dir) / fname
    fpath.write_bytes(content)
    return str(fpath)


def _project_template_path() -> Path:
    # backend/services/busy_export.py -> backend -> project root
    return Path(__file__).resolve().parents[2] / "Book1.xlsx"


def export_busy_purchase_bills_xlsx(rows: list, template_path: Path = None) -> bytes:
    """
    Export Busy purchase bills in XLSX format using Book1-style column order.
    Expected row keys:
      vch_series, vch_bill_date, vch_bill_no, sale_purc_type, party_name,
      mc_name, item_name, quantity, unit, price, itc_eligibility_type, narration
    """
    tpath = template_path or _project_template_path()
    if tpath.exists():
        wb = load_workbook(tpath)
        ws = wb[wb.sheetnames[0]]
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

    # Keep template headers, but clear all data rows from row 2 onward.
    if ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row - 1)

    # If template header row is empty or malformed, set required headers explicitly.
    existing_headers = [ws.cell(row=1, column=i + 1).value for i in range(len(BUSY_XLSX_HEADERS))]
    if existing_headers != BUSY_XLSX_HEADERS:
        for i, h in enumerate(BUSY_XLSX_HEADERS, start=1):
            ws.cell(row=1, column=i).value = h

    for r in rows:
        ws.append([
            r.get("vch_series", "Main"),
            r.get("vch_bill_date"),
            r.get("vch_bill_no"),
            r.get("sale_purc_type"),
            r.get("party_name"),
            r.get("mc_name", "Main Store"),
            r.get("item_name"),
            r.get("quantity"),
            r.get("unit", "QUINTAL"),
            r.get("price"),
            r.get("itc_eligibility_type", "Input Goods/Services"),
            r.get("narration"),
        ])

    out = BytesIO()
    wb.save(out)
    return out.getvalue()
