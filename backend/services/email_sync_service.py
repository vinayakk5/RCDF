"""
Email ingest sync service.
Fetches attachments over IMAP and creates PendingIngest rows for review.
"""

import email
import asyncio
import hashlib
import imaplib
import logging
import os
import re
from openpyxl import load_workbook
from datetime import datetime, timedelta, timezone
from email.header import decode_header
from email.utils import parseaddr, parsedate_to_datetime
from pathlib import Path
from typing import List, Optional

from sqlalchemy import or_, select, text
from sqlalchemy.orm import Session

from database import get_settings
from models import DocumentType, IngestSource, IngestStatus, PendingIngest
from services.ocr_service import classify_document_type, extract_document_by_type

log = logging.getLogger(__name__)

_EXCEL_EXTS = {".xlsx", ".xls"}
_ALLOWED_EXTS = {".pdf", ".png", ".jpg", ".jpeg", ".webp", *_EXCEL_EXTS}


def _env_flag(name: str, default: bool = False) -> bool:
    raw_default = "true" if default else "false"
    raw = str(os.getenv(name, raw_default) or raw_default).strip().lower()
    return raw in {"1", "true", "yes", "on"}


def _normalize_sender_email(value: str) -> str:
    return str(value or "").strip().lower()


def _allowed_sender_set(raw: Optional[str]) -> set:
    values = []
    for part in str(raw or "").split(","):
        s = _normalize_sender_email(part)
        if s:
            values.append(s)
    return set(values)


def _email_hint_document_type(subject: str, file_name: str) -> Optional[str]:
    text = f"{subject or ''} {file_name or ''}".lower()
    if re.search(r"\bquality\s*report\b|\blab\s*report\b|\banalysis\s*report\b|\bmoisture\s*report\b|\btest\s*report\b|\bcoa\b|certificate\s+of\s+analysis", text):
        return DocumentType.not_classified.value
    if re.search(r"\b(?:plant\s*)?(?:unload(?:ing)?|unlod(?:ing)?)\b|\brm\s*unload", text):
        return DocumentType.plant_unloading.value
    if re.search(r"\breject(?:ion|ed)?\b|partial\s*reject|complete\s*reject|short\s*receipt|quality\s*issue", text):
        return DocumentType.rejection_notice.value
    if re.search(r"\bpurchase\s*order\b|\bpo\s*(?:no|number)?\b|\bpo[-_/ ]?\d+", text):
        return DocumentType.purchase_order.value
    if re.search(r"\bnotice\s+inviting\s+tender\b|\bnit\b|\btender\b", text):
        return DocumentType.tender_notice.value
    return None


def _looks_like_unloading_sheet_hint(subject: str, file_name: str) -> bool:
    text = f"{subject or ''} {file_name or ''}".lower()
    return bool(re.search(r"\b(?:plant\s*)?(?:unload(?:ing)?|unlod(?:ing)?)\b|\brm\s*unload", text))


def _excel_cell_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    return str(value).strip()


def _excel_norm_token(value) -> str:
    return re.sub(r"[^a-z0-9]+", "", _excel_cell_text(value).lower())


def _excel_cell_float(value) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    raw = _excel_cell_text(value).replace(",", "")
    if not raw:
        return None
    m = re.search(r"-?\d+(?:\.\d+)?", raw)
    if not m:
        return None
    try:
        return float(m.group(0))
    except Exception:
        return None


def _excel_cell_int(value) -> Optional[int]:
    f = _excel_cell_float(value)
    if f is None:
        return None
    try:
        return int(round(f))
    except Exception:
        return None


def _excel_cell_date_iso(value) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()

    # openpyxl may yield date values as python date objects
    if hasattr(value, "isoformat") and not isinstance(value, str):
        try:
            return str(value.isoformat())
        except Exception:
            pass

    s = _excel_cell_text(value)
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%d.%m.%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except Exception:
            continue
    return None


def _pick_excel_unloading_header(ws) -> tuple[Optional[int], dict]:
    best_row = None
    best_map = {}
    best_score = -1

    max_scan = min(int(ws.max_row or 0), 80)
    for ridx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True), start=1):
        col_map = {}
        score = 0
        for cidx, cell in enumerate(row):
            n = _excel_norm_token(cell)
            if not n:
                continue

            if "date" in n or n in {"dt", "entry"}:
                col_map.setdefault("date", cidx)
                score += 1
            if "truck" in n or ("vehicle" in n and "type" not in n):
                col_map.setdefault("truck_number", cidx)
                score += 3
            if n.startswith("ws") or "wsno" in n or "wsnumber" in n:
                col_map.setdefault("ws_no", cidx)
                score += 2
            if "bag" in n:
                col_map.setdefault("no_of_bags", cidx)
                score += 2
            if ("net" in n and ("qty" in n or "quantity" in n or "weight" in n or "wt" in n)) or n in {"netqty", "netweight"}:
                col_map.setdefault("net_qty_mt", cidx)
                score += 4
            if "received" in n and ("qty" in n or "quantity" in n or "weight" in n or "wt" in n):
                col_map.setdefault("received_qty_mt", cidx)
                score += 2
            if "total" in n and ("qty" in n or "quantity" in n or "weight" in n or "wt" in n):
                col_map.setdefault("total_qty_mt", cidx)
                score += 1

            if n.startswith("rm") and ("no" in n or "number" in n or n == "rm"):
                col_map.setdefault("rm_number", cidx)
                score += 1
            if "item" in n or "material" in n:
                col_map.setdefault("item_name", cidx)
                score += 1
            if "party" in n or "vendor" in n or "supplier" in n:
                col_map.setdefault("party_name", cidx)
                score += 1
            if (n.startswith("po") and ("no" in n or "number" in n or n == "po")) or "workorder" in n:
                col_map.setdefault("po_number", cidx)
                score += 1
            if "plant" in n:
                col_map.setdefault("plant_name", cidx)
                score += 1

        has_core = bool(
            col_map.get("truck_number") is not None
            or col_map.get("net_qty_mt") is not None
            or col_map.get("received_qty_mt") is not None
        )
        if has_core and score > best_score:
            best_score = score
            best_row = ridx
            best_map = col_map

    if best_row is None or best_score < 3:
        return None, {}
    return best_row, best_map


def _extract_plant_unloading_from_excel(path: str, subject: str, file_name: str) -> dict:
    wb = None
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        return {
            "source": "excel_local",
            "document_type": DocumentType.plant_unloading.value,
            "rows": [],
            "high_confidence": False,
            "manual_required": True,
            "error": f"excel_read_failed: {str(e)[:220]}",
        }

    try:
        ws = wb[wb.sheetnames[0]]
        header_row, col_map = _pick_excel_unloading_header(ws)
        if not header_row:
            return {
                "source": "excel_local",
                "document_type": DocumentType.plant_unloading.value,
                "rows": [],
                "high_confidence": False,
                "manual_required": True,
                "error": "excel_header_not_found: could not detect unloading columns",
            }

        def _cell(row_vals, key):
            cidx = col_map.get(key)
            if cidx is None:
                return None
            if cidx >= len(row_vals):
                return None
            return row_vals[cidx]

        rows = []
        blank_streak = 0
        max_rows = min(int(ws.max_row or 0), header_row + 5000)
        for row_vals in ws.iter_rows(min_row=header_row + 1, max_row=max_rows, values_only=True):
            date_iso = _excel_cell_date_iso(_cell(row_vals, "date"))
            truck = _excel_cell_text(_cell(row_vals, "truck_number"))
            ws_no = _excel_cell_text(_cell(row_vals, "ws_no"))
            no_of_bags = _excel_cell_int(_cell(row_vals, "no_of_bags"))
            net_qty = _excel_cell_float(_cell(row_vals, "net_qty_mt"))
            received_qty = _excel_cell_float(_cell(row_vals, "received_qty_mt"))
            total_qty = _excel_cell_float(_cell(row_vals, "total_qty_mt"))

            rm_number = _excel_cell_text(_cell(row_vals, "rm_number"))
            item_name = _excel_cell_text(_cell(row_vals, "item_name"))
            party_name = _excel_cell_text(_cell(row_vals, "party_name"))
            po_number = _excel_cell_text(_cell(row_vals, "po_number"))
            plant_name = _excel_cell_text(_cell(row_vals, "plant_name"))

            has_any = any([
                date_iso,
                truck,
                ws_no,
                no_of_bags is not None,
                net_qty is not None,
                received_qty is not None,
                total_qty is not None,
                rm_number,
                item_name,
                party_name,
            ])
            if not has_any:
                blank_streak += 1
                if blank_streak >= 30 and rows:
                    break
                continue
            blank_streak = 0

            if not truck and net_qty is None and received_qty is None and total_qty is None and no_of_bags is None:
                continue

            row = {}
            if date_iso:
                row["date"] = date_iso
            if truck:
                row["truck_number"] = truck
            if ws_no:
                row["ws_no"] = ws_no
            if no_of_bags is not None:
                row["no_of_bags"] = no_of_bags
            if received_qty is not None:
                row["received_qty_mt"] = received_qty
            if net_qty is not None:
                row["net_qty_mt"] = net_qty
            if total_qty is not None:
                row["total_qty_mt"] = total_qty
            if rm_number:
                row["rm_number"] = rm_number
            if item_name:
                row["item_name"] = item_name
            if party_name:
                row["party_name"] = party_name
            if po_number:
                row["po_number"] = po_number
            if plant_name:
                row["plant_name"] = plant_name

            rows.append(row)

        def _first_value(key: str) -> Optional[str]:
            for r in rows:
                val = r.get(key)
                if val is None:
                    continue
                sval = str(val).strip()
                if sval:
                    return sval
            return None

        rm_number = _first_value("rm_number")
        if not rm_number:
            hint_text = f"{subject or ''} {file_name or ''}"
            m = re.search(r"\bRM[\s\-_/]*([A-Z0-9][A-Z0-9\-_/]{1,30})\b", hint_text, re.IGNORECASE)
            if m:
                rm_number = f"RM-{m.group(1).upper()}"

        payload = {
            "source": "excel_local",
            "document_type": DocumentType.plant_unloading.value,
            "rm_number": rm_number,
            "item_name": _first_value("item_name"),
            "party_name": _first_value("party_name"),
            "plant_name": _first_value("plant_name"),
            "po_number": _first_value("po_number"),
            "sheet_date": _first_value("date"),
            "rows": rows,
            "confidence": 0.92 if rows else 0.2,
            "high_confidence": bool(rows),
        }
        if not rows:
            payload["manual_required"] = True
            payload["error"] = "No usable unloading rows found in Excel attachment"
        return payload
    except Exception as e:
        return {
            "source": "excel_local",
            "document_type": DocumentType.plant_unloading.value,
            "rows": [],
            "high_confidence": False,
            "manual_required": True,
            "error": f"excel_parse_failed: {str(e)[:220]}",
        }
    finally:
        try:
            if wb is not None:
                wb.close()
        except Exception:
            pass


def _candidate_score(candidates: List[dict], doc_type: str) -> float:
    for cand in candidates or []:
        dt = str(cand.get("document_type") or cand.get("type") or "").strip().lower()
        if dt != str(doc_type or "").strip().lower():
            continue
        try:
            return float(cand.get("score") or 0)
        except Exception:
            return 0.0
    return 0.0


def _short_text(value: Optional[str], limit: int = 220) -> str:
    txt = str(value or "").replace("\n", " ").strip()
    if len(txt) <= max(12, int(limit)):
        return txt
    return txt[: max(9, int(limit) - 3)] + "..."


def _classifier_attempts_summary(candidates: List[dict], limit: int = 4) -> str:
    if not candidates:
        return "none"

    parts: List[str] = []
    for cand in candidates[: max(1, int(limit))]:
        if not isinstance(cand, dict):
            continue

        provider = str(cand.get("provider") or cand.get("source") or "unknown").strip()
        model = str(cand.get("model") or "").strip()
        doc_type = str(cand.get("document_type") or cand.get("type") or "").strip()

        score_raw = cand.get("confidence") if cand.get("confidence") is not None else cand.get("score")
        try:
            score_txt = f"{float(score_raw):.2f}"
        except Exception:
            score_txt = "?"

        label = provider
        if model:
            label = f"{label}/{model}"
        if doc_type:
            label = f"{label}->{doc_type}"
        label = f"{label}@{score_txt}"

        reason = str(cand.get("reason") or "").strip()
        if reason:
            label = f"{label} ({_short_text(reason, 90)})"
        parts.append(label)

    if not parts:
        return "none"
    if len(candidates) > limit:
        parts.append(f"+{len(candidates) - limit} more")
    return " | ".join(parts)


def _extractor_source_label(payload: dict) -> str:
    if not isinstance(payload, dict):
        return "unknown"
    source = str(payload.get("source") or "unknown").strip()
    model = str(
        payload.get("openrouter_model")
        or payload.get("mistral_model")
        or payload.get("github_model")
        or payload.get("sambanova_model")
        or payload.get("model")
        or ""
    ).strip()
    if model and model not in source:
        return f"{source}/{model}"
    return source


def _decode_mime_header(value: str) -> str:
    if not value:
        return ""
    parts = []
    for chunk, enc in decode_header(value):
        if isinstance(chunk, bytes):
            try:
                parts.append(chunk.decode(enc or "utf-8", errors="replace"))
            except Exception:
                parts.append(chunk.decode("utf-8", errors="replace"))
        else:
            parts.append(str(chunk))
    return "".join(parts).strip()


def _safe_filename(name: str, fallback: str) -> str:
    raw = _decode_mime_header(name or "") or fallback
    raw = raw.replace("\\", "_").replace("/", "_")
    cleaned = re.sub(r"\s+", " ", raw).strip()
    cleaned = re.sub(r"[^A-Za-z0-9._()\- ]", "_", cleaned)
    cleaned = cleaned.lstrip(".")
    if not cleaned:
        cleaned = fallback
    if len(cleaned) > 180:
        stem = Path(cleaned).stem[:140]
        ext = Path(cleaned).suffix[:20]
        cleaned = f"{stem}{ext}"
    return cleaned


def _imap_date(days: int) -> str:
    dt = datetime.utcnow() - timedelta(days=max(int(days), 0))
    return dt.strftime("%d-%b-%Y")


def list_configured_email_accounts(settings=None) -> List[dict]:
    """Return configured sync accounts as normalized dicts.

    Each account dict contains: email_user, email_pass, mailbox, host.
    Falls back to EMAIL_USER/EMAIL_PASS when EMAIL_SYNC_ACCOUNTS is empty.
    """
    settings = settings or get_settings()

    default_host = str(getattr(settings, "email_host", "imap.gmail.com") or "imap.gmail.com").strip() or "imap.gmail.com"
    default_mailbox = str(getattr(settings, "email_sync_mailbox", "INBOX") or "INBOX").strip() or "INBOX"
    raw = str(getattr(settings, "email_sync_accounts", "") or "").strip()

    out: List[dict] = []
    seen = set()

    if raw:
        entries = [e.strip() for e in raw.split(";") if e.strip()]
        for ent in entries:
            parts = [p.strip() for p in ent.split("|")]
            if not parts:
                continue
            email_user = str(parts[0] if len(parts) > 0 else "").strip()
            email_pass = str(parts[1] if len(parts) > 1 else "").strip()
            mailbox = str(parts[2] if len(parts) > 2 else "").strip() or default_mailbox
            host = str(parts[3] if len(parts) > 3 else "").strip() or default_host

            if not email_user or not email_pass:
                continue

            key = (email_user.lower(), mailbox.upper(), host.lower())
            if key in seen:
                continue
            seen.add(key)
            out.append({
                "email_user": email_user,
                "email_pass": email_pass,
                "mailbox": mailbox,
                "host": host,
            })

    if not out:
        email_user = str(getattr(settings, "email_user", "") or "").strip()
        email_pass = str(getattr(settings, "email_pass", "") or "").strip()
        if email_user and email_pass:
            out.append({
                "email_user": email_user,
                "email_pass": email_pass,
                "mailbox": default_mailbox,
                "host": default_host,
            })

    return out


def _parse_email_received_at(msg) -> Optional[datetime]:
    raw = msg.get("Date") if msg else None
    if not raw:
        return None
    try:
        dt = parsedate_to_datetime(raw)
        if not dt:
            return None
        if dt.tzinfo is not None:
            dt = dt.astimezone(timezone.utc).replace(tzinfo=None)
        return dt.replace(tzinfo=None)
    except Exception:
        return None


def get_email_sync_checkpoint(db: Session, mailbox: str, email_user: str) -> dict:
    row = db.execute(
        text(
            """
            SELECT mailbox, email_user, last_uid, last_run_at, last_status,
                   last_error, last_scanned, last_created, last_duplicates
            FROM email_sync_checkpoints
            WHERE mailbox = :mailbox AND email_user = :email_user
            LIMIT 1
            """
        ),
        {"mailbox": mailbox, "email_user": email_user},
    ).mappings().first()
    if not row:
        return {
            "mailbox": mailbox,
            "email_user": email_user,
            "last_uid": 0,
            "last_run_at": None,
            "last_status": None,
            "last_error": None,
            "last_scanned": 0,
            "last_created": 0,
            "last_duplicates": 0,
        }
    return {
        "mailbox": row.get("mailbox"),
        "email_user": row.get("email_user"),
        "last_uid": int(row.get("last_uid") or 0),
        "last_run_at": str(row.get("last_run_at")) if row.get("last_run_at") else None,
        "last_status": row.get("last_status"),
        "last_error": row.get("last_error"),
        "last_scanned": int(row.get("last_scanned") or 0),
        "last_created": int(row.get("last_created") or 0),
        "last_duplicates": int(row.get("last_duplicates") or 0),
    }


def _upsert_email_sync_checkpoint(
    db: Session,
    *,
    mailbox: str,
    email_user: str,
    last_uid: int,
    status: str,
    error_text: Optional[str],
    scanned: int,
    created: int,
    duplicates: int,
) -> None:
    db.execute(
        text(
            """
            INSERT INTO email_sync_checkpoints (
                mailbox, email_user, last_uid, last_run_at, last_status,
                last_error, last_scanned, last_created, last_duplicates
            ) VALUES (
                :mailbox, :email_user, :last_uid, NOW(), :last_status,
                :last_error, :last_scanned, :last_created, :last_duplicates
            )
            ON DUPLICATE KEY UPDATE
                last_uid = VALUES(last_uid),
                last_run_at = VALUES(last_run_at),
                last_status = VALUES(last_status),
                last_error = VALUES(last_error),
                last_scanned = VALUES(last_scanned),
                last_created = VALUES(last_created),
                last_duplicates = VALUES(last_duplicates)
            """
        ),
        {
            "mailbox": mailbox,
            "email_user": email_user,
            "last_uid": int(last_uid or 0),
            "last_status": status[:30],
            "last_error": (error_text or "")[:1000] if error_text else None,
            "last_scanned": int(scanned or 0),
            "last_created": int(created or 0),
            "last_duplicates": int(duplicates or 0),
        },
    )


def _insert_email_sync_log(
    db: Session,
    *,
    sync_reason: str,
    email_user: str,
    mailbox: str,
    imap_uid: Optional[int],
    message_id: Optional[str],
    from_address: Optional[str],
    subject: Optional[str],
    received_at: Optional[datetime],
    status: str,
    attachments_total: int = 0,
    attachments_created: int = 0,
    attachments_duplicates: int = 0,
    attachments_skipped: int = 0,
    note: Optional[str] = None,
) -> None:
    try:
        with db.begin_nested():
            db.execute(
                text(
                    """
                    INSERT INTO email_sync_logs (
                        sync_reason, email_user, mailbox, imap_uid, message_id,
                        from_address, subject, received_at, status,
                        attachments_total, attachments_created, attachments_duplicates,
                        attachments_skipped, note, synced_at
                    ) VALUES (
                        :sync_reason, :email_user, :mailbox, :imap_uid, :message_id,
                        :from_address, :subject, :received_at, :status,
                        :attachments_total, :attachments_created, :attachments_duplicates,
                        :attachments_skipped, :note, NOW()
                    )
                    """
                ),
                {
                    "sync_reason": str(sync_reason or "manual")[:20],
                    "email_user": str(email_user or "")[:200],
                    "mailbox": str(mailbox or "INBOX")[:100],
                    "imap_uid": int(imap_uid) if imap_uid is not None else None,
                    "message_id": (str(message_id or "")[:255] or None),
                    "from_address": (str(from_address or "")[:255] or None),
                    "subject": (str(subject or "")[:500] or None),
                    "received_at": received_at,
                    "status": str(status or "unknown")[:30],
                    "attachments_total": int(attachments_total or 0),
                    "attachments_created": int(attachments_created or 0),
                    "attachments_duplicates": int(attachments_duplicates or 0),
                    "attachments_skipped": int(attachments_skipped or 0),
                    "note": (str(note or "")[:1000] or None),
                },
            )
    except Exception as e:
        # Logging should never break ingest. Keep sync path resilient.
        log.warning(
            "email_sync_logs insert skipped (user=%s mailbox=%s uid=%s): %s",
            str(email_user or "")[:80],
            str(mailbox or "INBOX")[:40],
            imap_uid,
            str(e)[:200],
        )


async def sync_email_pending_ingests(
    db: Session,
    *,
    limit: int = 20,
    since_days: int = 14,
    unread_only: bool = True,
    mailbox: str = "INBOX",
    mark_seen: bool = False,
    start_uid: Optional[int] = None,
    update_checkpoint: bool = True,
    ignore_duplicates: bool = False,
    ignore_sender_allowlist: bool = False,
    email_user_override: Optional[str] = None,
    email_pass_override: Optional[str] = None,
    mailbox_override: Optional[str] = None,
    host_override: Optional[str] = None,
    sync_reason: str = "manual",
) -> dict:
    settings = get_settings()
    # allow overriding host/credentials when syncing multiple accounts
    host = str((host_override or settings.email_host) or "imap.gmail.com").strip()
    user = str((email_user_override or settings.email_user) or "").strip()
    password = email_pass_override or settings.email_pass or ""
    mailbox = str((mailbox_override or mailbox) or settings.email_sync_mailbox or "INBOX").strip() or "INBOX"
    configured_allowed_senders = _allowed_sender_set(getattr(settings, "email_sync_allowed_senders", ""))
    allowed_senders = set() if ignore_sender_allowlist else configured_allowed_senders
    ocr_verbose = _env_flag("EMAIL_SYNC_OCR_VERBOSE", False)
    log_skipped_sender = _env_flag("EMAIL_SYNC_LOG_SKIPPED_SENDERS", False)
    log_duplicate_events = _env_flag("EMAIL_SYNC_LOG_DUPLICATES", False)

    if not user or not password:
        raise RuntimeError("Email sync is not configured. Set EMAIL_USER and EMAIL_PASS in backend .env")

    dest_dir = Path(settings.upload_dir) / "pending_ingests"
    dest_dir.mkdir(parents=True, exist_ok=True)

    summary = {
        "ok": True,
        "mailbox": mailbox,
        "start_uid": int(start_uid or 0),
        "end_uid": int(start_uid or 0),
        "scanned_messages": 0,
        "created": 0,
        "duplicates": 0,
        "skipped": 0,
        "skipped_sender": 0,
        "errors": [],
        "created_ids": [],
        "allowlist_enabled": bool(allowed_senders),
        "allowed_senders": sorted(list(configured_allowed_senders)),
        "ignore_sender_allowlist": bool(ignore_sender_allowlist),
        "ignore_duplicates": bool(ignore_duplicates),
    }

    conn = None

    def _connect_imap_session() -> imaplib.IMAP4_SSL:
        c = imaplib.IMAP4_SSL(host)
        c.login(user, password)
        typ_local, _ = c.select(mailbox, readonly=(not mark_seen))
        if typ_local != "OK":
            raise RuntimeError(f"Failed to open mailbox '{mailbox}'")
        return c

    try:
        conn = _connect_imap_session()

        # Incremental mode uses UID checkpoint and scans only new messages.
        # Fallback mode uses unread/since filters.
        if start_uid and int(start_uid) > 0:
            search_terms: List[str] = ["UID", f"{int(start_uid) + 1}:*"]
            if unread_only:
                search_terms.append("UNSEEN")
        else:
            search_terms = []
            if unread_only:
                search_terms.append("UNSEEN")
            if since_days and since_days > 0:
                search_terms.extend(["SINCE", _imap_date(since_days)])
            if not search_terms:
                search_terms = ["ALL"]

        typ, data = conn.uid("search", None, *search_terms)
        if typ != "OK":
            raise RuntimeError("IMAP search failed")

        uid_bytes = (data[0] or b"").split()
        uid_pairs = []
        for u in uid_bytes:
            if not u:
                continue
            u_text = u.decode("utf-8", errors="ignore")
            try:
                uid_pairs.append((int(u_text), u_text))
            except Exception:
                continue
        uid_pairs.sort(key=lambda x: x[0])
        uids = [u_text for _, u_text in uid_pairs]
        if not uids:
            if update_checkpoint:
                _upsert_email_sync_checkpoint(
                    db,
                    mailbox=mailbox,
                    email_user=user,
                    last_uid=int(start_uid or 0),
                    status="ok",
                    error_text=None,
                    scanned=0,
                    created=0,
                    duplicates=0,
                )
                db.commit()
            return summary

        # For checkpoint-based runs, process oldest first so UID checkpoint advances safely.
        # For manual non-checkpoint runs, process newest first so recent/today emails are included within limit.
        scan_limit = int(limit or 0)
        if scan_limit < 0:
            scan_limit = 0
        if scan_limit == 0:
            scan_limit = len(uids)

        if update_checkpoint:
            target_uids = uids[:scan_limit]
        else:
            target_uids = list(reversed(uids))[:scan_limit]
        max_uid_seen = int(start_uid or 0)

        for uid in target_uids:
            uid_num = None
            try:
                uid_num = int(uid)
                if uid_num > max_uid_seen:
                    max_uid_seen = uid_num
            except Exception:
                pass

            summary["scanned_messages"] += 1

            typ = "NO"
            msg_data = None
            try:
                typ, msg_data = conn.uid("fetch", uid, "(RFC822)")
            except imaplib.IMAP4.abort as fetch_abort:
                # Long OCR processing can let IMAP idle out; reconnect and retry once.
                log.warning(
                    "IMAP fetch aborted for uid=%s user=%s mailbox=%s; retrying after reconnect: %s",
                    uid,
                    user,
                    mailbox,
                    str(fetch_abort)[:180],
                )
                try:
                    try:
                        conn.close()
                    except Exception:
                        pass
                    try:
                        conn.logout()
                    except Exception:
                        pass
                    conn = _connect_imap_session()
                    typ, msg_data = conn.uid("fetch", uid, "(RFC822)")
                except Exception as reconnect_ex:
                    _insert_email_sync_log(
                        db,
                        sync_reason=sync_reason,
                        email_user=user,
                        mailbox=mailbox,
                        imap_uid=uid_num,
                        message_id=None,
                        from_address=None,
                        subject=None,
                        received_at=None,
                        status="fetch_error",
                        note=f"imap reconnect failed: {str(reconnect_ex)[:220]}",
                    )
                    summary["errors"].append({"uid": uid, "error": f"imap reconnect failed: {str(reconnect_ex)[:180]}"})
                    continue
            except Exception as fetch_ex:
                _insert_email_sync_log(
                    db,
                    sync_reason=sync_reason,
                    email_user=user,
                    mailbox=mailbox,
                    imap_uid=uid_num,
                    message_id=None,
                    from_address=None,
                    subject=None,
                    received_at=None,
                    status="fetch_error",
                    note=f"fetch exception: {str(fetch_ex)[:220]}",
                )
                summary["errors"].append({"uid": uid, "error": f"fetch exception: {str(fetch_ex)[:180]}"})
                continue

            if typ != "OK" or not msg_data:
                _insert_email_sync_log(
                    db,
                    sync_reason=sync_reason,
                    email_user=user,
                    mailbox=mailbox,
                    imap_uid=uid_num,
                    message_id=None,
                    from_address=None,
                    subject=None,
                    received_at=None,
                    status="fetch_error",
                    note="fetch failed",
                )
                summary["errors"].append({"uid": uid, "error": "fetch failed"})
                continue

            raw_msg = None
            for part in msg_data:
                if isinstance(part, tuple) and len(part) >= 2 and isinstance(part[1], (bytes, bytearray)):
                    raw_msg = bytes(part[1])
                    break
            if not raw_msg:
                _insert_email_sync_log(
                    db,
                    sync_reason=sync_reason,
                    email_user=user,
                    mailbox=mailbox,
                    imap_uid=uid_num,
                    message_id=None,
                    from_address=None,
                    subject=None,
                    received_at=None,
                    status="payload_error",
                    note="empty RFC822 payload",
                )
                summary["errors"].append({"uid": uid, "error": "empty RFC822 payload"})
                continue

            msg = email.message_from_bytes(raw_msg)
            from_raw = msg.get("From") or ""
            from_addr = parseaddr(from_raw)[1] or from_raw
            from_addr_norm = _normalize_sender_email(from_addr)
            subject = _decode_mime_header(msg.get("Subject") or "")
            message_id = (msg.get("Message-ID") or f"uid-{uid}").strip()
            received_at = _parse_email_received_at(msg)

            if allowed_senders and from_addr_norm not in allowed_senders:
                summary["skipped"] += 1
                summary["skipped_sender"] += 1
                if log_skipped_sender:
                    log.info(
                        "Email sync skipped sender uid=%s from=%s (not in allowlist)",
                        uid,
                        from_addr_norm,
                    )
                _insert_email_sync_log(
                    db,
                    sync_reason=sync_reason,
                    email_user=user,
                    mailbox=mailbox,
                    imap_uid=uid_num,
                    message_id=message_id,
                    from_address=from_addr,
                    subject=subject,
                    received_at=received_at,
                    status="skipped_sender",
                    note="sender not in allowlist",
                )
                continue

            attachment_found = False
            attachment_index = 0
            created_in_msg = 0
            duplicates_in_msg = 0
            skipped_in_msg = 0
            attachment_total = 0
            msg_errors = []

            for part in msg.walk():
                if part.get_content_maintype() == "multipart":
                    continue

                filename = part.get_filename()
                disposition = str(part.get("Content-Disposition") or "").lower()
                if not filename and "attachment" not in disposition:
                    continue

                attachment_total += 1

                attachment_index += 1
                fallback = f"attachment_{attachment_index}.bin"
                safe_name = _safe_filename(filename or fallback, fallback)
                ext = Path(safe_name).suffix.lower()
                if ext not in _ALLOWED_EXTS:
                    summary["skipped"] += 1
                    skipped_in_msg += 1
                    continue

                blob = part.get_payload(decode=True)
                if not blob:
                    summary["skipped"] += 1
                    skipped_in_msg += 1
                    continue

                attachment_found = True
                digest = hashlib.sha256(blob).hexdigest()
                source_key_raw = f"{message_id}|{safe_name}"
                source_message_id = hashlib.sha1(source_key_raw.encode("utf-8", errors="ignore")).hexdigest()

                existing = db.execute(
                    select(PendingIngest.id).where(
                        PendingIngest.source == IngestSource.email,
                        or_(
                            PendingIngest.source_message_id == source_message_id,
                            PendingIngest.file_hash == digest,
                        ),
                    ).limit(1)
                ).scalars().first()
                if existing:
                    if not ignore_duplicates:
                        summary["duplicates"] += 1
                        duplicates_in_msg += 1
                        if log_duplicate_events:
                            log.info(
                                "Email OCR skipped duplicate uid=%s file=%s existing_pending_ingest_id=%s",
                                uid,
                                safe_name,
                                existing,
                            )
                        continue
                    if log_duplicate_events:
                        log.info(
                            "Email OCR duplicate bypass enabled uid=%s file=%s existing_pending_ingest_id=%s",
                            uid,
                            safe_name,
                            existing,
                        )

                try:
                    stamped_name = f"email_{datetime.now().strftime('%Y%m%d%H%M%S')}_{uid}_{safe_name}"
                    abs_path = dest_dir / stamped_name
                    abs_path.write_bytes(blob)
                    rel_path = str(Path("pending_ingests") / stamped_name)
                except Exception as ex:
                    summary["errors"].append({"uid": uid, "file": safe_name, "error": f"write_failed: {str(ex)}"})
                    msg_errors.append(f"write_failed: {str(ex)}")
                    skipped_in_msg += 1
                    continue

                try:
                    is_excel_attachment = ext in _EXCEL_EXTS
                    cls_name_hint = f"{safe_name} {subject}".strip()
                    hinted_doc_type = _email_hint_document_type(subject, safe_name)
                    if is_excel_attachment and not hinted_doc_type and _looks_like_unloading_sheet_hint(subject, safe_name):
                        hinted_doc_type = DocumentType.plant_unloading.value

                    cls_candidates: List[dict] = []
                    if is_excel_attachment:
                        doc_type = hinted_doc_type or DocumentType.not_classified.value
                        cls_conf = 0.86 if doc_type == DocumentType.plant_unloading.value else (0.55 if hinted_doc_type else 0.3)
                        cls_candidates = [{
                            "document_type": doc_type,
                            "score": round(float(cls_conf), 4),
                            "source": "excel_hint",
                        }]
                    else:
                        try:
                            doc_type, cls_conf, cls_candidates = await asyncio.wait_for(
                                asyncio.to_thread(classify_document_type, str(abs_path), cls_name_hint),
                                timeout=45,
                            )
                        except Exception as cls_ex:
                            # Do not block sync creation on model/rate-limit failures.
                            log.warning(
                                "Email OCR classifier failed uid=%s file=%s: %s",
                                uid,
                                safe_name,
                                _short_text(str(cls_ex), 220),
                            )
                            doc_type = hinted_doc_type or DocumentType.not_classified.value
                            cls_conf = 0.35 if hinted_doc_type else 0.2
                            cls_candidates = [{
                                "document_type": doc_type,
                                "score": round(float(cls_conf), 4),
                                "source": "classifier_fallback",
                                "reason": f"classifier_failed: {str(cls_ex)[:180]}",
                            }]

                    if hinted_doc_type:
                        predicted_type = str(doc_type or "").strip().lower()
                        hinted_score = _candidate_score(cls_candidates or [], hinted_doc_type)
                        predicted_score = _candidate_score(cls_candidates or [], predicted_type)
                        low_conf = float(cls_conf or 0) < 0.62
                        weak_gap = hinted_score >= max(0.6, predicted_score - 0.35)

                        if hinted_doc_type == DocumentType.not_classified.value:
                            if low_conf or predicted_type in {
                                DocumentType.purchase_bill.value,
                                DocumentType.rejection_notice.value,
                                DocumentType.plant_unloading.value,
                            }:
                                doc_type = hinted_doc_type
                                cls_conf = max(float(cls_conf or 0), 0.7)
                                cls_candidates = (cls_candidates or []) + [{
                                    "document_type": hinted_doc_type,
                                    "score": round(max(hinted_score, predicted_score, 0.0) + 0.01, 4),
                                    "source": "email_subject_hint",
                                }]

                        # Email subjects often carry explicit PO/Rejection words even when OCR is noisy.
                        if (
                            predicted_type in {DocumentType.purchase_bill.value, DocumentType.tender_notice.value}
                            and hinted_doc_type in {DocumentType.purchase_order.value, DocumentType.rejection_notice.value}
                            and (low_conf or weak_gap)
                        ):
                            doc_type = hinted_doc_type
                            cls_conf = max(float(cls_conf or 0), 0.62)
                            cls_candidates = (cls_candidates or []) + [{
                                "document_type": hinted_doc_type,
                                "score": round(max(hinted_score, predicted_score, 0.0) + 0.01, 4),
                                "source": "email_subject_hint",
                            }]

                    if ocr_verbose:
                        log.info(
                            "Email OCR classify uid=%s file=%s -> type=%s conf=%.2f hint=%s attempts=%s",
                            uid,
                            safe_name,
                            doc_type,
                            float(cls_conf or 0),
                            hinted_doc_type or "-",
                            _classifier_attempts_summary(cls_candidates),
                        )
                    else:
                        log.info(
                            "Email OCR classify uid=%s file=%s -> type=%s conf=%.2f",
                            uid,
                            safe_name,
                            doc_type,
                            float(cls_conf or 0),
                        )

                    try:
                        doc_enum = DocumentType(doc_type)
                    except Exception:
                        doc_enum = DocumentType.not_classified
                        cls_candidates = (cls_candidates or []) + [{"type": str(doc_type), "score": 0.0}]

                    try:
                        if doc_enum == DocumentType.not_classified:
                            extracted = {
                                "document_type": doc_enum.value,
                                "source": "classifier",
                                "high_confidence": False,
                                "manual_required": True,
                                "note": "Not classified into predefined document types; manual review required.",
                            }
                        elif is_excel_attachment:
                            if doc_enum == DocumentType.plant_unloading:
                                extracted = await asyncio.wait_for(
                                    extract_document_by_type(str(abs_path), doc_enum.value),
                                    timeout=120,
                                )
                            else:
                                extracted = {
                                    "document_type": doc_enum.value,
                                    "source": "excel_local",
                                    "high_confidence": False,
                                    "manual_required": True,
                                    "note": "Excel extraction currently supports plant unloading sheets; manual review required.",
                                }
                        else:
                            extracted = await asyncio.wait_for(
                                extract_document_by_type(str(abs_path), doc_enum.value),
                                timeout=120,
                            )
                        if not isinstance(extracted, dict):
                            extracted = {"error": "invalid extractor response", "high_confidence": False}
                    except Exception as extract_ex:
                        log.warning(
                            "Email OCR extract failed uid=%s file=%s doc_type=%s: %s",
                            uid,
                            safe_name,
                            doc_enum.value,
                            _short_text(str(extract_ex), 220),
                        )
                        extracted = {
                            "document_type": doc_enum.value,
                            "error": f"extract_failed: {str(extract_ex)[:200]}",
                            "high_confidence": False,
                        }

                    source_label = _extractor_source_label(extracted)
                    fields_found = extracted.get("fields_found") if isinstance(extracted, dict) else None
                    high_conf = bool(extracted.get("high_confidence")) if isinstance(extracted, dict) else False
                    provider_errors = extracted.get("provider_errors") if isinstance(extracted, dict) else None
                    if ocr_verbose and isinstance(provider_errors, list):
                        for perr in provider_errors[:3]:
                            if str(perr or "").strip():
                                log.warning(
                                    "Email OCR provider issue uid=%s file=%s source=%s: %s",
                                    uid,
                                    safe_name,
                                    source_label,
                                    _short_text(str(perr), 220),
                                )
                        if len(provider_errors) > 3:
                            log.warning(
                                "Email OCR provider issue uid=%s file=%s source=%s: +%s more",
                                uid,
                                safe_name,
                                source_label,
                                len(provider_errors) - 3,
                            )

                    provider_attempts = extracted.get("provider_attempts") if isinstance(extracted, dict) else None
                    if ocr_verbose and isinstance(provider_attempts, list) and provider_attempts:
                        max_attempt_logs = 8
                        for idx, att in enumerate(provider_attempts[:max_attempt_logs], start=1):
                            if not isinstance(att, dict):
                                continue
                            att_provider = str(att.get("provider") or "unknown").strip()
                            att_source = str(att.get("source") or att_provider).strip()
                            att_status = str(att.get("status") or "unknown").strip()
                            att_high_conf = bool(att.get("high_confidence"))
                            att_score = att.get("signal_score")
                            att_score_txt = "?"
                            try:
                                att_score_txt = f"{float(att_score):.2f}"
                            except Exception:
                                pass

                            att_error = _short_text(str(att.get("error") or ""), 220)
                            att_preview = _short_text(str(att.get("response_preview") or ""), 320)

                            if att_error:
                                log.warning(
                                    "Email OCR API[%s] uid=%s file=%s provider=%s source=%s status=%s high_conf=%s score=%s error=%s response=%s",
                                    idx,
                                    uid,
                                    safe_name,
                                    att_provider,
                                    att_source,
                                    att_status,
                                    att_high_conf,
                                    att_score_txt,
                                    att_error,
                                    att_preview,
                                )
                            else:
                                log.info(
                                    "Email OCR API[%s] uid=%s file=%s provider=%s source=%s status=%s high_conf=%s score=%s response=%s",
                                    idx,
                                    uid,
                                    safe_name,
                                    att_provider,
                                    att_source,
                                    att_status,
                                    att_high_conf,
                                    att_score_txt,
                                    att_preview,
                                )

                        if len(provider_attempts) > max_attempt_logs:
                            log.info(
                                "Email OCR API trace uid=%s file=%s: +%s more attempts",
                                uid,
                                safe_name,
                                len(provider_attempts) - max_attempt_logs,
                            )

                    err_text = str(extracted.get("error") or "").strip() if isinstance(extracted, dict) else ""
                    if err_text:
                        log.warning(
                            "Email OCR result uid=%s file=%s doc_type=%s source=%s ERROR=%s",
                            uid,
                            safe_name,
                            doc_enum.value,
                            source_label,
                            _short_text(err_text, 260),
                        )
                    else:
                        if ocr_verbose:
                            log.info(
                                "Email OCR result uid=%s file=%s doc_type=%s source=%s high_conf=%s fields=%s",
                                uid,
                                safe_name,
                                doc_enum.value,
                                source_label,
                                high_conf,
                                fields_found,
                            )
                        else:
                            log.info(
                                "Email OCR result uid=%s file=%s doc_type=%s source=%s high_conf=%s",
                                uid,
                                safe_name,
                                doc_enum.value,
                                source_label,
                                high_conf,
                            )

                    unclear_fields = extracted.get("unclear_fields") if isinstance(extracted.get("unclear_fields"), list) else []
                    review_note = f"Email sync: {subject[:120]}" if subject else "Email sync"

                    row = PendingIngest(
                        company_id=None,
                        main_tender_id=None,
                        tender_id=None,
                        source=IngestSource.email,
                        source_address=from_addr[:200] if from_addr else None,
                        source_account=user[:200] if user else None,
                        source_message_id=source_message_id[:100],
                        file_name=safe_name,
                        file_path=rel_path,
                        file_hash=digest,
                        document_type=doc_enum,
                        classifier_confidence=float(cls_conf or 0),
                        classifier_candidates=cls_candidates or [],
                        extracted_payload=extracted,
                        unclear_fields=unclear_fields,
                        status=IngestStatus.pending,
                        review_notes=review_note,
                    )
                    db.add(row)
                    db.flush()

                    summary["created"] += 1
                    summary["created_ids"].append(row.id)
                    created_in_msg += 1
                except Exception as ex:
                    summary["errors"].append({"uid": uid, "file": safe_name, "error": str(ex)})
                    msg_errors.append(str(ex))
                    skipped_in_msg += 1
                    log.exception("Email attachment ingest failed for uid=%s file=%s", uid, safe_name)

            if not attachment_found:
                summary["skipped"] += 1

            if attachment_total == 0:
                msg_status = "no_attachment"
                msg_note = None
            elif created_in_msg > 0 and (duplicates_in_msg > 0 or skipped_in_msg > 0 or msg_errors):
                msg_status = "partial"
                msg_note = f"created={created_in_msg}, dup={duplicates_in_msg}, skipped={skipped_in_msg}"
            elif created_in_msg > 0:
                msg_status = "created"
                msg_note = None
            elif duplicates_in_msg > 0 and not msg_errors:
                msg_status = "duplicate"
                msg_note = f"duplicates={duplicates_in_msg}"
            elif msg_errors:
                msg_status = "error"
                msg_note = "; ".join(msg_errors)[:1000]
            else:
                msg_status = "skipped"
                msg_note = f"skipped={skipped_in_msg}"

            _insert_email_sync_log(
                db,
                sync_reason=sync_reason,
                email_user=user,
                mailbox=mailbox,
                imap_uid=uid_num,
                message_id=message_id,
                from_address=from_addr,
                subject=subject,
                received_at=received_at,
                status=msg_status,
                attachments_total=attachment_total,
                attachments_created=created_in_msg,
                attachments_duplicates=duplicates_in_msg,
                attachments_skipped=skipped_in_msg,
                note=msg_note,
            )

            if mark_seen and unread_only and created_in_msg > 0:
                try:
                    conn.uid("store", uid, "+FLAGS", "(\\Seen)")
                except Exception:
                    summary["errors"].append({"uid": uid, "error": "failed to mark seen"})

        summary["end_uid"] = max_uid_seen
        if update_checkpoint:
            _upsert_email_sync_checkpoint(
                db,
                mailbox=mailbox,
                email_user=user,
                last_uid=max_uid_seen,
                status="ok" if not summary["errors"] else "partial",
                error_text=None if not summary["errors"] else f"{len(summary['errors'])} item errors",
                scanned=summary["scanned_messages"],
                created=summary["created"],
                duplicates=summary["duplicates"],
            )
        db.commit()
        return summary
    except Exception as e:
        if update_checkpoint:
            try:
                db.rollback()
                _upsert_email_sync_checkpoint(
                    db,
                    mailbox=mailbox,
                    email_user=user,
                    last_uid=int(start_uid or 0),
                    status="failed",
                    error_text=str(e),
                    scanned=summary.get("scanned_messages", 0),
                    created=summary.get("created", 0),
                    duplicates=summary.get("duplicates", 0),
                )
                db.commit()
            except Exception:
                db.rollback()
        else:
            db.rollback()
        summary["ok"] = False
        raise
    finally:
        if conn is not None:
            try:
                conn.close()
            except Exception:
                pass
            try:
                conn.logout()
            except Exception:
                pass


async def sync_all_email_accounts(
    db: Session,
    *,
    limit: int = 20,
    since_days: int = 14,
    unread_only: bool = True,
    mark_seen: bool = False,
    start_uid: Optional[int] = None,
    update_checkpoint: bool = True,
    ignore_duplicates: bool = False,
    ignore_sender_allowlist: bool = False,
    account_filter: Optional[str] = None,
    sync_reason: str = "manual",
):
    """Sync across configured email accounts with per-account checkpoints."""
    settings = get_settings()
    accounts = list_configured_email_accounts(settings)
    if not accounts:
        raise RuntimeError("Email sync is not configured. Set EMAIL_SYNC_ACCOUNTS or EMAIL_USER/EMAIL_PASS in backend .env")

    flt = str(account_filter or "").strip().lower()
    if flt and flt not in {"all", "*"}:
        accounts = [a for a in accounts if str(a.get("email_user") or "").strip().lower() == flt]
        if not accounts:
            raise RuntimeError(f"No configured email account matched: {account_filter}")

    per_account = []
    overall = {
        "ok": True,
        "per_account": per_account,
        "created_total": 0,
        "scanned_total": 0,
        "duplicates_total": 0,
        "skipped_sender_total": 0,
        "errors_total": 0,
        "created": 0,
        "scanned_messages": 0,
        "duplicates": 0,
        "skipped_sender": 0,
    }

    for account in accounts:
        acct_user = str(account.get("email_user") or "").strip()
        acct_pass = str(account.get("email_pass") or "").strip()
        acct_mailbox = str(account.get("mailbox") or "INBOX").strip() or "INBOX"
        acct_host = str(account.get("host") or "imap.gmail.com").strip() or "imap.gmail.com"

        if update_checkpoint and (start_uid is None or int(start_uid or 0) <= 0):
            cp = get_email_sync_checkpoint(db, mailbox=acct_mailbox, email_user=acct_user)
            acct_start_uid = int(cp.get("last_uid") or 0)
        else:
            acct_start_uid = int(start_uid or 0)

        try:
            res = await sync_email_pending_ingests(
                db,
                limit=limit,
                since_days=since_days,
                unread_only=unread_only,
                mark_seen=mark_seen,
                start_uid=acct_start_uid,
                update_checkpoint=update_checkpoint,
                ignore_duplicates=ignore_duplicates,
                ignore_sender_allowlist=ignore_sender_allowlist,
                email_user_override=acct_user,
                email_pass_override=acct_pass,
                mailbox_override=acct_mailbox,
                host_override=acct_host,
                sync_reason=sync_reason,
            )
            per_account.append({
                "account": acct_user,
                "mailbox": acct_mailbox,
                "host": acct_host,
                "start_uid": acct_start_uid,
                "result": res,
            })
            overall["created_total"] += int(res.get("created") or 0)
            overall["scanned_total"] += int(res.get("scanned_messages") or 0)
            overall["duplicates_total"] += int(res.get("duplicates") or 0)
            overall["skipped_sender_total"] += int(res.get("skipped_sender") or 0)
            overall["errors_total"] += len(res.get("errors") or [])
        except Exception as e:
            per_account.append({
                "account": acct_user,
                "mailbox": acct_mailbox,
                "host": acct_host,
                "start_uid": acct_start_uid,
                "error": str(e),
            })
            overall["ok"] = False

    overall["created"] = overall["created_total"]
    overall["scanned_messages"] = overall["scanned_total"]
    overall["duplicates"] = overall["duplicates_total"]
    overall["skipped_sender"] = overall["skipped_sender_total"]
    return overall
