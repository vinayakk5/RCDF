"""
RCDF Supply — FastAPI Backend
All routes for all 9 modules.
Run: uvicorn main:app --reload --port 8000
"""
import os, logging, re, json, hashlib, shutil, asyncio, httpx
from pathlib import Path
from datetime import date, datetime, timedelta
from typing import Optional, List
from difflib import SequenceMatcher
from contextlib import asynccontextmanager
from io import BytesIO
import time
from fastapi import FastAPI, Depends, HTTPException, UploadFile, File, Form, Query
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, Response
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy.orm import Session
from sqlalchemy import select, func, and_, or_, desc, text
from pydantic import BaseModel
from openpyxl import load_workbook

from database import get_db, get_settings, init_dirs, engine, SessionLocal
from models import *

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(name)s — %(message)s")
log = logging.getLogger(__name__)
logging.getLogger("httpx").setLevel(logging.WARNING)
logging.getLogger("httpcore").setLevel(logging.WARNING)


def _ensure_runtime_schema() -> None:
    """Create new runtime tables if missing (lightweight alternative to migrations)."""
    def _column_exists(conn, table_name: str, column_name: str) -> bool:
        q = text("""
            SELECT COUNT(*)
            FROM information_schema.COLUMNS
            WHERE TABLE_SCHEMA = DATABASE()
              AND TABLE_NAME = :table_name
              AND COLUMN_NAME = :column_name
        """)
        return (conn.execute(q, {"table_name": table_name, "column_name": column_name}).scalar() or 0) > 0

    def _index_exists(conn, table_name: str, index_name: str) -> bool:
        q = text("""
            SELECT COUNT(*)
            FROM information_schema.STATISTICS
            WHERE TABLE_SCHEMA = DATABASE()
              AND TABLE_NAME = :table_name
              AND INDEX_NAME = :index_name
        """)
        return (conn.execute(q, {"table_name": table_name, "index_name": index_name}).scalar() or 0) > 0

    def _ensure_column(conn, table_name: str, column_name: str, ddl: str) -> None:
        if not _column_exists(conn, table_name, column_name):
            conn.execute(text(f"ALTER TABLE {table_name} ADD COLUMN {column_name} {ddl}"))

    def _ensure_index(conn, table_name: str, index_name: str, ddl: str) -> None:
        if not _index_exists(conn, table_name, index_name):
            conn.execute(text(f"ALTER TABLE {table_name} ADD INDEX {index_name} {ddl}"))

    with engine.begin() as conn:
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS companies (
                id INT AUTO_INCREMENT PRIMARY KEY,
                name VARCHAR(200) NOT NULL UNIQUE,
                code VARCHAR(50) NULL UNIQUE,
                is_active BOOLEAN DEFAULT TRUE,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP
            )
        """))
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS main_tenders (
                id INT AUTO_INCREMENT PRIMARY KEY,
                company_id INT NOT NULL,
                tender_code VARCHAR(100) NOT NULL,
                title VARCHAR(200) NULL,
                notes TEXT NULL,
                is_active BOOLEAN DEFAULT TRUE,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                UNIQUE KEY uq_main_tenders_company_code (company_id, tender_code),
                INDEX idx_main_tenders_company (company_id),
                FOREIGN KEY (company_id) REFERENCES companies(id)
            )
        """))

        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS plant_unloading_masters (
                id INT AUTO_INCREMENT PRIMARY KEY,
                rm_number VARCHAR(100) NOT NULL,
                rm_number_norm VARCHAR(100) NULL,
                rm_number_base VARCHAR(100) NULL,
                item_name VARCHAR(100) NOT NULL,
                party_name VARCHAR(200) NOT NULL,
                plant_id INT NULL,
                plant_name VARCHAR(100) NULL,
                assignment_status VARCHAR(30) DEFAULT 'pending',
                assignment_reason TEXT NULL,
                assignment_confidence DECIMAL(4,3) NULL,
                mapping_source VARCHAR(30) NULL,
                requires_manual_assignment BOOLEAN DEFAULT FALSE,
                is_manual_override BOOLEAN DEFAULT FALSE,
                manual_assigned_by VARCHAR(100) NULL,
                manual_assigned_at DATETIME NULL,
                po_number VARCHAR(100) NULL,
                notes TEXT NULL,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                INDEX idx_pum_key (rm_number, item_name, party_name)
            )
        """))
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS plant_unloading_entries (
                id INT AUTO_INCREMENT PRIMARY KEY,
                master_id INT NOT NULL,
                image_path VARCHAR(500) NULL,
                source VARCHAR(20) DEFAULT 'web',
                ws_no VARCHAR(50) NULL,
                entry_date DATE NOT NULL,
                truck_number VARCHAR(30) NOT NULL,
                no_of_bags INT NULL,
                received_qty_mt DECIMAL(10,3) NULL,
                net_qty_mt DECIMAL(10,3) NOT NULL,
                total_qty_mt DECIMAL(12,3) NULL,
                item_name VARCHAR(100) NULL,
                status ENUM('pending','flagged','approved','linked','rejected') DEFAULT 'pending',
                reviewed_by VARCHAR(100) NULL,
                reviewed_at DATETIME NULL,
                receipt_id INT NULL,
                receipt_created BOOLEAN DEFAULT FALSE,
                dedupe_key VARCHAR(255) NOT NULL,
                ocr_source ENUM('paddle','gemini','manual') DEFAULT 'paddle',
                ocr_confidence DECIMAL(4,3) NULL,
                ocr_raw_json TEXT NULL,
                notes TEXT NULL,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                UNIQUE KEY uq_pue_dedupe (dedupe_key),
                INDEX idx_pue_master (master_id),
                INDEX idx_pue_status (status),
                FOREIGN KEY (master_id) REFERENCES plant_unloading_masters(id) ON DELETE CASCADE,
                FOREIGN KEY (receipt_id) REFERENCES plant_receipts(id) ON DELETE SET NULL
            )
        """))

        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS pending_ingests (
                id INT AUTO_INCREMENT PRIMARY KEY,
                company_id INT NULL,
                main_tender_id INT NULL,
                tender_id INT NULL,
                source ENUM('web','telegram','email','manual') DEFAULT 'web',
                source_address VARCHAR(200) NULL,
                source_message_id VARCHAR(100) NULL,
                file_name VARCHAR(255) NOT NULL,
                file_path VARCHAR(500) NOT NULL,
                file_hash VARCHAR(64) NULL,
                document_type ENUM('purchase_bill','tender_notice','purchase_order','rejection_notice','plant_unloading','not_classified') NOT NULL,
                classifier_confidence DECIMAL(4,3) NULL,
                classifier_candidates JSON NULL,
                extracted_payload JSON NULL,
                unclear_fields JSON NULL,
                status ENUM('pending','approved','rejected','processed') DEFAULT 'pending',
                assigned_company_id INT NULL,
                assigned_main_tender_id INT NULL,
                assigned_tender_id INT NULL,
                review_notes TEXT NULL,
                reviewed_by VARCHAR(100) NULL,
                reviewed_at DATETIME NULL,
                action_status VARCHAR(30) DEFAULT 'pending',
                action_error TEXT NULL,
                action_payload JSON NULL,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                INDEX idx_pending_ingests_status (status),
                INDEX idx_pending_ingests_doc_type (document_type),
                INDEX idx_pending_ingests_scope (company_id, main_tender_id, tender_id),
                INDEX idx_pending_ingests_file_hash (file_hash),
                FOREIGN KEY (company_id) REFERENCES companies(id),
                FOREIGN KEY (main_tender_id) REFERENCES main_tenders(id),
                FOREIGN KEY (tender_id) REFERENCES tenders(id),
                FOREIGN KEY (assigned_company_id) REFERENCES companies(id),
                FOREIGN KEY (assigned_main_tender_id) REFERENCES main_tenders(id),
                FOREIGN KEY (assigned_tender_id) REFERENCES tenders(id)
            )
        """))

        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS purchase_orders (
                id INT AUTO_INCREMENT PRIMARY KEY,
                company_id INT NULL,
                main_tender_id INT NULL,
                tender_id INT NULL,
                po_number VARCHAR(100) NOT NULL,
                po_date DATE NULL,
                seller_name VARCHAR(200) NULL,
                buyer_name VARCHAR(200) NULL,
                buyer_email VARCHAR(200) NULL,
                plant_id INT NULL,
                plant_name VARCHAR(100) NULL,
                total_amount DECIMAL(14,2) NULL,
                line_items JSON NULL,
                status ENUM('draft','approved','cancelled') DEFAULT 'draft',
                source VARCHAR(20) DEFAULT 'ingest',
                source_doc_path VARCHAR(500) NULL,
                source_pending_id INT NULL UNIQUE,
                notes TEXT NULL,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                INDEX idx_purchase_orders_scope (company_id, main_tender_id, tender_id),
                INDEX idx_purchase_orders_po (po_number),
                FOREIGN KEY (company_id) REFERENCES companies(id),
                FOREIGN KEY (main_tender_id) REFERENCES main_tenders(id),
                FOREIGN KEY (tender_id) REFERENCES tenders(id),
                FOREIGN KEY (plant_id) REFERENCES plants(id),
                FOREIGN KEY (source_pending_id) REFERENCES pending_ingests(id)
            )
        """))

        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS email_sync_checkpoints (
                id INT AUTO_INCREMENT PRIMARY KEY,
                mailbox VARCHAR(100) NOT NULL,
                email_user VARCHAR(200) NOT NULL,
                last_uid BIGINT DEFAULT 0,
                last_run_at DATETIME NULL,
                last_status VARCHAR(30) NULL,
                last_error VARCHAR(1000) NULL,
                last_scanned INT DEFAULT 0,
                last_created INT DEFAULT 0,
                last_duplicates INT DEFAULT 0,
                UNIQUE KEY uq_email_sync_checkpoint (mailbox, email_user)
            )
        """))

        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS email_sync_logs (
                id BIGINT AUTO_INCREMENT PRIMARY KEY,
                synced_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                sync_reason VARCHAR(20) NULL,
                email_user VARCHAR(200) NOT NULL,
                mailbox VARCHAR(100) NOT NULL,
                imap_uid BIGINT NULL,
                message_id VARCHAR(255) NULL,
                from_address VARCHAR(255) NULL,
                subject VARCHAR(500) NULL,
                received_at DATETIME NULL,
                status VARCHAR(30) NOT NULL,
                attachments_total INT DEFAULT 0,
                attachments_created INT DEFAULT 0,
                attachments_duplicates INT DEFAULT 0,
                attachments_skipped INT DEFAULT 0,
                note VARCHAR(1000) NULL,
                INDEX idx_email_sync_logs_time (synced_at),
                INDEX idx_email_sync_logs_user_time (email_user, synced_at),
                INDEX idx_email_sync_logs_status (status)
            )
        """))

        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS busy_party_mappings (
                id INT AUTO_INCREMENT PRIMARY KEY,
                company_id INT NOT NULL,
                source_party_name VARCHAR(200) NOT NULL,
                busy_party_name VARCHAR(200) NOT NULL,
                sale_purc_type_override VARCHAR(30) NULL,
                notes TEXT NULL,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                UNIQUE KEY uq_busy_party_map_company_source (company_id, source_party_name),
                INDEX idx_busy_party_map_company (company_id),
                FOREIGN KEY (company_id) REFERENCES companies(id)
            )
        """))

        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS busy_party_master (
                id INT AUTO_INCREMENT PRIMARY KEY,
                company_id INT NULL,
                busy_party_name VARCHAR(220) NOT NULL,
                alias VARCHAR(220) NULL,
                parent_group VARCHAR(150) NULL,
                dealer_type VARCHAR(80) NULL,
                gstin VARCHAR(20) NULL,
                filing_frequency VARCHAR(40) NULL,
                state_code VARCHAR(2) NULL,
                state_name VARCHAR(100) NULL,
                station VARCHAR(120) NULL,
                name_normalized VARCHAR(260) NOT NULL,
                source_file VARCHAR(255) NULL,
                is_active BOOLEAN DEFAULT TRUE,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                INDEX idx_busy_party_master_scope (company_id, is_active),
                INDEX idx_busy_party_master_name (name_normalized),
                INDEX idx_busy_party_master_gstin (gstin),
                FOREIGN KEY (company_id) REFERENCES companies(id)
            )
        """))

        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS busy_staging_bills (
                id INT AUTO_INCREMENT PRIMARY KEY,
                company_id INT NULL,
                source VARCHAR(20) DEFAULT 'web',
                image_path VARCHAR(500) NULL,
                file_hash VARCHAR(64) NULL,
                broker_name VARCHAR(200) NULL,
                vehicle_number VARCHAR(20) NULL,
                material_name VARCHAR(100) NULL,
                qty_mt DECIMAL(10,2) NULL,
                rate_per_mt DECIMAL(10,2) NULL,
                total_amount DECIMAL(14,2) NULL,
                bill_date DATE NULL,
                bill_number VARCHAR(100) NULL,
                plant_name VARCHAR(100) NULL,
                ocr_source ENUM('paddle','gemini','manual') DEFAULT 'paddle',
                ocr_confidence DECIMAL(4,3) NULL,
                ocr_raw_text TEXT NULL,
                is_handwritten BOOLEAN DEFAULT FALSE,
                validation_amount BOOLEAN DEFAULT FALSE,
                validation_vehicle BOOLEAN DEFAULT FALSE,
                validation_material BOOLEAN DEFAULT FALSE,
                busy_exported BOOLEAN DEFAULT FALSE,
                busy_exported_at DATETIME NULL,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                UNIQUE KEY uq_busy_staging_bills_file_hash (file_hash),
                INDEX idx_busy_staging_bills_company (company_id),
                INDEX idx_busy_staging_bills_exported (busy_exported),
                INDEX idx_busy_staging_bills_created (created_at),
                FOREIGN KEY (company_id) REFERENCES companies(id)
            )
        """))
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS whatsapp_configs (
                id INT AUTO_INCREMENT PRIMARY KEY,
                company_id INT NULL,
                is_enabled BOOLEAN DEFAULT TRUE,
                auto_reply BOOLEAN DEFAULT TRUE,
                whitelisted_jids JSON NULL,
                monitored_groups JSON NULL,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                INDEX idx_whatsapp_configs_company (company_id),
                FOREIGN KEY (company_id) REFERENCES companies(id)
            )
        """))

        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS whatsapp_logs (
                id INT AUTO_INCREMENT PRIMARY KEY,
                company_id INT NULL,
                message_id VARCHAR(100) NULL,
                chat_jid VARCHAR(120) NULL,
                chat_name VARCHAR(200) NULL,
                sender_jid VARCHAR(120) NULL,
                sender_name VARCHAR(200) NULL,
                sender_phone VARCHAR(30) NULL,
                is_group BOOLEAN DEFAULT FALSE,
                doc_type VARCHAR(50) DEFAULT 'purchase_bill',
                media_path VARCHAR(500) NULL,
                raw_text TEXT NULL,
                ocr_result JSON NULL,
                matched_id INT NULL,
                status VARCHAR(30) DEFAULT 'processed',
                error_message TEXT NULL,
                reply_sent TEXT NULL,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                INDEX idx_whatsapp_logs_msg (message_id),
                INDEX idx_whatsapp_logs_chat (chat_jid),
                INDEX idx_whatsapp_logs_created (created_at),
                FOREIGN KEY (company_id) REFERENCES companies(id)
            )
        """))

        # Reconciliation columns for dispatch and receipts (safe additive changes).
        _ensure_column(conn, "dispatches", "material_id", "INT NULL")
        _ensure_column(conn, "dispatches", "material_name", "VARCHAR(100) NULL")
        _ensure_column(conn, "dispatches", "consumed_qty_qtl", "DECIMAL(10,2) DEFAULT 0")

        # Context lineage columns.
        _ensure_column(conn, "sproxx_cycles", "company_id", "INT NULL")

        _ensure_column(conn, "tenders", "company_id", "INT NULL")
        _ensure_column(conn, "tenders", "main_tender_id", "INT NULL")
        _ensure_column(conn, "tenders", "fulfilled_qty_mt", "DECIMAL(12,3) DEFAULT 0")

        _ensure_column(conn, "deals", "company_id", "INT NULL")
        _ensure_column(conn, "deals", "main_tender_id", "INT NULL")

        _ensure_column(conn, "bills", "company_id", "INT NULL")
        _ensure_column(conn, "bills", "main_tender_id", "INT NULL")
        _ensure_column(conn, "bills", "tender_id", "INT NULL")
        _ensure_column(conn, "bills", "unloading_match_entry_id", "INT NULL")
        _ensure_column(conn, "bills", "unloading_match_master_id", "INT NULL")
        _ensure_column(conn, "bills", "unloading_match_method", "VARCHAR(40) NULL")
        _ensure_column(conn, "bills", "whatsapp_msg_id", "VARCHAR(100) NULL")
        _ensure_column(conn, "brokers", "whatsapp_phone", "VARCHAR(30) NULL")

        _ensure_column(conn, "dispatches", "company_id", "INT NULL")
        _ensure_column(conn, "dispatches", "main_tender_id", "INT NULL")
        _ensure_column(conn, "dispatches", "tender_id", "INT NULL")

        _ensure_column(conn, "plant_receipts", "matched_dispatch_id", "INT NULL")
        _ensure_column(conn, "plant_receipts", "material_id", "INT NULL")
        _ensure_column(conn, "plant_receipts", "material_name", "VARCHAR(100) NULL")
        _ensure_column(conn, "plant_receipts", "received_qty_qtl", "DECIMAL(10,2) NULL")
        _ensure_column(conn, "plant_receipts", "matched_qty_qtl", "DECIMAL(10,2) DEFAULT 0")
        _ensure_column(conn, "plant_receipts", "match_status", "VARCHAR(20) DEFAULT 'unmatched'")
        _ensure_column(conn, "plant_receipts", "match_reason", "TEXT NULL")
        _ensure_column(conn, "plant_receipts", "match_applied_at", "DATETIME NULL")
        _ensure_column(conn, "plant_receipts", "rm_number", "VARCHAR(100) NULL")
        _ensure_column(conn, "plant_receipts", "party_name", "VARCHAR(200) NULL")
        _ensure_column(conn, "plant_receipts", "po_number", "VARCHAR(100) NULL")
        _ensure_column(conn, "plant_receipts", "company_id", "INT NULL")
        _ensure_column(conn, "plant_receipts", "main_tender_id", "INT NULL")
        _ensure_column(conn, "plant_receipts", "tender_id", "INT NULL")

        _ensure_column(conn, "plant_unloading_masters", "company_id", "INT NULL")
        _ensure_column(conn, "plant_unloading_masters", "main_tender_id", "INT NULL")
        _ensure_column(conn, "plant_unloading_masters", "tender_id", "INT NULL")
        _ensure_column(conn, "plant_unloading_masters", "rm_number_norm", "VARCHAR(100) NULL")
        _ensure_column(conn, "plant_unloading_masters", "rm_number_base", "VARCHAR(100) NULL")
        _ensure_column(conn, "plant_unloading_masters", "assignment_status", "VARCHAR(30) DEFAULT 'pending'")
        _ensure_column(conn, "plant_unloading_masters", "assignment_reason", "TEXT NULL")
        _ensure_column(conn, "plant_unloading_masters", "assignment_confidence", "DECIMAL(4,3) NULL")
        _ensure_column(conn, "plant_unloading_masters", "mapping_source", "VARCHAR(30) NULL")
        _ensure_column(conn, "plant_unloading_masters", "requires_manual_assignment", "BOOLEAN DEFAULT FALSE")
        _ensure_column(conn, "plant_unloading_masters", "is_manual_override", "BOOLEAN DEFAULT FALSE")
        _ensure_column(conn, "plant_unloading_masters", "manual_assigned_by", "VARCHAR(100) NULL")
        _ensure_column(conn, "plant_unloading_masters", "manual_assigned_at", "DATETIME NULL")

        _ensure_column(conn, "plant_unloading_entries", "company_id", "INT NULL")
        _ensure_column(conn, "plant_unloading_entries", "main_tender_id", "INT NULL")
        _ensure_column(conn, "plant_unloading_entries", "tender_id", "INT NULL")

        _ensure_column(conn, "purchase_bills", "company_id", "INT NULL")
        _ensure_column(conn, "purchase_bills", "main_tender_id", "INT NULL")
        _ensure_column(conn, "purchase_bills", "tender_id", "INT NULL")

        _ensure_column(conn, "sales_bills", "company_id", "INT NULL")
        _ensure_column(conn, "sales_bills", "main_tender_id", "INT NULL")

        _ensure_column(conn, "payments", "company_id", "INT NULL")
        _ensure_column(conn, "payments", "main_tender_id", "INT NULL")
        _ensure_column(conn, "payments", "tender_id", "INT NULL")

        _ensure_column(conn, "busy_exports", "company_id", "INT NULL")
        _ensure_column(conn, "busy_exports", "main_tender_id", "INT NULL")
        _ensure_column(conn, "busy_exports", "created_at", "DATETIME DEFAULT CURRENT_TIMESTAMP")

        _ensure_column(conn, "dispatches", "is_deleted", "BOOLEAN DEFAULT FALSE")
        _ensure_column(conn, "plant_receipts", "is_deleted", "BOOLEAN DEFAULT FALSE")

        _ensure_index(conn, "dispatches", "idx_dispatch_match_keys", "(plant_id, material_id, is_deleted)")
        _ensure_index(conn, "plant_receipts", "idx_receipt_match_keys", "(plant_id, material_id, vehicle_number)")
        _ensure_index(conn, "tenders", "idx_tenders_scope", "(company_id, main_tender_id, created_at)")
        _ensure_index(conn, "deals", "idx_deals_scope", "(company_id, main_tender_id, tender_id)")
        _ensure_index(conn, "bills", "idx_bills_scope", "(company_id, main_tender_id, tender_id, status)")
        _ensure_index(conn, "dispatches", "idx_dispatch_scope", "(company_id, main_tender_id, tender_id, is_deleted)")
        _ensure_index(conn, "plant_receipts", "idx_receipt_scope", "(company_id, main_tender_id, tender_id, is_deleted)")

        # Extend enum safely for existing deployments.
        try:
            conn.execute(text("""
                ALTER TABLE pending_ingests
                MODIFY COLUMN document_type ENUM(
                    'purchase_bill','tender_notice','purchase_order','rejection_notice','plant_unloading','not_classified'
                ) NOT NULL
            """))
        except Exception as e:
            log.warning("Could not alter pending_ingests.document_type enum: %s", e)

        try:
            conn.execute(text("""
                ALTER TABLE bills
                MODIFY COLUMN source ENUM('telegram','whatsapp','web','email','manual') DEFAULT 'telegram'
            """))
        except Exception as e:
            log.warning("Could not alter bills.source enum: %s", e)

        try:
            conn.execute(text("""
                ALTER TABLE pending_ingests
                MODIFY COLUMN source ENUM('web','whatsapp','telegram','email','manual') DEFAULT 'web'
            """))
        except Exception as e:
            log.warning("Could not alter pending_ingests.source enum: %s", e)

        # Backfill lineage where parent linkage already exists.
        conn.execute(text("""
            UPDATE deals d
            JOIN tenders t ON d.tender_id = t.id
            SET d.company_id = COALESCE(d.company_id, t.company_id),
                d.main_tender_id = COALESCE(d.main_tender_id, t.main_tender_id)
            WHERE d.company_id IS NULL OR d.main_tender_id IS NULL
        """))
        conn.execute(text("""
            UPDATE bills b
            JOIN deals d ON b.deal_id = d.id
            SET b.company_id = COALESCE(b.company_id, d.company_id),
                b.main_tender_id = COALESCE(b.main_tender_id, d.main_tender_id),
                b.tender_id = COALESCE(b.tender_id, d.tender_id)
            WHERE b.company_id IS NULL OR b.main_tender_id IS NULL OR b.tender_id IS NULL
        """))
        conn.execute(text("""
            UPDATE dispatches ds
            JOIN deals d ON ds.deal_id = d.id
            SET ds.company_id = COALESCE(ds.company_id, d.company_id),
                ds.main_tender_id = COALESCE(ds.main_tender_id, d.main_tender_id),
                ds.tender_id = COALESCE(ds.tender_id, d.tender_id)
            WHERE ds.company_id IS NULL OR ds.main_tender_id IS NULL OR ds.tender_id IS NULL
        """))
        conn.execute(text("""
            UPDATE plant_receipts r
            JOIN dispatches ds ON r.dispatch_id = ds.id
            SET r.company_id = COALESCE(r.company_id, ds.company_id),
                r.main_tender_id = COALESCE(r.main_tender_id, ds.main_tender_id),
                r.tender_id = COALESCE(r.tender_id, ds.tender_id)
            WHERE r.company_id IS NULL OR r.main_tender_id IS NULL OR r.tender_id IS NULL
        """))
        conn.execute(text("""
            UPDATE purchase_bills pb
            JOIN bills b ON pb.bill_id = b.id
            SET pb.company_id = COALESCE(pb.company_id, b.company_id),
                pb.main_tender_id = COALESCE(pb.main_tender_id, b.main_tender_id),
                pb.tender_id = COALESCE(pb.tender_id, b.tender_id)
            WHERE pb.company_id IS NULL OR pb.main_tender_id IS NULL OR pb.tender_id IS NULL
        """))
        conn.execute(text("""
            UPDATE payments p
            JOIN purchase_bills pb ON p.purchase_bill_id = pb.id
            SET p.company_id = COALESCE(p.company_id, pb.company_id),
                p.main_tender_id = COALESCE(p.main_tender_id, pb.main_tender_id),
                p.tender_id = COALESCE(p.tender_id, pb.tender_id)
            WHERE p.company_id IS NULL OR p.main_tender_id IS NULL OR p.tender_id IS NULL
        """))
        conn.execute(text("""
            UPDATE plant_unloading_entries e
            JOIN plant_unloading_masters m ON e.master_id = m.id
            SET e.company_id = COALESCE(e.company_id, m.company_id),
                e.main_tender_id = COALESCE(e.main_tender_id, m.main_tender_id),
                e.tender_id = COALESCE(e.tender_id, m.tender_id)
            WHERE e.company_id IS NULL OR e.main_tender_id IS NULL OR e.tender_id IS NULL
        """))

        # Canonicalize legacy plant spelling in master table.
        has_kaladera = conn.execute(text("SELECT COUNT(*) FROM plants WHERE LOWER(name) = 'kaladera'"))
        has_kaladera = int(has_kaladera.scalar() or 0) > 0
        if not has_kaladera:
            conn.execute(text("UPDATE plants SET name = 'Kaladera' WHERE LOWER(name) = 'kaladers'"))


def normalize_rm_number(raw: Optional[str]) -> str:
    """Normalize RM numbers (e.g. 'RM 270', 'RM-270', 'RM270', 'RM 200 A') -> canonical format ('RM-270', 'RM-200A')."""
    if not raw:
        return ""
    s = str(raw).strip().upper()
    s = re.sub(r"^(NOTICE|INVITING|TENDERS?|FOR|SUPPLY|OF|RAW|MATERIALS?|TENDER|ID|NO|REF|NO\.?|\:|\s)+", "", s)
    m = re.search(r"RM\s*[-_]?\s*(\d+)\s*[-_]?\s*([A-Z0-9]*)", s)
    if m:
        num = m.group(1)
        suffix = m.group(2).strip()
        if suffix:
            return f"RM-{num}{suffix}"
        return f"RM-{num}"
    m2 = re.search(r"\b(\d+)\s*([A-Z]?)\b", s)
    if m2:
        num = m2.group(1)
        suffix = m2.group(2).strip()
        if suffix:
            return f"RM-{num}{suffix}"
        return f"RM-{num}"
    return s


def _email_sync_runs_per_day() -> int:
    settings = get_settings()
    try:
        runs = int(settings.email_sync_runs_per_day or 4)
    except Exception:
        runs = 4
    return max(1, min(runs, 24))


def _email_sync_interval_seconds() -> int:
    runs = _email_sync_runs_per_day()
    # 3/day ~= 8h, 4/day ~= 6h
    return max(1800, int(86400 / runs))


_email_sync_lock = asyncio.Lock()


async def _run_email_sync_once(reason: str = "auto") -> dict:
    from services.email_sync_service import (
        get_email_sync_checkpoint,
        list_configured_email_accounts,
        sync_all_email_accounts,
        sync_email_pending_ingests,
    )

    settings = get_settings()
    accounts = list_configured_email_accounts(settings)
    if not accounts:
        return {"ok": False, "reason": reason, "skipped": "email credentials not configured"}

    limit = max(1, min(int(settings.email_sync_limit or 80), 400))
    # Auto-sync should scan both read and unread emails.
    unread_only = False
    mark_seen = bool(settings.email_sync_mark_seen)

    with SessionLocal() as db:
        if len(accounts) > 1:
            result = await sync_all_email_accounts(
                db,
                limit=limit,
                since_days=5,
                unread_only=unread_only,
                mark_seen=mark_seen,
                start_uid=None,
                update_checkpoint=True,
                ignore_duplicates=False,
                sync_reason=reason,
            )
        else:
            acct = accounts[0]
            mailbox = str(acct.get("mailbox") or settings.email_sync_mailbox or "INBOX").strip() or "INBOX"
            user = str(acct.get("email_user") or "").strip()
            cp = get_email_sync_checkpoint(db, mailbox=mailbox, email_user=user)
            start_uid = int(cp.get("last_uid") or 0)
            result = await sync_email_pending_ingests(
                db,
                limit=limit,
                since_days=5,
                unread_only=unread_only,
                mailbox=mailbox,
                mark_seen=mark_seen,
                start_uid=start_uid,
                update_checkpoint=True,
                ignore_duplicates=False,
                email_user_override=user,
                email_pass_override=str(acct.get("email_pass") or ""),
                host_override=str(acct.get("host") or settings.email_host or "imap.gmail.com"),
                sync_reason=reason,
            )
        result["reason"] = reason
        return result


async def _run_email_sync_once_locked(reason: str = "auto") -> dict:
    async with _email_sync_lock:
        return await _run_email_sync_once(reason=reason)


async def _email_sync_scheduler_loop(stop_event: asyncio.Event) -> None:
    # Stagger initial run so API starts quickly.
    try:
        await asyncio.wait_for(stop_event.wait(), timeout=8)
        return
    except asyncio.TimeoutError:
        pass

    while not stop_event.is_set():
        settings = get_settings()
        if bool(settings.email_sync_auto_enabled):
            try:
                result = await _run_email_sync_once_locked(reason="auto")
                log.info(
                    "Email auto-sync completed: scanned=%s created=%s duplicates=%s start_uid=%s end_uid=%s",
                    result.get("scanned_messages"),
                    result.get("created"),
                    result.get("duplicates"),
                    result.get("start_uid"),
                    result.get("end_uid"),
                )
            except Exception as e:
                log.warning("Email auto-sync failed: %s", e)

        try:
            await asyncio.wait_for(stop_event.wait(), timeout=_email_sync_interval_seconds())
        except asyncio.TimeoutError:
            continue

# ── App setup ──────────────────────────────────────────────────────────────

@asynccontextmanager
async def lifespan(app: FastAPI):
    init_dirs()
    _ensure_runtime_schema()

    stop_event = asyncio.Event()
    app.state.email_sync_stop_event = stop_event
    app.state.email_sync_task = asyncio.create_task(_email_sync_scheduler_loop(stop_event))

    log.info("RCDF Supply backend started")
    try:
        yield
    finally:
        stop_event.set()
        task = getattr(app.state, "email_sync_task", None)
        if task:
            task.cancel()
            try:
                await task
            except Exception:
                pass

app = FastAPI(title="RCDF Supply Operations", version="1.0.0", lifespan=lifespan)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# Serve frontend
frontend_dir = Path(__file__).parent.parent / "frontend"
if frontend_dir.exists():
    app.mount("/static", StaticFiles(directory=str(frontend_dir)), name="static")

@app.get("/", response_class=FileResponse)
def root():
    return str(frontend_dir / "index.html")


# ── Pydantic schemas ───────────────────────────────────────────────────────

class TenderCreate(BaseModel):
    company_id:      Optional[int] = None
    main_tender_id:  Optional[int] = None
    cycle_id:        int
    tender_number:   str
    plant_id:        int
    material_id:     int
    tender_mt:       float
    week1_target_mt: float
    week1_deadline:  date
    week2_deadline:  date
    penalty_pct:     float = 20.0
    notes:           Optional[str] = None

class DealCreate(BaseModel):
    company_id:   Optional[int] = None
    main_tender_id: Optional[int] = None
    tender_id:   int
    broker_id:   int
    material_id: int
    deal_mt:     float
    rate_per_mt: float
    notes:       Optional[str] = None

class BillUpdate(BaseModel):
    deal_id:        Optional[int]   = None
    broker_name:    Optional[str]   = None
    vehicle_number: Optional[str]   = None
    material_name:  Optional[str]   = None
    qty_mt:         Optional[float] = None
    rate_per_mt:    Optional[float] = None
    total_amount:   Optional[float] = None
    bill_date:      Optional[date]  = None
    bill_number:    Optional[str]   = None
    plant_name:     Optional[str]   = None
    status:         Optional[str]   = None
    notes:          Optional[str]   = None


class PlantUnloadingUpdate(BaseModel):
    ws_no:            Optional[str] = None
    entry_date:       Optional[date] = None
    truck_number:     Optional[str] = None
    no_of_bags:       Optional[int] = None
    received_qty_mt:  Optional[float] = None
    net_qty_mt:       Optional[float] = None
    total_qty_mt:     Optional[float] = None
    item_name:        Optional[str] = None
    status:           Optional[str] = None
    reviewed_by:      Optional[str] = None
    notes:            Optional[str] = None


class PlantUnloadingMasterUpdate(BaseModel):
    company_id:      Optional[int] = None
    main_tender_id:  Optional[int] = None
    tender_id:       Optional[int] = None
    rm_number:       Optional[str] = None
    item_name:       Optional[str] = None
    party_name:      Optional[str] = None
    plant_name:      Optional[str] = None
    po_number:       Optional[str] = None
    notes:           Optional[str] = None


class IdListPayload(BaseModel):
    ids: List[int]


class BusyPartyMappingUpsert(BaseModel):
    source_party_name: str
    busy_party_name: str
    sale_purc_type_override: Optional[str] = None
    notes: Optional[str] = None

class ReceiptCreate(BaseModel):
    vehicle_number:   str
    plant_id:         int
    receipt_date:     date
    accepted_mt:      float
    rejected_mt:      float = 0
    rejection_reason: Optional[str] = None
    dispatch_id:      Optional[int] = None
    material_id:      Optional[int] = None
    material_name:    Optional[str] = None

class DispatchCreate(BaseModel):
    bill_id:        Optional[int] = None
    deal_id:        Optional[int] = None
    material_id:    Optional[int] = None
    material_name:  Optional[str] = None
    vehicle_number: str
    dispatch_date:  date
    qty_mt:         float
    plant_id:       int
    driver_name:    Optional[str] = None
    driver_phone:   Optional[str] = None
    status:         Optional[str] = None

class DispatchUpdate(BaseModel):
    bill_id:        Optional[int] = None
    deal_id:        Optional[int] = None
    material_id:    Optional[int] = None
    material_name:  Optional[str] = None
    vehicle_number: Optional[str] = None
    dispatch_date:  Optional[date] = None
    qty_mt:         Optional[float] = None
    plant_id:       Optional[int] = None
    driver_name:    Optional[str] = None
    driver_phone:   Optional[str] = None
    status:         Optional[str] = None

class ReceiptUpdate(BaseModel):
    vehicle_number:   Optional[str] = None
    plant_id:         Optional[int] = None
    receipt_date:     Optional[date] = None
    accepted_mt:      Optional[float] = None
    rejected_mt:      Optional[float] = None
    rejection_reason: Optional[str] = None
    dispatch_id:      Optional[int] = None
    material_id:      Optional[int] = None
    material_name:    Optional[str] = None


class ReceiptManualMatch(BaseModel):
    dispatch_id: int

class PaymentCreate(BaseModel):
    purchase_bill_id: int
    amount:           float
    payment_mode:     str = "neft"
    payment_date:     Optional[date] = None
    reference_no:     Optional[str]  = None
    notes:            Optional[str]  = None

class MarketPriceCreate(BaseModel):
    material_id:  int
    price_date:   date
    price_per_mt: float
    market:       Optional[str] = None
    notes:        Optional[str] = None

class BrokerCreate(BaseModel):
    name:    str
    phone:   Optional[str] = None
    gstin:   Optional[str] = None
    address: Optional[str] = None

class CycleCreate(BaseModel):
    company_id:  Optional[int] = None
    name:        str
    cycle_start: date
    cycle_end:   date
    week1_end:   date
    week2_end:   date


class CompanyCreate(BaseModel):
    name: str
    code: Optional[str] = None


class MainTenderCreate(BaseModel):
    company_id: int
    tender_code: str
    title: Optional[str] = None
    notes: Optional[str] = None


class PendingIngestUpdate(BaseModel):
    document_type: Optional[str] = None
    extracted_payload: Optional[dict] = None
    unclear_fields: Optional[List[str]] = None
    review_notes: Optional[str] = None
    classifier_confidence: Optional[float] = None
    status: Optional[str] = None


class PendingIngestAssign(BaseModel):
    company_id: Optional[int] = None
    main_tender_id: Optional[int] = None
    tender_id: Optional[int] = None


class PendingIngestApprove(BaseModel):
    operator: Optional[str] = None
    review_notes: Optional[str] = None
    auto_action: bool = False


class PendingIngestReparse(BaseModel):
    reclassify: bool = False
    document_type: Optional[str] = None
    ocr_engine: Optional[str] = None


class PendingIngestBulkReparse(BaseModel):
    ids: List[int]
    reclassify: bool = False
    ocr_engine: Optional[str] = None


class BusyStagingReparse(BaseModel):
    ocr_engine: Optional[str] = None


class BusyStagingBulkReparse(BaseModel):
    ids: List[int]
    ocr_engine: Optional[str] = None


class PurchaseOrderCreate(BaseModel):
    company_id: Optional[int] = None
    main_tender_id: Optional[int] = None
    tender_id: Optional[int] = None
    po_number: str
    po_date: Optional[date] = None
    seller_name: Optional[str] = None
    buyer_name: Optional[str] = None
    buyer_email: Optional[str] = None
    plant_name: Optional[str] = None
    total_amount: Optional[float] = None
    line_items: Optional[List[dict]] = None
    notes: Optional[str] = None


class PurchaseOrderUpdate(BaseModel):
    po_number: Optional[str] = None
    po_date: Optional[date] = None
    seller_name: Optional[str] = None
    buyer_name: Optional[str] = None
    buyer_email: Optional[str] = None
    plant_name: Optional[str] = None
    total_amount: Optional[float] = None
    line_items: Optional[List[dict]] = None
    status: Optional[str] = None
    notes: Optional[str] = None

# ── Helper ─────────────────────────────────────────────────────────────────

def _tender_to_dict(t: Tender, db: Session = None) -> dict:
    accepted = sum(float(d.accepted_mt or 0) for d in t.deals if d.status != DealStatus.cancelled)
    dispatched = sum(float(d.dispatched_mt or 0) for d in t.deals if d.status != DealStatus.cancelled)
    deal_mt = sum(float(d.deal_mt) for d in t.deals if d.status != DealStatus.cancelled)
    fulfilled_from_unloading = float(t.fulfilled_qty_mt or 0)
    effective_fulfilled = fulfilled_from_unloading if fulfilled_from_unloading > 0 else accepted
    tender_mt = float(t.tender_mt)
    w1_target = float(t.week1_target_mt)

    # Compute billed progress (sum of PurchaseBill.qty_mt) per deal and tender when db provided
    billed_by_deal = {}
    total_billed = 0
    if db is not None:
        for d in t.deals:
            if d.status == DealStatus.cancelled:
                billed_by_deal[d.id] = 0
                continue
            billed = db.execute(select(func.sum(PurchaseBill.qty_mt)).where(PurchaseBill.deal_id == d.id)).scalar() or 0
            billed = float(billed)
            billed_by_deal[d.id] = billed
            total_billed += billed

    deals_list = []
    for d in t.deals:
        deal_dict = {
            "id": d.id, "deal_number": d.deal_number,
            "broker": d.broker.name, "broker_id": d.broker_id,
            "deal_mt": float(d.deal_mt), "rate_per_mt": float(d.rate_per_mt),
            "dispatched_mt": float(d.dispatched_mt or 0),
            "accepted_mt": float(d.accepted_mt or 0),
            "rejected_mt": float(d.rejected_mt or 0),
            "status": d.status.value,
        }
        if db is not None:
            deal_dict["billed_mt"] = round(billed_by_deal.get(d.id, 0), 2)
            # remaining to bill for this deal
            deal_dict["bill_remaining_mt"] = round(float(d.deal_mt) - deal_dict["billed_mt"], 2)
        deals_list.append(deal_dict)

    return {
        "id": t.id, "tender_number": t.tender_number,
        "company_id": t.company_id,
        "company_name": t.company.name if getattr(t, "company", None) else None,
        "main_tender_id": t.main_tender_id,
        "main_tender_code": t.main_tender.tender_code if getattr(t, "main_tender", None) else None,
        "plant": t.plant.name, "plant_id": t.plant_id,
        "material": t.material.name, "material_id": t.material_id,
        "cycle": t.cycle.name if t.cycle else "",
        "tender_mt": tender_mt,
        "week1_target_mt": w1_target,
        "week1_deadline": str(t.week1_deadline),
        "week2_deadline": str(t.week2_deadline),
        "status": t.status.value,
        "penalty_pct": float(t.penalty_pct),
        "notes": t.notes,
        "deals": deals_list,
        "summary": {
            "total_deal_mt":      round(deal_mt, 2),
            "total_dispatched_mt":round(dispatched, 2),
            "total_accepted_mt_dispatch": round(accepted, 2),
            "total_fulfilled_from_unloading_mt": round(fulfilled_from_unloading, 2),
            "total_fulfilled_mt": round(effective_fulfilled, 2),
            "total_accepted_mt":  round(effective_fulfilled, 2),
            "total_billed_mt":    round(total_billed, 2),
            "accepted_pct":       round(effective_fulfilled / tender_mt * 100, 1) if tender_mt else 0,
            "week1_pct":          round(effective_fulfilled / w1_target * 100, 1) if w1_target else 0,
            "remaining_mt":       round(tender_mt - effective_fulfilled, 2),
            "bill_remaining_mt":  round(tender_mt - total_billed, 2),
        }
    }

def _bill_to_dict(b: Bill) -> dict:
    # Safely resolve relationships — may be unloaded if FK is NULL
    try:
        broker_name = b.broker.name if b.broker_id and b.broker else b.broker_name
    except Exception:
        broker_name = b.broker_name
    try:
        material_name = b.material.name if b.material_id and b.material else b.material_name
    except Exception:
        material_name = b.material_name
    try:
        plant_name = b.plant.name if b.plant_id and b.plant else b.plant_name
    except Exception:
        plant_name = b.plant_name
    tender_number = None
    try:
        tender_number = b.tender.tender_number if b.tender_id and b.tender else None
    except Exception:
        tender_number = None
    return {
        "id": b.id, "status": b.status.value, "source": b.source.value,
        "company_id": b.company_id,
        "main_tender_id": b.main_tender_id,
        "tender_id": b.tender_id,
        "tender_number": tender_number,
        "tender_tag": (f"T{b.tender_id}" if b.tender_id else None),
        "broker_name": broker_name,
        "broker_id": b.broker_id, "deal_id": b.deal_id,
        "vehicle_number": b.vehicle_number,
        "material": material_name,
        "qty_mt": float(b.qty_mt) if b.qty_mt else None,
        "rate_per_mt": float(b.rate_per_mt) if b.rate_per_mt else None,
        "total_amount": float(b.total_amount) if b.total_amount else None,
        "bill_date": str(b.bill_date) if b.bill_date else None,
        "bill_number": b.bill_number,
        "plant": plant_name,
        "plant_id": b.plant_id,
        "ocr_source": b.ocr_source.value if b.ocr_source else None,
        "ocr_confidence": float(b.ocr_confidence) if b.ocr_confidence else None,
        "is_handwritten": b.is_handwritten,
        "validation_amount": b.validation_amount,
        "validation_vehicle": b.validation_vehicle,
        "validation_material": b.validation_material,
        "image_path": b.image_path,
        "notes": b.notes,
        "unloading_match_entry_id": b.unloading_match_entry_id,
        "unloading_match_master_id": b.unloading_match_master_id,
        "unloading_match_method": b.unloading_match_method,
        "created_at": str(b.created_at),
    }


def _busy_staging_to_dict(b: BusyStagingBill) -> dict:
    return {
        "id": b.id,
        "company_id": b.company_id,
        "source": b.source,
        "image_path": b.image_path,
        "file_hash": b.file_hash,
        "broker_name": b.broker_name,
        "vehicle_number": b.vehicle_number,
        "material_name": b.material_name,
        "qty_mt": float(b.qty_mt) if b.qty_mt is not None else None,
        "rate_per_mt": float(b.rate_per_mt) if b.rate_per_mt is not None else None,
        "total_amount": float(b.total_amount) if b.total_amount is not None else None,
        "bill_date": str(b.bill_date) if b.bill_date else None,
        "bill_number": b.bill_number,
        "plant_name": b.plant_name,
        "ocr_source": b.ocr_source.value if b.ocr_source else None,
        "ocr_confidence": float(b.ocr_confidence) if b.ocr_confidence is not None else None,
        "ocr_raw_text": b.ocr_raw_text,
        "is_handwritten": bool(b.is_handwritten),
        "validation_amount": bool(b.validation_amount),
        "validation_vehicle": bool(b.validation_vehicle),
        "validation_material": bool(b.validation_material),
        "busy_exported": bool(b.busy_exported),
        "busy_exported_at": str(b.busy_exported_at) if b.busy_exported_at else None,
        "created_at": str(b.created_at) if b.created_at else None,
        "updated_at": str(b.updated_at) if b.updated_at else None,
    }


def _resolve_bill_match_keys(
    db: Session,
    b: Bill,
) -> tuple[str, Optional[int], Optional[int], str]:
    truck_key = _norm_vehicle(b.vehicle_number)
    if not truck_key:
        return "", None, None, ""

    plant_id = b.plant_id
    if not plant_id and b.plant_name:
        pid, pname = _resolve_plant(db, b.plant_name)
        if pid:
            b.plant_id = pid
            plant_id = pid
        if pname:
            b.plant_name = pname

    material_id, material_name = _resolve_material(
        db,
        material_id=b.material_id,
        material_name=b.material_name,
    )
    if material_id and b.material_id != material_id:
        b.material_id = material_id
    if material_name and not _norm_text(b.material_name):
        b.material_name = material_name

    material_key = _norm_text(material_name or b.material_name)
    return truck_key, plant_id, material_id, material_key


def _lineage_from_unloading_entry(entry: PlantUnloadingEntry) -> tuple[Optional[int], Optional[int], Optional[int]]:
    master = entry.master
    company_id = entry.company_id if entry.company_id is not None else (master.company_id if master else None)
    main_tender_id = entry.main_tender_id if entry.main_tender_id is not None else (master.main_tender_id if master else None)
    tender_id = entry.tender_id if entry.tender_id is not None else (master.tender_id if master else None)
    return company_id, main_tender_id, tender_id


def _occupied_unloading_entry_ids_for_bill_matching(
    db: Session,
    *,
    exclude_bill_id: Optional[int] = None,
) -> set[int]:
    q = select(Bill.unloading_match_entry_id).where(
        Bill.unloading_match_entry_id != None,
        Bill.deal_id == None,
        Bill.status != BillStatus.rejected,
    )
    if exclude_bill_id is not None:
        q = q.where(Bill.id != exclude_bill_id)
    values = db.execute(q).scalars().all()
    return {int(v) for v in values if v is not None}


def _unloading_entry_claims_for_bill_matching(db: Session) -> dict[int, set[int]]:
    q = select(Bill.id, Bill.unloading_match_entry_id).where(
        Bill.unloading_match_entry_id != None,
        Bill.deal_id == None,
        Bill.status != BillStatus.rejected,
    )
    rows = db.execute(q).all()
    claims: dict[int, set[int]] = {}
    for bill_id, entry_id in rows:
        if bill_id is None or entry_id is None:
            continue
        eid = int(entry_id)
        bid = int(bill_id)
        if eid not in claims:
            claims[eid] = set()
        claims[eid].add(bid)
    return claims


def _find_unloading_entry_match_for_bill(
    db: Session,
    b: Bill,
    *,
    occupied_entry_ids: Optional[set[int]] = None,
) -> Optional[PlantUnloadingEntry]:
    if not b:
        return None

    truck_key, plant_id, bill_material_id, bill_material_key = _resolve_bill_match_keys(db, b)
    if not truck_key or not (bill_material_id or bill_material_key):
        return None

    if occupied_entry_ids is None:
        occupied_entry_ids = _occupied_unloading_entry_ids_for_bill_matching(db, exclude_bill_id=b.id)

    normalized_truck_expr = func.upper(
        func.replace(
            func.replace(
                func.replace(func.coalesce(PlantUnloadingEntry.truck_number, ""), " ", ""),
                "-",
                "",
            ),
            "/",
            "",
        )
    )

    q = (
        select(PlantUnloadingEntry)
        .join(PlantUnloadingMaster, PlantUnloadingEntry.master_id == PlantUnloadingMaster.id)
        .where(
            PlantUnloadingEntry.status != BillStatus.rejected,
            normalized_truck_expr == truck_key,
        )
        .order_by(desc(PlantUnloadingEntry.entry_date), desc(PlantUnloadingEntry.created_at), desc(PlantUnloadingEntry.id))
    )
    if plant_id is not None:
        q = q.where(PlantUnloadingMaster.plant_id == plant_id)

    candidates = db.execute(q).scalars().all()
    if not candidates:
        return None

    material_cache: dict[str, tuple[Optional[int], str]] = {}
    matched_entries: list[PlantUnloadingEntry] = []
    best_entry = None
    best_score = None

    for entry in candidates:
        if _norm_vehicle(entry.truck_number) != truck_key:
            continue

        entry_item = entry.item_name or (entry.master.item_name if entry.master else None)
        cache_key = _norm_text(entry_item)
        cached = material_cache.get(cache_key)
        if cached is None:
            entry_material_id, entry_material_name = _resolve_material(db, material_name=entry_item)
            cached = (entry_material_id, _norm_text(entry_material_name or entry_item))
            material_cache[cache_key] = cached

        entry_material_id, entry_material_key = cached
        if bill_material_id and entry_material_id is not None:
            material_matches = (bill_material_id == entry_material_id)
        else:
            material_matches = bool(bill_material_key and bill_material_key == entry_material_key)
        if not material_matches:
            continue
        if occupied_entry_ids and (entry.id in occupied_entry_ids):
            continue

        matched_entries.append(entry)

        status_rank = 2 if entry.status == BillStatus.approved else (1 if entry.status == BillStatus.linked else 0)
        lineage_rank = 1 if (entry.tender_id or (entry.master and entry.master.tender_id)) else 0
        if b.bill_date and entry.entry_date:
            proximity_rank = -abs((entry.entry_date - b.bill_date).days)
        else:
            proximity_rank = 0

        score = (
            status_rank,
            lineage_rank,
            proximity_rank,
            entry.entry_date or date.min,
            entry.created_at or datetime.min,
            entry.id or 0,
        )
        if best_score is None or score > best_score:
            best_score = score
            best_entry = entry

    if not matched_entries:
        return None

    # If bill plant is unknown, allow fallback only when material+truck points to a
    # single unloading plant; this avoids cross-plant mis-assignment.
    if plant_id is None:
        matched_plant_ids = {
            e.master.plant_id
            for e in matched_entries
            if e.master and e.master.plant_id is not None
        }
        if len(matched_plant_ids) > 1:
            return None
        if len(matched_plant_ids) == 1:
            resolved_pid = next(iter(matched_plant_ids))
            if b.plant_id != resolved_pid:
                b.plant_id = resolved_pid
            if best_entry and best_entry.master and best_entry.master.plant_name:
                b.plant_name = best_entry.master.plant_name

    return best_entry


def _apply_unloading_match_to_bill(
    db: Session,
    b: Bill,
    *,
    occupied_entry_ids: Optional[set[int]] = None,
) -> bool:
    if not b or b.deal_id:
        return False

    matched_entry = _find_unloading_entry_match_for_bill(db, b, occupied_entry_ids=occupied_entry_ids)
    had_auto_match = bool(b.unloading_match_entry_id or b.unloading_match_master_id or b.unloading_match_method)
    changed = False

    if matched_entry:
        company_id, main_tender_id, tender_id = _lineage_from_unloading_entry(matched_entry)

        if b.company_id != company_id:
            b.company_id = company_id
            changed = True
        if b.main_tender_id != main_tender_id:
            b.main_tender_id = main_tender_id
            changed = True
        if b.tender_id != tender_id:
            b.tender_id = tender_id
            changed = True

        if b.unloading_match_entry_id != matched_entry.id:
            b.unloading_match_entry_id = matched_entry.id
            changed = True
        if b.unloading_match_master_id != matched_entry.master_id:
            b.unloading_match_master_id = matched_entry.master_id
            changed = True
        if b.unloading_match_method != "truck_plant_material":
            b.unloading_match_method = "truck_plant_material"
            changed = True
        return changed

    if had_auto_match:
        if b.company_id is not None:
            b.company_id = None
            changed = True
        if b.main_tender_id is not None:
            b.main_tender_id = None
            changed = True
        if b.tender_id is not None:
            b.tender_id = None
            changed = True
        if b.unloading_match_entry_id is not None:
            b.unloading_match_entry_id = None
            changed = True
        if b.unloading_match_master_id is not None:
            b.unloading_match_master_id = None
            changed = True
        if b.unloading_match_method is not None:
            b.unloading_match_method = None
            changed = True

    return changed


def _reconcile_unloading_match_for_bill_ids(db: Session, bill_ids: List[Optional[int]]) -> int:
    unique_ids = sorted({int(bid) for bid in (bill_ids or []) if bid})
    if not unique_ids:
        return 0

    bills: List[Bill] = []
    for bill_id in unique_ids:
        bill = db.get(Bill, bill_id)
        if bill is not None:
            bills.append(bill)
    if not bills:
        return 0

    bills.sort(
        key=lambda b: (
            b.bill_date or date.min,
            b.created_at or datetime.min,
            b.id or 0,
        ),
        reverse=False,
    )

    claims = _unloading_entry_claims_for_bill_matching(db)
    changed = 0
    for bill in bills:
        current_entry = int(bill.unloading_match_entry_id) if bill.unloading_match_entry_id is not None else None
        if current_entry is not None:
            claimers = claims.get(current_entry)
            if claimers is not None:
                claimers.discard(int(bill.id))
                if not claimers:
                    claims.pop(current_entry, None)

        occupied_entry_ids = set(claims.keys())

        if _apply_unloading_match_to_bill(db, bill, occupied_entry_ids=occupied_entry_ids):
            changed += 1

        if bill.unloading_match_entry_id is not None:
            eid = int(bill.unloading_match_entry_id)
            if eid not in claims:
                claims[eid] = set()
            claims[eid].add(int(bill.id))
    return changed


def _auto_assign_bill_to_deal_and_dispatch(db: Session, bill: Bill) -> Optional[Deal]:
    """Auto-assign uploaded bill to matching active Deal, Tender, and create Dispatch in_transit."""
    if not bill or bill.deal_id:
        return bill.deal if bill else None

    # 1. Resolve Broker
    broker_id = bill.broker_id
    if not broker_id and bill.broker_name:
        b_obj = db.execute(select(Broker).where(func.lower(Broker.name) == bill.broker_name.strip().lower())).scalar_one_or_none()
        if not b_obj:
            b_res = db.execute(select(Broker).where(func.lower(Broker.name).contains(bill.broker_name.strip().lower()))).first()
            if b_res: b_obj = b_res[0]
        if b_obj:
            broker_id = b_obj.id
            bill.broker_id = b_obj.id

    # 2. Resolve Plant
    plant_id = bill.plant_id
    if not plant_id and bill.plant_name:
        p_id, _ = _resolve_plant(db, bill.plant_name)
        if p_id:
            plant_id = p_id
            bill.plant_id = p_id

    # 3. Resolve Material
    material_id = bill.material_id
    if not material_id and bill.material_name:
        m_obj = db.execute(select(Material).where(func.lower(Material.name) == bill.material_name.strip().lower())).scalar_one_or_none()
        if not m_obj:
            m_res = db.execute(select(Material).where(func.lower(Material.name).contains(bill.material_name.strip().lower()))).first()
            if m_res: m_obj = m_res[0]
        if m_obj:
            material_id = m_obj.id
            bill.material_id = m_obj.id

    # 4. Search candidate active Deals
    q = select(Deal).join(Tender, Deal.tender_id == Tender.id)
    if broker_id:
        q = q.where(Deal.broker_id == broker_id)
    if plant_id:
        q = q.where(Tender.plant_id == plant_id)
    if material_id:
        q = q.where(Deal.material_id == material_id)
    
    q = q.where(Deal.status.in_([DealStatus.active, DealStatus.partial]))
    q = q.order_by(desc(Deal.created_at))
    candidate_deals = db.execute(q).scalars().all()

    if not candidate_deals and (broker_id or plant_id or material_id):
        q2 = select(Deal).join(Tender, Deal.tender_id == Tender.id)
        if broker_id:
            q2 = q2.where(Deal.broker_id == broker_id)
        elif plant_id:
            q2 = q2.where(Tender.plant_id == plant_id)
        q2 = q2.where(Deal.status.in_([DealStatus.active, DealStatus.partial]))
        q2 = q2.order_by(desc(Deal.created_at))
        candidate_deals = db.execute(q2).scalars().all()

    if candidate_deals:
        matched_deal = candidate_deals[0]
        bill.deal_id = matched_deal.id
        bill.tender_id = matched_deal.tender_id
        bill.main_tender_id = matched_deal.main_tender_id
        bill.company_id = matched_deal.company_id
        bill.broker_id = matched_deal.broker_id
        bill.material_id = matched_deal.material_id
        bill.status = BillStatus.linked
        db.flush()

        # 5. Create or update Dispatch in_transit for this bill
        if bill.vehicle_number and bill.qty_mt:
            existing_disp = db.execute(
                select(Dispatch).where(Dispatch.bill_id == bill.id, Dispatch.is_deleted == False)
            ).scalar_one_or_none()
            if not existing_disp and matched_deal.tender:
                disp = Dispatch(
                    company_id=bill.company_id,
                    main_tender_id=bill.main_tender_id,
                    tender_id=bill.tender_id,
                    bill_id=bill.id,
                    deal_id=matched_deal.id,
                    material_id=bill.material_id,
                    material_name=bill.material_name,
                    vehicle_number=bill.vehicle_number,
                    dispatch_date=bill.bill_date or date.today(),
                    qty_mt=bill.qty_mt,
                    plant_id=matched_deal.tender.plant_id,
                    status=DispatchStatus.in_transit,
                    notes=f"Auto-created from Bill {bill.bill_number or bill.id}"
                )
                db.add(disp)
                db.flush()
        return matched_deal

    return None


def _reconcile_unloading_match_for_all_bills(db: Session) -> int:
    bills = db.execute(
        select(Bill)
        .where(Bill.deal_id == None)
        .order_by(Bill.bill_date, Bill.created_at, Bill.id)
    ).scalars().all()

    claims = _unloading_entry_claims_for_bill_matching(db)
    changed = 0
    for bill in bills:
        current_entry = int(bill.unloading_match_entry_id) if bill.unloading_match_entry_id is not None else None
        if current_entry is not None:
            claimers = claims.get(current_entry)
            if claimers is not None:
                claimers.discard(int(bill.id))
                if not claimers:
                    claims.pop(current_entry, None)

        occupied_entry_ids = set(claims.keys())

        if _apply_unloading_match_to_bill(db, bill, occupied_entry_ids=occupied_entry_ids):
            changed += 1

        if bill.unloading_match_entry_id is not None:
            eid = int(bill.unloading_match_entry_id)
            if eid not in claims:
                claims[eid] = set()
            claims[eid].add(int(bill.id))
    return changed


def _norm_text(v: Optional[str]) -> str:
    return re.sub(r'\s+', ' ', str(v or '').strip().lower())


def _norm_code(v: Optional[str]) -> str:
    return re.sub(r'[^a-z0-9]', '', str(v or '').strip().lower())


_NULLISH_TEXT = {
    "", "null", "none", "nil", "na", "n/a", "nan", "-", "--", "unknown",
}


def _clean_optional_text(v: Optional[str]) -> Optional[str]:
    s = str(v or "").strip()
    if not s:
        return None
    if _norm_text(s) in _NULLISH_TEXT:
        return None
    return s


def _to_float(v) -> Optional[float]:
    if v is None:
        return None
    s = str(v).strip().replace(',', '')
    if not s or s.lower() in ('none', 'null', 'na', 'n/a', '-'):
        return None
    try:
        return float(s)
    except Exception:
        return None


def _to_int(v) -> Optional[int]:
    f = _to_float(v)
    return int(f) if f is not None else None


_PLANT_ALIASES = {
    "kaladera": "kaladera",
    "kaladers": "kaladera",
    "ajmer": "ajmer",
    "jodhpur": "jodhpur",
    "nadbai": "nadbai",
    "bikaner": "bikaner",
    "pali": "pali",
    "lambiyan": "lambiyan",
    "lambiya": "lambiyan",
}


def _norm_plant_hint(v: Optional[str]) -> str:
    s = _norm_text(v)
    s = s.replace("cattle feed plant", " ")
    s = re.sub(r"\bcfp\b", " ", s)
    s = re.sub(r"[^a-z0-9\s]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _resolve_plant(db: Session, plant_name: Optional[str]) -> tuple[Optional[int], Optional[str]]:
    plant_clean = _clean_optional_text(plant_name)
    if not plant_clean:
        return None, None
    normalized = _norm_text(plant_clean)
    if not normalized:
        return None, None

    # Exact name (case-insensitive)
    pl = db.execute(select(Plant).where(func.lower(Plant.name) == normalized)).scalar_one_or_none()
    if pl:
        return pl.id, pl.name

    # Code match (AJM/JDH/...)
    code = _norm_code(plant_clean)
    if code:
        by_code = db.execute(select(Plant).where(func.lower(func.coalesce(Plant.code, "")) == code)).scalar_one_or_none()
        if by_code:
            return by_code.id, by_code.name

    # Alias + fuzzy tokens (handles forms like "CFP AJMER", "CATTLE FEED PLANT KALADERA")
    hint = _norm_plant_hint(plant_clean)
    alias_target = None
    for key, target in _PLANT_ALIASES.items():
        if key in hint:
            alias_target = target
            break

    plants = db.execute(select(Plant)).scalars().all()
    if alias_target:
        for p in plants:
            if _norm_text(p.name) == alias_target:
                return p.id, p.name

    for p in plants:
        pnorm = _norm_plant_hint(p.name)
        if hint and (hint == pnorm or hint in pnorm or pnorm in hint):
            return p.id, p.name

    # Final fallback: always choose the best fuzzy plant match from the master table.
    best = None
    best_score = -1.0
    hint_code = _norm_code(hint)
    for p in plants:
        pnorm = _norm_plant_hint(p.name)
        pcode = _norm_code(pnorm)
        if not pcode:
            continue
        ratio = SequenceMatcher(None, hint_code, pcode).ratio()
        hint_tokens = set(hint.split())
        p_tokens = set(pnorm.split())
        overlap = (len(hint_tokens & p_tokens) / float(len(p_tokens))) if p_tokens else 0.0
        score = max(ratio, overlap)
        if score > best_score:
            best_score = score
            best = p

    # Do not force a random plant on weak/noisy hints.
    if best is not None and best_score >= 0.60:
        return best.id, best.name

    return None, plant_clean


def _pick_unloading_master_without_plant(
    candidates: List[PlantUnloadingMaster],
    truck_number: Optional[str],
) -> Optional[PlantUnloadingMaster]:
    if not candidates:
        return None
    if len(candidates) == 1:
        return candidates[0]

    truck_code = _norm_code(truck_number)
    ranked = []
    for cand in candidates:
        entries = list(cand.entries or [])
        entry_count = len(entries)
        approved_entries = [e for e in entries if e.status == BillStatus.approved]
        approved_count = len(approved_entries)
        truck_hits = 0
        truck_hits_approved = 0
        if truck_code:
            truck_hits = sum(1 for e in entries if _norm_code(e.truck_number) == truck_code)
            truck_hits_approved = sum(1 for e in approved_entries if _norm_code(e.truck_number) == truck_code)
        ranked.append((
            truck_hits_approved,
            approved_count,
            truck_hits,
            entry_count,
            1 if cand.is_manual_override else 0,
            cand.updated_at or datetime.min,
            cand,
        ))

    ranked.sort(key=lambda x: (x[0], x[1], x[2], x[3], x[4], x[5]), reverse=True)
    top = ranked[0]
    second = ranked[1]

    # Strongest signal: approved-history truck matches.
    if top[0] > second[0]:
        return top[6]
    if top[0] > 0 and top[1] >= second[1]:
        return top[6]

    # Next signal: approved-history volume (avoid recently-created pending duplicates).
    if top[1] > second[1]:
        return top[6]

    # Fallback truck signal when approved counts are tied at zero.
    if top[2] > second[2]:
        return top[6]
    if top[2] > 0 and top[3] >= second[3]:
        return top[6]

    # Without truck signal, require a clear historical dominance.
    if top[2] == 0 and top[3] >= second[3] + 2:
        return top[6]

    return None


_MATERIAL_ALIASES = {
    "domc": "Domc",
    "doms": "Domc",
    "dorb": "Dorb",
    "maize": "Maize",
    "rice ddgs": "Rice DDGS",
    "ddgs": "Rice DDGS",
}


def _canonical_material_name(name: Optional[str]) -> Optional[str]:
    if not name:
        return None
    normalized = _norm_text(name)
    if not normalized:
        return None
    return _MATERIAL_ALIASES.get(normalized, str(name).strip())


def _resolve_material(
    db: Session,
    material_id: Optional[int] = None,
    material_name: Optional[str] = None,
) -> tuple[Optional[int], Optional[str]]:
    if material_id:
        m = db.get(Material, material_id)
        if m:
            return m.id, m.name

    canonical = _canonical_material_name(material_name)
    if not canonical:
        return None, (str(material_name).strip() if material_name else None)

    m = db.execute(select(Material).where(func.lower(Material.name) == canonical.lower())).scalar_one_or_none()
    if m:
        return m.id, m.name

    # fallback fuzzy: compare normalized names
    mats = db.execute(select(Material)).scalars().all()
    c_norm = _norm_text(canonical)
    for mm in mats:
        if _norm_text(mm.name) == c_norm:
            return mm.id, mm.name
    return None, canonical


def _norm_vehicle(v: Optional[str]) -> str:
    return re.sub(r"[^A-Za-z0-9]", "", str(v or "")).upper()


def _validate_vehicle_number(v: Optional[str], field_name: str = "vehicle_number") -> None:
    # Keep API validation aligned with DB String(20) to avoid SQL 500s.
    s = str(v or "").strip()
    if not s:
        raise HTTPException(400, f"{field_name} is required")
    if len(s) > 20:
        raise HTTPException(400, f"{field_name} must be 20 characters or fewer")


def _dispatch_remaining_qtl(d: Dispatch) -> float:
    qty = float(d.qty_mt or 0)
    consumed = float(d.consumed_qty_qtl or 0)
    return max(qty - consumed, 0.0)


def _recalculate_deal_dispatched_mt(db: Session, deal_ids: List[Optional[int]]) -> None:
    ids = sorted({int(x) for x in (deal_ids or []) if x})
    if not ids:
        return

    for deal_id in ids:
        deal = db.get(Deal, deal_id)
        if not deal:
            continue
        total = db.execute(
            select(func.coalesce(func.sum(Dispatch.qty_mt), 0)).where(
                Dispatch.deal_id == deal_id,
                Dispatch.is_deleted == False,
            )
        ).scalar() or 0
        deal.dispatched_mt = round(float(total), 3)


def _active_receipt_for_dispatch(
    db: Session,
    dispatch_id: int,
    exclude_receipt_id: Optional[int] = None,
) -> Optional[PlantReceipt]:
    rows = _active_receipts_for_dispatch(
        db,
        dispatch_id=dispatch_id,
        exclude_receipt_id=exclude_receipt_id,
    )
    if len(rows) > 1:
        log.warning(
            "Dispatch %s has multiple active receipts (%s); strict 1:1 violated",
            dispatch_id,
            len(rows),
        )
    return rows[0] if rows else None


def _active_receipts_for_dispatch(
    db: Session,
    dispatch_id: int,
    exclude_receipt_id: Optional[int] = None,
) -> List[PlantReceipt]:
    q = select(PlantReceipt).where(
        PlantReceipt.dispatch_id == dispatch_id,
        PlantReceipt.is_deleted == False,
    )
    if exclude_receipt_id:
        q = q.where(PlantReceipt.id != exclude_receipt_id)
    return db.execute(q.order_by(desc(PlantReceipt.created_at))).scalars().all()


def _active_dispatches_for_bill(
    db: Session,
    bill_id: int,
    exclude_dispatch_id: Optional[int] = None,
) -> List[Dispatch]:
    q = select(Dispatch).where(
        Dispatch.bill_id == bill_id,
        Dispatch.is_deleted == False,
    )
    if exclude_dispatch_id:
        q = q.where(Dispatch.id != exclude_dispatch_id)
    return db.execute(q.order_by(desc(Dispatch.created_at))).scalars().all()


def _assert_bill_dispatch_1to1(
    db: Session,
    bill_id: Optional[int],
    exclude_dispatch_id: Optional[int] = None,
    allow_existing: bool = False,
) -> None:
    if not bill_id:
        return
    rows = _active_dispatches_for_bill(db, bill_id, exclude_dispatch_id=exclude_dispatch_id)
    if len(rows) > 1:
        raise HTTPException(
            409,
            "Data integrity violation: bill has multiple active dispatches. Resolve duplicates before continuing.",
        )
    if rows and not allow_existing:
        raise HTTPException(
            400,
            f"Bill already has dispatch #{rows[0].id}. Only one dispatch is allowed per bill.",
        )


def _assert_dispatch_receipt_1to1(
    db: Session,
    dispatch_id: Optional[int],
    exclude_receipt_id: Optional[int] = None,
    allow_existing: bool = False,
) -> None:
    if not dispatch_id:
        return
    rows = _active_receipts_for_dispatch(db, dispatch_id, exclude_receipt_id=exclude_receipt_id)
    if len(rows) > 1:
        raise HTTPException(
            409,
            "Data integrity violation: dispatch has multiple active receipts. Resolve duplicates before continuing.",
        )
    if rows and not allow_existing:
        raise HTTPException(
            400,
            f"Dispatch #{dispatch_id} already has an active receipt. Only one receipt is allowed per dispatch.",
        )


def _norm_party_name(v: Optional[str]) -> str:
    return re.sub(r"\s+", " ", str(v or "").strip().lower())


_STATE_CODE_TO_NAME = {
    "01": "Jammu and Kashmir",
    "02": "Himachal Pradesh",
    "03": "Punjab",
    "04": "Chandigarh",
    "05": "Uttarakhand",
    "06": "Haryana",
    "07": "Delhi",
    "08": "Rajasthan",
    "09": "Uttar Pradesh",
    "10": "Bihar",
    "11": "Sikkim",
    "12": "Arunachal Pradesh",
    "13": "Nagaland",
    "14": "Manipur",
    "15": "Mizoram",
    "16": "Tripura",
    "17": "Meghalaya",
    "18": "Assam",
    "19": "West Bengal",
    "20": "Jharkhand",
    "21": "Odisha",
    "22": "Chhattisgarh",
    "23": "Madhya Pradesh",
    "24": "Gujarat",
    "26": "Dadra and Nagar Haveli and Daman and Diu",
    "27": "Maharashtra",
    "29": "Karnataka",
    "30": "Goa",
    "31": "Lakshadweep",
    "32": "Kerala",
    "33": "Tamil Nadu",
    "34": "Puducherry",
    "35": "Andaman and Nicobar Islands",
    "36": "Telangana",
    "37": "Andhra Pradesh",
    "38": "Ladakh",
}
_STATE_NAME_TO_CODE = {_norm_party_name(v): k for k, v in _STATE_CODE_TO_NAME.items()}


def _clean_gstin(v: Optional[str]) -> str:
    s = re.sub(r"[^A-Za-z0-9]", "", str(v or "").strip().upper())
    return s[:15] if s else ""


def _state_from_gstin(gstin: Optional[str]) -> tuple[Optional[str], Optional[str]]:
    g = _clean_gstin(gstin)
    if len(g) < 2 or not g[:2].isdigit():
        return None, None
    code = g[:2]
    return code, _STATE_CODE_TO_NAME.get(code)


def _state_code_from_text(v: Optional[str]) -> Optional[str]:
    txt = _norm_party_name(v)
    if not txt:
        return None
    if len(txt) == 2 and txt.isdigit() and txt in _STATE_CODE_TO_NAME:
        return txt
    if txt in _STATE_NAME_TO_CODE:
        return _STATE_NAME_TO_CODE[txt]
    for name, code in _STATE_NAME_TO_CODE.items():
        if name and name in txt:
            return code
    return None


def _extract_station_hint(v: Optional[str]) -> Optional[str]:
    s = _norm_party_name(v)
    if not s:
        return None
    for delim in ("-", "/", "|"):
        if delim in s:
            part = s.rsplit(delim, 1)[-1].strip()
            if re.fullmatch(r"[a-z ]{2,40}", part):
                return part
    parts = [p.strip() for p in s.split(",") if p.strip()]
    if len(parts) >= 2 and re.fullmatch(r"[a-z ]{2,40}", parts[-1]):
        return parts[-1]
    return None


def _normalize_party_key(v: Optional[str]) -> str:
    s = _norm_party_name(v)
    if not s:
        return ""
    s = s.replace("&", " and ")
    s = re.sub(r"[^a-z0-9\s]", " ", s)
    s = re.sub(
        r"\b(m/s|ms|traders?|trading|enterprises?|industr(?:y|ies)|agency|agencies|co|company|pvt|private|ltd|limited)\b",
        " ",
        s,
    )
    s = re.sub(r"\s+", " ", s).strip()
    return s or _norm_party_name(v)


def _infer_bill_source_state(bill: Bill) -> tuple[Optional[str], Optional[str]]:
    broker = bill.broker if bill else None
    code, state_name = _state_from_gstin(broker.gstin if broker else None)
    if code:
        return code, state_name

    address = str((broker.address if broker else "") or "")
    code = _state_code_from_text(address)
    if code:
        return code, _STATE_CODE_TO_NAME.get(code)
    return None, None


def _infer_bill_source_station(bill: Bill, source_party_name: Optional[str]) -> Optional[str]:
    broker = bill.broker if bill else None
    station = _extract_station_hint((broker.address if broker else None) or "")
    if station:
        return station
    return _extract_station_hint(source_party_name)


def _score_busy_party_candidate(
    source_name: str,
    source_key: str,
    source_state_code: Optional[str],
    source_station: Optional[str],
    source_gstin: Optional[str],
    row: BusyPartyMaster,
) -> float:
    cand_name = str(row.busy_party_name or "").strip()
    if not cand_name:
        return 0.0

    cand_key = str(row.name_normalized or "").strip() or _normalize_party_key(cand_name)
    if not cand_key:
        return 0.0

    src_raw = _norm_party_name(source_name)
    cand_raw = _norm_party_name(cand_name)
    ratio_key = SequenceMatcher(None, source_key, cand_key).ratio() if source_key and cand_key else 0.0
    ratio_raw = SequenceMatcher(None, src_raw, cand_raw).ratio() if src_raw and cand_raw else 0.0

    src_tokens = set(source_key.split()) if source_key else set()
    cand_tokens = set(cand_key.split()) if cand_key else set()
    overlap = (len(src_tokens & cand_tokens) / float(len(src_tokens | cand_tokens))) if (src_tokens and cand_tokens) else 0.0

    name_score = max(ratio_key, ratio_raw, overlap)
    if source_key and cand_key and source_key == cand_key:
        name_score = 1.0
    elif source_key and cand_key and (source_key in cand_key or cand_key in source_key):
        name_score = max(name_score, 0.93)

    cand_state = str(row.state_code or "").strip()
    if source_state_code and cand_state:
        state_score = 1.0 if source_state_code == cand_state else 0.0
    elif source_state_code or cand_state:
        state_score = 0.35
    else:
        state_score = 0.5

    cand_station = _norm_party_name(row.station) if row.station else _extract_station_hint(cand_name)
    if source_station and cand_station:
        if source_station == cand_station or source_station in cand_station or cand_station in source_station:
            station_score = 1.0
        else:
            station_score = 0.0
    elif source_station or cand_station:
        if source_station and source_station in cand_raw:
            station_score = 0.9
        else:
            station_score = 0.35
    else:
        station_score = 0.5

    cand_gstin = _clean_gstin(row.gstin)
    src_gstin = _clean_gstin(source_gstin)
    if src_gstin and cand_gstin:
        gst_score = 1.0 if src_gstin == cand_gstin else 0.0
    elif src_gstin or cand_gstin:
        gst_score = 0.3
    else:
        gst_score = 0.5

    return round((name_score * 0.60 + state_score * 0.20 + station_score * 0.10 + gst_score * 0.10) * 100.0, 2)


def _suggest_busy_party(
    db: Session,
    company_id: Optional[int],
    source_party_name: Optional[str],
    source_state_code: Optional[str] = None,
    source_station: Optional[str] = None,
    source_gstin: Optional[str] = None,
) -> Optional[dict]:
    source_name = str(source_party_name or "").strip()
    source_key = _normalize_party_key(source_name)
    if not source_name or not source_key:
        return None

    q = select(BusyPartyMaster).where(BusyPartyMaster.is_active == True)
    if company_id is not None:
        q = q.where(or_(BusyPartyMaster.company_id == company_id, BusyPartyMaster.company_id == None))
    else:
        q = q.where(BusyPartyMaster.company_id == None)

    rows = db.execute(q).scalars().all()
    if not rows:
        return None

    ranked = []
    for r in rows:
        score = _score_busy_party_candidate(
            source_name=source_name,
            source_key=source_key,
            source_state_code=source_state_code,
            source_station=source_station,
            source_gstin=source_gstin,
            row=r,
        )
        if score < 35:
            continue
        ranked.append({
            "id": r.id,
            "party_name": r.busy_party_name,
            "score": score,
            "state_code": r.state_code,
            "state_name": r.state_name,
            "station": r.station,
            "gstin": r.gstin,
            "scope_rank": 1 if (company_id is not None and r.company_id == company_id) else 0,
        })

    if not ranked:
        return None

    ranked.sort(key=lambda x: (x["score"], x["scope_rank"]), reverse=True)
    best = ranked[0]
    second = ranked[1] if len(ranked) > 1 else None
    ambiguous = bool(second and best["score"] >= 70 and abs(best["score"] - second["score"]) <= 3)
    auto_pick = bool(best["score"] >= 78 and not ambiguous)

    return {
        "party_name": best["party_name"],
        "score": best["score"],
        "state_code": best.get("state_code"),
        "state_name": best.get("state_name"),
        "station": best.get("station"),
        "gstin": best.get("gstin"),
        "ambiguous": ambiguous,
        "auto_pick": auto_pick,
        "candidates": ranked[:5],
    }


def _read_busy_party_master_rows(raw_bytes: bytes, source_file: Optional[str] = None) -> List[dict]:
    if not raw_bytes:
        return []

    try:
        wb = load_workbook(BytesIO(raw_bytes), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(400, f"Could not read Busy file. Ensure it is an Excel workbook. Error: {str(e)[:180]}")

    try:
        ws = wb[wb.sheetnames[0]]
        preview = list(ws.iter_rows(min_row=1, max_row=min(30, ws.max_row), values_only=True))

        header_row_idx = None
        headers = None
        for idx, row in enumerate(preview, start=1):
            vals = [_norm_party_name(v) for v in row if str(v or "").strip()]
            if not vals:
                continue
            has_name = any(v == "name" or v.startswith("name") for v in vals)
            has_gst = any("gst" in v for v in vals)
            has_parent = any("parent group" in v or v == "group" for v in vals)
            if has_name and (has_gst or has_parent):
                header_row_idx = idx
                headers = row
                break

        if not header_row_idx or not headers:
            raise HTTPException(400, "Could not detect header row in Busy file. Expected columns like Name/GSTIN/Parent Group.")

        def _find_col(*aliases: str) -> Optional[int]:
            for i, h in enumerate(headers):
                hs = _norm_party_name(h)
                if not hs:
                    continue
                for a in aliases:
                    aa = _norm_party_name(a)
                    if hs == aa or hs.startswith(aa) or aa in hs:
                        return i
            return None

        c_name = _find_col("name", "party name", "account name")
        c_alias = _find_col("alias")
        c_parent = _find_col("parent group", "group")
        c_dealer = _find_col("type of dealer", "dealer type")
        c_gstin = _find_col("gstin", "gst no", "gst")
        c_freq = _find_col("filing frequency", "frequency")
        c_state = _find_col("state")
        c_station = _find_col("station", "city", "town", "location", "place")

        if c_name is None:
            raise HTTPException(400, "Busy file does not contain a usable Name column")

        out_by_key = {}

        for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
            name = str(row[c_name]).strip() if c_name < len(row) and row[c_name] is not None else ""
            if not name:
                continue
            if _norm_party_name(name) in {"name", "list of accounts"}:
                continue

            alias = str(row[c_alias]).strip() if c_alias is not None and c_alias < len(row) and row[c_alias] is not None else None
            parent_group = str(row[c_parent]).strip() if c_parent is not None and c_parent < len(row) and row[c_parent] is not None else None
            dealer_type = str(row[c_dealer]).strip() if c_dealer is not None and c_dealer < len(row) and row[c_dealer] is not None else None
            gstin = _clean_gstin(row[c_gstin] if c_gstin is not None and c_gstin < len(row) else None)
            filing_frequency = str(row[c_freq]).strip() if c_freq is not None and c_freq < len(row) and row[c_freq] is not None else None
            station = str(row[c_station]).strip() if c_station is not None and c_station < len(row) and row[c_station] is not None else None

            state_code, state_name = _state_from_gstin(gstin)
            if not state_code and c_state is not None and c_state < len(row):
                code_from_text = _state_code_from_text(row[c_state])
                if code_from_text:
                    state_code = code_from_text
                    state_name = _STATE_CODE_TO_NAME.get(code_from_text)

            station = station or _extract_station_hint(name)
            normalized = _normalize_party_key(name)
            if not normalized:
                continue

            key = (normalized, gstin or "")
            payload = {
                "busy_party_name": name,
                "alias": alias,
                "parent_group": parent_group,
                "dealer_type": dealer_type,
                "gstin": gstin or None,
                "filing_frequency": filing_frequency,
                "state_code": state_code,
                "state_name": state_name,
                "station": station,
                "name_normalized": normalized,
                "source_file": source_file,
                "is_active": True,
            }

            existing = out_by_key.get(key)
            if not existing:
                out_by_key[key] = payload
            else:
                # Prefer richer rows if duplicates are present.
                if (not existing.get("gstin")) and payload.get("gstin"):
                    out_by_key[key] = payload

        return list(out_by_key.values())
    finally:
        try:
            wb.close()
        except Exception:
            pass


_ALLOWED_SALE_PURC_TYPES = {"L/GST-ItemWise", "I/GST-ItemWise"}


def _normalize_sale_purc_override(v: Optional[str]) -> Optional[str]:
    s = str(v or "").strip()
    if not s or s.lower() == "auto":
        return None
    if s not in _ALLOWED_SALE_PURC_TYPES:
        raise HTTPException(400, "sale_purc_type_override must be L/GST-ItemWise, I/GST-ItemWise, or empty")
    return s


def _find_busy_party_mapping(db: Session, company_id: Optional[int], source_party_name: Optional[str]) -> Optional[BusyPartyMapping]:
    if not company_id:
        return None
    source = str(source_party_name or "").strip()
    if not source:
        return None

    exact = db.execute(
        select(BusyPartyMapping).where(
            BusyPartyMapping.company_id == company_id,
            func.lower(BusyPartyMapping.source_party_name) == source.lower(),
        )
    ).scalar_one_or_none()
    if exact:
        return exact

    # Fallback to normalized whitespace matching when source names vary in spacing.
    normalized = _norm_party_name(source)
    all_rows = db.execute(
        select(BusyPartyMapping).where(BusyPartyMapping.company_id == company_id)
    ).scalars().all()
    for row in all_rows:
        if _norm_party_name(row.source_party_name) == normalized:
            return row
    return None


def _find_busy_party_master_by_name(
    db: Session,
    company_id: Optional[int],
    busy_party_name: Optional[str],
) -> Optional[BusyPartyMaster]:
    name = str(busy_party_name or "").strip()
    if not name:
        return None

    q = select(BusyPartyMaster).where(
        BusyPartyMaster.is_active == True,
        func.lower(BusyPartyMaster.busy_party_name) == name.lower(),
    )
    if company_id is not None:
        q = q.where(or_(BusyPartyMaster.company_id == company_id, BusyPartyMaster.company_id == None))
    else:
        q = q.where(BusyPartyMaster.company_id == None)

    rows = db.execute(q).scalars().all()
    if not rows:
        return None
    if company_id is not None:
        rows.sort(key=lambda r: 0 if r.company_id == company_id else 1)
    return rows[0]


def _resolve_sale_purc_type_for_busy(
    bill: Bill,
    mapping: Optional[BusyPartyMapping],
    party_state_code: Optional[str] = None,
) -> str:
    override = _normalize_sale_purc_override(mapping.sale_purc_type_override if mapping else None)
    if override:
        return override

    if party_state_code and party_state_code.isdigit():
        return "L/GST-ItemWise" if party_state_code == "08" else "I/GST-ItemWise"

    gstin = str((bill.broker.gstin if bill.broker else "") or "").strip().upper()
    if len(gstin) >= 2 and gstin[:2].isdigit():
        return "L/GST-ItemWise" if gstin[:2] == "08" else "I/GST-ItemWise"

    addr = str((bill.broker.address if bill.broker else "") or "").strip().lower()
    if addr:
        if "rajasthan" in addr or " rj " in f" {addr} " or "raj." in addr:
            return "L/GST-ItemWise"
        return "I/GST-ItemWise"

    # User-confirmed default when state cannot be determined.
    return "L/GST-ItemWise"


def _resolve_sale_purc_type_for_busy_staging(
    mapping: Optional[BusyPartyMapping],
    party_state_code: Optional[str] = None,
) -> str:
    override = _normalize_sale_purc_override(mapping.sale_purc_type_override if mapping else None)
    if override:
        return override

    if party_state_code and party_state_code.isdigit():
        return "L/GST-ItemWise" if party_state_code == "08" else "I/GST-ItemWise"

    return "L/GST-ItemWise"


def _collect_exported_busy_bill_ids(
    db: Session,
    company_id: Optional[int] = None,
    main_tender_id: Optional[int] = None,
) -> set[int]:
    clauses = ["export_type = :export_type", "file_path LIKE :file_pattern"]
    params = {
        "export_type": "purchase_bill",
        "file_pattern": "%purchase_bills_busy_%",
    }
    if company_id is not None:
        clauses.append("company_id = :company_id")
        params["company_id"] = company_id
    if main_tender_id is not None:
        clauses.append("main_tender_id = :main_tender_id")
        params["main_tender_id"] = main_tender_id

    sql = "SELECT record_ids FROM busy_exports WHERE " + " AND ".join(clauses)
    try:
        rows = db.execute(text(sql), params).mappings().all()
    except Exception as e:
        log.warning("busy_exports record_ids lookup skipped: %s", e)
        return set()

    out: set[int] = set()

    def _extract_bill_id(v) -> Optional[int]:
        if isinstance(v, int):
            return int(v)
        if isinstance(v, str) and v.strip().isdigit():
            return int(v.strip())
        if isinstance(v, dict):
            candidate = v.get("bill_id") or v.get("id")
            try:
                return int(candidate) if candidate is not None else None
            except Exception:
                return None
        return None

    for row in rows:
        vals_raw = row.get("record_ids")
        vals = vals_raw if isinstance(vals_raw, list) else []
        if isinstance(vals_raw, str):
            try:
                parsed = json.loads(vals_raw)
                if isinstance(parsed, list):
                    vals = parsed
            except Exception:
                vals = []
        for v in vals:
            bill_id = _extract_bill_id(v)
            if bill_id is not None:
                out.add(bill_id)
    return out


def _build_busy_ready_row_for_staging(
    db: Session,
    row: BusyStagingBill,
) -> tuple[Optional[dict], Optional[str]]:
    if not row:
        return None, "missing_busy_staging_bill"

    base = _busy_staging_to_dict(row)
    source_party_name = str(row.broker_name or "").strip()
    source_station = _extract_station_hint(source_party_name)

    mapping = _find_busy_party_mapping(db, row.company_id, source_party_name)
    mapped_party_name = mapping.busy_party_name if mapping else None
    party_name = str((mapped_party_name or source_party_name) or "").strip() or "UNKNOWN PARTY"
    party_state_code = None
    mapping_status = "manual" if mapping else "unmapped"
    mapping_confidence = 100.0 if mapping else None
    mapping_candidates = []
    suggested_party_name = None

    if mapping:
        matched_master = _find_busy_party_master_by_name(db, row.company_id, mapping.busy_party_name)
        if matched_master and matched_master.state_code:
            party_state_code = matched_master.state_code
    else:
        suggestion = _suggest_busy_party(
            db,
            company_id=row.company_id,
            source_party_name=source_party_name,
            source_state_code=None,
            source_station=source_station,
            source_gstin=None,
        )
        if suggestion:
            suggested_party_name = str(suggestion.get("party_name") or "").strip() or None
            mapping_confidence = float(suggestion.get("score") or 0.0)
            mapping_candidates = suggestion.get("candidates") or []
            party_state_code = suggestion.get("state_code") or None
            if suggestion.get("auto_pick") and suggested_party_name:
                mapped_party_name = suggested_party_name
                party_name = suggested_party_name
                mapping_status = "fuzzy"
            elif suggestion.get("ambiguous"):
                mapping_status = "ambiguous"
            else:
                mapping_status = "suggested"

    item_name = str(row.material_name or "").strip()
    qty = float(row.qty_mt or 0)
    price = float(row.rate_per_mt or 0)
    vch_bill_date = row.bill_date

    base.update({
        "source_party_name": source_party_name,
        "mapped_party_name": mapped_party_name,
        "suggested_party_name": suggested_party_name,
        "mapping_status": mapping_status,
        "mapping_confidence": mapping_confidence,
        "mapping_candidates": mapping_candidates,
        "source_state_code": None,
        "source_state_name": None,
        "source_station": source_station,
        "sale_purc_type_override": mapping.sale_purc_type_override if mapping else None,

        "bill_number": row.bill_number,
        "vehicle_number": row.vehicle_number,
        "material_name": item_name,
        "qty": qty,
        "price": price,
        "bill_date": str(vch_bill_date) if vch_bill_date else None,

        "vch_series": "Main",
        "vch_bill_date": vch_bill_date,
        "vch_bill_no": row.bill_number or str(row.id),
        "sale_purc_type": _resolve_sale_purc_type_for_busy_staging(mapping, party_state_code=party_state_code),
        "party_name": party_name,
        "mc_name": "Main Store",
        "item_name": item_name,
        "quantity": qty,
        "unit": "QUINTAL",
        "price_value": price,
        "itc_eligibility_type": "Input Goods/Services",
        "narration": row.vehicle_number or "",
    })
    return base, None


def _collect_busy_staging_rows(
    db: Session,
    company_id: Optional[int] = None,
    include_exported: bool = True,
    only_ids: Optional[set[int]] = None,
) -> tuple[List[dict], List[dict]]:
    q = select(BusyStagingBill).order_by(desc(BusyStagingBill.created_at))
    q = _apply_scope_filters(q, BusyStagingBill, company_id)
    if only_ids:
        q = q.where(BusyStagingBill.id.in_(list(only_ids)))

    rows: List[dict] = []
    rejected: List[dict] = []
    for row in db.execute(q).scalars().all():
        payload, reason = _build_busy_ready_row_for_staging(db, row)
        if payload:
            if (not include_exported) and payload.get("busy_exported"):
                continue
            rows.append(payload)
        else:
            rejected.append({"busy_staging_bill_id": row.id if row else None, "reason": reason})
    return rows, rejected


def _build_busy_ready_row_for_bill(
    db: Session,
    bill: Bill,
    exported_bill_ids: Optional[set[int]] = None,
) -> tuple[Optional[dict], Optional[str]]:
    if not bill:
        return None, "missing_bill"
    if not (bill.company_id and bill.main_tender_id and bill.tender_id):
        return None, "missing_tender_linkage"

    dispatches = _active_dispatches_for_bill(db, bill.id)
    if len(dispatches) != 1:
        return None, f"dispatch_count={len(dispatches)}"
    dispatch = dispatches[0]

    receipts = _active_receipts_for_dispatch(db, dispatch.id)
    if len(receipts) != 1:
        return None, f"receipt_count={len(receipts)}"
    receipt = receipts[0]
    if receipt.match_status not in ("auto", "manual"):
        return None, f"receipt_match_status={receipt.match_status or 'unknown'}"

    pb = db.execute(select(PurchaseBill).where(PurchaseBill.bill_id == bill.id)).scalar_one_or_none()
    if pb and pb.status == PurchaseBillStatus.cancelled:
        return None, "purchase_bill_cancelled"

    source_party_name = str(
        bill.broker_name
        or (bill.broker.name if bill.broker else "")
        or (pb.broker.name if pb and pb.broker else "")
        or ""
    ).strip()

    source_gstin = _clean_gstin((bill.broker.gstin if bill.broker else None) or (pb.broker.gstin if pb and pb.broker else None))
    source_state_code, source_state_name = _infer_bill_source_state(bill)
    source_station = _infer_bill_source_station(bill, source_party_name)

    mapping = _find_busy_party_mapping(db, bill.company_id, source_party_name)
    mapped_party_name = mapping.busy_party_name if mapping else None
    party_name = str((mapped_party_name or source_party_name) or "").strip() or "UNKNOWN PARTY"
    party_state_code = None
    mapping_status = "manual" if mapping else "unmapped"
    mapping_confidence = 100.0 if mapping else None
    mapping_candidates = []
    suggested_party_name = None

    if mapping:
        matched_master = _find_busy_party_master_by_name(db, bill.company_id, mapping.busy_party_name)
        if matched_master and matched_master.state_code:
            party_state_code = matched_master.state_code
    else:
        suggestion = _suggest_busy_party(
            db,
            company_id=bill.company_id,
            source_party_name=source_party_name,
            source_state_code=source_state_code,
            source_station=source_station,
            source_gstin=source_gstin,
        )
        if suggestion:
            suggested_party_name = str(suggestion.get("party_name") or "").strip() or None
            mapping_confidence = float(suggestion.get("score") or 0.0)
            mapping_candidates = suggestion.get("candidates") or []
            party_state_code = suggestion.get("state_code") or None
            if suggestion.get("auto_pick") and suggested_party_name:
                mapped_party_name = suggested_party_name
                party_name = suggested_party_name
                mapping_status = "fuzzy"
            elif suggestion.get("ambiguous"):
                mapping_status = "ambiguous"
            else:
                mapping_status = "suggested"

    vch_bill_date = receipt.receipt_date or bill.bill_date or (pb.bill_date if pb else None)
    item_name = str(
        bill.material_name
        or (bill.material.name if bill.material else "")
        or ""
    ).strip()
    qty = float(pb.qty_mt or 0) if pb and pb.qty_mt is not None else float(bill.qty_mt or dispatch.qty_mt or 0)
    price = float(pb.rate_per_mt or 0) if pb and pb.rate_per_mt is not None else float(bill.rate_per_mt or 0)
    busy_exported = bool(pb.busy_exported) if pb else False
    if not busy_exported and exported_bill_ids and bill.id in exported_bill_ids:
        busy_exported = True

    row = {
        "purchase_bill_id": pb.id if pb else None,
        "bill_id": bill.id,
        "company_id": bill.company_id,
        "main_tender_id": bill.main_tender_id,
        "tender_id": bill.tender_id,
        "deal_id": bill.deal_id or (pb.deal_id if pb else None) or dispatch.deal_id,
        "company_name": bill.company.name if bill.company else None,
        "main_tender_code": bill.main_tender.tender_code if bill.main_tender else None,
        "sub_tender_number": bill.tender.tender_number if bill.tender else None,
        "deal_number": bill.deal.deal_number if bill.deal else None,
        "dispatch_id": dispatch.id,
        "receipt_id": receipt.id,
        "status": pb.status.value if pb else bill.status.value,
        "busy_exported": busy_exported,

        "source_party_name": source_party_name,
        "mapped_party_name": mapped_party_name,
        "suggested_party_name": suggested_party_name,
        "mapping_status": mapping_status,
        "mapping_confidence": mapping_confidence,
        "mapping_candidates": mapping_candidates,
        "source_state_code": source_state_code,
        "source_state_name": source_state_name,
        "source_station": source_station,
        "sale_purc_type_override": mapping.sale_purc_type_override if mapping else None,

        "bill_number": bill.bill_number,
        "vehicle_number": bill.vehicle_number or dispatch.vehicle_number,
        "material_name": item_name,
        "qty": qty,
        "price": price,
        "receipt_date": str(vch_bill_date) if vch_bill_date else None,

        # Export payload fields
        "vch_series": "Main",
        "vch_bill_date": vch_bill_date,
        "vch_bill_no": bill.bill_number or (pb.pb_number if pb else None) or str(bill.id),
        "sale_purc_type": _resolve_sale_purc_type_for_busy(bill, mapping, party_state_code=party_state_code),
        "party_name": party_name,
        "mc_name": "Main Store",
        "item_name": item_name,
        "quantity": qty,
        "unit": "QUINTAL",
        "price_value": price,
        "itc_eligibility_type": "Input Goods/Services",
        "narration": bill.vehicle_number or dispatch.vehicle_number or "",
    }
    return row, None


def _busy_export_lineage_from_row(row: dict) -> dict:
    return {
        "bill_id": row.get("bill_id"),
        "purchase_bill_id": row.get("purchase_bill_id"),
        "company_id": row.get("company_id"),
        "company_name": row.get("company_name"),
        "main_tender_id": row.get("main_tender_id"),
        "main_tender_code": row.get("main_tender_code"),
        "tender_id": row.get("tender_id"),
        "sub_tender_number": row.get("sub_tender_number"),
        "deal_id": row.get("deal_id"),
        "deal_number": row.get("deal_number"),
        "dispatch_id": row.get("dispatch_id"),
        "receipt_id": row.get("receipt_id"),
    }


def _collect_busy_ready_bill_rows(
    db: Session,
    company_id: Optional[int] = None,
    main_tender_id: Optional[int] = None,
    tender_id: Optional[int] = None,
    include_exported: bool = False,
    only_bill_ids: Optional[set[int]] = None,
) -> tuple[List[dict], List[dict]]:
    q = select(Bill).order_by(desc(Bill.created_at))
    q = _apply_scope_filters(q, Bill, company_id, main_tender_id, tender_id)
    if only_bill_ids:
        q = q.where(Bill.id.in_(list(only_bill_ids)))

    exported_bill_ids = _collect_exported_busy_bill_ids(db, company_id=company_id, main_tender_id=main_tender_id)

    rows: List[dict] = []
    rejected: List[dict] = []
    for bill in db.execute(q).scalars().all():
        row, reason = _build_busy_ready_row_for_bill(db, bill, exported_bill_ids=exported_bill_ids)
        if row:
            if (not include_exported) and row.get("busy_exported"):
                continue
            rows.append(row)
        else:
            rejected.append({"bill_id": bill.id, "reason": reason})
    return rows, rejected


def _refresh_dispatch_status(db: Session, dispatch_id: Optional[int]) -> None:
    if not dispatch_id:
        return
    d = db.get(Dispatch, dispatch_id)
    if not d:
        return
    r = _active_receipt_for_dispatch(db, dispatch_id)
    if not r:
        d.status = DispatchStatus.in_transit
        return
    acc = float(r.accepted_mt or 0)
    rej = float(r.rejected_mt or 0)
    if rej == 0 and acc > 0:
        d.status = DispatchStatus.accepted
    elif acc == 0 and rej > 0:
        d.status = DispatchStatus.rejected
    else:
        d.status = DispatchStatus.partial


def _apply_receipt_delta(
    db: Session,
    dispatch_id: Optional[int],
    delta_accepted: float,
    delta_rejected: float,
) -> None:
    if not dispatch_id:
        return
    d = db.get(Dispatch, dispatch_id)
    if not d:
        raise HTTPException(404, "Linked dispatch not found")

    delta_total = float(delta_accepted or 0) + float(delta_rejected or 0)
    new_consumed = float(d.consumed_qty_qtl or 0) + delta_total
    if new_consumed < -1e-9:
        raise HTTPException(400, "Invalid reconciliation: consumed quantity would become negative")
    if new_consumed - float(d.qty_mt or 0) > 1e-9:
        raise HTTPException(400, "Receipt quantity exceeds dispatch remaining quantity")

    d.consumed_qty_qtl = max(round(new_consumed, 3), 0)

    if d.deal_id:
        deal = db.get(Deal, d.deal_id)
        if deal:
            new_acc = float(deal.accepted_mt or 0) + float(delta_accepted or 0)
            new_rej = float(deal.rejected_mt or 0) + float(delta_rejected or 0)
            if new_acc < -1e-9 or new_rej < -1e-9:
                raise HTTPException(400, "Invalid reconciliation: deal accepted/rejected would become negative")
            deal.accepted_mt = max(round(new_acc, 3), 0)
            deal.rejected_mt = max(round(new_rej, 3), 0)


def _enrich_dispatch_material(db: Session, d: Dispatch) -> None:
    mid = d.material_id
    mname = d.material_name

    if not mid and d.bill_id:
        b = db.get(Bill, d.bill_id)
        if b:
            mid = mid or b.material_id
            mname = mname or b.material_name

    if (not mid or not mname) and d.deal_id:
        deal = db.get(Deal, d.deal_id)
        if deal and deal.material_id:
            mid = mid or deal.material_id
            if not mname and deal.material:
                mname = deal.material.name

    rid, rname = _resolve_material(db, mid, mname)
    d.material_id = rid
    d.material_name = rname


def _enrich_receipt_material(db: Session, r: PlantReceipt) -> None:
    mid = r.material_id
    mname = r.material_name
    if (not mid or not mname) and r.dispatch_id:
        d = db.get(Dispatch, r.dispatch_id)
        if d:
            _enrich_dispatch_material(db, d)
            mid = mid or d.material_id
            mname = mname or d.material_name
    rid, rname = _resolve_material(db, mid, mname)
    r.material_id = rid
    r.material_name = rname


def _match_receipt_to_dispatch(db: Session, r: PlantReceipt, manual: bool = False) -> Optional[int]:
    # If manually linked, trust dispatch_id and mark manual.
    if manual and r.dispatch_id:
        r.matched_dispatch_id = r.dispatch_id
        r.match_status = "manual"
        r.match_reason = "Linked manually by operator"
        return r.dispatch_id

    _enrich_receipt_material(db, r)
    if not r.vehicle_number or not r.plant_id or not r.material_id:
        r.match_status = "unmatched"
        r.match_reason = "Missing matching keys (truck/plant/material)"
        return None

    truck = _norm_vehicle(r.vehicle_number)
    candidates = db.execute(
        select(Dispatch).where(
            Dispatch.is_deleted == False,
            Dispatch.plant_id == r.plant_id,
            Dispatch.material_id == r.material_id,
        )
    ).scalars().all()

    filtered: List[Dispatch] = []
    for d in candidates:
        if _norm_vehicle(d.vehicle_number) != truck:
            continue
        if _dispatch_remaining_qtl(d) <= 0:
            continue
        other_receipt = _active_receipt_for_dispatch(db, d.id, exclude_receipt_id=r.id)
        if other_receipt:
            continue
        # Date Sanity Rule: Receipt date must be on or after dispatch date
        if r.receipt_date and d.dispatch_date and r.receipt_date < d.dispatch_date:
            r.match_status = "flagged_date"
            r.match_reason = f"Receipt date ({r.receipt_date}) is earlier than dispatch date ({d.dispatch_date})"
            continue
        filtered.append(d)

    if len(filtered) == 1:
        d = filtered[0]
        r.dispatch_id = d.id
        r.matched_dispatch_id = d.id
        r.match_status = "auto"
        r.match_reason = "Auto-matched on truck + plant + material"
        return d.id
    if len(filtered) > 1:
        r.match_status = "ambiguous"
        r.match_reason = f"Multiple dispatch candidates ({len(filtered)}) on truck + plant + material"
        return None

    r.match_status = "unmatched"
    r.match_reason = "No open dispatch found on truck + plant + material"
    return None


def _validate_receipt_dispatch_keys(r: PlantReceipt, d: Dispatch) -> None:
    if _norm_vehicle(r.vehicle_number) != _norm_vehicle(d.vehicle_number):
        raise HTTPException(400, "Dispatch vehicle does not match receipt vehicle")
    if r.plant_id and d.plant_id and r.plant_id != d.plant_id:
        raise HTTPException(400, "Dispatch plant does not match receipt plant")
    if r.material_id and d.material_id and r.material_id != d.material_id:
        raise HTTPException(400, "Dispatch material does not match receipt material")


def _soft_delete_receipt_with_revert(db: Session, r: PlantReceipt) -> None:
    if not r or r.is_deleted:
        return
    old_dispatch_id = r.dispatch_id
    old_acc = float(r.accepted_mt or 0)
    old_rej = float(r.rejected_mt or 0)
    _apply_receipt_delta(db, old_dispatch_id, -old_acc, -old_rej)
    r.is_deleted = True
    r.deleted_at = datetime.now()
    db.flush()
    _refresh_dispatch_status(db, old_dispatch_id)


def _manual_match_candidates_for_receipt(db: Session, r: PlantReceipt) -> List[Dispatch]:
    _enrich_receipt_material(db, r)
    if not r.vehicle_number or not r.plant_id:
        return []

    q = select(Dispatch).where(
        Dispatch.is_deleted == False,
        Dispatch.plant_id == r.plant_id,
    )
    if r.material_id:
        q = q.where(Dispatch.material_id == r.material_id)

    truck = _norm_vehicle(r.vehicle_number)
    out: List[Dispatch] = []
    for d in db.execute(q.order_by(desc(Dispatch.dispatch_date), desc(Dispatch.id))).scalars().all():
        _enrich_dispatch_material(db, d)
        if _norm_vehicle(d.vehicle_number) != truck:
            continue
        if _dispatch_remaining_qtl(d) <= 0:
            continue
        existing = _active_receipt_for_dispatch(db, d.id, exclude_receipt_id=r.id)
        if existing:
            continue
        out.append(d)
    return out


def _auto_match_open_receipt_for_dispatch(db: Session, d: Dispatch) -> Optional[int]:
    if not d or d.is_deleted:
        return None

    _enrich_dispatch_material(db, d)
    if not d.vehicle_number or not d.plant_id or not d.material_id:
        return None

    if _active_receipt_for_dispatch(db, d.id):
        return None

    # Guard against double-application if legacy data has consumed quantity without active receipt.
    if float(d.consumed_qty_qtl or 0) > 1e-9:
        return None

    if _dispatch_remaining_qtl(d) <= 0:
        return None

    truck = _norm_vehicle(d.vehicle_number)
    candidates = db.execute(
        select(PlantReceipt).where(
            PlantReceipt.is_deleted == False,
            PlantReceipt.dispatch_id.is_(None),
            PlantReceipt.plant_id == d.plant_id,
            PlantReceipt.material_id == d.material_id,
        )
    ).scalars().all()

    eligible: List[PlantReceipt] = []
    for r in candidates:
        if _norm_vehicle(r.vehicle_number) != truck:
            continue
        eligible.append(r)

    if len(eligible) != 1:
        if len(eligible) > 1:
            log.info(
                "Dispatch %s auto-match skipped due to ambiguous receipts (%s) on truck + plant + material",
                d.id,
                len(eligible),
            )
        return None

    r = eligible[0]
    _enrich_receipt_material(db, r)
    _validate_receipt_dispatch_keys(r, d)

    try:
        _apply_receipt_delta(db, d.id, float(r.accepted_mt or 0), float(r.rejected_mt or 0))
    except HTTPException as exc:
        r.dispatch_id = None
        r.matched_dispatch_id = None
        r.matched_qty_qtl = 0
        r.match_applied_at = None
        r.match_status = "unmatched"
        r.match_reason = f"Auto-match not applied: {exc.detail}"
        return None

    _stamp_lineage(r, d.company_id, d.main_tender_id, d.tender_id)
    r.dispatch_id = d.id
    r.matched_dispatch_id = d.id
    r.matched_qty_qtl = float(r.accepted_mt or 0) + float(r.rejected_mt or 0)
    r.match_applied_at = datetime.now()
    r.match_status = "auto"
    r.match_reason = "Auto-matched on truck + plant + material (dispatch created/updated)"
    # SessionLocal uses autoflush=False; flush now so immediate status refresh sees this link.
    db.flush()
    return r.id


def _sync_receipt_from_unloading_entry(db: Session, e: PlantUnloadingEntry) -> Optional[int]:
    if not e or not e.receipt_id:
        return None

    r = db.get(PlantReceipt, e.receipt_id)
    if not r or r.is_deleted:
        e.receipt_id = None
        e.receipt_created = False
        return None

    master = e.master
    old_dispatch_id = r.dispatch_id
    old_accepted = float(r.accepted_mt or 0)
    old_rejected = float(r.rejected_mt or 0)

    _apply_receipt_delta(db, old_dispatch_id, -old_accepted, -old_rejected)

    accepted = _effective_unloading_qty_mt(e.net_qty_mt, e.received_qty_mt)
    received = _to_float(e.received_qty_mt)
    if received is None:
        received = accepted
    rejected = max(received - accepted, 0)

    r.vehicle_number = e.truck_number
    _validate_vehicle_number(r.vehicle_number)
    r.receipt_date = e.entry_date
    r.accepted_mt = accepted
    r.rejected_mt = rejected
    r.received_qty_qtl = received

    if master:
        r.rm_number = master.rm_number
        r.party_name = master.party_name
        r.po_number = master.po_number

    item_name = e.item_name or (master.item_name if master else None)
    material_id, material_name = _resolve_material(db, material_name=item_name)
    r.material_id = material_id
    r.material_name = material_name

    plant_id = master.plant_id if master else None
    if not plant_id and master and master.plant_name:
        pid, pname = _resolve_plant(db, master.plant_name)
        if pid:
            master.plant_id = pid
            master.plant_name = pname
            plant_id = pid
    if plant_id:
        r.plant_id = plant_id

    company_id_ctx = e.company_id if e.company_id is not None else (master.company_id if master else None)
    main_tender_id_ctx = e.main_tender_id if e.main_tender_id is not None else (master.main_tender_id if master else None)
    tender_id_ctx = e.tender_id if e.tender_id is not None else (master.tender_id if master else None)
    _stamp_lineage(r, company_id_ctx, main_tender_id_ctx, tender_id_ctx)

    keep_dispatch_id = None
    if old_dispatch_id:
        d_old = db.get(Dispatch, old_dispatch_id)
        if d_old and not d_old.is_deleted:
            _enrich_dispatch_material(db, d_old)
            existing = _active_receipt_for_dispatch(db, d_old.id, exclude_receipt_id=r.id)
            if not existing:
                try:
                    _validate_receipt_dispatch_keys(r, d_old)
                    keep_dispatch_id = d_old.id
                except HTTPException:
                    keep_dispatch_id = None

    if keep_dispatch_id:
        r.dispatch_id = keep_dispatch_id
        if r.match_status != "manual":
            r.match_status = "auto"
            r.match_reason = "Auto-matched on truck + plant + material"
    else:
        r.dispatch_id = None
        r.matched_dispatch_id = None
        r.matched_qty_qtl = 0
        r.match_applied_at = None
        _match_receipt_to_dispatch(db, r, manual=False)

    if r.dispatch_id:
        try:
            _apply_receipt_delta(db, r.dispatch_id, accepted, rejected)
            r.matched_dispatch_id = r.dispatch_id
            r.matched_qty_qtl = accepted + rejected
            r.match_applied_at = datetime.now()
        except HTTPException as exc:
            r.dispatch_id = None
            r.matched_dispatch_id = None
            r.matched_qty_qtl = 0
            r.match_applied_at = None
            r.match_status = "unmatched"
            r.match_reason = f"Auto-match not applied: {exc.detail}"

    db.flush()
    _refresh_dispatch_status(db, old_dispatch_id)
    _refresh_dispatch_status(db, r.dispatch_id)
    return r.id


def _get_or_create_unloading_master(
    db: Session,
    rm_number: Optional[str],
    item_name: Optional[str],
    party_name: Optional[str],
    plant_name: Optional[str],
    po_number: Optional[str],
    company_id: Optional[int] = None,
    main_tender_id: Optional[int] = None,
    tender_id: Optional[int] = None,
    truck_number: Optional[str] = None,
) -> PlantUnloadingMaster:
    rm = _clean_optional_text(rm_number) or "UNKNOWN"
    item = _clean_optional_text(item_name) or "UNKNOWN"
    party_raw = _clean_optional_text(party_name) or "UNKNOWN"
    plant_hint = _clean_optional_text(plant_name)
    po_clean = _clean_optional_text(po_number)
    party = _normalize_unloading_party_name(db, party_raw, company_id)

    rm_norm, rm_base, _rm_suffix = _normalize_rm_components(rm)
    rm_n = _norm_text(rm)
    rm_norm_n = _norm_text(rm_norm)
    rm_base_n = _norm_text(rm_base)
    item_n = _norm_text(item)
    party_n = _norm_text(party)
    party_raw_n = _norm_text(party_raw)

    assignment = _resolve_unloading_assignment(
        db,
        rm_number=rm,
        item_name=item,
        plant_name=plant_hint,
        company_id_hint=company_id,
        main_tender_id_hint=main_tender_id,
        tender_id_hint=tender_id,
    )
    scope_company = assignment.get("company_id") if assignment.get("company_id") is not None else company_id
    scope_main = assignment.get("main_tender_id") if assignment.get("main_tender_id") is not None else main_tender_id
    scope_tender = assignment.get("tender_id") if assignment.get("tender_id") is not None else tender_id

    plant_id_hint = None
    if plant_hint:
        plant_id_hint, _plant_name_resolved = _resolve_plant(db, plant_hint)

    rm_match_parts = []
    if rm_norm_n:
        rm_match_parts.append(func.lower(func.coalesce(PlantUnloadingMaster.rm_number_norm, "")) == rm_norm_n)
    if rm_n:
        rm_match_parts.append(func.lower(func.coalesce(PlantUnloadingMaster.rm_number, "")) == rm_n)

    existing_q = select(PlantUnloadingMaster).where(
        func.lower(func.coalesce(PlantUnloadingMaster.item_name, "")) == item_n,
        func.lower(func.coalesce(PlantUnloadingMaster.party_name, "")) == party_n,
        or_(*rm_match_parts),
    )
    existing_q = _apply_unloading_master_scope_hints(existing_q, scope_company, scope_main, scope_tender)
    existing_candidates = db.execute(
        existing_q.order_by(desc(PlantUnloadingMaster.is_manual_override), desc(PlantUnloadingMaster.updated_at))
    ).scalars().all()
    existing = None
    if existing_candidates:
        if not plant_id_hint and len(existing_candidates) > 1:
            existing = _pick_unloading_master_without_plant(existing_candidates, truck_number)
        else:
            existing = existing_candidates[0]

    # Backward-compat: reuse legacy masters keyed by pre-normalized party label.
    if not existing and party_raw_n and party_raw_n != party_n:
        legacy_q = select(PlantUnloadingMaster).where(
            func.lower(func.coalesce(PlantUnloadingMaster.item_name, "")) == item_n,
            func.lower(func.coalesce(PlantUnloadingMaster.party_name, "")) == party_raw_n,
            or_(*rm_match_parts),
        )
        legacy_q = _apply_unloading_master_scope_hints(legacy_q, scope_company, scope_main, scope_tender)
        legacy_candidates = db.execute(
            legacy_q.order_by(desc(PlantUnloadingMaster.is_manual_override), desc(PlantUnloadingMaster.updated_at))
        ).scalars().all()
        if legacy_candidates:
            if not plant_id_hint and len(legacy_candidates) > 1:
                existing = _pick_unloading_master_without_plant(legacy_candidates, truck_number)
            else:
                existing = legacy_candidates[0]

    # If plant is missing/unresolved and scoped matching fails, broaden search to
    # existing registers in same company/main scope and use historical row hints.
    if not existing and not plant_id_hint:
        broad_q = select(PlantUnloadingMaster).where(
            func.lower(func.coalesce(PlantUnloadingMaster.item_name, "")) == item_n,
            func.lower(func.coalesce(PlantUnloadingMaster.party_name, "")) == party_n,
            or_(*rm_match_parts),
        )
        if scope_company is not None:
            broad_q = broad_q.where(or_(PlantUnloadingMaster.company_id == scope_company, PlantUnloadingMaster.company_id == None))
        if scope_main is not None:
            broad_q = broad_q.where(or_(PlantUnloadingMaster.main_tender_id == scope_main, PlantUnloadingMaster.main_tender_id == None))
        broad_candidates = db.execute(broad_q).scalars().all()

        if (not broad_candidates) and party_raw_n and party_raw_n != party_n:
            broad_legacy_q = select(PlantUnloadingMaster).where(
                func.lower(func.coalesce(PlantUnloadingMaster.item_name, "")) == item_n,
                func.lower(func.coalesce(PlantUnloadingMaster.party_name, "")) == party_raw_n,
                or_(*rm_match_parts),
            )
            if scope_company is not None:
                broad_legacy_q = broad_legacy_q.where(or_(PlantUnloadingMaster.company_id == scope_company, PlantUnloadingMaster.company_id == None))
            if scope_main is not None:
                broad_legacy_q = broad_legacy_q.where(or_(PlantUnloadingMaster.main_tender_id == scope_main, PlantUnloadingMaster.main_tender_id == None))
            broad_candidates = db.execute(broad_legacy_q).scalars().all()

        picked = _pick_unloading_master_without_plant(broad_candidates, truck_number)
        if picked is not None:
            existing = picked

    # Re-upload reuse: attach to highly similar existing register in same scope.
    if not existing:
        reuse_parts = []
        if rm_norm_n:
            reuse_parts.append(func.lower(func.coalesce(PlantUnloadingMaster.rm_number_norm, "")) == rm_norm_n)
        if rm_base_n:
            reuse_parts.append(func.lower(func.coalesce(PlantUnloadingMaster.rm_number_base, "")) == rm_base_n)
        if rm_n:
            reuse_parts.append(func.lower(func.coalesce(PlantUnloadingMaster.rm_number, "")) == rm_n)

        if reuse_parts:
            cand_q = select(PlantUnloadingMaster).where(or_(*reuse_parts))
            cand_q = _apply_unloading_master_scope_hints(cand_q, scope_company, scope_main, scope_tender)
            candidates = db.execute(cand_q).scalars().all()

            best = None
            best_score = 0.0
            for cand in candidates:
                item_score = _name_similarity(item, cand.item_name)
                party_score = max(_name_similarity(party, cand.party_name), _name_similarity(party_raw, cand.party_name))
                cand_plant = cand.plant.name if cand.plant else cand.plant_name
                if plant_hint and cand_plant:
                    plant_score = _name_similarity(plant_hint, cand_plant)
                elif not plant_hint and not cand_plant:
                    plant_score = 0.6
                else:
                    plant_score = 0.25

                score = (item_score * 0.55) + (party_score * 0.35) + (plant_score * 0.10)
                if cand.is_manual_override:
                    score += 0.05
                if score > best_score:
                    best_score = score
                    best = cand

            if best is not None and best_score >= 0.84:
                existing = best

    if existing:
        if party and _norm_text(existing.party_name) != party_n:
            existing.party_name = party
        if po_clean and not existing.po_number:
            existing.po_number = po_clean
        if plant_hint and not existing.plant_name:
            pid, pname = _resolve_plant(db, plant_hint)
            existing.plant_id = pid or existing.plant_id
            existing.plant_name = pname or existing.plant_name
        if rm_norm and not existing.rm_number_norm:
            existing.rm_number_norm = rm_norm
        if rm_base and not existing.rm_number_base:
            existing.rm_number_base = rm_base

        # Preserve manual overrides and existing confidently assigned registers.
        if existing.is_manual_override and existing.tender_id:
            existing.assignment_status = "assigned"
            existing.requires_manual_assignment = False
            existing.mapping_source = existing.mapping_source or "manual"
        elif not (existing.assignment_status == "assigned" and existing.tender_id):
            _apply_unloading_assignment_to_master(existing, assignment)
            _sync_unloading_lineage_from_master(db, existing)

        return existing

    pid, pname = _resolve_plant(db, plant_hint)
    m = PlantUnloadingMaster(
        rm_number=rm,
        rm_number_norm=rm_norm,
        rm_number_base=rm_base,
        item_name=item,
        party_name=party,
        plant_id=pid,
        plant_name=pname,
        po_number=po_clean,
    )
    db.add(m)
    db.flush()
    _apply_unloading_assignment_to_master(m, assignment)
    _sync_unloading_lineage_from_master(db, m)
    db.flush()
    return m


def _build_unloading_dedupe_key(
    master_id: int,
    ws_no: Optional[str],
    entry_date: date,
    truck_number: Optional[str],
    net_qty: Optional[float] = None,
    received_qty: Optional[float] = None,
) -> str:
    # Duplicate identity rule: same register + same truck + same effective weight.
    truck_key = _norm_code(truck_number)
    weight = _effective_unloading_qty_mt(_to_float(net_qty), _to_float(received_qty))
    weight_key = f"{weight:.3f}" if weight > 0 else "na"
    return f"{master_id}|{truck_key}|{weight_key}"


def _find_duplicate_unloading_row_for_master(
    existing_rows: List[PlantUnloadingEntry],
    *,
    truck_number: Optional[str],
    net_qty: Optional[float],
    received_qty: Optional[float],
    exclude_entry_id: Optional[int] = None,
) -> Optional[PlantUnloadingEntry]:
    truck_key = _norm_code(truck_number)
    if not truck_key:
        return None

    incoming_weight = _effective_unloading_qty_mt(net_qty, received_qty)
    if incoming_weight <= 0:
        return None

    for ex in existing_rows:
        if exclude_entry_id is not None and ex.id == exclude_entry_id:
            continue
        if truck_key != _norm_code(ex.truck_number):
            continue

        ex_weight = _effective_unloading_qty_mt(
            _to_float(ex.net_qty_mt),
            _to_float(ex.received_qty_mt),
        )
        if ex_weight <= 0:
            continue

        if abs(incoming_weight - ex_weight) <= 0.02:
            return ex

    return None


def _is_duplicate_unloading_row_for_master(
    existing_rows: List[PlantUnloadingEntry],
    *,
    ws_no: Optional[str],
    entry_date: date,
    truck_number: Optional[str],
    net_qty: Optional[float],
    received_qty: Optional[float],
    no_of_bags: Optional[int],
) -> bool:
    return _find_duplicate_unloading_row_for_master(
        existing_rows,
        truck_number=truck_number,
        net_qty=net_qty,
        received_qty=received_qty,
    ) is not None


def _unloading_master_merge_group_key(m: PlantUnloadingMaster) -> tuple:
    rm_norm, _rm_base, _rm_suffix = _normalize_rm_components(m.rm_number)
    rm_key = _norm_text(m.rm_number_norm or rm_norm or m.rm_number)
    item_key = _norm_text(m.item_name)
    party_key = _norm_text(m.party_name)
    return (m.company_id, m.main_tender_id, rm_key, item_key, party_key)


def _should_preserve_unloading_source_master(
    source: PlantUnloadingMaster,
    target: PlantUnloadingMaster,
    source_entries: List[PlantUnloadingEntry],
    target_entries: List[PlantUnloadingEntry],
) -> bool:
    source_approved = sum(1 for e in source_entries if e.status == BillStatus.approved)
    target_approved = sum(1 for e in target_entries if e.status == BillStatus.approved)

    # Protect manually curated lineage registers from auto-merge into another tender.
    if source.is_manual_override and source.tender_id and source.tender_id != target.tender_id:
        return True

    # If both sides already have approved history on different subtenders, assume
    # they represent distinct operational registers and do not auto-merge.
    if (
        source.tender_id
        and target.tender_id
        and source.tender_id != target.tender_id
        and source_approved > 0
        and target_approved > 0
    ):
        return True

    return False


def _prefer_source_unloading_entry(source_entry: PlantUnloadingEntry, target_entry: PlantUnloadingEntry) -> bool:
    source_has_receipt = bool(source_entry.receipt_id)
    target_has_receipt = bool(target_entry.receipt_id)

    # Never replace an entry already linked to a receipt; it can hold dispatch lineage.
    if target_has_receipt:
        return False
    if source_has_receipt and not target_has_receipt:
        return True

    source_approved = source_entry.status == BillStatus.approved
    target_approved = target_entry.status == BillStatus.approved
    if source_approved != target_approved:
        return source_approved

    source_reviewed = bool(source_entry.reviewed_at)
    target_reviewed = bool(target_entry.reviewed_at)
    if source_reviewed != target_reviewed:
        return source_reviewed

    return False


def _merge_unloading_duplicate_masters(
    db: Session,
    *,
    company_id: Optional[int] = None,
    main_tender_id: Optional[int] = None,
    rm_number: Optional[str] = None,
    dry_run: bool = True,
    delete_empty_masters: bool = True,
) -> dict:
    q = select(PlantUnloadingMaster)
    if company_id is not None:
        q = q.where(PlantUnloadingMaster.company_id == company_id)
    if main_tender_id is not None:
        q = q.where(PlantUnloadingMaster.main_tender_id == main_tender_id)

    rm_norm, rm_base, _rm_suffix = _normalize_rm_components(rm_number)
    if rm_number:
        rm_raw_n = _norm_text(rm_number)
        rm_norm_n = _norm_text(rm_norm)
        rm_base_n = _norm_text(rm_base)
        parts = []
        if rm_raw_n:
            parts.append(func.lower(func.coalesce(PlantUnloadingMaster.rm_number, "")) == rm_raw_n)
        if rm_norm_n:
            parts.append(func.lower(func.coalesce(PlantUnloadingMaster.rm_number_norm, "")) == rm_norm_n)
        if rm_base_n:
            parts.append(func.lower(func.coalesce(PlantUnloadingMaster.rm_number_base, "")) == rm_base_n)
        if parts:
            q = q.where(or_(*parts))

    masters = db.execute(q.order_by(desc(PlantUnloadingMaster.updated_at), desc(PlantUnloadingMaster.id))).scalars().all()

    groups = {}
    for m in masters:
        key = _unloading_master_merge_group_key(m)
        groups.setdefault(key, []).append(m)

    stats = {
        "groups_considered": 0,
        "groups_skipped_ambiguous": 0,
        "groups_with_actions": 0,
        "source_masters_considered": 0,
        "source_masters_skipped_protected": 0,
        "source_masters_merged": 0,
        "entries_moved": 0,
        "entries_deleted_as_duplicate": 0,
        "entries_target_replaced": 0,
        "entries_conflicted_receipt": 0,
        "empty_masters_candidates": 0,
        "empty_masters_deleted": 0,
    }
    impacted_tender_ids = set()
    details = []

    for key, candidates in groups.items():
        if len(candidates) < 2:
            continue

        stats["groups_considered"] += 1
        target = _pick_unloading_master_without_plant(candidates, None)
        if target is None:
            stats["groups_skipped_ambiguous"] += 1
            details.append({
                "group": {
                    "company_id": key[0],
                    "main_tender_id": key[1],
                    "rm": key[2],
                    "item": key[3],
                    "party": key[4],
                },
                "target_master_id": None,
                "source_master_ids": [m.id for m in candidates],
                "note": "ambiguous_target",
            })
            continue

        target_entries = list(target.entries or [])
        group_actions = {
            "entries_moved": 0,
            "entries_deleted_as_duplicate": 0,
            "entries_target_replaced": 0,
            "entries_conflicted_receipt": 0,
            "source_masters_merged": 0,
            "source_masters_skipped_protected": 0,
            "empty_masters_candidates": 0,
            "empty_masters_deleted": 0,
        }

        source_ids = []
        for source in candidates:
            if source.id == target.id:
                continue

            source_ids.append(source.id)
            stats["source_masters_considered"] += 1
            source_entries = list(source.entries or [])

            if _should_preserve_unloading_source_master(source, target, source_entries, target_entries):
                stats["source_masters_skipped_protected"] += 1
                group_actions["source_masters_skipped_protected"] += 1
                continue

            source_remaining_est = len(source_entries)
            source_changes = 0

            for e in source_entries:
                target_key = _build_unloading_dedupe_key(
                    target.id,
                    e.ws_no,
                    e.entry_date,
                    e.truck_number,
                    _to_float(e.net_qty_mt),
                    _to_float(e.received_qty_mt),
                )
                target_rows = db.execute(
                    select(PlantUnloadingEntry).where(PlantUnloadingEntry.master_id == target.id)
                ).scalars().all()
                dup = _find_duplicate_unloading_row_for_master(
                    target_rows,
                    truck_number=e.truck_number,
                    net_qty=_to_float(e.net_qty_mt),
                    received_qty=_to_float(e.received_qty_mt),
                    exclude_entry_id=e.id,
                )

                if dup:
                    source_has_receipt = bool(e.receipt_id)
                    target_has_receipt = bool(dup.receipt_id)

                    if source_has_receipt and target_has_receipt:
                        stats["entries_conflicted_receipt"] += 1
                        group_actions["entries_conflicted_receipt"] += 1
                        continue

                    keep_source = _prefer_source_unloading_entry(e, dup)
                    if keep_source:
                        stats["entries_target_replaced"] += 1
                        group_actions["entries_target_replaced"] += 1
                        source_remaining_est -= 1
                        source_changes += 1
                        impacted_tender_ids.update([e.tender_id, dup.tender_id, source.tender_id, target.tender_id])

                        if not dry_run:
                            db.delete(dup)
                            db.flush()
                            e.master_id = target.id
                            e.master = target
                            e.dedupe_key = target_key
                            _stamp_lineage(e, target.company_id, target.main_tender_id, target.tender_id)
                            if e.receipt_id:
                                _sync_receipt_from_unloading_entry(db, e)
                    else:
                        stats["entries_deleted_as_duplicate"] += 1
                        group_actions["entries_deleted_as_duplicate"] += 1
                        source_remaining_est -= 1
                        source_changes += 1
                        impacted_tender_ids.update([e.tender_id, source.tender_id])

                        if not dry_run:
                            db.delete(e)
                    continue

                stats["entries_moved"] += 1
                group_actions["entries_moved"] += 1
                source_remaining_est -= 1
                source_changes += 1
                impacted_tender_ids.update([e.tender_id, source.tender_id, target.tender_id])

                if not dry_run:
                    e.master_id = target.id
                    e.master = target
                    e.dedupe_key = target_key
                    _stamp_lineage(e, target.company_id, target.main_tender_id, target.tender_id)
                    if e.receipt_id:
                        _sync_receipt_from_unloading_entry(db, e)

            if source_changes > 0:
                stats["source_masters_merged"] += 1
                group_actions["source_masters_merged"] += 1

            if delete_empty_masters and source_remaining_est == 0:
                stats["empty_masters_candidates"] += 1
                group_actions["empty_masters_candidates"] += 1
                if not dry_run:
                    db.flush()
                    remaining = db.execute(
                        select(func.count()).select_from(PlantUnloadingEntry).where(PlantUnloadingEntry.master_id == source.id)
                    ).scalar() or 0
                    if remaining == 0:
                        db.delete(source)
                        stats["empty_masters_deleted"] += 1
                        group_actions["empty_masters_deleted"] += 1

        group_had_actions = any(group_actions[k] > 0 for k in (
            "entries_moved",
            "entries_deleted_as_duplicate",
            "entries_target_replaced",
            "source_masters_merged",
            "empty_masters_deleted",
        ))
        if group_had_actions:
            stats["groups_with_actions"] += 1
            if not dry_run:
                db.flush()
                _sync_unloading_lineage_from_master(db, target)

        details.append({
            "group": {
                "company_id": key[0],
                "main_tender_id": key[1],
                "rm": key[2],
                "item": key[3],
                "party": key[4],
            },
            "target_master_id": target.id,
            "source_master_ids": source_ids,
            "actions": group_actions,
        })

    return {
        "stats": stats,
        "impacted_tender_ids": sorted({int(tid) for tid in impacted_tender_ids if tid}),
        "groups": details,
    }


def _plant_unloading_to_dict(e: PlantUnloadingEntry) -> dict:
    m = e.master
    return {
        "id": e.id,
        "master_id": e.master_id,
        "company_id": e.company_id if e.company_id is not None else (m.company_id if m else None),
        "main_tender_id": e.main_tender_id if e.main_tender_id is not None else (m.main_tender_id if m else None),
        "tender_id": e.tender_id if e.tender_id is not None else (m.tender_id if m else None),
        "rm_number": m.rm_number if m else None,
        "rm_number_norm": m.rm_number_norm if m else None,
        "rm_number_base": m.rm_number_base if m else None,
        "item_name": e.item_name or (m.item_name if m else None),
        "party_name": m.party_name if m else None,
        "plant_id": (m.plant_id if m else None),
        "plant_name": (m.plant.name if m and m.plant else (m.plant_name if m else None)),
        "po_number": m.po_number if m else None,
        "assignment_status": m.assignment_status if m else None,
        "assignment_reason": m.assignment_reason if m else None,
        "assignment_confidence": float(m.assignment_confidence) if m and m.assignment_confidence is not None else None,
        "mapping_source": m.mapping_source if m else None,
        "requires_manual_assignment": bool(m.requires_manual_assignment) if m else False,
        "is_manual_override": bool(m.is_manual_override) if m else False,
        "manual_assigned_by": m.manual_assigned_by if m else None,
        "manual_assigned_at": str(m.manual_assigned_at) if m and m.manual_assigned_at else None,
        "ws_no": e.ws_no,
        "entry_date": str(e.entry_date) if e.entry_date else None,
        "truck_number": e.truck_number,
        "no_of_bags": e.no_of_bags,
        "received_qty_mt": float(e.received_qty_mt) if e.received_qty_mt is not None else None,
        "net_qty_mt": float(e.net_qty_mt) if e.net_qty_mt is not None else None,
        "total_qty_mt": float(e.total_qty_mt) if e.total_qty_mt is not None else None,
        "status": e.status.value if hasattr(e.status, 'value') else str(e.status),
        "reviewed_by": e.reviewed_by,
        "reviewed_at": str(e.reviewed_at) if e.reviewed_at else None,
        "receipt_id": e.receipt_id,
        "receipt_created": bool(e.receipt_created),
        "ocr_source": e.ocr_source.value if e.ocr_source else None,
        "ocr_confidence": float(e.ocr_confidence) if e.ocr_confidence is not None else None,
        "image_path": e.image_path,
        "notes": e.notes,
        "created_at": str(e.created_at) if e.created_at else None,
    }


def _pending_ingest_to_dict(p: PendingIngest) -> dict:
    return {
        "id": p.id,
        "company_id": p.company_id,
        "main_tender_id": p.main_tender_id,
        "tender_id": p.tender_id,
        "source": p.source.value if p.source else None,
        "source_address": p.source_address,
        "source_account": p.source_account,
        "source_message_id": p.source_message_id,
        "file_name": p.file_name,
        "file_path": p.file_path,
        "file_hash": p.file_hash,
        "document_type": p.document_type.value if p.document_type else None,
        "classifier_confidence": float(p.classifier_confidence) if p.classifier_confidence is not None else None,
        "classifier_candidates": p.classifier_candidates or [],
        "extracted_payload": p.extracted_payload or {},
        "unclear_fields": p.unclear_fields or [],
        "status": p.status.value if p.status else None,
        "assigned_company_id": p.assigned_company_id,
        "assigned_main_tender_id": p.assigned_main_tender_id,
        "assigned_tender_id": p.assigned_tender_id,
        "review_notes": p.review_notes,
        "reviewed_by": p.reviewed_by,
        "reviewed_at": str(p.reviewed_at) if p.reviewed_at else None,
        "action_status": p.action_status,
        "action_error": p.action_error,
        "action_payload": p.action_payload,
        "created_at": str(p.created_at) if p.created_at else None,
        "updated_at": str(p.updated_at) if p.updated_at else None,
    }


def _pending_abs_path(file_path: str) -> Path:
    p = Path(file_path)
    if p.is_absolute() and p.exists():
        return p
    settings = get_settings()
    candidates = [
        p,
        Path(settings.upload_dir) / p,
        Path(settings.upload_dir) / "bills" / p.name,
        Path(settings.upload_dir) / "whatsapp" / p.name,
        Path(__file__).parent / "whatsapp_bridge" / "uploads" / p.name,
        Path(__file__).parent / "uploads" / p.name,
        Path(__file__).parent / "uploads" / "whatsapp" / p.name,
    ]
    for c in candidates:
        if c.exists():
            return c.resolve()
    return (Path(settings.upload_dir) / p).resolve()


def _busy_staging_abs_path(image_path: str) -> Path:
    p = Path(image_path)
    if p.is_absolute():
        return p
    settings = get_settings()
    return (Path(settings.upload_dir) / "busy_staging_bills" / p).resolve()


def _email_uid_from_pending_path(file_path: Optional[str]) -> Optional[int]:
    name = Path(str(file_path or "")).name
    m = re.match(r"^email_\d{14}_(\d+)_", name)
    if not m:
        return None
    try:
        return int(m.group(1))
    except Exception:
        return None


def _rewind_email_checkpoint_for_deleted_ingest(db: Session, p: PendingIngest) -> None:
    if not p or p.source != IngestSource.email:
        return

    email_user = str(p.source_account or "").strip().lower()
    if not email_user:
        return

    uid = _email_uid_from_pending_path(p.file_path)
    if not uid:
        return

    target_uid = max(int(uid) - 1, 0)
    try:
        db.execute(
            text(
                """
                UPDATE email_sync_checkpoints
                SET last_uid = CASE WHEN last_uid > :target_uid THEN :target_uid ELSE last_uid END
                WHERE LOWER(email_user) = :email_user
                """
            ),
            {
                "target_uid": target_uid,
                "email_user": email_user,
            },
        )
    except Exception as ex:
        log.warning("Failed to rewind email checkpoint for deleted ingest #%s: %s", getattr(p, "id", None), ex)



def _is_duplicate_busy_staging(db: Session, vendor_name: Optional[str], bill_number: Optional[str], company_id: Optional[int] = None) -> bool:
    if not (vendor_name and bill_number):
        return False

    def normalize_vendor(s: str) -> str:
        s = str(s or "").strip().lower()
        s = re.sub(r'^m\s*/?\s*s\.?\s*', '', s)
        s = re.sub(r'[\.,\(\)]', '', s)
        s = re.sub(r'\s+', ' ', s)
        return s

    def normalize_bill_no(s: str) -> str:
        s = str(s or "").strip().lower()
        s = re.sub(r'^(invoice|inv|bill|no|no\.|no:|invoice\s*no)\s*[:#\-\s]*', '', s)
        s = re.sub(r'[^a-z0-9\-\/]', '', s)
        return s

    try:
        vendor_n = normalize_vendor(vendor_name)
        bill_n = normalize_bill_no(bill_number)

        try:
            sql_q = select(BusyStagingBill).where(
                func.lower(func.coalesce(BusyStagingBill.bill_number, '')) == (bill_n or '').lower(),
                func.lower(func.coalesce(BusyStagingBill.broker_name, '')) == (vendor_n or '').lower(),
            )
            if company_id is not None:
                sql_q = sql_q.where(BusyStagingBill.company_id == company_id)
            existing = db.execute(sql_q).scalar_one_or_none()
            if existing:
                return True
        except Exception:
            pass

        q = select(BusyStagingBill).where(BusyStagingBill.bill_number != None)
        if bill_n:
            q = q.where(func.lower(BusyStagingBill.bill_number).like(f"%{bill_n}%"))
        if company_id is not None:
            q = q.where(BusyStagingBill.company_id == company_id)
        candidates = db.execute(q.limit(50)).scalars().all()

        for c in candidates:
            if normalize_bill_no(c.bill_number) == bill_n and normalize_vendor(c.broker_name) == vendor_n:
                return True
    except Exception as e:
        log.exception(f"Busy staging duplicate check error: {e}")
    return False

async def _reparse_pending_ingest_record(
    db: Session,
    p: PendingIngest,
    *,
    reclassify: bool = False,
    document_type_override: Optional[str] = None,
    ocr_engine: Optional[str] = None,
) -> dict:
    from services.ocr_service import classify_document_type, extract_document_by_type

    abs_path = _pending_abs_path(p.file_path)
    if not abs_path.exists():
        raise HTTPException(404, f"Source file not found for ingest #{p.id}")

    file_name_hint = p.file_name or abs_path.name
    cls_conf = float(p.classifier_confidence or 0)
    cls_candidates = p.classifier_candidates or []
    engine = str(ocr_engine or "").strip().lower()
    if engine in {"", "auto"}:
        engine = None
    allowed_engines = {"groq", "mistral", "openrouter", "gemini", "azure", "github"}
    if engine and engine not in allowed_engines:
        raise HTTPException(400, f"Invalid ocr_engine: {ocr_engine}")

    if document_type_override:
        try:
            doc_enum = DocumentType(str(document_type_override).strip())
        except Exception:
            raise HTTPException(400, f"Invalid document_type: {document_type_override}")
    elif reclassify or not p.document_type:
        hint_text = f"{file_name_hint} {p.review_notes or ''}".strip()
        doc_type, cls_conf, cls_candidates = classify_document_type(str(abs_path), hint_text)
        try:
            doc_enum = DocumentType(doc_type)
        except Exception:
            doc_enum = DocumentType.not_classified
            cls_candidates = (cls_candidates or []) + [{"type": str(doc_type), "score": 0.0}]
    else:
        doc_enum = p.document_type

    if doc_enum == DocumentType.not_classified:
        extracted = {
            "document_type": doc_enum.value,
            "source": "classifier",
            "high_confidence": False,
            "manual_required": True,
            "note": "Not classified into predefined types; route for manual review.",
        }
    else:
        extracted = await extract_document_by_type(str(abs_path), doc_enum.value, prefer_ocr_engine=engine)
        if not isinstance(extracted, dict):
            extracted = {"error": "invalid extractor response", "high_confidence": False}

    if isinstance(extracted, dict):
        extracted["reparse_ocr_engine"] = engine or "auto"

    unclear_fields = extracted.get("unclear_fields") if isinstance(extracted.get("unclear_fields"), list) else []

    p.document_type = doc_enum
    p.classifier_confidence = cls_conf
    p.classifier_candidates = cls_candidates or []
    p.extracted_payload = extracted
    p.unclear_fields = unclear_fields

    return {
        "id": p.id,
        "document_type": doc_enum.value,
        "classifier_confidence": float(cls_conf or 0),
        "unclear_count": len(unclear_fields),
        "extractor_source": extracted.get("source"),
        "ocr_engine": engine or "auto",
    }


def _to_bill_source(src: Optional[IngestSource]) -> BillSource:
    if src == IngestSource.telegram:
        return BillSource.telegram
    if src == IngestSource.email:
        return BillSource.email
    if src == IngestSource.manual:
        return BillSource.manual
    return BillSource.web


OWN_COMPANY_IDENTIFIERS = [
    {"name": "shree nath industries", "email": "gordhan.khandelwal@yahoo.com"},
    {"name": "shree vinayak trading company", "email": "vtcjaipur2012@gmail.com"},
    {"name": "shree ganpati enterprises", "email": "vinayakkhandelwal88@gmail.com"},
]

_PO_ALLOWED_MATERIALS = {"Maize", "Dorb", "Domc", "Rice DDGS"}
_PO_PLANT_DISPLAY = {
    "kaladera": "Kaladera",
    "ajmer": "Ajmer",
    "jodhpur": "Jodhpur",
    "nadbai": "Nadbai",
    "bikaner": "Bikaner",
    "pali": "Pali",
    "lambiyan": "Lambiyan",
}


def _name_similarity(a: Optional[str], b: Optional[str]) -> float:
    a_norm = _norm_text(a)
    b_norm = _norm_text(b)
    if not a_norm or not b_norm:
        return 0.0
    ratio = SequenceMatcher(None, _norm_code(a_norm), _norm_code(b_norm)).ratio()
    a_tokens = set(a_norm.split())
    b_tokens = set(b_norm.split())
    overlap = (len(a_tokens & b_tokens) / float(len(b_tokens))) if b_tokens else 0.0
    return max(ratio, overlap)


def _canonical_po_plant_name(name: Optional[str]) -> Optional[str]:
    hint = _norm_plant_hint(name)
    if not hint:
        return None
    for key, target in _PLANT_ALIASES.items():
        if key in hint:
            return _PO_PLANT_DISPLAY.get(target)
    return None


def _canonical_po_material_name(name: Optional[str]) -> Optional[str]:
    canonical = _canonical_material_name(name)
    if not canonical:
        return None
    norm = _norm_text(canonical)
    for fixed in _PO_ALLOWED_MATERIALS:
        if _norm_text(fixed) == norm:
            return fixed
    return None


def _normalize_rm_components(raw: Optional[str]) -> tuple[Optional[str], Optional[str], Optional[str]]:
    """
    Return (normalized_code, base_code, suffix) for RM-like tokens.

    Suffix-distinct policy is enforced: RM-282-A and RM-282-B are treated as
    separate main tender codes.
    """
    text = str(raw or "").strip().upper()
    if not text:
        return None, None, None

    m = re.search(r"\bRM\s*[-_/ ]*0*([0-9]{2,6})(?:\s*[-_/ ]*([A-Z]{1,3}))?\b", text)
    if not m:
        m = re.match(r"^0*([0-9]{2,6})(?:\s*[-_/ ]*([A-Z]{1,3}))?$", text)
    if not m:
        return None, None, None

    digits = m.group(1).lstrip("0") or "0"
    suffix = (str(m.group(2) or "").strip().upper() or None)
    base = f"RM-{digits}"
    normalized = f"{base}-{suffix}" if suffix else base
    return normalized, base, suffix


def _normalize_rm_main_tender_code(raw: Optional[str]) -> Optional[str]:
    normalized, _base, _suffix = _normalize_rm_components(raw)
    return normalized


def _effective_unloading_qty_mt(net_qty: Optional[float], received_qty: Optional[float]) -> float:
    net = _to_float(net_qty)
    received = _to_float(received_qty)
    if net is not None and net > 0:
        return net
    if received is not None:
        return received
    return float(net or 0)


def _resolve_unloading_assignment(
    db: Session,
    *,
    rm_number: Optional[str],
    item_name: Optional[str],
    plant_name: Optional[str],
    company_id_hint: Optional[int] = None,
    main_tender_id_hint: Optional[int] = None,
    tender_id_hint: Optional[int] = None,
) -> dict:
    rm_norm, rm_base, rm_suffix = _normalize_rm_components(rm_number)
    rm_norm_lower = _norm_text(rm_norm)
    plant_hint = _clean_optional_text(plant_name)
    plant_id_hint, plant_name_resolved_hint = _resolve_plant(db, plant_hint) if plant_hint else (None, None)

    resolved_company = company_id_hint
    resolved_main = main_tender_id_hint
    resolved_tender = tender_id_hint

    if any(v is not None for v in (company_id_hint, main_tender_id_hint, tender_id_hint)):
        resolved_company, resolved_main, resolved_tender = _resolve_context_ids(
            db,
            company_id=company_id_hint,
            main_tender_id=main_tender_id_hint,
            tender_id=tender_id_hint,
        )

    if resolved_tender is not None:
        t = db.get(Tender, resolved_tender)
        if t:
            if resolved_company is None:
                resolved_company = t.company_id
            if resolved_main is None:
                resolved_main = t.main_tender_id

        tender_hint_usable = t is not None
        if tender_hint_usable and not plant_hint:
            tender_hint_usable = False
        if tender_hint_usable and plant_hint and not plant_id_hint:
            tender_hint_usable = False
        if tender_hint_usable and t and plant_id_hint and t.plant_id and t.plant_id != plant_id_hint:
            tender_hint_usable = False
        if tender_hint_usable and t and rm_norm and t.main_tender and _norm_text(t.main_tender.tender_code) != rm_norm_lower:
            tender_hint_usable = False

        if tender_hint_usable:
            reason = "Assigned from provided subtender context."
            return {
                "company_id": resolved_company,
                "main_tender_id": resolved_main,
                "tender_id": resolved_tender,
                "rm_number_norm": rm_norm,
                "rm_number_base": rm_base,
                "assignment_status": "assigned",
                "assignment_reason": reason,
                "assignment_confidence": 1.0,
                "mapping_source": "context_tender",
                "requires_manual_assignment": False,
            }

        # Context subtender can be stale for email/manual imports when plant is
        # missing/noisy; fall back to RM+plant auto mapping instead of forcing it.
        resolved_tender = None

    selected_main = None
    main_candidates: List[MainTender] = []
    context_main = db.get(MainTender, resolved_main) if resolved_main is not None else None

    if rm_norm:
        q = select(MainTender).where(func.lower(MainTender.tender_code) == rm_norm_lower)
        if resolved_company is not None:
            q = q.where(MainTender.company_id == resolved_company)
        main_candidates = db.execute(q).scalars().all()

        if len(main_candidates) == 1:
            selected_main = main_candidates[0]
        elif len(main_candidates) > 1:
            return {
                "company_id": resolved_company,
                "main_tender_id": None,
                "tender_id": None,
                "rm_number_norm": rm_norm,
                "rm_number_base": rm_base,
                "assignment_status": "ambiguous",
                "assignment_reason": (
                    f"RM {rm_norm} matches multiple main tenders. Please assign manually."
                ),
                "assignment_confidence": 0.35,
                "mapping_source": "rm_auto",
                "requires_manual_assignment": True,
            }
        elif context_main and _norm_text(context_main.tender_code) == rm_norm_lower:
            selected_main = context_main
    elif context_main:
        selected_main = context_main

    if selected_main is not None:
        selected_company = selected_main.company_id
    else:
        selected_company = resolved_company

    if selected_main is None:
        if rm_norm and rm_suffix:
            reason = (
                f"No main tender found for suffix-distinct RM code {rm_norm}. "
                "Assign RM manually."
            )
        elif rm_norm:
            reason = f"No main tender found for RM code {rm_norm}."
        else:
            reason = "RM code was not detected from unloading register."
        return {
            "company_id": selected_company,
            "main_tender_id": resolved_main,
            "tender_id": None,
            "rm_number_norm": rm_norm,
            "rm_number_base": rm_base,
            "assignment_status": "unresolved",
            "assignment_reason": reason,
            "assignment_confidence": 0.0,
            "mapping_source": "rm_auto",
            "requires_manual_assignment": True,
        }

    plant_id, plant_resolved_name = plant_id_hint, plant_name_resolved_hint
    if not plant_id:
        reason = (
            f"RM {selected_main.tender_code} matched main tender, but plant could not be resolved. "
            "Assign subtender manually."
        )
        return {
            "company_id": selected_company,
            "main_tender_id": selected_main.id,
            "tender_id": None,
            "rm_number_norm": rm_norm,
            "rm_number_base": rm_base,
            "assignment_status": "unresolved",
            "assignment_reason": reason,
            "assignment_confidence": 0.45,
            "mapping_source": "rm_auto",
            "requires_manual_assignment": True,
        }

    tender_q = select(Tender).where(
        Tender.main_tender_id == selected_main.id,
        Tender.plant_id == plant_id,
        Tender.status != TenderStatus.cancelled,
    )
    if selected_company is not None:
        tender_q = tender_q.where(or_(Tender.company_id == selected_company, Tender.company_id == None))
    tender_candidates = db.execute(tender_q).scalars().all()

    material_id, _material_name = _resolve_material(db, material_name=item_name)
    used_material_filter = False
    if material_id and len(tender_candidates) > 1:
        filtered = [t for t in tender_candidates if t.material_id == material_id]
        if filtered:
            tender_candidates = filtered
            used_material_filter = True

    if len(tender_candidates) == 1:
        selected_tender = tender_candidates[0]
        reason = (
            f"Auto-mapped using RM {selected_main.tender_code} + plant {plant_resolved_name or plant_name}."
        )
        mapping_source = "rm_plant"
        confidence = 0.93
        if used_material_filter:
            reason = (
                f"Auto-mapped using RM {selected_main.tender_code} + plant {plant_resolved_name or plant_name} "
                "+ item material match."
            )
            mapping_source = "rm_plant_material"
            confidence = 0.97
        if resolved_main is not None and resolved_main != selected_main.id:
            reason = (
                f"{reason} Provided main_tender_id {resolved_main} was overridden by RM mapping "
                f"to {selected_main.tender_code}."
            )
            mapping_source = "rm_override"
        return {
            "company_id": selected_tender.company_id if selected_tender.company_id is not None else selected_company,
            "main_tender_id": selected_tender.main_tender_id,
            "tender_id": selected_tender.id,
            "rm_number_norm": rm_norm,
            "rm_number_base": rm_base,
            "assignment_status": "assigned",
            "assignment_reason": reason,
            "assignment_confidence": confidence,
            "mapping_source": mapping_source,
            "requires_manual_assignment": False,
        }

    if len(tender_candidates) > 1:
        reason = (
            f"RM {selected_main.tender_code} and plant {plant_resolved_name or plant_name} "
            "match multiple subtenders. Please assign manually."
        )
        return {
            "company_id": selected_company,
            "main_tender_id": selected_main.id,
            "tender_id": None,
            "rm_number_norm": rm_norm,
            "rm_number_base": rm_base,
            "assignment_status": "ambiguous",
            "assignment_reason": reason,
            "assignment_confidence": 0.4,
            "mapping_source": "rm_auto",
            "requires_manual_assignment": True,
        }

    reason = (
        f"No subtender found for RM {selected_main.tender_code} and plant {plant_resolved_name or plant_name}."
    )
    return {
        "company_id": selected_company,
        "main_tender_id": selected_main.id,
        "tender_id": None,
        "rm_number_norm": rm_norm,
        "rm_number_base": rm_base,
        "assignment_status": "unresolved",
        "assignment_reason": reason,
        "assignment_confidence": 0.4,
        "mapping_source": "rm_auto",
        "requires_manual_assignment": True,
    }


def _apply_unloading_assignment_to_master(
    master: PlantUnloadingMaster,
    assignment: dict,
    *,
    manual_override: bool = False,
    operator: Optional[str] = None,
) -> None:
    master.company_id = assignment.get("company_id")
    master.main_tender_id = assignment.get("main_tender_id")
    master.tender_id = assignment.get("tender_id")
    master.rm_number_norm = assignment.get("rm_number_norm")
    master.rm_number_base = assignment.get("rm_number_base")
    master.assignment_status = assignment.get("assignment_status") or ("assigned" if master.tender_id else "pending")
    master.assignment_reason = assignment.get("assignment_reason")
    master.assignment_confidence = assignment.get("assignment_confidence")
    master.mapping_source = "manual" if manual_override else assignment.get("mapping_source")
    master.requires_manual_assignment = bool(assignment.get("requires_manual_assignment"))
    if manual_override:
        master.is_manual_override = True
        master.manual_assigned_by = operator or master.manual_assigned_by
        master.manual_assigned_at = datetime.now()


def _sync_unloading_lineage_from_master(db: Session, master: PlantUnloadingMaster) -> None:
    cid = master.company_id
    mid = master.main_tender_id
    tid = master.tender_id

    for e in (master.entries or []):
        _stamp_lineage(e, cid, mid, tid)
        if e.receipt_id:
            r = db.get(PlantReceipt, e.receipt_id)
            if r and not r.is_deleted:
                if r.dispatch_id:
                    ds = db.get(Dispatch, r.dispatch_id)
                    if ds and not ds.is_deleted:
                        _stamp_lineage(r, ds.company_id, ds.main_tender_id, ds.tender_id)
                        continue
                _stamp_lineage(r, cid, mid, tid)


def _recalculate_tender_fulfilled_from_unloading(
    db: Session,
    tender_ids: Optional[List[Optional[int]]] = None,
    company_id: Optional[int] = None,
) -> int:
    ids = sorted({int(tid) for tid in (tender_ids or []) if tid})

    if ids:
        tenders = [db.get(Tender, tid) for tid in ids]
        tenders = [t for t in tenders if t is not None]
    else:
        tq = select(Tender)
        if company_id is not None:
            tq = tq.where(Tender.company_id == company_id)
        tenders = db.execute(tq).scalars().all()

    updated = 0
    for t in tenders:
        rows = db.execute(
            select(PlantUnloadingEntry.net_qty_mt, PlantUnloadingEntry.received_qty_mt).where(
                PlantUnloadingEntry.tender_id == t.id,
                PlantUnloadingEntry.status == BillStatus.approved,
            )
        ).all()

        total = 0.0
        for net_qty, received_qty in rows:
            total += _effective_unloading_qty_mt(net_qty, received_qty)
        total = round(total, 3)

        current = round(float(t.fulfilled_qty_mt or 0), 3)
        if current != total:
            t.fulfilled_qty_mt = total
            updated += 1
    return updated


def _count_tender_fulfilled_mismatches(db: Session, company_id: Optional[int] = None) -> int:
    tq = select(Tender)
    if company_id is not None:
        tq = tq.where(Tender.company_id == company_id)
    tenders = db.execute(tq).scalars().all()

    mismatches = 0
    for t in tenders:
        rows = db.execute(
            select(PlantUnloadingEntry.net_qty_mt, PlantUnloadingEntry.received_qty_mt).where(
                PlantUnloadingEntry.tender_id == t.id,
                PlantUnloadingEntry.status == BillStatus.approved,
            )
        ).all()
        total = 0.0
        for net_qty, received_qty in rows:
            total += _effective_unloading_qty_mt(net_qty, received_qty)
        total = round(total, 3)
        current = round(float(t.fulfilled_qty_mt or 0), 3)
        if current != total:
            mismatches += 1
    return mismatches


def _apply_unloading_master_scope_hints(q, company_id: Optional[int], main_tender_id: Optional[int], tender_id: Optional[int]):
    if tender_id is not None:
        q = q.where(or_(PlantUnloadingMaster.tender_id == tender_id, PlantUnloadingMaster.tender_id == None))
    if main_tender_id is not None:
        q = q.where(or_(PlantUnloadingMaster.main_tender_id == main_tender_id, PlantUnloadingMaster.main_tender_id == None))
    if company_id is not None:
        q = q.where(or_(PlantUnloadingMaster.company_id == company_id, PlantUnloadingMaster.company_id == None))
    return q


def _purchase_order_to_dict(po: PurchaseOrder) -> dict:
    return {
        "id": po.id,
        "company_id": po.company_id,
        "main_tender_id": po.main_tender_id,
        "tender_id": po.tender_id,
        "po_number": po.po_number,
        "po_date": str(po.po_date) if po.po_date else None,
        "seller_name": po.seller_name,
        "buyer_name": po.buyer_name,
        "buyer_email": po.buyer_email,
        "plant_id": po.plant_id,
        "plant_name": po.plant_name,
        "total_amount": float(po.total_amount) if po.total_amount is not None else None,
        "line_items": po.line_items or [],
        "status": po.status.value if po.status else None,
        "source": po.source,
        "source_doc_path": po.source_doc_path,
        "source_pending_id": po.source_pending_id,
        "notes": po.notes,
        "created_at": str(po.created_at) if po.created_at else None,
        "updated_at": str(po.updated_at) if po.updated_at else None,
    }


def _purchase_order_subtender_preview(db: Session, po: PurchaseOrder) -> dict:
    rows: List[dict] = []
    warnings: List[str] = []

    items = po.line_items if isinstance(po.line_items, list) else []
    main_code = _derive_main_tender_code(po.po_number)

    plant_id = po.plant_id
    plant_name = po.plant_name
    if not plant_id and plant_name:
        plant_id, plant_name = _resolve_plant(db, plant_name)

    cycle = _active_cycle_for_scope(db, po.company_id)
    cycle_name = cycle.name if cycle else None

    if not po.company_id:
        warnings.append("Company is not assigned. Assign company before approval.")
    if not po.main_tender_id:
        warnings.append("Main tender is not assigned. Assign main tender before approval.")
    if not cycle:
        warnings.append("No active cycle found. Run-action will auto-create one from PO supply dates.")
    if not plant_id:
        warnings.append("Plant could not be resolved from PO.")
    if not items:
        warnings.append("No line items found in this PO.")

    for idx, item in enumerate(items, start=1):
        if not isinstance(item, dict):
            continue

        qty = _to_float(item.get("approved_quantity_qtl") or item.get("quantity_qtl"))
        mat_hint = item.get("material_type") or item.get("material") or item.get("item")
        mat_id, mat_name = _resolve_material(db, material_name=mat_hint)

        if mat_id is None:
            warnings.append(f"Line {idx}: material not resolved ({mat_hint or 'unknown'}).")

        tender_number = None
        exists_id = None
        action = "manual"

        if plant_id and mat_id and qty and qty > 0:
            tender_number = _build_po_sub_tender_number(main_code, plant_name, mat_name, idx)
            exists_id = db.execute(
                select(Tender.id).where(
                    Tender.tender_number == tender_number,
                    Tender.company_id == po.company_id,
                    Tender.main_tender_id == po.main_tender_id,
                )
            ).scalar_one_or_none()
            action = "update" if exists_id else "create"

        rows.append({
            "line": idx,
            "material_id": mat_id,
            "material": mat_name or mat_hint,
            "qty_qtl": qty,
            "plant_id": plant_id,
            "plant_name": plant_name,
            "tender_number": tender_number,
            "existing_tender_id": exists_id,
            "action": action,
        })

    return {
        "purchase_order_id": po.id,
        "po_number": po.po_number,
        "company_id": po.company_id,
        "main_tender_id": po.main_tender_id,
        "main_code": main_code,
        "plant_id": plant_id,
        "plant_name": plant_name,
        "cycle_id": cycle.id if cycle else None,
        "cycle_name": cycle_name,
        "rows": rows,
        "warnings": warnings,
    }


def _detect_own_company(payload: dict) -> Optional[str]:
    name_hints: List[str] = []
    email_hints: List[str] = []

    for key in ["winner_party_name", "winner_name", "buyer_name", "company_name", "approved_party_name"]:
        val = str(payload.get(key) or "").strip()
        if val:
            name_hints.append(val)

    for key in ["winner_party_email", "winner_email", "buyer_email", "email"]:
        val = str(payload.get(key) or "").strip()
        if val:
            email_hints.append(val)

    items = payload.get("items") if isinstance(payload.get("items"), list) else []
    for item in items:
        if not isinstance(item, dict):
            continue
        line_party = str(item.get("approved_party_name") or item.get("winner_party_name") or "").strip()
        if line_party:
            name_hints.append(line_party)

    norm_names = [_norm_text(v) for v in name_hints if _norm_text(v)]
    norm_emails = [_norm_text(v) for v in email_hints if _norm_text(v)]

    # 1) strict email match
    for item in OWN_COMPANY_IDENTIFIERS:
        nm = _norm_text(item.get("name"))
        em = _norm_text(item.get("email"))
        if em and any(em == cand for cand in norm_emails):
            return item.get("name")

    # 2) exact/contains name match
    for item in OWN_COMPANY_IDENTIFIERS:
        nm = _norm_text(item.get("name"))
        if nm and any(nm in cand or cand in nm for cand in norm_names):
            return item.get("name")

    # 3) fuzzy best-match for OCR noise and small mistakes
    best_name = None
    best_score = 0.0
    for cand in norm_names:
        for item in OWN_COMPANY_IDENTIFIERS:
            score = _name_similarity(cand, item.get("name"))
            if score > best_score:
                best_score = score
                best_name = item.get("name")

    if best_name and best_score >= 0.62:
        return best_name
    return None


_OWN_COMPANY_CANONICAL_NAMES = [
    "Shree Nath Industries",
    "Shree Vinayak Trading Company",
    "Shree Ganpati Enterprises",
]

_OWN_COMPANY_ALIAS_MAP = {
    _norm_text("shree nath industries"): "Shree Nath Industries",
    _norm_text("shree vinayak trading company"): "Shree Vinayak Trading Company",
    _norm_text("shree vinayak trading co"): "Shree Vinayak Trading Company",
    _norm_text("shri vinayak trading company"): "Shree Vinayak Trading Company",
    _norm_text("shree ganpati enterprises"): "Shree Ganpati Enterprises",
    _norm_text("shree ganpati enterpriese"): "Shree Ganpati Enterprises",
}


def _canonical_own_company_name(raw_name: Optional[str], *, min_score: float = 0.58) -> Optional[str]:
    src = str(raw_name or "").strip()
    if not src:
        return None
    src_n = _norm_text(src)
    if not src_n:
        return None

    aliased = _OWN_COMPANY_ALIAS_MAP.get(src_n)
    if aliased:
        return aliased

    # Exact/contains shortcut for OCR fragments.
    for nm in _OWN_COMPANY_CANONICAL_NAMES:
        nmn = _norm_text(nm)
        if src_n == nmn or src_n in nmn or nmn in src_n:
            return nm

    best_name = None
    best_score = 0.0
    for nm in _OWN_COMPANY_CANONICAL_NAMES:
        score = _name_similarity(src, nm)
        if score > best_score:
            best_score = score
            best_name = nm

    if best_name and best_score >= min_score:
        return best_name
    return None


def _normalize_unloading_party_name(db: Session, party_name: Optional[str], company_id: Optional[int]) -> str:
    src = str(party_name or "").strip()
    detected = _canonical_own_company_name(src)
    if detected:
        return detected

    fallback_name = None
    if company_id is not None:
        company = db.get(Company, company_id)
        fallback_name = str(getattr(company, "name", "") or "").strip() or None
        fallback_detected = _canonical_own_company_name(fallback_name)
        if fallback_detected:
            return fallback_detected
        if fallback_name:
            return fallback_name

    # Force to nearest known own company for noisy OCR labels.
    if src:
        best = max(_OWN_COMPANY_CANONICAL_NAMES, key=lambda nm: _name_similarity(src, nm))
        return best

    return _OWN_COMPANY_CANONICAL_NAMES[0]


def _find_or_create_company(db: Session, company_name: Optional[str]) -> Optional[int]:
    name = str(company_name or "").strip()
    if not name:
        return None
    existing = db.execute(select(Company).where(func.lower(Company.name) == name.lower())).scalar_one_or_none()
    if existing:
        return existing.id
    c = Company(name=name)
    db.add(c)
    db.flush()
    return c.id


def _find_company_exact(db: Session, company_name: Optional[str]) -> Optional[int]:
    name = str(company_name or "").strip()
    if not name:
        return None
    existing = db.execute(select(Company).where(func.lower(Company.name) == name.lower())).scalar_one_or_none()
    return existing.id if existing else None


def _resolve_company_for_pending(
    db: Session,
    *,
    current_company_id: Optional[int],
    payload: dict,
    hints: Optional[List[Optional[str]]] = None,
    allow_create: bool = False,
) -> tuple[Optional[int], Optional[str]]:
    if current_company_id is not None:
        return current_company_id, None

    own_company = _detect_own_company(payload)
    if own_company:
        company_id = _find_or_create_company(db, own_company)
        if company_id is not None:
            return company_id, None

    raw_hints = hints or []
    clean_hints: List[str] = []
    for item in raw_hints:
        txt = str(item or "").strip()
        if txt and txt.lower() not in {"na", "n/a", "none", "null", "unknown"}:
            if txt not in clean_hints:
                clean_hints.append(txt)

    exact_match_ids: List[int] = []
    for txt in clean_hints:
        company_id = _find_company_exact(db, txt)
        if company_id and company_id not in exact_match_ids:
            exact_match_ids.append(company_id)

    if len(exact_match_ids) == 1:
        return exact_match_ids[0], None
    if len(exact_match_ids) > 1:
        return None, "Multiple company matches found. Please assign company manually."

    if allow_create:
        # Safety guard: only auto-create from reasonably descriptive names.
        create_hint = next((txt for txt in clean_hints if len(txt) >= 6 and len(txt.split()) >= 2), None)
        if create_hint:
            company_id = _find_or_create_company(db, create_hint)
            if company_id is not None:
                return company_id, None

    return None, "Could not safely resolve company. Please assign company/main tender manually."


def _find_or_create_main_tender(db: Session, company_id: int, tender_code: str, title: Optional[str] = None) -> int:
    code = str(tender_code or "").strip()
    if not code:
        code = "UNTITLED"
    existing = db.execute(
        select(MainTender).where(
            MainTender.company_id == company_id,
            func.lower(MainTender.tender_code) == code.lower(),
        )
    ).scalar_one_or_none()
    if existing:
        if title and not existing.title:
            existing.title = str(title).strip()
        return existing.id
    mt = MainTender(company_id=company_id, tender_code=code, title=(str(title).strip() if title else None))
    db.add(mt)
    db.flush()
    return mt.id


def _active_cycle_for_scope(db: Session, company_id: Optional[int]) -> Optional[SproxxCycle]:
    if company_id is not None:
        cyc = db.execute(
            select(SproxxCycle)
            .where(SproxxCycle.is_active == True, or_(SproxxCycle.company_id == company_id, SproxxCycle.company_id == None))
            .order_by(desc(SproxxCycle.company_id), desc(SproxxCycle.id))
        ).scalars().first()
        if cyc:
            return cyc
    return db.execute(select(SproxxCycle).where(SproxxCycle.is_active == True).order_by(desc(SproxxCycle.id))).scalars().first()


def _safe_tender_token(v: Optional[str], fallback: str = "NA") -> str:
    token = re.sub(r"[^A-Za-z0-9]+", "-", str(v or "").strip()).strip("-")
    token = re.sub(r"-+", "-", token)
    return (token or fallback)[:40]


def _build_po_sub_tender_number(main_code: str, plant_name: Optional[str], material_name: Optional[str], line_no: int) -> str:
    plant = _canonical_po_plant_name(plant_name)
    material = _canonical_po_material_name(material_name)
    if plant and material:
        # Required naming format: "RM-123 ITEM PLANT"
        return f"{str(main_code).strip()} {material} {plant}"[:50]
    fallback = f"{_safe_tender_token(main_code, 'RM')} L{line_no}"
    return fallback[:50]


def _extract_po_supply_window(payload: dict, po_date: Optional[date]) -> tuple[date, date]:
    start = _parse_optional_date(
        payload.get("supply_period_start")
        or payload.get("start_date")
        or payload.get("supply_start")
        or payload.get("period_start")
    )
    end = _parse_optional_date(
        payload.get("supply_period_end")
        or payload.get("end_date")
        or payload.get("supply_end")
        or payload.get("period_end")
    )

    if start is None:
        start = po_date or date.today()
    if end is None:
        end = start + timedelta(days=14)
    if end < start:
        end = start
    return start, end


def _ensure_cycle_for_po_action(db: Session, company_id: Optional[int], start_date: date, end_date: date) -> SproxxCycle:
    cycle = _active_cycle_for_scope(db, company_id)
    if cycle:
        return cycle

    week1_end = start_date + timedelta(days=7)
    if week1_end > end_date:
        week1_end = end_date

    if company_id is not None:
        db.query(SproxxCycle).filter(SproxxCycle.company_id == company_id, SproxxCycle.is_active == True).update(
            {"is_active": False}, synchronize_session=False
        )
    else:
        db.query(SproxxCycle).filter(SproxxCycle.company_id == None, SproxxCycle.is_active == True).update(
            {"is_active": False}, synchronize_session=False
        )

    auto_name = f"AUTO-{start_date.isoformat()}-{end_date.isoformat()}"
    cycle = SproxxCycle(
        company_id=company_id,
        name=auto_name[:100],
        cycle_start=start_date,
        cycle_end=end_date,
        week1_end=week1_end,
        week2_end=end_date,
        is_active=True,
    )
    db.add(cycle)
    db.flush()
    return cycle


def _apply_pending_action(db: Session, p: PendingIngest) -> dict:
    payload = p.extracted_payload or {}
    if not isinstance(payload, dict):
        payload = {}

    company_id = p.assigned_company_id if p.assigned_company_id is not None else p.company_id
    main_tender_id = p.assigned_main_tender_id if p.assigned_main_tender_id is not None else p.main_tender_id
    tender_id = p.assigned_tender_id if p.assigned_tender_id is not None else p.tender_id

    doc_type = p.document_type.value if p.document_type else ""

    if doc_type == "purchase_bill":
        plant_name = payload.get("destination_plant") or payload.get("plant_name")
        plant_id, plant_name_resolved = _resolve_plant(db, plant_name)
        bill_date = _parse_optional_date(payload.get("bill_date"))

        src_abs = _pending_abs_path(p.file_path)
        final_rel_path = p.file_path
        if src_abs.exists():
            settings = get_settings()
            bills_dir = Path(settings.upload_dir) / "bills"
            bills_dir.mkdir(parents=True, exist_ok=True)
            dst_name = f"ing_{datetime.now().strftime('%Y%m%d%H%M%S')}_{p.file_name}"
            dst_abs = bills_dir / dst_name
            shutil.move(str(src_abs), str(dst_abs))
            final_rel_path = str(Path("bills") / dst_name)

        b = Bill(
            company_id=company_id,
            main_tender_id=main_tender_id,
            tender_id=tender_id,
            source=_to_bill_source(p.source),
            image_path=final_rel_path,
            vehicle_number=payload.get("vehicle_number"),
            broker_name=payload.get("vendor_name") or payload.get("broker_name"),
            material_name=payload.get("material_type"),
            qty_mt=_to_float(payload.get("quantity_qtl")),
            rate_per_mt=_to_float(payload.get("rate_per_qtl")),
            total_amount=_to_float(payload.get("total_amount")),
            bill_date=bill_date,
            bill_number=payload.get("bill_number"),
            plant_id=plant_id,
            plant_name=plant_name_resolved,
            notes=f"Auto-created from pending ingest #{p.id}",
            ocr_source=OcrSource.paddle if payload.get("source") == "paddle" else OcrSource.gemini,
            ocr_confidence=_to_float(payload.get("confidence") or payload.get("ocr_confidence")),
            ocr_raw_text=payload.get("raw_text", ""),
            is_handwritten=bool(payload.get("is_handwritten") or False),
            validation_amount=bool(payload.get("validation_amount") or False),
            validation_vehicle=bool(payload.get("validation_vehicle") or False),
            validation_material=bool(payload.get("validation_material") or False),
            status=BillStatus.pending,
        )
        db.add(b)
        db.flush()
        _reconcile_unloading_match_for_bill_ids(db, [b.id])
        return {"created_bill_id": b.id, "document_type": doc_type}

    if doc_type == "plant_unloading":
        image_name = Path(p.file_path).name
        created_entries, duplicates, total_rows = _ingest_unloading_rows(
            db=db,
            extracted=payload,
            image_path=image_name,
            source=p.source.value if p.source else "web",
            company_id=company_id,
            main_tender_id=main_tender_id,
            tender_id=tender_id,
        )
        unresolved_reasons = []
        for e in created_entries:
            m = e.master
            if not m:
                continue
            if m.requires_manual_assignment or not m.tender_id:
                reason = str(m.assignment_reason or "Manual assignment required before approval").strip()
                if reason and reason not in unresolved_reasons:
                    unresolved_reasons.append(reason)

        return {
            "created_entry_ids": [e.id for e in created_entries],
            "duplicates": duplicates,
            "rows_received": total_rows,
            "document_type": doc_type,
            "manual_required": bool(unresolved_reasons),
            "reason": unresolved_reasons[0] if unresolved_reasons else None,
            "manual_reasons": unresolved_reasons,
        }

    if doc_type == "rejection_notice":
        vehicle = payload.get("vehicle_number") or payload.get("truck_number")
        if not vehicle:
            return {
                "document_type": doc_type,
                "manual_required": True,
                "reason": "rejection_notice action requires truck number",
            }

        plant_name = payload.get("plant_name") or payload.get("destination_plant")
        plant_id, _ = _resolve_plant(db, plant_name)
        if not plant_id:
            return {
                "document_type": doc_type,
                "manual_required": True,
                "reason": "rejection_notice action requires a resolvable plant",
            }

        accepted = 0.0
        rejected = _to_float(payload.get("rejected_qty_qtl"))
        if rejected is None:
            rejected = _to_float(payload.get("quantity_qtl"))
        if rejected is None:
            rejected = _to_float(payload.get("weight"))
        if rejected is None:
            rejected = 1000.0

        mat_id, mat_name = _resolve_material(db, material_name=payload.get("material_type") or payload.get("item_name"))

        rejection_date = _parse_optional_date(payload.get("rejection_date") or payload.get("notice_date") or payload.get("bill_date"))
        rejection_type = str(payload.get("rejection_type") or "").strip().lower()
        reason = str(payload.get("reason") or "").strip() or "Auto-created from rejection notice"
        if rejection_type in {"partial", "complete"}:
            reason = f"{reason} ({rejection_type})"

        receipt = PlantReceipt(
            company_id=company_id,
            main_tender_id=main_tender_id,
            tender_id=tender_id,
            dispatch_id=None,
            vehicle_number=str(vehicle).strip()[:20],
            plant_id=plant_id,
            receipt_date=rejection_date or date.today(),
            accepted_mt=accepted,
            rejected_mt=rejected,
            received_qty_qtl=accepted + rejected,
            matched_qty_qtl=0,
            match_status="unmatched",
            material_id=mat_id,
            material_name=mat_name,
            rm_number=payload.get("tender_rm_number") or payload.get("rm_number") or payload.get("tender_number"),
            po_number=payload.get("po_number"),
            party_name=payload.get("vendor_name"),
            rejection_reason=reason,
            source="ingest",
            email_raw=json.dumps(payload, ensure_ascii=False),
        )
        db.add(receipt)
        db.flush()
        return {"created_receipt_id": receipt.id, "document_type": doc_type}

    if doc_type == "tender_notice":
        tender_ref = payload.get("tender_rm_number") or payload.get("tender_number") or payload.get("rm_number")
        items = payload.get("items") if isinstance(payload.get("items"), list) else []
        return {
            "document_type": doc_type,
            "manual_required": True,
            "note": "Tender notice captured for user visibility only; no records auto-created.",
            "captured": {
                "tender_rm_number": tender_ref,
                "plant_name": payload.get("plant_name"),
                "supply_period_start": payload.get("supply_period_start") or payload.get("start_date"),
                "supply_period_end": payload.get("supply_period_end") or payload.get("end_date"),
                "items": items,
            },
        }

    if doc_type == "purchase_order":
        po_number = str(payload.get("po_number") or payload.get("bill_number") or "").strip()
        po_date_hint = _parse_optional_date(payload.get("po_date") or payload.get("bill_date")) or date.today()
        if not po_number:
            tender_hint = str(payload.get("tender_rm_number") or payload.get("rm_number") or "RM").strip()
            tender_token = re.sub(r"[^A-Za-z0-9]+", "", tender_hint.upper())[:30] or "RM"
            po_number = f"AUTO-{tender_token}-{po_date_hint.strftime('%Y%m%d')}-{p.id}"

        own_company = _detect_own_company(payload)
        if not own_company:
            return {
                "document_type": doc_type,
                "manual_required": True,
                "reason": "PO winner is not one of configured companies; skipped by business rule.",
            }

        resolved_company_id, company_issue = _resolve_company_for_pending(
            db,
            current_company_id=company_id,
            payload=payload,
            hints=[own_company, payload.get("winner_party_name"), payload.get("buyer_name"), payload.get("company_name")],
            allow_create=False,
        )
        if resolved_company_id is None:
            return {
                "document_type": doc_type,
                "manual_required": True,
                "reason": company_issue or "purchase_order action could not resolve company",
            }
        company_id = resolved_company_id

        tender_rm_raw = str(payload.get("tender_rm_number") or payload.get("rm_number") or payload.get("tender_number") or "").strip()
        main_code = _normalize_rm_main_tender_code(tender_rm_raw) or _normalize_rm_main_tender_code(po_number)
        if not main_code:
            return {
                "document_type": doc_type,
                "manual_required": True,
                "reason": "PO action requires RM code in format RM-<digits>",
            }
        tender_rm = main_code

        # PO flow must always anchor to the RM code from the PO payload, even if a stale
        # assigned main_tender_id is present from prior manual context selection.
        resolved_main_tender_id = _find_or_create_main_tender(db, company_id, main_code, title=f"PO {po_number}")
        if main_tender_id is not None and main_tender_id != resolved_main_tender_id:
            log.warning(
                "PO ingest #%s main_tender override: assigned_main_tender_id=%s replaced_by_rm=%s (%s)",
                p.id,
                main_tender_id,
                resolved_main_tender_id,
                main_code,
            )
        main_tender_id = resolved_main_tender_id

        p.company_id = company_id
        p.main_tender_id = main_tender_id

        plant_name = payload.get("plant_name") or payload.get("destination_plant")
        plant_id, plant_name_resolved = _resolve_plant(db, plant_name)

        if not plant_id:
            return {
                "document_type": doc_type,
                "manual_required": True,
                "reason": "PO action requires a resolvable plant",
            }

        po_plant_name = _canonical_po_plant_name(plant_name_resolved or plant_name)
        if not po_plant_name:
            return {
                "document_type": doc_type,
                "manual_required": True,
                "reason": "PO plant must match one of configured plants",
            }

        po_date = _parse_optional_date(payload.get("po_date") or payload.get("bill_date")) or po_date_hint
        supply_start, supply_end = _extract_po_supply_window(payload, po_date)
        items = payload.get("items") if isinstance(payload.get("items"), list) else []
        normalized_items = []
        skipped_lines: List[int] = []
        for idx, item in enumerate(items, start=1):
            if not isinstance(item, dict):
                continue
            mname = item.get("material_type") or item.get("material") or item.get("item")
            qty = _to_float(item.get("approved_quantity_qtl") or item.get("quantity_qtl"))
            rate = _to_float(item.get("approved_rate_per_qtl") or item.get("rate_per_qtl"))
            amt = _to_float(item.get("line_amount") or item.get("amount"))
            approved_party = item.get("approved_party_name") or payload.get("winner_party_name") or own_company

            line_payload = {
                "approved_party_name": approved_party,
                "winner_party_name": payload.get("winner_party_name"),
                "buyer_name": payload.get("buyer_name"),
                "winner_party_email": payload.get("winner_party_email"),
                "buyer_email": payload.get("buyer_email"),
            }
            if not _detect_own_company(line_payload):
                skipped_lines.append(idx)
                continue

            _, mname_resolved = _resolve_material(db, material_name=mname)
            material_fixed = _canonical_po_material_name(mname_resolved or mname)
            if not material_fixed:
                skipped_lines.append(idx)
                continue
            if qty is None or qty <= 0:
                skipped_lines.append(idx)
                continue
            normalized_items.append({
                "material_type": material_fixed,
                "approved_quantity_qtl": qty,
                "approved_rate_per_qtl": rate,
                "approved_party_name": approved_party,
                "line_amount": amt,
                # Backward-compatible aliases used by some UI helpers.
                "quantity_qtl": qty,
                "rate_per_qtl": rate,
            })

        if not normalized_items:
            return {
                "document_type": doc_type,
                "manual_required": True,
                "reason": "PO has no eligible won items for configured companies.",
                "skipped_lines": skipped_lines,
            }

        po = db.execute(
            select(PurchaseOrder).where(PurchaseOrder.source_pending_id == p.id)
        ).scalar_one_or_none()
        if not po:
            po = PurchaseOrder(source_pending_id=p.id)
            db.add(po)

        po.company_id = company_id
        po.main_tender_id = main_tender_id
        po.po_number = po_number
        po.po_date = po_date
        po.seller_name = payload.get("seller_name") or payload.get("vendor_name")
        po.buyer_name = payload.get("winner_party_name") or payload.get("buyer_name") or own_company
        po.buyer_email = payload.get("winner_party_email") or payload.get("buyer_email")
        po.plant_id = plant_id
        po.plant_name = po_plant_name
        po.total_amount = _to_float(payload.get("total_amount"))
        po.line_items = normalized_items
        po.status = PurchaseOrderStatus.approved
        po.source = "ingest"
        po.source_doc_path = p.file_path
        po.notes = (
            f"Auto-created from pending ingest #{p.id}; tender_rm={tender_rm}; "
            f"supply_start={supply_start.isoformat()}; supply_end={supply_end.isoformat()}"
        )
        db.flush()

        # Business rule: create/update sub-tenders only from PO (not from NIT).
        created_tender_ids: List[int] = []
        cycle = _ensure_cycle_for_po_action(db, company_id, supply_start, supply_end)
        if cycle and plant_id and normalized_items:
            for idx, item in enumerate(normalized_items, start=1):
                qty = _to_float(item.get("approved_quantity_qtl") or item.get("quantity_qtl"))
                if qty is None or qty <= 0:
                    continue
                mat_id, mat_name = _resolve_material(db, material_name=item.get("material_type"))
                if not mat_id:
                    continue
                tender_no = _build_po_sub_tender_number(main_code, po_plant_name, mat_name, idx)
                t = db.execute(
                    select(Tender).where(
                        Tender.tender_number == tender_no,
                        Tender.company_id == company_id,
                        Tender.main_tender_id == main_tender_id,
                    )
                ).scalar_one_or_none()
                week1_deadline = supply_start + timedelta(days=7)
                if week1_deadline > supply_end:
                    week1_deadline = supply_end
                if not t:
                    t = Tender(
                        company_id=company_id,
                        main_tender_id=main_tender_id,
                        cycle_id=cycle.id,
                        tender_number=tender_no,
                        plant_id=plant_id,
                        material_id=mat_id,
                        tender_mt=qty,
                        week1_target_mt=round(qty / 2.0, 3),
                        week1_deadline=week1_deadline,
                        week2_deadline=supply_end,
                        status=TenderStatus.active,
                        notes=f"Auto-created from PO {po_number}",
                    )
                    db.add(t)
                    db.flush()
                else:
                    # keep latest quantity and mark active
                    t.tender_mt = qty
                    t.week1_target_mt = round(qty / 2.0, 3)
                    t.week1_deadline = week1_deadline
                    t.week2_deadline = supply_end
                    t.cycle_id = cycle.id
                    t.status = TenderStatus.active
                created_tender_ids.append(t.id)

            if created_tender_ids:
                po.tender_id = created_tender_ids[0]
                p.tender_id = created_tender_ids[0]

        return {
            "document_type": doc_type,
            "purchase_order_id": po.id,
            "company_id": company_id,
            "main_tender_id": main_tender_id,
            "created_tender_ids": created_tender_ids,
            "supply_period_start": supply_start.isoformat(),
            "supply_period_end": supply_end.isoformat(),
            "cycle_id": cycle.id if cycle else None,
            "skipped_lines": skipped_lines,
            "note": "Sub-tenders created/updated from PO details only.",
        }

    if doc_type == "not_classified":
        return {
            "document_type": doc_type,
            "manual_required": True,
            "reason": "Document was not classified into predefined types and needs manual review.",
        }

    raise HTTPException(400, f"Unsupported document type for action: {doc_type}")


def _ingest_unloading_rows(
    db: Session,
    extracted: dict,
    image_path: str,
    source: str = "web",
    company_id: Optional[int] = None,
    main_tender_id: Optional[int] = None,
    tender_id: Optional[int] = None,
) -> tuple[list[PlantUnloadingEntry], int, int]:
    rows = extracted.get("rows") or []
    if not isinstance(rows, list):
        rows = []

    created_entries: list[PlantUnloadingEntry] = []
    duplicates = 0
    total_rows = len(rows)
    master_rows_cache: dict[int, List[PlantUnloadingEntry]] = {}

    header_rm = _clean_optional_text(extracted.get("rm_number"))
    header_item = _clean_optional_text(extracted.get("item_name") or extracted.get("item"))
    header_party = _clean_optional_text(extracted.get("party_name") or extracted.get("vendor_name"))
    header_plant = _clean_optional_text(extracted.get("plant_name") or extracted.get("destination_plant"))
    header_po = _clean_optional_text(extracted.get("po_number"))

    for row in rows:
        if not isinstance(row, dict):
            continue
        row_date = _parse_optional_date(row.get("date") or row.get("entry_date")) or date.today()
        truck = str(row.get("truck_number") or row.get("truck_no") or "").strip()
        ws_no = str(row.get("ws_no") or row.get("ws_number") or "").strip() or None
        no_of_bags = _to_int(row.get("no_of_bags"))
        received_qty = _to_float(row.get("received_qty_mt") or row.get("received_quantity_mt") or row.get("received_qty"))
        net_qty = _to_float(row.get("net_qty_mt") or row.get("net_quantity_mt") or row.get("net_weight"))
        total_qty = _to_float(row.get("total_qty_mt") or row.get("total_qty"))

        row_item = _clean_optional_text(row.get("item_name") or row.get("item") or header_item) or "UNKNOWN"
        row_party = _clean_optional_text(row.get("party_name") or header_party) or "UNKNOWN"
        row_rm = _clean_optional_text(row.get("rm_number") or header_rm) or "UNKNOWN"
        row_plant = _clean_optional_text(row.get("plant_name")) or header_plant
        row_po = _clean_optional_text(row.get("po_number") or header_po)

        master = _get_or_create_unloading_master(
            db=db,
            rm_number=row_rm,
            item_name=row_item,
            party_name=row_party,
            plant_name=row_plant,
            po_number=row_po,
            company_id=company_id,
            main_tender_id=main_tender_id,
            tender_id=tender_id,
            truck_number=truck,
        )

        cached_rows = master_rows_cache.get(master.id)
        if cached_rows is None:
            cached_rows = db.execute(
                select(PlantUnloadingEntry).where(PlantUnloadingEntry.master_id == master.id)
            ).scalars().all()
            master_rows_cache[master.id] = cached_rows

        if not truck:
            # If truck number is missing, keep deterministic placeholder for dedupe check.
            truck = f"UNKNOWN-{row.get('sno') or ws_no or row_date.isoformat()}"

        dedupe_key = _build_unloading_dedupe_key(
            master.id,
            ws_no,
            row_date,
            truck,
            net_qty,
            received_qty,
        )
        existing = db.execute(select(PlantUnloadingEntry).where(PlantUnloadingEntry.dedupe_key == dedupe_key)).scalar_one_or_none()
        if existing:
            duplicates += 1
            continue

        if _is_duplicate_unloading_row_for_master(
            cached_rows,
            ws_no=ws_no,
            entry_date=row_date,
            truck_number=truck,
            net_qty=net_qty,
            received_qty=received_qty,
            no_of_bags=no_of_bags,
        ):
            duplicates += 1
            continue

        needs_review = net_qty is None or _norm_code(truck).startswith("unknown")
        e = PlantUnloadingEntry(
            company_id=master.company_id,
            main_tender_id=master.main_tender_id,
            tender_id=master.tender_id,
            master_id=master.id,
            image_path=image_path,
            source=source,
            ws_no=ws_no,
            entry_date=row_date,
            truck_number=truck,
            no_of_bags=no_of_bags,
            received_qty_mt=received_qty,
            net_qty_mt=net_qty if net_qty is not None else (received_qty if received_qty is not None else 0),
            total_qty_mt=total_qty,
            item_name=row_item,
            status=BillStatus.flagged if needs_review else BillStatus.pending,
            dedupe_key=dedupe_key,
            ocr_source=OcrSource.paddle if extracted.get("source") == "paddle" else OcrSource.gemini,
            ocr_confidence=_to_float(extracted.get("confidence") or extracted.get("ocr_confidence")),
            ocr_raw_json=json.dumps({"header": {k: extracted.get(k) for k in ["rm_number", "item_name", "party_name", "plant_name", "po_number"]}, "row": row}, ensure_ascii=False),
        )
        db.add(e)
        created_entries.append(e)
        cached_rows.append(e)

    db.flush()
    _reconcile_unloading_match_for_all_bills(db)
    db.commit()
    for e in created_entries:
        db.refresh(e)

    return created_entries, duplicates, total_rows


def _parse_optional_date(val):
    if val is None:
        return None
    s = str(val).strip()
    if not s or s.lower() in ('none', 'null', 'n/a'):
        return None
    # If it's already a date
    if isinstance(val, date):
        return val
    # Try ISO first
    try:
        return date.fromisoformat(s)
    except Exception:
        pass
    # Try common formats
    from datetime import datetime as _dt
    for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%d %b %Y", "%d %B %Y", "%Y/%m/%d", "%d.%m.%Y"):
        try:
            return _dt.strptime(s, fmt).date()
        except Exception:
            continue
    return None


def _normalize_ocr_engine(ocr_engine: Optional[str]) -> Optional[str]:
    engine = str(ocr_engine or "").strip().lower()
    if engine in {"", "auto"}:
        return None
    allowed_engines = {"groq", "mistral", "openrouter", "gemini", "azure", "github"}
    if engine not in allowed_engines:
        raise HTTPException(400, f"Invalid ocr_engine: {ocr_engine}")
    return engine


def _apply_extracted_to_busy_staging(b: BusyStagingBill, extracted: dict) -> None:
    plant_name_extracted = extracted.get("destination_plant") or extracted.get("plant_name")
    bill_date_parsed = _parse_optional_date(extracted.get("bill_date"))

    def _num(v):
        try:
            return float(v)
        except Exception:
            return None

    b.vehicle_number = extracted.get("vehicle_number")
    b.broker_name = extracted.get("vendor_name") or extracted.get("broker_name") or None
    b.material_name = extracted.get("material_type")
    b.qty_mt = _num(extracted.get("quantity_qtl"))
    b.rate_per_mt = _num(extracted.get("rate_per_qtl"))
    b.total_amount = _num(extracted.get("total_amount"))
    b.bill_date = bill_date_parsed
    b.bill_number = extracted.get("bill_number")
    b.plant_name = plant_name_extracted

    src = str(extracted.get("source") or "").strip().lower()
    if src == "paddle":
        b.ocr_source = OcrSource.paddle
    elif src == "manual":
        b.ocr_source = OcrSource.manual
    else:
        b.ocr_source = OcrSource.gemini

    b.ocr_confidence = extracted.get("confidence") or extracted.get("ocr_confidence")
    b.ocr_raw_text = extracted.get("raw_text") or extracted.get("ocr_raw_text") or ""
    b.is_handwritten = extracted.get("is_handwritten", False)
    b.validation_amount = extracted.get("validation_amount", False)
    b.validation_vehicle = extracted.get("validation_vehicle", False)
    b.validation_material = extracted.get("validation_material", False)


def _apply_scope_filters(q, model, company_id: Optional[int] = None, main_tender_id: Optional[int] = None, tender_id: Optional[int] = None):
    if company_id is not None and hasattr(model, "company_id"):
        q = q.where(getattr(model, "company_id") == company_id)
    if main_tender_id is not None and hasattr(model, "main_tender_id"):
        q = q.where(getattr(model, "main_tender_id") == main_tender_id)
    if tender_id is not None and hasattr(model, "tender_id"):
        q = q.where(getattr(model, "tender_id") == tender_id)
    return q


def _assert_record_scope(
    record,
    company_id: Optional[int] = None,
    main_tender_id: Optional[int] = None,
    tender_id: Optional[int] = None,
    label: str = "resource",
) -> None:
    if record is None:
        raise HTTPException(404, f"{label} not found")
    if company_id is not None and getattr(record, "company_id", None) != company_id:
        raise HTTPException(404, f"{label} not found")
    if main_tender_id is not None and getattr(record, "main_tender_id", None) != main_tender_id:
        raise HTTPException(404, f"{label} not found")
    if tender_id is not None and getattr(record, "tender_id", None) != tender_id:
        raise HTTPException(404, f"{label} not found")


def _resolve_context_ids(
    db: Session,
    company_id: Optional[int] = None,
    main_tender_id: Optional[int] = None,
    tender_id: Optional[int] = None,
) -> tuple[Optional[int], Optional[int], Optional[int]]:
    resolved_company = company_id
    resolved_main = main_tender_id
    resolved_tender = tender_id

    if resolved_tender is not None:
        tc, tm, _ = _lineage_from_tender(db, resolved_tender)
        if tc is None and tm is None:
            raise HTTPException(404, "Tender not found")
        if resolved_company is not None and tc is not None and resolved_company != tc:
            raise HTTPException(400, "company_id does not match tender lineage")
        if resolved_main is not None and tm is not None and resolved_main != tm:
            raise HTTPException(400, "main_tender_id does not match tender lineage")
        resolved_company = resolved_company if resolved_company is not None else tc
        resolved_main = resolved_main if resolved_main is not None else tm

    if resolved_main is not None:
        mt = db.get(MainTender, resolved_main)
        if not mt:
            raise HTTPException(404, "Main tender not found")
        if resolved_company is not None and resolved_company != mt.company_id:
            raise HTTPException(400, "company_id does not match main tender")
        resolved_company = resolved_company if resolved_company is not None else mt.company_id

    if resolved_company is not None and not db.get(Company, resolved_company):
        raise HTTPException(404, "Company not found")

    return resolved_company, resolved_main, resolved_tender


def _lineage_from_tender(db: Session, tender_id: Optional[int]) -> tuple[Optional[int], Optional[int], Optional[int]]:
    if not tender_id:
        return None, None, None
    t = db.get(Tender, tender_id)
    if not t:
        return None, None, None
    return t.company_id, t.main_tender_id, t.id


def _lineage_from_deal(db: Session, deal_id: Optional[int]) -> tuple[Optional[int], Optional[int], Optional[int]]:
    if not deal_id:
        return None, None, None
    d = db.get(Deal, deal_id)
    if not d:
        return None, None, None
    company_id = d.company_id
    main_tender_id = d.main_tender_id
    tender_id = d.tender_id
    if tender_id and (company_id is None or main_tender_id is None):
        tc, tm, _ = _lineage_from_tender(db, tender_id)
        company_id = company_id if company_id is not None else tc
        main_tender_id = main_tender_id if main_tender_id is not None else tm
    if tender_id:
        tc, tm, _ = _lineage_from_tender(db, tender_id)
        if tc is not None and company_id is not None and tc != company_id:
            raise HTTPException(400, "company_id does not match tender lineage")
        if tm is not None and main_tender_id is not None and tm != main_tender_id:
            raise HTTPException(400, "main_tender_id does not match tender lineage")
    return company_id, main_tender_id, tender_id


def _stamp_lineage(obj, company_id: Optional[int], main_tender_id: Optional[int], tender_id: Optional[int]) -> None:
    if hasattr(obj, "company_id"):
        obj.company_id = company_id
    if hasattr(obj, "main_tender_id"):
        obj.main_tender_id = main_tender_id
    if hasattr(obj, "tender_id"):
        obj.tender_id = tender_id


def _company_to_dict(c: Company) -> dict:
    return {
        "id": c.id,
        "name": c.name,
        "code": c.code,
        "is_active": bool(c.is_active),
    }


def _main_tender_to_dict(mt: MainTender) -> dict:
    return {
        "id": mt.id,
        "company_id": mt.company_id,
        "tender_code": mt.tender_code,
        "title": mt.title,
        "notes": mt.notes,
        "is_active": bool(mt.is_active),
        "created_at": str(mt.created_at) if mt.created_at else None,
    }


def _derive_main_tender_code(tender_number: Optional[str]) -> str:
    norm = normalize_rm_number(tender_number)
    if norm:
        return norm
    s = str(tender_number or "").strip().upper()
    if not s:
        return "UNTITLED"
    m = re.match(r"^([A-Z]+[-/]?\d+)", s)
    if m:
        return m.group(1).replace("/", "-")
    parts = re.split(r"[-\s]+", s)
    if len(parts) >= 2 and parts[0].isalpha() and parts[1].isdigit():
        return f"{parts[0]}-{parts[1]}"
    return parts[0] if parts else s


def _is_duplicate_bill(db: Session, vendor_name: Optional[str], vehicle_number: Optional[str], bill_number: Optional[str], operator: Optional[str] = None) -> bool:
    """Return True if a bill exists in DB matching vendor_name and bill_number.
    Vehicle number is ignored for duplicate detection.
    Matching strategy:
    - Normalize vendor and bill strings (strip common prefixes/punctuation, lowercase).
    - Quick SQL equality fast-path, then fall back to candidate scans with Python normalization.
    - If `operator` is provided, ensure `Bill.reviewed_by` equals it.
    """
    if not (vendor_name and bill_number):
        return False

    def normalize_vendor(s: str) -> str:
        s = str(s or "").strip().lower()
        s = re.sub(r'^m\s*/?\s*s\.?\s*', '', s)
        s = re.sub(r'[\.\,\(\)]', '', s)
        s = re.sub(r'\s+', ' ', s)
        return s

    def normalize_bill_no(s: str) -> str:
        s = str(s or "").strip().lower()
        s = re.sub(r'^(invoice|inv|bill|no|no\.|no:|invoice\s*no)\s*[:#\-\s]*', '', s)
        s = re.sub(r'[^a-z0-9\-\/]', '', s)
        return s

    # vehicle normalization removed — vehicle is not part of duplicate detection

    try:
        vendor_n = normalize_vendor(vendor_name)
        bill_n = normalize_bill_no(bill_number)
        log.debug(f"Duplicate check -- vendor_n='{vendor_n}', bill_n='{bill_n}', operator='{operator}'")

        # Quick exact normalized match in SQL first (fast path)
        try:
            sql_q = select(Bill).where(
                func.lower(func.coalesce(Bill.bill_number, '')) == (bill_n or '').lower(),
                func.lower(func.coalesce(Bill.broker_name, '')) == (vendor_n or '').lower(),
            )
            if operator:
                sql_q = sql_q.where(func.coalesce(Bill.reviewed_by, '') == operator)
            existing = db.execute(sql_q).scalar_one_or_none()
            if existing:
                log.info(f"Duplicate detected by SQL fast path against bill id={existing.id}")
                return True
        except Exception:
            pass

        # Candidate bills: match bill_number OR broker_name roughly
        candidates = db.execute(
            select(Bill).where(
                (Bill.bill_number != None) & (
                    func.lower(Bill.bill_number).like(f"%{bill_n}%")
                )
            ).limit(50)
        ).scalars().all()

        # Also include bills where broker_name or linked Broker.name matches
        broker_cands = db.execute(select(Bill).where(Bill.broker_name != None).limit(50)).scalars().all()
        candidates.extend(broker_cands)

        # Deduplicate candidate list
        seen_ids = set()
        uniq = []
        for c in candidates:
            if c.id in seen_ids: continue
            seen_ids.add(c.id); uniq.append(c)

        for c in uniq:
            if operator and (c.reviewed_by or '') != operator:
                continue
            # Normalize stored values
            stored_vendor = normalize_vendor(c.broker_name or (c.broker.name if c.broker else ''))
            stored_bill = normalize_bill_no(c.bill_number or '')
            log.debug(f"Duplicate candidates: bill_id={c.id} vendor='{stored_vendor}' bill='{stored_bill}'")
            if stored_vendor == vendor_n and stored_bill == bill_n:
                log.info(f"Duplicate detected against bill id={c.id}")
                return True

        return False
    except Exception:
        return False

# ── DASHBOARD SUMMARY ──────────────────────────────────────────────────────

@app.get("/api/companies")
def list_companies(db: Session = Depends(get_db)):
    rows = db.execute(select(Company).where(Company.is_active == True).order_by(Company.name)).scalars().all()
    return [_company_to_dict(c) for c in rows]


@app.post("/api/companies")
def create_company(data: CompanyCreate, db: Session = Depends(get_db)):
    c = Company(name=data.name.strip(), code=(data.code.strip() if data.code else None))
    db.add(c)
    db.commit()
    db.refresh(c)
    return _company_to_dict(c)


@app.get("/api/main-tenders")
def list_main_tenders(company_id: Optional[int] = None, db: Session = Depends(get_db)):
    q = select(MainTender).where(MainTender.is_active == True).order_by(desc(MainTender.created_at))
    if company_id is not None:
        q = q.where(MainTender.company_id == company_id)
    rows = db.execute(q).scalars().all()
    return [_main_tender_to_dict(mt) for mt in rows]


@app.post("/api/main-tenders")
def create_main_tender(data: MainTenderCreate, db: Session = Depends(get_db)):
    if not db.get(Company, data.company_id):
        raise HTTPException(404, "Company not found")
    mt = MainTender(
        company_id=data.company_id,
        tender_code=data.tender_code.strip(),
        title=(data.title.strip() if data.title else None),
        notes=(data.notes.strip() if data.notes else None),
    )
    db.add(mt)
    db.commit()
    db.refresh(mt)
    return _main_tender_to_dict(mt)


@app.post("/api/admin/backfill-lineage")
def admin_backfill_lineage(
    company_id: Optional[int] = Query(None),
    fill_missing_company: bool = Query(False),
    dry_run: bool = Query(False),
    db: Session = Depends(get_db),
):
    if company_id is not None and not db.get(Company, company_id):
        raise HTTPException(404, "Company not found")

    stats = {
        "tenders_company_filled": 0,
        "main_tenders_created": 0,
        "tenders_main_filled": 0,
        "deals_filled": 0,
        "bills_filled": 0,
        "dispatches_filled": 0,
        "receipts_filled": 0,
        "purchase_bills_filled": 0,
        "payments_filled": 0,
        "unloading_masters_filled": 0,
        "unloading_entries_filled": 0,
        "unloading_masters_rm_normalized": 0,
        "unloading_masters_assignment_refreshed": 0,
        "unloading_masters_manual_required": 0,
        "tenders_fulfilled_recomputed": 0,
    }

    t_q = select(Tender)
    if company_id is not None:
        t_q = t_q.where((Tender.company_id == company_id) | (Tender.company_id == None))
    tenders = db.execute(t_q).scalars().all()

    mt_cache = {}
    for t in tenders:
        if t.company_id is None and fill_missing_company and company_id is not None:
            stats["tenders_company_filled"] += 1
            if not dry_run:
                t.company_id = company_id

        if t.company_id is None:
            continue
        if company_id is not None and t.company_id != company_id:
            continue

        if t.main_tender_id is None:
            code = _derive_main_tender_code(t.tender_number)
            key = (t.company_id, code)
            mt = mt_cache.get(key)
            if mt is None:
                mt = db.execute(
                    select(MainTender).where(MainTender.company_id == t.company_id, MainTender.tender_code == code)
                ).scalar_one_or_none()
                if mt is None:
                    stats["main_tenders_created"] += 1
                    if not dry_run:
                        mt = MainTender(company_id=t.company_id, tender_code=code, title=f"{code} master")
                        db.add(mt)
                        db.flush()
                mt_cache[key] = mt
            stats["tenders_main_filled"] += 1
            if not dry_run and mt is not None:
                t.main_tender_id = mt.id

    d_q = select(Deal)
    if company_id is not None:
        d_q = d_q.where((Deal.company_id == company_id) | (Deal.company_id == None))
    for d in db.execute(d_q).scalars().all():
        t = db.get(Tender, d.tender_id)
        if not t:
            continue
        if company_id is not None and t.company_id != company_id:
            continue
        needs = (d.company_id != t.company_id) or (d.main_tender_id != t.main_tender_id)
        if needs:
            stats["deals_filled"] += 1
            if not dry_run:
                d.company_id = t.company_id
                d.main_tender_id = t.main_tender_id

    b_q = select(Bill)
    if company_id is not None:
        b_q = b_q.where((Bill.company_id == company_id) | (Bill.company_id == None))
    for b in db.execute(b_q).scalars().all():
        lc = lm = lt = None
        if b.deal_id:
            lc, lm, lt = _lineage_from_deal(db, b.deal_id)
        elif b.tender_id:
            lc, lm, lt = _lineage_from_tender(db, b.tender_id)
        if company_id is not None and lc != company_id:
            continue
        needs = (b.company_id != lc) or (b.main_tender_id != lm) or (lt is not None and b.tender_id != lt)
        if needs:
            stats["bills_filled"] += 1
            if not dry_run:
                b.company_id = lc
                b.main_tender_id = lm
                if lt is not None:
                    b.tender_id = lt

    ds_q = select(Dispatch)
    if company_id is not None:
        ds_q = ds_q.where((Dispatch.company_id == company_id) | (Dispatch.company_id == None))
    for ds in db.execute(ds_q).scalars().all():
        lc, lm, lt = _lineage_from_deal(db, ds.deal_id)
        if company_id is not None and lc != company_id:
            continue
        needs = (ds.company_id != lc) or (ds.main_tender_id != lm) or (ds.tender_id != lt)
        if needs:
            stats["dispatches_filled"] += 1
            if not dry_run:
                ds.company_id = lc
                ds.main_tender_id = lm
                ds.tender_id = lt

    r_q = select(PlantReceipt)
    if company_id is not None:
        r_q = r_q.where((PlantReceipt.company_id == company_id) | (PlantReceipt.company_id == None))
    for r in db.execute(r_q).scalars().all():
        lc = lm = lt = None
        if r.dispatch_id:
            ds = db.get(Dispatch, r.dispatch_id)
            if ds:
                lc, lm, lt = ds.company_id, ds.main_tender_id, ds.tender_id
        else:
            # Keep lineage on dispatchless receipts (e.g., unmatched unloading receipts)
            # and avoid forcing them to NULL on every backfill pass.
            continue
        if lc is None and lm is None and lt is None:
            continue
        if company_id is not None and lc != company_id:
            continue
        needs = (r.company_id != lc) or (r.main_tender_id != lm) or (r.tender_id != lt)
        if needs:
            stats["receipts_filled"] += 1
            if not dry_run:
                r.company_id = lc
                r.main_tender_id = lm
                r.tender_id = lt

    pb_q = select(PurchaseBill)
    if company_id is not None:
        pb_q = pb_q.where((PurchaseBill.company_id == company_id) | (PurchaseBill.company_id == None))
    for pb in db.execute(pb_q).scalars().all():
        b = db.get(Bill, pb.bill_id)
        if not b:
            continue
        if company_id is not None and b.company_id != company_id:
            continue
        needs = (pb.company_id != b.company_id) or (pb.main_tender_id != b.main_tender_id) or (pb.tender_id != b.tender_id)
        if needs:
            stats["purchase_bills_filled"] += 1
            if not dry_run:
                pb.company_id = b.company_id
                pb.main_tender_id = b.main_tender_id
                pb.tender_id = b.tender_id

    p_q = select(Payment)
    if company_id is not None:
        p_q = p_q.where((Payment.company_id == company_id) | (Payment.company_id == None))
    for p in db.execute(p_q).scalars().all():
        pb = db.get(PurchaseBill, p.purchase_bill_id)
        if not pb:
            continue
        if company_id is not None and pb.company_id != company_id:
            continue
        needs = (p.company_id != pb.company_id) or (p.main_tender_id != pb.main_tender_id) or (p.tender_id != pb.tender_id)
        if needs:
            stats["payments_filled"] += 1
            if not dry_run:
                p.company_id = pb.company_id
                p.main_tender_id = pb.main_tender_id
                p.tender_id = pb.tender_id

    m_q = select(PlantUnloadingMaster)
    if company_id is not None:
        m_q = m_q.where((PlantUnloadingMaster.company_id == company_id) | (PlantUnloadingMaster.company_id == None))
    for m in db.execute(m_q).scalars().all():
        if company_id is not None and m.company_id not in (None, company_id):
            continue

        old_lineage = (m.company_id, m.main_tender_id, m.tender_id)
        old_assign = (
            m.assignment_status,
            bool(m.requires_manual_assignment),
            m.mapping_source,
            _norm_text(m.assignment_reason),
        )

        rm_norm, rm_base, _suffix = _normalize_rm_components(m.rm_number)
        if _norm_text(m.rm_number_norm) != _norm_text(rm_norm) or _norm_text(m.rm_number_base) != _norm_text(rm_base):
            stats["unloading_masters_rm_normalized"] += 1
            if not dry_run:
                m.rm_number_norm = rm_norm
                m.rm_number_base = rm_base

        assignment = _resolve_unloading_assignment(
            db,
            rm_number=m.rm_number,
            item_name=m.item_name,
            plant_name=m.plant_name,
            company_id_hint=m.company_id,
            main_tender_id_hint=m.main_tender_id,
            tender_id_hint=m.tender_id,
        )

        if m.is_manual_override and m.tender_id:
            final_lineage = old_lineage
            final_manual_required = False
            final_assign = ("assigned", False, m.mapping_source or "manual", _norm_text(m.assignment_reason))
            if not dry_run:
                m.assignment_status = "assigned"
                m.requires_manual_assignment = False
                m.mapping_source = m.mapping_source or "manual"
        else:
            final_lineage = (
                assignment.get("company_id"),
                assignment.get("main_tender_id"),
                assignment.get("tender_id"),
            )
            final_manual_required = bool(assignment.get("requires_manual_assignment"))
            final_assign = (
                assignment.get("assignment_status") or "pending",
                final_manual_required,
                assignment.get("mapping_source"),
                _norm_text(assignment.get("assignment_reason")),
            )
            if not dry_run:
                _apply_unloading_assignment_to_master(m, assignment)
                _sync_unloading_lineage_from_master(db, m)

        if old_lineage != final_lineage and final_lineage[0] is not None:
            stats["unloading_masters_filled"] += 1

        if old_assign != final_assign:
            stats["unloading_masters_assignment_refreshed"] += 1

        if final_manual_required:
            stats["unloading_masters_manual_required"] += 1

    e_q = select(PlantUnloadingEntry)
    if company_id is not None:
        e_q = e_q.where((PlantUnloadingEntry.company_id == company_id) | (PlantUnloadingEntry.company_id == None))
    for e in db.execute(e_q).scalars().all():
        m = db.get(PlantUnloadingMaster, e.master_id)
        if not m:
            continue
        if company_id is not None and m.company_id != company_id:
            continue
        needs = (e.company_id != m.company_id) or (e.main_tender_id != m.main_tender_id) or (e.tender_id != m.tender_id)
        if needs:
            stats["unloading_entries_filled"] += 1
            if not dry_run:
                e.company_id = m.company_id
                e.main_tender_id = m.main_tender_id
                e.tender_id = m.tender_id

    if not dry_run:
        db.flush()
        stats["tenders_fulfilled_recomputed"] = _recalculate_tender_fulfilled_from_unloading(db, company_id=company_id)
        db.commit()
    else:
        stats["tenders_fulfilled_recomputed"] = _count_tender_fulfilled_mismatches(db, company_id=company_id)
    return {"ok": True, "dry_run": dry_run, "stats": stats}


@app.post("/api/admin/merge-unloading-duplicates")
def admin_merge_unloading_duplicates(
    company_id: Optional[int] = Query(None),
    main_tender_id: Optional[int] = Query(None),
    rm_number: Optional[str] = Query(None),
    dry_run: bool = Query(True),
    delete_empty_masters: bool = Query(True),
    db: Session = Depends(get_db),
):
    if company_id is not None and not db.get(Company, company_id):
        raise HTTPException(404, "Company not found")

    if main_tender_id is not None:
        mt = db.get(MainTender, main_tender_id)
        if not mt:
            raise HTTPException(404, "Main tender not found")
        if company_id is not None and mt.company_id != company_id:
            raise HTTPException(400, "main_tender_id does not belong to company_id")
        if company_id is None:
            company_id = mt.company_id

    report = _merge_unloading_duplicate_masters(
        db,
        company_id=company_id,
        main_tender_id=main_tender_id,
        rm_number=rm_number,
        dry_run=dry_run,
        delete_empty_masters=delete_empty_masters,
    )

    impacted_tender_ids = report.get("impacted_tender_ids") or []
    stats = report.get("stats") or {}

    if not dry_run:
        db.flush()
        stats["tenders_fulfilled_recomputed"] = _recalculate_tender_fulfilled_from_unloading(
            db,
            tender_ids=impacted_tender_ids,
        )
        db.commit()
    else:
        stats["tenders_fulfilled_recomputed"] = 0

    report["stats"] = stats
    return {
        "ok": True,
        "dry_run": dry_run,
        "company_id": company_id,
        "main_tender_id": main_tender_id,
        "rm_number": rm_number,
        "delete_empty_masters": delete_empty_masters,
        **report,
    }

@app.get("/api/dashboard")
def dashboard_summary(
    cycle_id: Optional[int] = None,
    company_id: Optional[int] = None,
    main_tender_id: Optional[int] = None,
    tender_id: Optional[int] = None,
    db: Session = Depends(get_db),
):
    s = get_settings()

    # Active cycle
    cycle = db.execute(select(SproxxCycle).where(SproxxCycle.is_active == True)).scalar_one_or_none()

    q = select(Tender)
    q = _apply_scope_filters(q, Tender, company_id, main_tender_id, tender_id)
    if cycle_id:
        q = q.where(Tender.cycle_id == cycle_id)
    elif cycle:
        q = q.where(Tender.cycle_id == cycle.id)
    tenders = db.execute(q).scalars().all()

    total_tender = total_dispatched = total_accepted = 0
    at_risk = penalty_risk = 0
    for t in tenders:
        for d in t.deals:
            if d.status == DealStatus.cancelled: continue
            total_dispatched += float(d.dispatched_mt or 0)
            total_accepted   += float(d.accepted_mt or 0)
        total_tender += float(t.tender_mt)
        summary = _tender_to_dict(t)["summary"]
        if summary["accepted_pct"] < 50: at_risk += 1
        if summary["week1_pct"] < 80:    penalty_risk += 1

    # Bills stats
    bill_counts = {s.value: 0 for s in BillStatus}
    bill_count_q = select(Bill.status, func.count(Bill.id)).group_by(Bill.status)
    bill_count_q = _apply_scope_filters(bill_count_q, Bill, company_id, main_tender_id, tender_id)
    for row in db.execute(bill_count_q).all():
        bill_counts[row[0].value] = row[1]

    # Payments pending
    pending_q = select(func.sum(PurchaseBill.total_amount)).where(PurchaseBill.status == PurchaseBillStatus.draft)
    pending_q = _apply_scope_filters(pending_q, PurchaseBill, company_id, main_tender_id, tender_id)
    pending_payments = db.execute(pending_q).scalar() or 0

    # Outstanding sales
    outstanding_q = select(func.sum(SalesBill.total_amount)).where(SalesBill.status.in_([SalesBillStatus.sent, SalesBillStatus.overdue]))
    outstanding_q = _apply_scope_filters(outstanding_q, SalesBill, company_id, main_tender_id)
    outstanding_sales = db.execute(outstanding_q).scalar() or 0

    return {
        "cycle": {"id": cycle.id, "name": cycle.name} if cycle else None,
        "tenders": {
            "total": len(tenders),
            "tender_mt": round(total_tender, 2),
            "dispatched_mt": round(total_dispatched, 2),
            "accepted_mt": round(total_accepted, 2),
            "accepted_pct": round(total_accepted / total_tender * 100, 1) if total_tender else 0,
            "at_risk": at_risk,
            "penalty_risk": penalty_risk,
        },
        "bills": bill_counts,
        "payments": {
            "pending_purchase": round(float(pending_payments), 2),
            "outstanding_sales": round(float(outstanding_sales), 2),
        }
    }

# ── SPROXX CYCLES ──────────────────────────────────────────────────────────

@app.get("/api/cycles")
def list_cycles(company_id: Optional[int] = None, db: Session = Depends(get_db)):
    q = select(SproxxCycle).order_by(desc(SproxxCycle.cycle_start))
    if company_id is not None:
        q = q.where(or_(SproxxCycle.company_id == company_id, SproxxCycle.company_id == None))
    cycles = db.execute(q).scalars().all()
    return [{"id": c.id, "name": c.name, "cycle_start": str(c.cycle_start),
             "cycle_end": str(c.cycle_end), "week1_end": str(c.week1_end),
             "week2_end": str(c.week2_end), "is_active": c.is_active,
             "company_id": c.company_id} for c in cycles]

@app.post("/api/cycles")
def create_cycle(data: CycleCreate, company_id: Optional[int] = Query(None), db: Session = Depends(get_db)):
    payload = data.model_dump()
    payload["company_id"] = payload.get("company_id") if payload.get("company_id") is not None else company_id
    cycle = SproxxCycle(**payload)
    db.add(cycle); db.commit(); db.refresh(cycle)
    return {"id": cycle.id, "name": cycle.name}

@app.patch("/api/cycles/{cycle_id}/activate")
def activate_cycle(cycle_id: int, db: Session = Depends(get_db)):
    db.execute(select(SproxxCycle).where(SproxxCycle.is_active == True).execution_options(
        synchronize_session="fetch"))
    db.query(SproxxCycle).update({"is_active": False})
    db.query(SproxxCycle).filter_by(id=cycle_id).update({"is_active": True})
    db.commit()
    return {"ok": True}

# ── TENDERS ────────────────────────────────────────────────────────────────

@app.get("/api/tenders/nits")
def get_tender_invitations(q: Optional[str] = None, company_id: Optional[int] = None, db: Session = Depends(get_db)):
    """Get latest tender invitation (NIT) and searchable history of all NITs."""
    stmt = select(PendingIngest).where(PendingIngest.document_type == DocumentType.tender_notice)
    if company_id:
        stmt = stmt.where((PendingIngest.company_id == company_id) | (PendingIngest.assigned_company_id == company_id))
    stmt = stmt.order_by(desc(PendingIngest.created_at))
    ingests = db.execute(stmt).scalars().all()

    nit_records = []
    for ing in ingests:
        payload = ing.extracted_payload or ing.classifier_candidates or {}
        tender_rm = normalize_rm_number(
            payload.get("tender_rm_number") or payload.get("rm_number") or payload.get("tender_number") or ing.file_name
        )
        
        items = payload.get("items") if isinstance(payload.get("items"), list) else []
        plant_name = payload.get("plant_name") or "All Plants"
        bidding_deadline = payload.get("bidding_deadline") or payload.get("end_date")
        tender_date = payload.get("tender_date") or payload.get("start_date") or (str(ing.created_at.date()) if ing.created_at else "")

        processed_items = []
        for item in items:
            if isinstance(item, dict):
                processed_items.append({
                    "item_name": item.get("item_name") or item.get("name") or "Raw Material",
                    "asked_qty_mt": float(item.get("asked_qty_mt") or item.get("qty_mt") or 0),
                    "plant_name": item.get("plant_name") or plant_name,
                    "rate_per_mt": float(item.get("rate_per_mt")) if item.get("rate_per_mt") else None
                })
        
        rec = {
            "id": ing.id,
            "rm_number": tender_rm,
            "raw_rm_number": payload.get("tender_rm_number") or payload.get("rm_number"),
            "file_name": ing.file_name,
            "file_path": ing.file_path,
            "status": ing.status.value if ing.status else "pending",
            "plant_name": plant_name,
            "tender_date": str(tender_date),
            "bidding_deadline": str(bidding_deadline) if bidding_deadline else None,
            "items": processed_items,
            "total_asked_mt": sum(i["asked_qty_mt"] for i in processed_items),
            "created_at": str(ing.created_at) if ing.created_at else None,
        }

        if q:
            q_lower = q.lower().strip()
            searchable = f"{rec['rm_number']} {rec['plant_name']} {rec['file_name']} {' '.join(i['item_name'] for i in rec['items'])}".lower()
            if q_lower not in searchable:
                continue

        nit_records.append(rec)

    latest = nit_records[0] if nit_records else None
    return {
        "latest": latest,
        "history": nit_records,
        "total_count": len(nit_records)
    }


@app.get("/api/tenders")
def list_tenders(cycle_id: Optional[int] = None, plant_id: Optional[int] = None,
                 status: Optional[str] = None, company_id: Optional[int] = None,
                 main_tender_id: Optional[int] = None, db: Session = Depends(get_db)):
    q = select(Tender).order_by(desc(Tender.created_at))
    q = _apply_scope_filters(q, Tender, company_id, main_tender_id)
    if cycle_id:  q = q.where(Tender.cycle_id  == cycle_id)
    if plant_id:  q = q.where(Tender.plant_id  == plant_id)
    if status:    q = q.where(Tender.status     == status)
    tenders = db.execute(q).scalars().all()
    return [_tender_to_dict(t, db) for t in tenders]

@app.post("/api/tenders")
def create_tender(data: TenderCreate, company_id: Optional[int] = Query(None),
                  main_tender_id: Optional[int] = Query(None), db: Session = Depends(get_db)):
    payload = data.model_dump()
    mt_id = payload.get("main_tender_id") or main_tender_id
    company_id = payload.get("company_id") if payload.get("company_id") is not None else company_id
    company_id, mt_id, _ = _resolve_context_ids(db, company_id=company_id, main_tender_id=mt_id)
    payload["company_id"] = company_id
    payload["main_tender_id"] = mt_id
    t = Tender(**payload)
    t.status = TenderStatus.active
    db.add(t); db.commit(); db.refresh(t)
    return _tender_to_dict(t)

@app.get("/api/tenders/{tender_id}")
def get_tender(tender_id: int, company_id: Optional[int] = None,
               main_tender_id: Optional[int] = None, db: Session = Depends(get_db)):
    t = db.get(Tender, tender_id)
    _assert_record_scope(t, company_id, main_tender_id, label="Tender")
    return _tender_to_dict(t, db)

@app.patch("/api/tenders/{tender_id}/status")
def update_tender_status(tender_id: int, status: str, company_id: Optional[int] = None,
                         main_tender_id: Optional[int] = None, db: Session = Depends(get_db)):
    t = db.get(Tender, tender_id)
    _assert_record_scope(t, company_id, main_tender_id, label="Tender")
    t.status = TenderStatus(status)
    db.commit()
    return {"ok": True}

# ── DEALS ──────────────────────────────────────────────────────────────────

@app.get("/api/deals")
def list_deals(tender_id: Optional[int] = None, broker_id: Optional[int] = None,
               company_id: Optional[int] = None, main_tender_id: Optional[int] = None,
               db: Session = Depends(get_db)):
    from sqlalchemy.orm import joinedload

    q = select(Deal).options(
        joinedload(Deal.tender).joinedload(Tender.plant),
        joinedload(Deal.tender).joinedload(Tender.material),
        joinedload(Deal.broker),
        joinedload(Deal.material),
    ).order_by(desc(Deal.created_at))
    q = _apply_scope_filters(q, Deal, company_id, main_tender_id, tender_id)
    if tender_id: q = q.where(Deal.tender_id == tender_id)
    if broker_id: q = q.where(Deal.broker_id == broker_id)
    deals = db.execute(q).scalars().all()
    rows = []
    for d in deals:
        tender = d.tender
        tender_plant = None
        tender_item = None
        if tender:
            tender_plant = tender.plant.name if getattr(tender, "plant", None) else None
            tender_item = tender.material.name if getattr(tender, "material", None) else None
        subtender_name = f"{tender_plant} - {tender_item}" if tender_plant and tender_item else None

        rows.append({
            "id": d.id,
            "deal_number": d.deal_number,
            "company_id": d.company_id,
            "main_tender_id": d.main_tender_id,
            "tender_id": d.tender_id,
            "tender_plant": tender_plant,
            "tender_item": tender_item,
            "subtender_name": subtender_name,
            "broker": d.broker.name,
            "broker_id": d.broker_id,
            "material": d.material.name,
            "deal_mt": float(d.deal_mt),
            "rate_per_mt": float(d.rate_per_mt),
            "total_value": round(float(d.deal_mt) * float(d.rate_per_mt), 2),
            "dispatched_mt": float(d.dispatched_mt or 0),
            "accepted_mt": float(d.accepted_mt or 0),
            "rejected_mt": float(d.rejected_mt or 0),
            "status": d.status.value,
            "bills_count": len(d.bills),
        })
    return rows

@app.post("/api/deals")
def create_deal(data: DealCreate, company_id: Optional[int] = Query(None),
                main_tender_id: Optional[int] = Query(None), db: Session = Depends(get_db)):
    # Auto-generate deal number
    count = db.execute(select(func.count(Deal.id))).scalar() or 0
    deal_no = f"D-{datetime.now().strftime('%y%m')}-{count+1:03d}"
    payload = data.model_dump()
    cid = payload.get("company_id") if payload.get("company_id") is not None else company_id
    mid = payload.get("main_tender_id") if payload.get("main_tender_id") is not None else main_tender_id
    cid, mid, _ = _resolve_context_ids(db, company_id=cid, main_tender_id=mid, tender_id=payload.get("tender_id"))
    payload["company_id"] = cid
    payload["main_tender_id"] = mid
    d = Deal(deal_number=deal_no, **payload)
    db.add(d); db.commit(); db.refresh(d)
    return {"id": d.id, "deal_number": d.deal_number}

# ── BILLS ──────────────────────────────────────────────────────────────────

@app.get("/api/bills")
def list_bills(status: Optional[str] = None, deal_id: Optional[int] = None,
               q_search: Optional[str] = Query(None, alias="q"),
               company_id: Optional[int] = None,
               main_tender_id: Optional[int] = None,
               tender_id: Optional[int] = None,
               db: Session = Depends(get_db)):
    from sqlalchemy.orm import joinedload
    q = select(Bill).options(
        joinedload(Bill.broker), joinedload(Bill.material), joinedload(Bill.plant), joinedload(Bill.tender)
    ).order_by(desc(Bill.created_at))
    q = _apply_scope_filters(q, Bill, company_id, main_tender_id, tender_id)
    if status:   q = q.where(Bill.status == BillStatus(status))
    if deal_id:  q = q.where(Bill.deal_id == deal_id)
    bills = db.execute(q).scalars().unique().all()
    result = [_bill_to_dict(b) for b in bills]
    if q_search:
        qs = q_search.lower()
        result = [b for b in result if qs in str(b.get("broker_name","")).lower()
                  or qs in str(b.get("vehicle_number","")).lower()
                  or qs in str(b.get("material","")).lower()]
    return result

@app.get("/api/bills/{bill_id}")
def get_bill(bill_id: int, company_id: Optional[int] = None,
             main_tender_id: Optional[int] = None, tender_id: Optional[int] = None,
             db: Session = Depends(get_db)):
    b = db.get(Bill, bill_id)
    _assert_record_scope(b, company_id, main_tender_id, tender_id, label="Bill")
    return _bill_to_dict(b)

@app.patch("/api/bills/{bill_id}")
def update_bill(bill_id: int, data: BillUpdate, company_id: Optional[int] = None,
                main_tender_id: Optional[int] = None, tender_id: Optional[int] = None,
                db: Session = Depends(get_db)):
    b = db.get(Bill, bill_id)
    _assert_record_scope(b, company_id, main_tender_id, tender_id, label="Bill")
    # Track old linked values to keep related models in sync
    old_deal = b.deal_id

    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    # Handle status string → enum
    if "status" in update_data:
        update_data["status"] = BillStatus(update_data["status"])
    for k, v in update_data.items():
        setattr(b, k, v)

    if b.deal_id:
        lc, lm, lt = _lineage_from_deal(db, b.deal_id)
        _stamp_lineage(b, lc, lm, lt)
    elif b.tender_id:
        lc, lm, lt = _lineage_from_tender(db, b.tender_id)
        _stamp_lineage(b, lc, lm, lt)
    # Re-validate
    if b.qty_mt and b.rate_per_mt and b.total_amount:
        computed = float(b.qty_mt) * float(b.rate_per_mt)
        b.validation_amount = abs(computed - float(b.total_amount)) / float(b.total_amount) <= 0.05

    affected_deal_ids = {old_deal, b.deal_id}

    # Update related dispatch (if any) to reflect bill changes
    dispatches = _active_dispatches_for_bill(db, b.id)
    if len(dispatches) > 1:
        raise HTTPException(
            409,
            "Data integrity violation: bill has multiple active dispatches. Resolve duplicates before updating.",
        )
    disp = dispatches[0] if dispatches else None
    if disp:
        # if plant changed, ensure dispatch plant updated
        if b.plant_id:
            disp.plant_id = b.plant_id
        # update vehicle, qty and date
        disp.vehicle_number = b.vehicle_number or disp.vehicle_number
        disp.qty_mt = b.qty_mt or disp.qty_mt
        disp.dispatch_date = b.bill_date or disp.dispatch_date
        disp.material_id = b.material_id or disp.material_id
        disp.material_name = b.material_name or disp.material_name

        previous_dispatch_deal_id = disp.deal_id
        if b.deal_id and previous_dispatch_deal_id != b.deal_id:
            active_receipt = _active_receipt_for_dispatch(db, disp.id)
            moved_accepted = float(active_receipt.accepted_mt or 0) if active_receipt else 0.0
            moved_rejected = float(active_receipt.rejected_mt or 0) if active_receipt else 0.0

            if active_receipt and (moved_accepted or moved_rejected):
                _apply_receipt_delta(db, disp.id, -moved_accepted, -moved_rejected)

            disp.deal_id = b.deal_id
            lc, lm, lt = _lineage_from_deal(db, b.deal_id)
            _stamp_lineage(disp, lc, lm, lt)

            if active_receipt and (moved_accepted or moved_rejected):
                _apply_receipt_delta(db, disp.id, moved_accepted, moved_rejected)

        _enrich_dispatch_material(db, disp)
        if float(disp.consumed_qty_qtl or 0) - float(disp.qty_mt or 0) > 1e-9:
            raise HTTPException(400, "Bill quantity cannot be lower than linked dispatch consumed quantity")

        affected_deal_ids.update({previous_dispatch_deal_id, disp.deal_id})

    _recalculate_deal_dispatched_mt(db, list(affected_deal_ids))
    _reconcile_unloading_match_for_bill_ids(db, [b.id])

    db.commit()
    return _bill_to_dict(b)

@app.patch("/api/bills/{bill_id}/approve")
def approve_bill(bill_id: int, company_id: Optional[int] = None,
                 main_tender_id: Optional[int] = None, tender_id: Optional[int] = None,
                 db: Session = Depends(get_db)):
    b = db.get(Bill, bill_id)
    _assert_record_scope(b, company_id, main_tender_id, tender_id, label="Bill")
    b.status = BillStatus.linked if b.deal_id else BillStatus.approved
    b.reviewed_at = datetime.now()
    # Auto-create purchase bill
    existing_pb = db.execute(select(PurchaseBill).where(PurchaseBill.bill_id == bill_id)).scalar_one_or_none()
    if not existing_pb and b.deal_id:
        # Prefer explicit broker on the bill, otherwise take broker from the linked deal
        deal = db.get(Deal, b.deal_id)
        broker_id = b.broker_id or (deal.broker_id if deal else None)
        if not broker_id:
            # Do not attempt to insert a PurchaseBill with NULL broker_id — return friendly error
            raise HTTPException(400, "Cannot create purchase bill: broker unknown for this bill. Link a broker or ensure the deal has a broker.")

        count = db.execute(select(func.count(PurchaseBill.id))).scalar() or 0
        pb = PurchaseBill(
            company_id   = b.company_id,
            main_tender_id = b.main_tender_id,
            tender_id    = b.tender_id,
            bill_id      = b.id,
            deal_id      = b.deal_id,
            broker_id    = broker_id,
            pb_number    = f"PB-{datetime.now().strftime('%y%m')}-{count+1:03d}",
            qty_mt       = b.qty_mt,
            rate_per_mt  = b.rate_per_mt,
            total_amount = b.total_amount,
            bill_date    = b.bill_date or date.today(),
        )
        db.add(pb)
    # Auto-create a dispatch for this bill if it's linked to a deal and no dispatch exists
    # If plant_id is missing but plant_name is present, try to resolve it now
    if not b.plant_id and b.plant_name:
        try:
            pid, pname = _resolve_plant(db, b.plant_name)
            if pid:
                b.plant_id = pid
            if pname:
                b.plant_name = pname
        except Exception:
            pass

    existing_disps = _active_dispatches_for_bill(db, bill_id)
    if len(existing_disps) > 1:
        raise HTTPException(
            409,
            "Data integrity violation: bill has multiple active dispatches. Resolve duplicates before approval.",
        )
    existing_disp = existing_disps[0] if existing_disps else None
    if b.deal_id and not existing_disp and b.qty_mt and b.plant_id:
        d = Dispatch(
            company_id = b.company_id,
            main_tender_id = b.main_tender_id,
            tender_id = b.tender_id,
            bill_id = b.id,
            deal_id = b.deal_id,
            material_id = b.material_id,
            material_name = b.material_name,
            vehicle_number = b.vehicle_number or "",
            dispatch_date = b.bill_date or date.today(),
            qty_mt = b.qty_mt,
            plant_id = b.plant_id,
            status = DispatchStatus.in_transit,
        )
        _enrich_dispatch_material(db, d)
        db.add(d)
        db.flush()
        _auto_match_open_receipt_for_dispatch(db, d)
        _refresh_dispatch_status(db, d.id)
        _recalculate_deal_dispatched_mt(db, [b.deal_id])
    db.commit()
    return _bill_to_dict(b)


@app.post("/api/bills/ensure-dispatches")
def ensure_dispatches(company_id: Optional[int] = None, main_tender_id: Optional[int] = None,
                      tender_id: Optional[int] = None, db: Session = Depends(get_db)):
    created = []
    affected_deal_ids = set()
    q = select(Bill).where(Bill.status.in_([BillStatus.approved, BillStatus.linked]))
    q = _apply_scope_filters(q, Bill, company_id, main_tender_id, tender_id)
    bills = db.execute(q).scalars().all()
    for b in bills:
        # skip if already has a dispatch
        existing_disps = _active_dispatches_for_bill(db, b.id)
        if len(existing_disps) > 1:
            log.warning(
                "Skipping auto-dispatch for bill %s due to 1:1 violation (%s active dispatches)",
                b.id,
                len(existing_disps),
            )
            continue
        if existing_disps:
            continue
        if not b.deal_id or not b.qty_mt:
            continue
        # resolve plant
        plant_id = b.plant_id
        if not plant_id and b.plant_name:
            pid, pname = _resolve_plant(db, b.plant_name)
            if pid:
                plant_id = pid
            if pname:
                b.plant_name = pname
        if not plant_id: continue
        d = Dispatch(
            company_id = b.company_id,
            main_tender_id = b.main_tender_id,
            tender_id = b.tender_id,
            bill_id = b.id,
            deal_id = b.deal_id,
            material_id = b.material_id,
            material_name = b.material_name,
            vehicle_number = b.vehicle_number or "",
            dispatch_date = b.bill_date or date.today(),
            qty_mt = b.qty_mt,
            plant_id = plant_id,
            status = DispatchStatus.in_transit,
        )
        _enrich_dispatch_material(db, d)
        db.add(d)
        db.flush()
        matched_receipt_id = _auto_match_open_receipt_for_dispatch(db, d)
        _refresh_dispatch_status(db, d.id)
        affected_deal_ids.add(b.deal_id)
        created.append({'bill_id': b.id, 'dispatch_id': d.id, 'auto_matched_receipt_id': matched_receipt_id})
    _recalculate_deal_dispatched_mt(db, list(affected_deal_ids))
    db.commit()
    return {'created': created, 'count': len(created)}

@app.patch("/api/bills/{bill_id}/link/{deal_id}")
def link_bill_to_deal(bill_id: int, deal_id: int, company_id: Optional[int] = None,
                      main_tender_id: Optional[int] = None, tender_id: Optional[int] = None,
                      db: Session = Depends(get_db)):
    b = db.get(Bill, bill_id)
    _assert_record_scope(b, company_id, main_tender_id, tender_id, label="Bill")
    dc, dm, dt = _lineage_from_deal(db, deal_id)
    if company_id is not None and dc != company_id:
        raise HTTPException(400, "Deal is outside selected company scope")
    if main_tender_id is not None and dm != main_tender_id:
        raise HTTPException(400, "Deal is outside selected main tender scope")
    if tender_id is not None and dt != tender_id:
        raise HTTPException(400, "Deal is outside selected tender scope")

    old_deal_id = b.deal_id
    affected_deal_ids = {old_deal_id, deal_id}

    b.deal_id = deal_id
    _stamp_lineage(b, dc, dm, dt)
    if b.status == BillStatus.approved:
        b.status = BillStatus.linked

    # If bill is approved/linked and no dispatch exists, create one
    existing_disps = _active_dispatches_for_bill(db, bill_id)
    if len(existing_disps) > 1:
        raise HTTPException(
            409,
            "Data integrity violation: bill has multiple active dispatches. Resolve duplicates before linking.",
        )
    existing_disp = existing_disps[0] if existing_disps else None

    if existing_disp and existing_disp.deal_id != deal_id:
        active_receipt = _active_receipt_for_dispatch(db, existing_disp.id)
        moved_accepted = float(active_receipt.accepted_mt or 0) if active_receipt else 0.0
        moved_rejected = float(active_receipt.rejected_mt or 0) if active_receipt else 0.0

        if active_receipt and (moved_accepted or moved_rejected):
            _apply_receipt_delta(db, existing_disp.id, -moved_accepted, -moved_rejected)

        affected_deal_ids.add(existing_disp.deal_id)
        existing_disp.deal_id = deal_id
        _stamp_lineage(existing_disp, dc, dm, dt)

        if active_receipt and (moved_accepted or moved_rejected):
            _apply_receipt_delta(db, existing_disp.id, moved_accepted, moved_rejected)

    if b.status in (BillStatus.approved, BillStatus.linked) and not existing_disp and b.qty_mt and (b.plant_id or b.plant_name):
        plant_id = b.plant_id
        if not plant_id and b.plant_name:
            pid, pname = _resolve_plant(db, b.plant_name)
            if pid:
                plant_id = pid
            if pname:
                b.plant_name = pname
        if plant_id:
            d = Dispatch(
                company_id = b.company_id,
                main_tender_id = b.main_tender_id,
                tender_id = b.tender_id,
                bill_id = b.id,
                deal_id = b.deal_id or deal_id,
                material_id = b.material_id,
                material_name = b.material_name,
                vehicle_number = b.vehicle_number or "",
                dispatch_date = b.bill_date or date.today(),
                qty_mt = b.qty_mt,
                plant_id = plant_id,
                status = DispatchStatus.in_transit,
            )
            _enrich_dispatch_material(db, d)
            db.add(d)
            db.flush()
            _auto_match_open_receipt_for_dispatch(db, d)
            _refresh_dispatch_status(db, d.id)

            _recalculate_deal_dispatched_mt(db, list(affected_deal_ids))
    db.commit()
    return _bill_to_dict(b)

# Bill upload (web portal)
@app.post("/api/bills/upload")
async def upload_bill(
    file:    UploadFile = File(...),
    source:  str = Form("web"),
    operator: Optional[str] = Form(None),
    company_id: Optional[int] = Query(None),
    main_tender_id: Optional[int] = Query(None),
    tender_id: Optional[int] = Query(None),
    db: Session = Depends(get_db)
):
    from services.ocr_service import extract_bill
    settings = get_settings()
    dest_dir = Path(settings.upload_dir) / "bills"
    dest_dir.mkdir(parents=True, exist_ok=True)
    fname = f"web_{datetime.now().strftime('%Y%m%d%H%M%S')}_{file.filename}"
    fpath = dest_dir / fname
    fpath.write_bytes(await file.read())

    extracted = await extract_bill(str(fpath))
    if not isinstance(extracted, dict):
        extracted = {}

    company_id, main_tender_id, tender_id = _resolve_context_ids(
        db, company_id=company_id, main_tender_id=main_tender_id, tender_id=tender_id
    )

    # Stop here when OCR failed with no usable fields to avoid creating empty rows.
    signal_keys = ["vendor_name", "broker_name", "vehicle_number", "bill_number", "material_type", "quantity_qtl", "total_amount"]
    has_signal = any(extracted.get(k) not in (None, "", "None", "null") for k in signal_keys)
    ocr_error = str(extracted.get("error") or "")
    if not has_signal and ocr_error:
        if "free-models-per-day" in ocr_error.lower():
            raise HTTPException(429, "OpenRouter free daily limit reached. Try again after reset or enter bill manually.")
        raise HTTPException(503, f"OCR extraction failed: {ocr_error}")

    vendor_name_ex = extracted.get("vendor_name") or extracted.get("broker_name")
    bn = extracted.get("bill_number")
    if _is_duplicate_busy_staging(db, vendor_name_ex, bn, company_id=company_id):
        try:
            if fpath.exists():
                fpath.unlink()
        except Exception:
            pass
        raise HTTPException(400, "Duplicate bill")

    # Log extracted key fields to help diagnose duplicate-check issues
    try:
        log.info(f"Upload extracted fields: vendor='{extracted.get('vendor_name') or extracted.get('broker_name')}', vehicle='{extracted.get('vehicle_number')}', bill_number='{extracted.get('bill_number')}'")
    except Exception:
        log.info("Upload extracted fields: <failed to render>")

    # Prevent duplicate bills: vendor_name + vehicle_number + bill_number (+ operator when provided)
    vendor_name_ex = extracted.get("vendor_name") or extracted.get("broker_name")
    veh_ex = extracted.get("vehicle_number")
    bn = extracted.get("bill_number")
    is_dup = _is_duplicate_bill(db, vendor_name_ex, veh_ex, bn, operator)
    log.info(f"Duplicate check result: {is_dup} for bill_number='{bn}' vendor='{vendor_name_ex}' vehicle='{veh_ex}'")
    if is_dup:
        raise HTTPException(400, "Duplicate bill")
    else:
        # Diagnostic: list possible candidate bills in DB for debugging
        try:
            candidates = []
            if bn:
                candidates.extend(db.execute(select(Bill).where(func.lower(Bill.bill_number) == (str(bn) if bn is not None else '').strip().lower())).scalars().all())
            if veh_ex:
                veh_norm = (str(veh_ex) if veh_ex is not None else '').strip().replace(' ', '')
                candidates.extend(db.execute(select(Bill).where(func.replace(func.coalesce(Bill.vehicle_number, ''), ' ', '') == veh_norm)).scalars().all())
            if vendor_name_ex:
                candidates.extend(db.execute(select(Bill).where(func.lower(func.coalesce(Bill.broker_name, '')) == (str(vendor_name_ex) if vendor_name_ex is not None else '').strip().lower())).scalars().all())
            # linked Broker by name
            if vendor_name_ex:
                candidates.extend(db.execute(select(Bill).join(Broker, Bill.broker_id == Broker.id).where(func.lower(Broker.name) == (str(vendor_name_ex) if vendor_name_ex is not None else '').strip().lower())).scalars().all())
            # Deduplicate
            seen = set(); uniq = []
            for c in candidates:
                if c.id in seen: continue
                seen.add(c.id); uniq.append(c)
            for c in uniq:
                log.info(f"Candidate bill id={c.id} broker_name='{c.broker_name}' broker_id='{c.broker_id}' linked_broker='{getattr(c.broker,'name',None)}' vehicle='{c.vehicle_number}' bill_number='{c.bill_number}' reviewed_by='{c.reviewed_by}' created_at='{c.created_at}'")
        except Exception as e:
            log.exception(f"Error while listing duplicate candidates: {e}")

    # Try to resolve plant name -> plant_id when OCR provides a destination plant
    plant_name_extracted = extracted.get("destination_plant") or extracted.get("plant_name")
    plant_id_resolved = None
    if plant_name_extracted:
        try:
            pid, pname = _resolve_plant(db, plant_name_extracted)
            if pid:
                plant_id_resolved = pid
            if pname:
                plant_name_extracted = pname
        except Exception:
            plant_id_resolved = None

    bill_date_parsed = _parse_optional_date(extracted.get("bill_date"))

    # coerce numeric fields
    def _num(v):
        try:
            return float(v)
        except Exception:
            return None

    b = Bill(
        company_id     = company_id,
        main_tender_id = main_tender_id,
        tender_id      = tender_id,
        source         = BillSource.web,
        image_path     = fname,
        vehicle_number = extracted.get("vehicle_number"),
        broker_name    = extracted.get("vendor_name") or extracted.get("broker_name") or None,
        material_name  = extracted.get("material_type"),
        qty_mt         = _num(extracted.get("quantity_qtl")),   # stored as Qtl
        rate_per_mt    = _num(extracted.get("rate_per_qtl")),   # rate per Qtl
        total_amount   = _num(extracted.get("total_amount")),
        bill_date      = bill_date_parsed,
        bill_number    = extracted.get("bill_number"),
        plant_name     = plant_name_extracted,
        plant_id       = plant_id_resolved,
        notes          = f"Transport: {extracted.get('transport_company','')} | Broker: {extracted.get('broker_name','')} | GSTIN: {extracted.get('gstin_vendor','')}".strip(" |"),
        ocr_source     = OcrSource.paddle if extracted.get("source") == "paddle" else OcrSource.gemini,
        ocr_confidence = extracted.get("confidence") or extracted.get("ocr_confidence"),
        ocr_raw_text   = extracted.get("raw_text", ""),
        is_handwritten = extracted.get("is_handwritten", False),
        validation_amount   = extracted.get("validation_amount", False),
        validation_vehicle  = extracted.get("validation_vehicle", False),
        validation_material = extracted.get("validation_material", False),
        status         = BillStatus.flagged if extracted.get("needs_review") else BillStatus.pending,
    )
    db.add(b)
    db.flush()
    _auto_assign_bill_to_deal_and_dispatch(db, b)
    _reconcile_unloading_match_for_bill_ids(db, [b.id])
    db.commit()
    db.refresh(b)
    return _bill_to_dict(b)


@app.post("/api/bills/bulk-upload")
async def bulk_upload_bill(
    files: List[UploadFile] = File(...),
    source: str = Form("web"),
    operator: Optional[str] = Form(None),
    company_id: Optional[int] = Query(None),
    main_tender_id: Optional[int] = Query(None),
    tender_id: Optional[int] = Query(None),
    db: Session = Depends(get_db)
):
    from services.ocr_service import extract_bill
    settings = get_settings()
    dest_dir = Path(settings.upload_dir) / "bills"
    dest_dir.mkdir(parents=True, exist_ok=True)
    created = []
    duplicates = 0
    failed = []
    stopped_due_to_rate_limit = False
    for file in files:
        fname = f"web_{datetime.now().strftime('%Y%m%d%H%M%S')}_{file.filename}"
        fpath = dest_dir / fname
        fpath.write_bytes(await file.read())
        extracted = await extract_bill(str(fpath))
        if not isinstance(extracted, dict):
            extracted = {}

        company_id, main_tender_id, tender_id = _resolve_context_ids(
            db, company_id=company_id, main_tender_id=main_tender_id, tender_id=tender_id
        )
        if tender_id:
            tc, tm, _ = _lineage_from_tender(db, tender_id)
            if tc is not None and company_id is not None and tc != company_id:
                raise HTTPException(400, "company_id does not match tender lineage")
            if tm is not None and main_tender_id is not None and tm != main_tender_id:
                raise HTTPException(400, "main_tender_id does not match tender lineage")
        try:
            log.info(f"Bulk upload extracted for '{file.filename}': vendor='{extracted.get('vendor_name') or extracted.get('broker_name')}', vehicle='{extracted.get('vehicle_number')}', bill_number='{extracted.get('bill_number')}'")
        except Exception:
            log.info(f"Bulk upload extracted for '{file.filename}': <failed to render>")

        signal_keys = ["vendor_name", "broker_name", "vehicle_number", "bill_number", "material_type", "quantity_qtl", "total_amount"]
        has_signal = any(extracted.get(k) not in (None, "", "None", "null") for k in signal_keys)
        ocr_error = str(extracted.get("error") or "")
        if not has_signal and ocr_error:
            failed.append({"file": file.filename, "error": ocr_error})
            log.warning(f"Bulk upload: OCR failed for '{file.filename}' -> {ocr_error}")
            if "free-models-per-day" in ocr_error.lower():
                stopped_due_to_rate_limit = True
                break
            continue

        time.sleep(1)
        # Resolve plant name to plant_id when possible
        plant_name_extracted = extracted.get("destination_plant") or extracted.get("plant_name")
        plant_id_resolved = None
        if plant_name_extracted:
            try:
                pid, pname = _resolve_plant(db, plant_name_extracted)
                if pid:
                    plant_id_resolved = pid
                if pname:
                    plant_name_extracted = pname
            except Exception:
                plant_id_resolved = None

        bill_date_parsed = _parse_optional_date(extracted.get("bill_date"))
        def _num(v):
            try:
                return float(v)
            except Exception:
                return None

        # Duplicate check for this extracted bill
        vendor_name_ex = extracted.get("vendor_name") or extracted.get("broker_name")
        veh_ex = extracted.get("vehicle_number")
        bn = extracted.get("bill_number")
        is_dup = _is_duplicate_bill(db, vendor_name_ex, veh_ex, bn, operator)
        if is_dup:
            log.info(f"Bulk upload: skipping duplicate '{file.filename}' -> bill_number='{bn}' vendor='{vendor_name_ex}' vehicle='{veh_ex}'")
            duplicates += 1
            continue
        else:
            try:
                # Diagnostic candidates for bulk as well
                cand = db.execute(select(Bill).where(func.lower(Bill.bill_number) == (str(bn) if bn is not None else '').strip().lower())).scalars().all() if bn else []
                for c in cand:
                    log.info(f"Bulk candidate match by bill_number: bill_id={c.id} broker_name='{c.broker_name}' vehicle='{c.vehicle_number}' bill_number='{c.bill_number}'")
            except Exception:
                pass

        b = Bill(
            company_id     = company_id,
            main_tender_id = main_tender_id,
            tender_id      = tender_id,
            source         = BillSource.web,
            image_path     = fname,
            vehicle_number = extracted.get("vehicle_number"),
            broker_name    = extracted.get("vendor_name") or extracted.get("broker_name") or None,
            material_name  = extracted.get("material_type"),
            qty_mt         = _num(extracted.get("quantity_qtl")),
            rate_per_mt    = _num(extracted.get("rate_per_qtl")),
            total_amount   = _num(extracted.get("total_amount")),
            bill_date      = bill_date_parsed,
            bill_number    = extracted.get("bill_number"),
            plant_name     = plant_name_extracted,
            plant_id       = plant_id_resolved,
            notes          = f"Transport: {extracted.get('transport_company','')} | Broker: {extracted.get('broker_name','')} | GSTIN: {extracted.get('gstin_vendor','')}".strip(" |"),
            ocr_source     = OcrSource.paddle if extracted.get("source") == "paddle" else OcrSource.gemini,
            ocr_confidence = extracted.get("confidence") or extracted.get("ocr_confidence"),
            ocr_raw_text   = extracted.get("raw_text", ""),
            is_handwritten = extracted.get("is_handwritten", False),
            validation_amount   = extracted.get("validation_amount", False),
            validation_vehicle  = extracted.get("validation_vehicle", False),
            validation_material = extracted.get("validation_material", False),
            status         = BillStatus.flagged if extracted.get("needs_review") else BillStatus.pending,
        )
        db.add(b)
        db.flush()
        _reconcile_unloading_match_for_bill_ids(db, [b.id])
        db.commit()
        db.refresh(b)
        created.append(_bill_to_dict(b))
    result = {"created": created, "duplicates": duplicates}
    if failed:
        result["failed"] = failed
    if stopped_due_to_rate_limit:
        result["stopped_due_to_rate_limit"] = True
    return result


# Busy staging bills (isolated Busy export workflow)
@app.post("/api/busy-staging/upload")
async def upload_busy_staging_bill(
    file: UploadFile = File(...),
    source: str = Form("web"),
    company_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
):
    from services.ocr_service import extract_bill

    if company_id is not None and not db.get(Company, company_id):
        raise HTTPException(404, "Company not found")

    settings = get_settings()
    dest_dir = Path(settings.upload_dir) / "busy_staging_bills"
    dest_dir.mkdir(parents=True, exist_ok=True)

    content = await file.read()
    digest = hashlib.sha256(content).hexdigest()
    dup = db.execute(select(BusyStagingBill).where(BusyStagingBill.file_hash == digest)).scalar_one_or_none()
    if dup:
        raise HTTPException(400, "Duplicate bill")

    fname = f"busy_{datetime.now().strftime('%Y%m%d%H%M%S')}_{file.filename}"
    fpath = dest_dir / fname
    fpath.write_bytes(content)

    extracted = await extract_bill(str(fpath))
    if not isinstance(extracted, dict):
        extracted = {}

    signal_keys = ["vendor_name", "broker_name", "vehicle_number", "bill_number", "material_type", "quantity_qtl", "total_amount"]
    has_signal = any(extracted.get(k) not in (None, "", "None", "null") for k in signal_keys)
    ocr_error = str(extracted.get("error") or "")
    if not has_signal and ocr_error:
        if "free-models-per-day" in ocr_error.lower():
            raise HTTPException(429, "OpenRouter free daily limit reached. Try again after reset or enter bill manually.")
        raise HTTPException(503, f"OCR extraction failed: {ocr_error}")

    src_value = str(source or "web").strip().lower() or "web"
    b = BusyStagingBill(
        company_id=company_id,
        source=src_value,
        image_path=fname,
        file_hash=digest,
    )
    _apply_extracted_to_busy_staging(b, extracted)
    db.add(b)
    db.commit()
    db.refresh(b)

    payload, _reason = _build_busy_ready_row_for_staging(db, b)
    return payload or _busy_staging_to_dict(b)


@app.post("/api/busy-staging/bulk-upload")
async def bulk_upload_busy_staging_bill(
    files: List[UploadFile] = File(...),
    source: str = Form("web"),
    company_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
):
    from services.ocr_service import extract_bill

    if company_id is not None and not db.get(Company, company_id):
        raise HTTPException(404, "Company not found")

    settings = get_settings()
    dest_dir = Path(settings.upload_dir) / "busy_staging_bills"
    dest_dir.mkdir(parents=True, exist_ok=True)

    created: List[dict] = []
    duplicates = 0
    failed: List[dict] = []
    stopped_due_to_rate_limit = False
    src_value = str(source or "web").strip().lower() or "web"

    for file in files:
        content = await file.read()
        digest = hashlib.sha256(content).hexdigest()
        dup = db.execute(select(BusyStagingBill).where(BusyStagingBill.file_hash == digest)).scalar_one_or_none()
        if dup:
            duplicates += 1
            continue

        fname = f"busy_{datetime.now().strftime('%Y%m%d%H%M%S')}_{file.filename}"
        fpath = dest_dir / fname
        fpath.write_bytes(content)

        extracted = await extract_bill(str(fpath))
        if not isinstance(extracted, dict):
            extracted = {}

        signal_keys = ["vendor_name", "broker_name", "vehicle_number", "bill_number", "material_type", "quantity_qtl", "total_amount"]
        has_signal = any(extracted.get(k) not in (None, "", "None", "null") for k in signal_keys)
        ocr_error = str(extracted.get("error") or "")
        if not has_signal and ocr_error:
            failed.append({"file": file.filename, "error": ocr_error})
            if "free-models-per-day" in ocr_error.lower():
                stopped_due_to_rate_limit = True
                break
            continue

        vendor_name_ex = extracted.get("vendor_name") or extracted.get("broker_name")
        bn = extracted.get("bill_number")
        if _is_duplicate_busy_staging(db, vendor_name_ex, bn, company_id=company_id):
            try:
                if fpath.exists():
                    fpath.unlink()
            except Exception:
                pass
            duplicates += 1
            continue

        time.sleep(1)
        b = BusyStagingBill(
            company_id=company_id,
            source=src_value,
            image_path=fname,
            file_hash=digest,
        )
        _apply_extracted_to_busy_staging(b, extracted)
        db.add(b)
        db.commit()
        db.refresh(b)

        payload, _reason = _build_busy_ready_row_for_staging(db, b)
        created.append(payload or _busy_staging_to_dict(b))

    result = {"created": created, "duplicates": duplicates}
    if failed:
        result["failed"] = failed
    if stopped_due_to_rate_limit:
        result["stopped_due_to_rate_limit"] = True
    return result


@app.get("/api/busy-staging/bills")
def list_busy_staging_bills(
    include_exported: bool = Query(True),
    company_id: Optional[int] = None,
    q_search: Optional[str] = Query(None, alias="q"),
    db: Session = Depends(get_db),
):
    rows, _rejected = _collect_busy_staging_rows(
        db,
        company_id=company_id,
        include_exported=include_exported,
    )
    if q_search:
        qs = q_search.lower().strip()
        rows = [
            r for r in rows
            if qs in str(r.get("broker_name") or "").lower()
            or qs in str(r.get("vehicle_number") or "").lower()
            or qs in str(r.get("material_name") or "").lower()
            or qs in str(r.get("bill_number") or "").lower()
        ]
    return rows


@app.post("/api/busy-staging/{staging_id}/reparse")
async def reparse_busy_staging_bill(
    staging_id: int,
    data: BusyStagingReparse,
    company_id: Optional[int] = None,
    db: Session = Depends(get_db),
):
    from services.ocr_service import extract_document_by_type

    b = db.get(BusyStagingBill, staging_id)
    _assert_record_scope(b, company_id, label="Busy staging bill")

    abs_path = _busy_staging_abs_path(b.image_path)
    if not abs_path.exists():
        raise HTTPException(404, f"Source file not found for busy staging bill #{b.id}")

    engine = _normalize_ocr_engine(data.ocr_engine)
    extracted = await extract_document_by_type(str(abs_path), "purchase_bill", prefer_ocr_engine=engine)
    if not isinstance(extracted, dict):
        extracted = {}

    signal_keys = ["vendor_name", "broker_name", "vehicle_number", "bill_number", "material_type", "quantity_qtl", "total_amount"]
    has_signal = any(extracted.get(k) not in (None, "", "None", "null") for k in signal_keys)
    ocr_error = str(extracted.get("error") or "")
    if not has_signal and ocr_error:
        if "free-models-per-day" in ocr_error.lower():
            raise HTTPException(429, "OpenRouter free daily limit reached. Try again after reset or enter bill manually.")
        raise HTTPException(503, f"OCR extraction failed: {ocr_error}")

    _apply_extracted_to_busy_staging(b, extracted)
    db.commit()
    db.refresh(b)

    payload, _reason = _build_busy_ready_row_for_staging(db, b)
    if not payload:
        payload = _busy_staging_to_dict(b)
    payload["reparse"] = {
        "ocr_engine": engine or "auto",
        "source": extracted.get("source"),
        "high_confidence": extracted.get("high_confidence"),
        "error": extracted.get("error"),
    }
    return payload


@app.post("/api/busy-staging/bulk-reparse")
async def bulk_reparse_busy_staging_bills(
    data: BusyStagingBulkReparse,
    company_id: Optional[int] = None,
    db: Session = Depends(get_db),
):
    from services.ocr_service import extract_document_by_type

    ids = []
    for i in (data.ids or []):
        try:
            val = int(i)
        except Exception:
            continue
        if val not in ids:
            ids.append(val)
    if not ids:
        raise HTTPException(400, "No ids provided")

    q = select(BusyStagingBill).where(BusyStagingBill.id.in_(ids))
    q = _apply_scope_filters(q, BusyStagingBill, company_id)
    rows = db.execute(q).scalars().all()
    row_by_id = {r.id: r for r in rows}

    engine = _normalize_ocr_engine(data.ocr_engine)
    reparsed: List[dict] = []
    failed: List[dict] = []
    missing_ids: List[int] = []

    for row_id in ids:
        b = row_by_id.get(row_id)
        if not b:
            missing_ids.append(row_id)
            continue
        try:
            abs_path = _busy_staging_abs_path(b.image_path)
            if not abs_path.exists():
                raise HTTPException(404, f"Source file not found for busy staging bill #{b.id}")

            extracted = await extract_document_by_type(str(abs_path), "purchase_bill", prefer_ocr_engine=engine)
            if not isinstance(extracted, dict):
                extracted = {}

            signal_keys = ["vendor_name", "broker_name", "vehicle_number", "bill_number", "material_type", "quantity_qtl", "total_amount"]
            has_signal = any(extracted.get(k) not in (None, "", "None", "null") for k in signal_keys)
            ocr_error = str(extracted.get("error") or "")
            if not has_signal and ocr_error:
                raise HTTPException(503, f"OCR extraction failed: {ocr_error}")

            _apply_extracted_to_busy_staging(b, extracted)
            reparsed.append({"id": b.id, "source": extracted.get("source")})
        except HTTPException as he:
            failed.append({"id": row_id, "error": str(he.detail)})
        except Exception as e:
            failed.append({"id": row_id, "error": str(e)})

    db.commit()
    return {
        "ok": len(failed) == 0,
        "requested": len(ids),
        "reparsed": len(reparsed),
        "failed": failed,
        "missing_ids": missing_ids,
        "items": reparsed,
    }


@app.post("/api/busy-staging/{staging_id}/delete")
def delete_busy_staging_bill(
    staging_id: int,
    company_id: Optional[int] = None,
    db: Session = Depends(get_db),
):
    b = db.get(BusyStagingBill, staging_id)
    _assert_record_scope(b, company_id, label="Busy staging bill")

    abs_path = _busy_staging_abs_path(b.image_path)
    db.delete(b)
    db.commit()

    try:
        if abs_path.exists():
            abs_path.unlink()
    except Exception:
        pass
    return {"ok": True}


@app.post("/api/busy-staging/bulk-delete")
def bulk_delete_busy_staging_bills(
    data: IdListPayload,
    company_id: Optional[int] = None,
    db: Session = Depends(get_db),
):
    ids = []
    for i in (data.ids or []):
        try:
            val = int(i)
        except Exception:
            continue
        if val not in ids:
            ids.append(val)
    if not ids:
        raise HTTPException(400, "No ids provided")

    q = select(BusyStagingBill).where(BusyStagingBill.id.in_(ids))
    q = _apply_scope_filters(q, BusyStagingBill, company_id)
    rows = db.execute(q).scalars().all()
    found_ids = {r.id for r in rows}
    missing_ids = [i for i in ids if i not in found_ids]

    file_paths: List[Path] = []
    deleted = 0
    for b in rows:
        try:
            file_paths.append(_busy_staging_abs_path(b.image_path))
        except Exception:
            pass
        db.delete(b)
        deleted += 1

    db.commit()

    for p in file_paths:
        try:
            if p.exists():
                p.unlink()
        except Exception:
            pass

    return {
        "ok": len(missing_ids) == 0,
        "deleted": deleted,
        "missing_ids": missing_ids,
    }


@app.post("/api/ingest/upload")
async def upload_pending_ingest(
    file: UploadFile = File(...),
    source: str = Form("web"),
    source_address: Optional[str] = Form(None),
    source_message_id: Optional[str] = Form(None),
    document_type: Optional[str] = Form(None),
    operator: Optional[str] = Form(None),
    company_id: Optional[int] = Query(None),
    main_tender_id: Optional[int] = Query(None),
    tender_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
):
    from services.ocr_service import classify_document_type, extract_document_by_type

    company_id, main_tender_id, tender_id = _resolve_context_ids(
        db, company_id=company_id, main_tender_id=main_tender_id, tender_id=tender_id
    )

    settings = get_settings()
    dest_dir = Path(settings.upload_dir) / "pending_ingests"
    dest_dir.mkdir(parents=True, exist_ok=True)

    content = await file.read()
    digest = hashlib.sha256(content).hexdigest()
    fname = f"ing_{datetime.now().strftime('%Y%m%d%H%M%S')}_{file.filename}"
    fpath = dest_dir / fname
    fpath.write_bytes(content)
    rel_path = str(Path("pending_ingests") / fname)

    cls_conf = 1.0
    cls_candidates = []
    if document_type:
        try:
            doc_type = DocumentType(document_type).value
        except Exception:
            raise HTTPException(400, f"Invalid document_type: {document_type}")
    else:
        doc_type, cls_conf, cls_candidates = classify_document_type(str(fpath), file.filename)

    extracted = await extract_document_by_type(str(fpath), doc_type)
    if not isinstance(extracted, dict):
        extracted = {"error": "invalid extractor response", "high_confidence": False}

    src_value = str(source or "web").strip().lower()
    try:
        ingest_source = IngestSource(src_value)
    except Exception:
        ingest_source = IngestSource.web

    unclear_fields = extracted.get("unclear_fields") if isinstance(extracted.get("unclear_fields"), list) else []

    p = PendingIngest(
        company_id=company_id,
        main_tender_id=main_tender_id,
        tender_id=tender_id,
        source=ingest_source,
        source_address=source_address,
        source_message_id=source_message_id,
        file_name=file.filename,
        file_path=rel_path,
        file_hash=digest,
        document_type=DocumentType(doc_type),
        classifier_confidence=cls_conf,
        classifier_candidates=cls_candidates,
        extracted_payload=extracted,
        unclear_fields=unclear_fields,
        status=IngestStatus.pending,
        review_notes=(f"Uploaded by {operator}" if operator else None),
    )
    db.add(p)
    db.commit()
    db.refresh(p)
    return _pending_ingest_to_dict(p)


@app.get("/api/ingest/email/config")
def ingest_email_config(db: Session = Depends(get_db)):
    from services.email_sync_service import get_email_sync_checkpoint, list_configured_email_accounts

    settings = get_settings()
    accounts = list_configured_email_accounts(settings)

    def _mask_user(user: str) -> str:
        user = str(user or "")
        if not user:
            return ""
        if len(user) <= 3:
            return "*" * len(user)
        return f"{user[:2]}***{user[-1]}"

    account_rows = []
    for acc in accounts:
        acc_user = str(acc.get("email_user") or "").strip()
        acc_mailbox = str(acc.get("mailbox") or settings.email_sync_mailbox or "INBOX").strip() or "INBOX"
        cp = get_email_sync_checkpoint(db, mailbox=acc_mailbox, email_user=acc_user)
        account_rows.append({
            "email_user": acc_user,
            "email_user_masked": _mask_user(acc_user),
            "mailbox": acc_mailbox,
            "email_host": str(acc.get("host") or settings.email_host or "imap.gmail.com"),
            "checkpoint": cp,
        })

    primary = account_rows[0] if account_rows else None
    checkpoint = primary.get("checkpoint") if primary else None
    runs_per_day = _email_sync_runs_per_day()
    return {
        "configured": bool(account_rows),
        "account_count": len(account_rows),
        "accounts": account_rows,
        "email_host": primary.get("email_host") if primary else settings.email_host,
        "email_user": primary.get("email_user_masked") if primary else "",
        "mailbox": primary.get("mailbox") if primary else (str(settings.email_sync_mailbox or "INBOX").strip() or "INBOX"),
        "auto_enabled": bool(settings.email_sync_auto_enabled),
        "runs_per_day": runs_per_day,
        "interval_hours": round(24.0 / runs_per_day, 2),
        "sync_limit": int(settings.email_sync_limit or 80),
        "unread_only": bool(settings.email_sync_unread_only),
        "mark_seen": bool(settings.email_sync_mark_seen),
        "checkpoint": checkpoint,
    }


@app.post("/api/ingest/email/sync")
async def sync_ingest_from_email(
    limit: Optional[int] = Query(None, ge=0, le=2000),
    since_days: Optional[int] = Query(None, ge=0, le=180),
    unread_only: Optional[bool] = Query(None),
    mark_seen: Optional[bool] = Query(None),
    mailbox: Optional[str] = Query(None),
    email_account: Optional[str] = Query(None),
    use_checkpoint: bool = Query(False),
    db: Session = Depends(get_db),
):
    from services.email_sync_service import (
        get_email_sync_checkpoint,
        list_configured_email_accounts,
        sync_all_email_accounts,
        sync_email_pending_ingests,
    )

    settings = get_settings()
    accounts = list_configured_email_accounts(settings)
    if not accounts:
        raise HTTPException(400, "Email sync is not configured. Set EMAIL_SYNC_ACCOUNTS or EMAIL_USER/EMAIL_PASS in backend .env")

    requested_account = str(email_account or "").strip()
    requested_account_norm = requested_account.lower()
    selected_all_accounts = not requested_account_norm or requested_account_norm in {"all", "*"}

    limit_value = int(limit if limit is not None else (settings.email_sync_limit or 80))
    since_days_value = int(since_days if since_days is not None else 14)
    unread_only_value = bool(unread_only if unread_only is not None else settings.email_sync_unread_only)
    mark_seen_value = bool(mark_seen if mark_seen is not None else settings.email_sync_mark_seen)

    try:
        async with _email_sync_lock:
            if selected_all_accounts and len(accounts) > 1:
                result = await sync_all_email_accounts(
                    db,
                    limit=limit_value,
                    since_days=since_days_value,
                    unread_only=unread_only_value,
                    mark_seen=mark_seen_value,
                    start_uid=None,
                    update_checkpoint=use_checkpoint,
                    ignore_duplicates=False,
                    sync_reason="manual",
                )
            else:
                if selected_all_accounts:
                    target = accounts[0]
                else:
                    target = next((a for a in accounts if str(a.get("email_user") or "").strip().lower() == requested_account_norm), None)
                    if target is None:
                        raise HTTPException(400, f"Unknown email account: {requested_account}")

                mailbox_value = str(mailbox or target.get("mailbox") or settings.email_sync_mailbox or "INBOX").strip() or "INBOX"
                start_uid = 0
                if use_checkpoint:
                    cp = get_email_sync_checkpoint(db, mailbox=mailbox_value, email_user=str(target.get("email_user") or ""))
                    start_uid = int(cp.get("last_uid") or 0)

                result = await sync_email_pending_ingests(
                    db,
                    limit=limit_value,
                    since_days=since_days_value,
                    unread_only=unread_only_value,
                    mark_seen=mark_seen_value,
                    mailbox=mailbox_value,
                    start_uid=start_uid,
                    update_checkpoint=use_checkpoint,
                    ignore_duplicates=False,
                    email_user_override=str(target.get("email_user") or ""),
                    email_pass_override=str(target.get("email_pass") or ""),
                    host_override=str(target.get("host") or settings.email_host or "imap.gmail.com"),
                    sync_reason="manual",
                )
        result["mode"] = "incremental" if use_checkpoint else "manual"
        result["requested_account"] = requested_account or "all"
        result["effective_limit"] = limit_value
        result["effective_since_days"] = since_days_value
        result["effective_unread_only"] = unread_only_value
        result["effective_mark_seen"] = mark_seen_value
        return result
    except HTTPException:
        raise
    except RuntimeError as e:
        raise HTTPException(400, str(e))
    except Exception as e:
        raise HTTPException(500, f"Email sync failed: {e}")


@app.get("/api/ingest/email/logs")
def list_ingest_email_logs(
    limit: int = Query(200, ge=1, le=1000),
    since_days: int = Query(14, ge=0, le=365),
    email_account: Optional[str] = Query(None),
    mailbox: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    filters = []
    params = {"limit": int(limit)}

    if since_days > 0:
        filters.append("synced_at >= DATE_SUB(NOW(), INTERVAL :since_days DAY)")
        params["since_days"] = int(since_days)
    if email_account:
        filters.append("LOWER(email_user) = :email_user")
        params["email_user"] = str(email_account).strip().lower()
    if mailbox:
        filters.append("mailbox = :mailbox")
        params["mailbox"] = str(mailbox).strip()
    if status:
        filters.append("status = :status")
        params["status"] = str(status).strip()

    sql = """
        SELECT id, synced_at, sync_reason, email_user, mailbox, imap_uid, message_id,
               from_address, subject, received_at, status,
               attachments_total, attachments_created, attachments_duplicates,
               attachments_skipped, note
        FROM email_sync_logs
    """
    if filters:
        sql += " WHERE " + " AND ".join(filters)
    sql += " ORDER BY synced_at DESC LIMIT :limit"

    rows = db.execute(text(sql), params).mappings().all()
    return [
        {
            "id": int(r.get("id") or 0),
            "synced_at": str(r.get("synced_at")) if r.get("synced_at") else None,
            "sync_reason": r.get("sync_reason"),
            "email_user": r.get("email_user"),
            "mailbox": r.get("mailbox"),
            "imap_uid": int(r.get("imap_uid")) if r.get("imap_uid") is not None else None,
            "message_id": r.get("message_id"),
            "from_address": r.get("from_address"),
            "subject": r.get("subject"),
            "received_at": str(r.get("received_at")) if r.get("received_at") else None,
            "status": r.get("status"),
            "attachments_total": int(r.get("attachments_total") or 0),
            "attachments_created": int(r.get("attachments_created") or 0),
            "attachments_duplicates": int(r.get("attachments_duplicates") or 0),
            "attachments_skipped": int(r.get("attachments_skipped") or 0),
            "note": r.get("note"),
        }
        for r in rows
    ]


@app.get("/api/ingest/pending")
def list_pending_ingests(
    status: Optional[str] = None,
    document_type: Optional[str] = None,
    include_not_classified: bool = Query(False),
    source: Optional[str] = None,
    q_search: Optional[str] = Query(None, alias="q"),
    company_id: Optional[int] = None,
    main_tender_id: Optional[int] = None,
    tender_id: Optional[int] = None,
    db: Session = Depends(get_db),
):
    q = select(PendingIngest).order_by(desc(PendingIngest.created_at))
    q = _apply_scope_filters(q, PendingIngest, company_id, main_tender_id, tender_id)

    if status:
        try:
            q = q.where(PendingIngest.status == IngestStatus(status))
        except Exception:
            raise HTTPException(400, f"Invalid status: {status}")
    if document_type:
        try:
            q = q.where(PendingIngest.document_type == DocumentType(document_type))
        except Exception:
            raise HTTPException(400, f"Invalid document_type: {document_type}")
    elif not include_not_classified:
        q = q.where(PendingIngest.document_type != DocumentType.not_classified)
    if source:
        try:
            q = q.where(PendingIngest.source == IngestSource(source))
        except Exception:
            raise HTTPException(400, f"Invalid source: {source}")

    rows = db.execute(q).scalars().all()
    out = [_pending_ingest_to_dict(r) for r in rows]
    if q_search:
        s = q_search.lower()
        out = [
            r for r in out
            if s in str(r.get("file_name") or "").lower()
            or s in str(r.get("document_type") or "").lower()
            or s in str(r.get("source_address") or "").lower()
            or s in json.dumps(r.get("extracted_payload") or {}, ensure_ascii=False).lower()
        ]
    return out


@app.get("/api/ingest/{ingest_id}")
def get_pending_ingest(
    ingest_id: int,
    company_id: Optional[int] = None,
    main_tender_id: Optional[int] = None,
    tender_id: Optional[int] = None,
    db: Session = Depends(get_db),
):
    p = db.get(PendingIngest, ingest_id)
    _assert_record_scope(p, company_id, main_tender_id, tender_id, label="Pending ingest")
    return _pending_ingest_to_dict(p)


@app.post("/api/ingest/{ingest_id}/update")
def update_pending_ingest(
    ingest_id: int,
    data: PendingIngestUpdate,
    operator: Optional[str] = Query(None),
    company_id: Optional[int] = None,
    main_tender_id: Optional[int] = None,
    tender_id: Optional[int] = None,
    db: Session = Depends(get_db),
):
    p = db.get(PendingIngest, ingest_id)
    _assert_record_scope(p, company_id, main_tender_id, tender_id, label="Pending ingest")

    patch = data.model_dump(exclude_unset=True)
    if "document_type" in patch and patch["document_type"] is not None:
        try:
            p.document_type = DocumentType(str(patch["document_type"]).strip())
        except Exception:
            raise HTTPException(400, f"Invalid document_type: {patch['document_type']}")
    if "extracted_payload" in patch and patch["extracted_payload"] is not None:
        p.extracted_payload = patch["extracted_payload"]
    if "unclear_fields" in patch and patch["unclear_fields"] is not None:
        p.unclear_fields = patch["unclear_fields"]
    if "review_notes" in patch:
        p.review_notes = patch.get("review_notes")
    if "classifier_confidence" in patch and patch["classifier_confidence"] is not None:
        p.classifier_confidence = patch["classifier_confidence"]
    if "status" in patch and patch["status"] is not None:
        try:
            p.status = IngestStatus(str(patch["status"]).strip())
        except Exception:
            raise HTTPException(400, f"Invalid status: {patch['status']}")

    if operator:
        p.reviewed_by = operator
        p.reviewed_at = datetime.now()

    db.commit()
    db.refresh(p)
    return _pending_ingest_to_dict(p)


@app.post("/api/ingest/{ingest_id}/assign")
def assign_pending_ingest(
    ingest_id: int,
    data: PendingIngestAssign,
    company_id: Optional[int] = None,
    main_tender_id: Optional[int] = None,
    tender_id: Optional[int] = None,
    db: Session = Depends(get_db),
):
    p = db.get(PendingIngest, ingest_id)
    _assert_record_scope(p, company_id, main_tender_id, tender_id, label="Pending ingest")

    assigned_company, assigned_main, assigned_tender = _resolve_context_ids(
        db,
        company_id=data.company_id,
        main_tender_id=data.main_tender_id,
        tender_id=data.tender_id,
    )
    p.assigned_company_id = assigned_company
    p.assigned_main_tender_id = assigned_main
    p.assigned_tender_id = assigned_tender

    db.commit()
    db.refresh(p)
    return _pending_ingest_to_dict(p)


@app.post("/api/ingest/{ingest_id}/approve")
def approve_pending_ingest(
    ingest_id: int,
    data: PendingIngestApprove,
    company_id: Optional[int] = None,
    main_tender_id: Optional[int] = None,
    tender_id: Optional[int] = None,
    db: Session = Depends(get_db),
):
    p = db.get(PendingIngest, ingest_id)
    _assert_record_scope(p, company_id, main_tender_id, tender_id, label="Pending ingest")

    p.status = IngestStatus.approved
    p.reviewed_at = datetime.now()
    if data.operator:
        p.reviewed_by = data.operator
    if data.review_notes:
        p.review_notes = data.review_notes

    # Strict approval gate: approving marks as reviewed only.
    # Action application must be explicitly triggered via run-action endpoint.
    p.action_status = "approved"
    p.action_error = None

    if data.auto_action:
        p.review_notes = (p.review_notes or "") + "\n[info] auto_action ignored; action requires explicit run-action call."

    db.commit()
    db.refresh(p)
    return _pending_ingest_to_dict(p)


@app.post("/api/ingest/{ingest_id}/run-action")
def run_pending_ingest_action(
    ingest_id: int,
    operator: Optional[str] = Query(None),
    review_notes: Optional[str] = Query(None),
    company_id: Optional[int] = None,
    main_tender_id: Optional[int] = None,
    tender_id: Optional[int] = None,
    db: Session = Depends(get_db),
):
    p = db.get(PendingIngest, ingest_id)
    _assert_record_scope(p, company_id, main_tender_id, tender_id, label="Pending ingest")

    if p.status != IngestStatus.approved:
        raise HTTPException(400, "Pending ingest must be approved before running action")

    if operator:
        p.reviewed_by = operator
    if review_notes:
        p.review_notes = review_notes

    try:
        action_payload = _apply_pending_action(db, p)
        p.action_payload = action_payload
        if isinstance(action_payload, dict) and action_payload.get("manual_required"):
            p.action_status = "manual"
        else:
            p.action_status = "processed"
            p.status = IngestStatus.processed
        p.action_error = None
    except Exception as e:
        p.action_status = "failed"
        p.action_error = str(e)

    db.commit()
    db.refresh(p)
    return _pending_ingest_to_dict(p)


@app.post("/api/ingest/{ingest_id}/delete")
def delete_pending_ingest(
    ingest_id: int,
    company_id: Optional[int] = None,
    main_tender_id: Optional[int] = None,
    tender_id: Optional[int] = None,
    db: Session = Depends(get_db),
):
    p = db.get(PendingIngest, ingest_id)
    _assert_record_scope(p, company_id, main_tender_id, tender_id, label="Pending ingest")

    abs_path = _pending_abs_path(p.file_path)
    _rewind_email_checkpoint_for_deleted_ingest(db, p)
    db.query(PurchaseOrder).filter(PurchaseOrder.source_pending_id == p.id).update(
        {"source_pending_id": None}, synchronize_session=False
    )
    db.delete(p)
    db.commit()

    try:
        if abs_path.exists():
            abs_path.unlink()
    except Exception:
        pass
    return {"ok": True}


@app.post("/api/ingest/{ingest_id}/reparse")
async def reparse_pending_ingest(
    ingest_id: int,
    data: PendingIngestReparse,
    operator: Optional[str] = Query(None),
    review_notes: Optional[str] = Query(None),
    company_id: Optional[int] = None,
    main_tender_id: Optional[int] = None,
    tender_id: Optional[int] = None,
    db: Session = Depends(get_db),
):
    p = db.get(PendingIngest, ingest_id)
    _assert_record_scope(p, company_id, main_tender_id, tender_id, label="Pending ingest")

    meta = await _reparse_pending_ingest_record(
        db,
        p,
        reclassify=bool(data.reclassify),
        document_type_override=data.document_type,
        ocr_engine=data.ocr_engine,
    )
    if operator:
        p.reviewed_by = operator
        p.reviewed_at = datetime.now()
    if review_notes:
        p.review_notes = review_notes
    p.action_error = None

    db.commit()
    db.refresh(p)

    out = _pending_ingest_to_dict(p)
    out["reparse"] = meta
    return out


@app.post("/api/ingest/bulk-reparse")
async def bulk_reparse_pending_ingests(
    data: PendingIngestBulkReparse,
    operator: Optional[str] = Query(None),
    review_notes: Optional[str] = Query(None),
    company_id: Optional[int] = None,
    main_tender_id: Optional[int] = None,
    tender_id: Optional[int] = None,
    db: Session = Depends(get_db),
):
    ids = []
    for i in (data.ids or []):
        try:
            val = int(i)
        except Exception:
            continue
        if val not in ids:
            ids.append(val)
    if not ids:
        raise HTTPException(400, "No ids provided")

    q = select(PendingIngest).where(PendingIngest.id.in_(ids))
    q = _apply_scope_filters(q, PendingIngest, company_id, main_tender_id, tender_id)
    rows = db.execute(q).scalars().all()
    row_by_id = {r.id: r for r in rows}

    reparsed: List[dict] = []
    failed: List[dict] = []
    missing_ids: List[int] = []

    for ingest_id in ids:
        p = row_by_id.get(ingest_id)
        if not p:
            missing_ids.append(ingest_id)
            continue
        try:
            meta = await _reparse_pending_ingest_record(
                db,
                p,
                reclassify=bool(data.reclassify),
                document_type_override=None,
                ocr_engine=data.ocr_engine,
            )
            if operator:
                p.reviewed_by = operator
                p.reviewed_at = datetime.now()
            if review_notes:
                p.review_notes = review_notes
            p.action_error = None
            reparsed.append(meta)
        except HTTPException as he:
            failed.append({"id": ingest_id, "error": str(he.detail)})
        except Exception as e:
            failed.append({"id": ingest_id, "error": str(e)})

    db.commit()

    return {
        "ok": len(failed) == 0,
        "requested": len(ids),
        "reparsed": len(reparsed),
        "failed": failed,
        "missing_ids": missing_ids,
        "items": reparsed,
    }


@app.post("/api/ingest/bulk-delete")
def bulk_delete_pending_ingests(
    data: IdListPayload,
    company_id: Optional[int] = None,
    main_tender_id: Optional[int] = None,
    tender_id: Optional[int] = None,
    db: Session = Depends(get_db),
):
    ids = []
    for i in (data.ids or []):
        try:
            val = int(i)
        except Exception:
            continue
        if val not in ids:
            ids.append(val)
    if not ids:
        raise HTTPException(400, "No ids provided")

    q = select(PendingIngest).where(PendingIngest.id.in_(ids))
    q = _apply_scope_filters(q, PendingIngest, company_id, main_tender_id, tender_id)
    rows = db.execute(q).scalars().all()
    found_ids = {r.id for r in rows}
    missing_ids = [i for i in ids if i not in found_ids]

    file_paths: List[Path] = []
    deleted = 0
    unlinked_purchase_orders = 0
    for p in rows:
        _rewind_email_checkpoint_for_deleted_ingest(db, p)
        try:
            file_paths.append(_pending_abs_path(p.file_path))
        except Exception:
            pass
        unlinked_purchase_orders += int(
            db.query(PurchaseOrder)
            .filter(PurchaseOrder.source_pending_id == p.id)
            .update({"source_pending_id": None}, synchronize_session=False)
            or 0
        )
        db.delete(p)
        deleted += 1

    db.commit()

    files_deleted = 0
    file_delete_errors = 0
    for path_obj in file_paths:
        try:
            if path_obj.exists() and path_obj.is_file():
                path_obj.unlink()
                files_deleted += 1
        except Exception:
            file_delete_errors += 1

    return {
        "ok": True,
        "deleted": deleted,
        "unlinked_purchase_orders": unlinked_purchase_orders,
        "files_deleted": files_deleted,
        "file_delete_errors": file_delete_errors,
        "missing_ids": missing_ids,
    }


# ── PLANT UNLOADING DETAILS (sheet-style receipt ingestion) ───────────────

@app.get("/api/plant-unloading")
def list_plant_unloading(
    status: Optional[str] = None,
    master_id: Optional[int] = None,
    company_id: Optional[int] = None,
    main_tender_id: Optional[int] = None,
    tender_id: Optional[int] = None,
    q_search: Optional[str] = Query(None, alias="q"),
    db: Session = Depends(get_db),
):
    q = select(PlantUnloadingEntry).order_by(desc(PlantUnloadingEntry.created_at))
    q = _apply_scope_filters(q, PlantUnloadingEntry, company_id, main_tender_id, tender_id)
    if master_id:
        q = q.where(PlantUnloadingEntry.master_id == master_id)
    if status:
        q = q.where(PlantUnloadingEntry.status == BillStatus(status))
    entries = db.execute(q).scalars().all()
    result = [_plant_unloading_to_dict(e) for e in entries]
    if q_search:
        qs = q_search.lower()
        result = [
            r for r in result
            if qs in str(r.get("rm_number") or "").lower()
            or qs in str(r.get("item_name") or "").lower()
            or qs in str(r.get("party_name") or "").lower()
            or qs in str(r.get("truck_number") or "").lower()
            or qs in str(r.get("ws_no") or "").lower()
        ]
    return result


@app.get("/api/plant-unloading/masters")
def list_plant_unloading_masters(
    q_search: Optional[str] = Query(None, alias="q"),
    company_id: Optional[int] = None,
    main_tender_id: Optional[int] = None,
    tender_id: Optional[int] = None,
    db: Session = Depends(get_db),
):
    q = select(PlantUnloadingMaster).order_by(desc(PlantUnloadingMaster.updated_at))
    q = _apply_scope_filters(q, PlantUnloadingMaster, company_id, main_tender_id, tender_id)
    masters = db.execute(q).scalars().all()
    result = []
    for m in masters:
        entries = list(m.entries or [])
        if q_search:
            s = q_search.lower()
            if s not in f"{m.rm_number} {m.item_name} {m.party_name} {m.plant_name or ''}".lower():
                continue
        total_trucks = len(entries)
        total_net = sum(float(e.net_qty_mt or 0) for e in entries)
        approved = sum(1 for e in entries if e.status == BillStatus.approved)
        pending = sum(1 for e in entries if e.status in (BillStatus.pending, BillStatus.flagged))
        last_date = max((e.entry_date for e in entries if e.entry_date), default=None)
        result.append({
            "id": m.id,
            "company_id": m.company_id,
            "main_tender_id": m.main_tender_id,
            "tender_id": m.tender_id,
            "rm_number": m.rm_number,
            "rm_number_norm": m.rm_number_norm,
            "rm_number_base": m.rm_number_base,
            "item_name": m.item_name,
            "party_name": m.party_name,
            "plant_id": m.plant_id,
            "plant_name": m.plant.name if m.plant else m.plant_name,
            "assignment_status": m.assignment_status,
            "assignment_reason": m.assignment_reason,
            "assignment_confidence": float(m.assignment_confidence) if m.assignment_confidence is not None else None,
            "mapping_source": m.mapping_source,
            "requires_manual_assignment": bool(m.requires_manual_assignment),
            "is_manual_override": bool(m.is_manual_override),
            "manual_assigned_by": m.manual_assigned_by,
            "manual_assigned_at": str(m.manual_assigned_at) if m.manual_assigned_at else None,
            "po_number": m.po_number,
            "total_trucks": total_trucks,
            "total_net_qty_mt": round(total_net, 3),
            "approved_rows": approved,
            "pending_rows": pending,
            "last_entry_date": str(last_date) if last_date else None,
            "updated_at": str(m.updated_at) if m.updated_at else None,
        })
    return result


@app.get("/api/plant-unloading/masters/{master_id}/entries")
def list_master_entries(
    master_id: int,
    company_id: Optional[int] = None,
    main_tender_id: Optional[int] = None,
    tender_id: Optional[int] = None,
    db: Session = Depends(get_db),
):
    master = db.get(PlantUnloadingMaster, master_id)
    _assert_record_scope(master, company_id, main_tender_id, tender_id, label="Plant unloading register")
    entries = db.execute(
        select(PlantUnloadingEntry)
        .where(PlantUnloadingEntry.master_id == master_id)
        .order_by(desc(PlantUnloadingEntry.entry_date), desc(PlantUnloadingEntry.created_at))
    ).scalars().all()
    return [_plant_unloading_to_dict(e) for e in entries]


@app.patch("/api/plant-unloading/masters/{master_id}")
def update_plant_unloading_master(master_id: int, data: PlantUnloadingMasterUpdate,
                                  company_id: Optional[int] = None,
                                  main_tender_id: Optional[int] = None,
                                  tender_id: Optional[int] = None,
                                  operator: Optional[str] = Query(None),
                                  db: Session = Depends(get_db)):
    m = db.get(PlantUnloadingMaster, master_id)
    _assert_record_scope(m, company_id, main_tender_id, tender_id, label="Plant unloading register")

    old_tender_ids = {m.tender_id}
    payload = data.model_dump(exclude_unset=True)

    if "rm_number" in payload:
        rm = str(payload.get("rm_number") or "").strip()
        if not rm:
            raise HTTPException(400, "RM number cannot be empty")
        m.rm_number = rm
        m.rm_number_norm, m.rm_number_base, _ = _normalize_rm_components(rm)

    if "item_name" in payload:
        item = str(payload.get("item_name") or "").strip()
        if not item:
            raise HTTPException(400, "Item name cannot be empty")
        m.item_name = item
        for e in (m.entries or []):
            e.item_name = item

    if "party_name" in payload:
        party = str(payload.get("party_name") or "").strip()
        if not party:
            raise HTTPException(400, "Party name cannot be empty")
        m.party_name = _normalize_unloading_party_name(db, party, m.company_id)

    if "plant_name" in payload:
        plant_input = str(payload.get("plant_name") or "").strip()
        if plant_input:
            pid, pname = _resolve_plant(db, plant_input)
            m.plant_id = pid
            m.plant_name = pname
        else:
            m.plant_id = None
            m.plant_name = None

    if "po_number" in payload:
        po = str(payload.get("po_number") or "").strip()
        m.po_number = po or None

    if "notes" in payload:
        notes = str(payload.get("notes") or "").strip()
        m.notes = notes or None

    scope_override = any(k in payload for k in ("company_id", "main_tender_id", "tender_id"))
    recalc_assignment = scope_override or any(
        k in payload for k in ("rm_number", "item_name", "party_name", "plant_name")
    ) or bool(m.requires_manual_assignment) or not bool(m.tender_id)

    if scope_override:
        resolved_company, resolved_main, resolved_tender = _resolve_context_ids(
            db,
            company_id=payload.get("company_id"),
            main_tender_id=payload.get("main_tender_id"),
            tender_id=payload.get("tender_id"),
        )
        assignment = _resolve_unloading_assignment(
            db,
            rm_number=m.rm_number,
            item_name=m.item_name,
            plant_name=m.plant_name,
            company_id_hint=resolved_company,
            main_tender_id_hint=resolved_main,
            tender_id_hint=resolved_tender,
        )
        assignment["company_id"] = resolved_company
        assignment["main_tender_id"] = resolved_main
        assignment["tender_id"] = resolved_tender
        assignment["assignment_status"] = "assigned" if resolved_tender is not None else (assignment.get("assignment_status") or "pending")
        assignment["assignment_reason"] = "Assigned manually by operator."
        assignment["assignment_confidence"] = 1.0 if resolved_tender is not None else (assignment.get("assignment_confidence") or 0.75)
        assignment["mapping_source"] = "manual"
        assignment["requires_manual_assignment"] = False if resolved_tender is not None else bool(assignment.get("requires_manual_assignment"))
        _apply_unloading_assignment_to_master(m, assignment, manual_override=True, operator=operator)
    elif recalc_assignment:
        assignment = _resolve_unloading_assignment(
            db,
            rm_number=m.rm_number,
            item_name=m.item_name,
            plant_name=m.plant_name,
            company_id_hint=m.company_id,
            main_tender_id_hint=m.main_tender_id,
            tender_id_hint=m.tender_id,
        )
        _apply_unloading_assignment_to_master(m, assignment)

    _sync_unloading_lineage_from_master(db, m)

    if any(k in payload for k in ("rm_number", "item_name", "party_name", "plant_name", "po_number")) or scope_override:
        for e in (m.entries or []):
            if e.receipt_id:
                _sync_receipt_from_unloading_entry(db, e)

    db.flush()
    _reconcile_unloading_match_for_all_bills(db)
    _recalculate_tender_fulfilled_from_unloading(db, list(old_tender_ids | {m.tender_id}))
    db.commit()
    return {"ok": True, "master": {"id": m.id, "tender_id": m.tender_id, "assignment_status": m.assignment_status}}


@app.post("/api/plant-unloading/upload")
async def upload_plant_unloading_sheet(
    file: UploadFile = File(...),
    source: str = Form("web"),
    ocr_engine: Optional[str] = Form(None),
    operator: Optional[str] = Form(None),
    company_id: Optional[int] = Query(None),
    main_tender_id: Optional[int] = Query(None),
    tender_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
):
    from services.ocr_service import extract_plant_unloading_sheet

    settings = get_settings()
    dest_dir = Path(settings.upload_dir) / "plant_unloading"
    dest_dir.mkdir(parents=True, exist_ok=True)
    fname = f"upl_{datetime.now().strftime('%Y%m%d%H%M%S')}_{file.filename}"
    fpath = dest_dir / fname
    fpath.write_bytes(await file.read())

    company_id, main_tender_id, tender_id = _resolve_context_ids(
        db, company_id=company_id, main_tender_id=main_tender_id, tender_id=tender_id
    )

    selected_engine = str(ocr_engine or "").strip().lower()
    if selected_engine in {"", "auto"}:
        selected_engine = None
    allowed_unloading_engines = {"groq", "mistral", "gemini", "azure", "github"}
    if selected_engine and selected_engine not in allowed_unloading_engines:
        raise HTTPException(400, f"Invalid ocr_engine: {ocr_engine}")

    extracted = await extract_plant_unloading_sheet(str(fpath), preferred_provider=selected_engine)
    rows = extracted.get("rows") or []
    if not rows:
        err = extracted.get("error") or "No unloading rows were extracted from the uploaded sheet"
        log.warning(f"Plant unloading upload failed for '{file.filename}': {err}")
        raise HTTPException(400, err)

    created_entries, duplicates, total_rows = _ingest_unloading_rows(
        db=db,
        extracted=extracted,
        image_path=fname,
        source=source,
        company_id=company_id,
        main_tender_id=main_tender_id,
        tender_id=tender_id,
    )

    if operator:
        for e in created_entries:
            e.reviewed_by = operator
        db.commit()

    return {
        "created": [_plant_unloading_to_dict(e) for e in created_entries],
        "duplicates": duplicates,
        "rows_received": total_rows,
        "upload_details": {
            "file_name": file.filename,
            "ocr_engine": selected_engine or "auto",
            "extract_source": extracted.get("source"),
            "extract_confidence": _to_float(extracted.get("confidence")),
            "provider_attempts": extracted.get("provider_attempts") or [],
        },
        "message": f"Processed {total_rows} rows, created {len(created_entries)}, skipped {duplicates} duplicates",
    }


@app.post("/api/plant-unloading/bulk-upload")
async def bulk_upload_plant_unloading_sheet(
    files: List[UploadFile] = File(...),
    source: str = Form("web"),
    ocr_engine: Optional[str] = Form(None),
    operator: Optional[str] = Form(None),
    company_id: Optional[int] = Query(None),
    main_tender_id: Optional[int] = Query(None),
    tender_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
):
    from services.ocr_service import extract_plant_unloading_sheet

    settings = get_settings()
    dest_dir = Path(settings.upload_dir) / "plant_unloading"
    dest_dir.mkdir(parents=True, exist_ok=True)

    company_id, main_tender_id, tender_id = _resolve_context_ids(
        db, company_id=company_id, main_tender_id=main_tender_id, tender_id=tender_id
    )

    all_created = []
    total_duplicates = 0
    total_rows = 0
    upload_details = []
    selected_engine = str(ocr_engine or "").strip().lower()
    if selected_engine in {"", "auto"}:
        selected_engine = None
    allowed_unloading_engines = {"groq", "mistral", "gemini", "azure", "github"}
    if selected_engine and selected_engine not in allowed_unloading_engines:
        raise HTTPException(400, f"Invalid ocr_engine: {ocr_engine}")

    for file in files:
        fname = f"upl_{datetime.now().strftime('%Y%m%d%H%M%S')}_{file.filename}"
        fpath = dest_dir / fname
        fpath.write_bytes(await file.read())
        extracted = await extract_plant_unloading_sheet(str(fpath), preferred_provider=selected_engine)
        rows = extracted.get("rows") or []
        if not rows:
            err = extracted.get("error") or "No unloading rows were extracted from uploaded sheet"
            log.warning(f"Bulk plant upload: no rows for '{file.filename}' error='{err}'")
            upload_details.append({
                "file_name": file.filename,
                "rows_received": 0,
                "created": 0,
                "duplicates": 0,
                "ocr_engine": selected_engine or "auto",
                "extract_source": extracted.get("source"),
                "extract_confidence": _to_float(extracted.get("confidence")),
                "provider_attempts": extracted.get("provider_attempts") or [],
                "error": err,
            })
            continue
        created_entries, duplicates, rows = _ingest_unloading_rows(
            db=db,
            extracted=extracted,
            image_path=fname,
            source=source,
            company_id=company_id,
            main_tender_id=main_tender_id,
            tender_id=tender_id,
        )
        if operator and created_entries:
            for e in created_entries:
                e.reviewed_by = operator
            db.commit()
        all_created.extend(created_entries)
        total_duplicates += duplicates
        total_rows += rows
        upload_details.append({
            "file_name": file.filename,
            "rows_received": rows,
            "created": len(created_entries),
            "duplicates": duplicates,
            "ocr_engine": selected_engine or "auto",
            "extract_source": extracted.get("source"),
            "extract_confidence": _to_float(extracted.get("confidence")),
            "provider_attempts": extracted.get("provider_attempts") or [],
            "error": None,
        })

    return {
        "created": [_plant_unloading_to_dict(e) for e in all_created],
        "duplicates": total_duplicates,
        "rows_received": total_rows,
        "upload_details": upload_details,
        "message": f"Processed {total_rows} rows, created {len(all_created)}, skipped {total_duplicates} duplicates",
    }


@app.patch("/api/plant-unloading/{entry_id}")
def update_plant_unloading_entry(entry_id: int, data: PlantUnloadingUpdate,
                                 company_id: Optional[int] = None,
                                 main_tender_id: Optional[int] = None,
                                 tender_id: Optional[int] = None,
                                 db: Session = Depends(get_db)):
    e = db.get(PlantUnloadingEntry, entry_id)
    _assert_record_scope(e, company_id, main_tender_id, tender_id, label="Plant unloading entry")
    old_tender_id = e.tender_id
    update_data = {k: v for k, v in data.model_dump(exclude_unset=True).items()}
    if "status" in update_data and update_data["status"] is not None:
        update_data["status"] = BillStatus(update_data["status"])
    for k, v in update_data.items():
        setattr(e, k, v)

    if e.receipt_id:
        _sync_receipt_from_unloading_entry(db, e)

    db.flush()
    _reconcile_unloading_match_for_all_bills(db)
    _recalculate_tender_fulfilled_from_unloading(db, [old_tender_id, e.tender_id])
    db.commit()
    db.refresh(e)
    return _plant_unloading_to_dict(e)


@app.patch("/api/plant-unloading/{entry_id}/approve")
def approve_plant_unloading_entry(entry_id: int, operator: Optional[str] = Query(None),
                                  company_id: Optional[int] = None,
                                  main_tender_id: Optional[int] = None,
                                  tender_id: Optional[int] = None,
                                  db: Session = Depends(get_db)):
    e = db.get(PlantUnloadingEntry, entry_id)
    _assert_record_scope(e, company_id, main_tender_id, tender_id, label="Plant unloading entry")

    master = e.master
    if not master:
        raise HTTPException(404, "Plant unloading register not found")

    if not master.is_manual_override or not master.tender_id or master.requires_manual_assignment:
        assignment = _resolve_unloading_assignment(
            db,
            rm_number=master.rm_number,
            item_name=master.item_name,
            plant_name=master.plant_name,
            company_id_hint=master.company_id,
            main_tender_id_hint=master.main_tender_id,
            tender_id_hint=master.tender_id,
        )
        if not master.is_manual_override or not master.tender_id:
            _apply_unloading_assignment_to_master(master, assignment)
        else:
            master.requires_manual_assignment = bool(assignment.get("requires_manual_assignment"))
            master.assignment_reason = assignment.get("assignment_reason")

    if master.requires_manual_assignment or not master.tender_id:
        reason = master.assignment_reason or "Unable to map RM/plant to a unique subtender"
        raise HTTPException(400, f"Manual assignment required before approval: {reason}")

    if master.company_id is None or master.main_tender_id is None:
        raise HTTPException(400, "Manual assignment required before approval: company/main tender missing")

    _sync_unloading_lineage_from_master(db, master)
    _stamp_lineage(e, master.company_id, master.main_tender_id, master.tender_id)

    e.status = BillStatus.approved
    e.reviewed_at = datetime.now()
    if operator:
        e.reviewed_by = operator

    if not e.receipt_created:
        company_id_ctx = e.company_id if e.company_id is not None else (master.company_id if master else None)
        main_tender_id_ctx = e.main_tender_id if e.main_tender_id is not None else (master.main_tender_id if master else None)
        tender_id_ctx = e.tender_id if e.tender_id is not None else (master.tender_id if master else None)
        plant_id = master.plant_id if master else None
        if not plant_id and master and master.plant_name:
            pid, pname = _resolve_plant(db, master.plant_name)
            if pid:
                master.plant_id = pid
                master.plant_name = pname
                plant_id = pid
        if not plant_id:
            raise HTTPException(400, "Cannot create receipt: plant not resolved for this unloading register")

        accepted = _effective_unloading_qty_mt(e.net_qty_mt, e.received_qty_mt)
        received = _to_float(e.received_qty_mt)
        if received is None:
            received = accepted
        rejected = max(received - accepted, 0)
        material_id, material_name = _resolve_material(db, material_name=(e.item_name or (master.item_name if master else None)))
        receipt = PlantReceipt(
            company_id=company_id_ctx,
            main_tender_id=main_tender_id_ctx,
            tender_id=tender_id_ctx,
            dispatch_id=None,
            vehicle_number=e.truck_number,
            plant_id=plant_id,
            receipt_date=e.entry_date,
            accepted_mt=accepted,
            rejected_mt=rejected,
            received_qty_qtl=received,
            matched_qty_qtl=0,
            match_status="unmatched",
            material_id=material_id,
            material_name=material_name,
            rm_number=master.rm_number if master else None,
            party_name=master.party_name if master else None,
            po_number=master.po_number if master else None,
            rejection_reason=("Auto-created from unloading sheet" if rejected > 0 else None),
            source="portal",
        )
        db.add(receipt)
        db.flush()

        _match_receipt_to_dispatch(db, receipt, manual=False)
        if receipt.dispatch_id:
            try:
                _apply_receipt_delta(db, receipt.dispatch_id, accepted, rejected)
                receipt.matched_dispatch_id = receipt.dispatch_id
                receipt.matched_qty_qtl = accepted + rejected
                receipt.match_applied_at = datetime.now()
            except HTTPException as exc:
                receipt.dispatch_id = None
                receipt.matched_dispatch_id = None
                receipt.matched_qty_qtl = 0
                receipt.match_applied_at = None
                receipt.match_status = "unmatched"
                receipt.match_reason = f"Auto-match not applied: {exc.detail}"
            db.flush()
        _refresh_dispatch_status(db, receipt.dispatch_id)

        e.receipt_id = receipt.id
        e.receipt_created = True

    db.flush()
    _reconcile_unloading_match_for_all_bills(db)
    _recalculate_tender_fulfilled_from_unloading(db, [e.tender_id])
    db.commit()
    db.refresh(e)
    return _plant_unloading_to_dict(e)


@app.patch("/api/plant-unloading/{entry_id}/reject")
def reject_plant_unloading_entry(entry_id: int, operator: Optional[str] = Query(None),
                                 company_id: Optional[int] = None,
                                 main_tender_id: Optional[int] = None,
                                 tender_id: Optional[int] = None,
                                 db: Session = Depends(get_db)):
    e = db.get(PlantUnloadingEntry, entry_id)
    _assert_record_scope(e, company_id, main_tender_id, tender_id, label="Plant unloading entry")

    old_tender_id = e.tender_id
    was_approved = e.status == BillStatus.approved
    if was_approved and e.receipt_id:
        r = db.get(PlantReceipt, e.receipt_id)
        if r and not r.is_deleted:
            _soft_delete_receipt_with_revert(db, r)
        e.receipt_id = None
        e.receipt_created = False

    e.status = BillStatus.rejected
    e.reviewed_at = datetime.now()
    if operator:
        e.reviewed_by = operator

    db.flush()
    _reconcile_unloading_match_for_all_bills(db)
    _recalculate_tender_fulfilled_from_unloading(db, [old_tender_id, e.tender_id])
    db.commit()
    db.refresh(e)
    return _plant_unloading_to_dict(e)


@app.delete("/api/plant-unloading/{entry_id}")
def delete_plant_unloading_entry(entry_id: int, company_id: Optional[int] = None,
                                 main_tender_id: Optional[int] = None,
                                 tender_id: Optional[int] = None,
                                 db: Session = Depends(get_db)):
    e = db.get(PlantUnloadingEntry, entry_id)
    _assert_record_scope(e, company_id, main_tender_id, tender_id, label="Plant unloading entry")
    affected_tender_ids = {e.tender_id}

    if e.receipt_id:
        r = db.get(PlantReceipt, e.receipt_id)
        if r and not r.is_deleted:
            _soft_delete_receipt_with_revert(db, r)

    master_id = e.master_id
    db.delete(e)
    db.flush()

    # Keep masters clean: if no rows left, remove master.
    left = db.execute(select(func.count(PlantUnloadingEntry.id)).where(PlantUnloadingEntry.master_id == master_id)).scalar() or 0
    if left == 0:
        m = db.get(PlantUnloadingMaster, master_id)
        if m:
            affected_tender_ids.add(m.tender_id)
            db.delete(m)

    db.flush()
    _reconcile_unloading_match_for_all_bills(db)
    _recalculate_tender_fulfilled_from_unloading(db, list(affected_tender_ids))
    db.commit()
    return {"ok": True}


@app.post("/api/plant-unloading/bulk-delete")
def bulk_delete_plant_unloading_entries(data: IdListPayload, company_id: Optional[int] = None,
                                        main_tender_id: Optional[int] = None,
                                        tender_id: Optional[int] = None,
                                        db: Session = Depends(get_db)):
    ids = [int(i) for i in (data.ids or [])]
    if not ids:
        raise HTTPException(400, "No ids provided")

    q = select(PlantUnloadingEntry).where(PlantUnloadingEntry.id.in_(ids))
    q = _apply_scope_filters(q, PlantUnloadingEntry, company_id, main_tender_id, tender_id)
    rows = db.execute(q).scalars().all()
    found_ids = {r.id for r in rows}
    missing_ids = [i for i in ids if i not in found_ids]

    master_ids = set()
    affected_tender_ids = set()
    deleted_count = 0
    for e in rows:
        if e.receipt_id:
            r = db.get(PlantReceipt, e.receipt_id)
            if r and not r.is_deleted:
                _soft_delete_receipt_with_revert(db, r)
        master_ids.add(e.master_id)
        affected_tender_ids.add(e.tender_id)
        db.delete(e)
        deleted_count += 1

    db.flush()

    # Keep master table clean if a register has no rows left
    deleted_masters = 0
    for mid in master_ids:
        left = db.execute(select(func.count(PlantUnloadingEntry.id)).where(PlantUnloadingEntry.master_id == mid)).scalar() or 0
        if left == 0:
            m = db.get(PlantUnloadingMaster, mid)
            if m:
                affected_tender_ids.add(m.tender_id)
                db.delete(m)
                deleted_masters += 1

    db.flush()
    _reconcile_unloading_match_for_all_bills(db)
    _recalculate_tender_fulfilled_from_unloading(db, list(affected_tender_ids))
    db.commit()
    return {
        "ok": True,
        "deleted_rows": deleted_count,
        "deleted_masters": deleted_masters,
        "missing_ids": missing_ids,
    }


@app.delete("/api/plant-unloading/masters/{master_id}")
def delete_plant_unloading_master(master_id: int, company_id: Optional[int] = None,
                                  main_tender_id: Optional[int] = None,
                                  tender_id: Optional[int] = None,
                                  db: Session = Depends(get_db)):
    m = db.get(PlantUnloadingMaster, master_id)
    _assert_record_scope(m, company_id, main_tender_id, tender_id, label="Plant unloading register")

    entries = db.execute(select(PlantUnloadingEntry).where(PlantUnloadingEntry.master_id == master_id)).scalars().all()
    deleted_rows = len(entries)
    affected_tender_ids = {m.tender_id}

    for e in entries:
        if e.receipt_id:
            r = db.get(PlantReceipt, e.receipt_id)
            if r and not r.is_deleted:
                _soft_delete_receipt_with_revert(db, r)
        affected_tender_ids.add(e.tender_id)

    # Explicitly delete child rows first to avoid ORM setting master_id to NULL
    # on non-nullable foreign keys during parent delete.
    for e in entries:
        db.delete(e)

    db.flush()
    db.delete(m)
    db.flush()
    _reconcile_unloading_match_for_all_bills(db)
    _recalculate_tender_fulfilled_from_unloading(db, list(affected_tender_ids))
    db.commit()
    return {"ok": True, "deleted_rows": deleted_rows}

# ── TELEGRAM WEBHOOK ───────────────────────────────────────────────────────

@app.post("/api/telegram/webhook")
async def telegram_webhook(payload: dict, db: Session = Depends(get_db)):
    from services.telegram_service import handle_telegram_update
    await handle_telegram_update(payload, db)
    return {"ok": True}

@app.post("/api/telegram/set-webhook")
async def set_telegram_webhook():
    settings = get_settings()
    if not settings.telegram_token:
        raise HTTPException(400, "TELEGRAM_TOKEN not set in .env")
    async with __import__("httpx").AsyncClient() as client:
        r = await client.get(
            f"https://api.telegram.org/bot{settings.telegram_token}/setWebhook"
            f"?url={settings.webhook_url}/api/telegram/webhook"
        )
    return r.json()

# ── DISPATCH & RECEIPTS ────────────────────────────────────────────────────

@app.get("/api/dispatches")
def list_dispatches(plant_id: Optional[int] = None, status: Optional[str] = None,
                    company_id: Optional[int] = None, main_tender_id: Optional[int] = None,
                    tender_id: Optional[int] = None, db: Session = Depends(get_db)):
    q = select(Dispatch).where(Dispatch.is_deleted == False).order_by(desc(Dispatch.dispatch_date))
    q = _apply_scope_filters(q, Dispatch, company_id, main_tender_id, tender_id)
    if plant_id: q = q.where(Dispatch.plant_id == plant_id)
    if status:   q = q.where(Dispatch.status == DispatchStatus(status))
    disps = db.execute(q).scalars().all()
    result = []
    for d in disps:
        _enrich_dispatch_material(db, d)
        active_receipt = _active_receipt_for_dispatch(db, d.id)
        consumed = float(d.consumed_qty_qtl or 0)
        qty = float(d.qty_mt or 0)
        result.append({
            "id": d.id,
            "company_id": d.company_id,
            "main_tender_id": d.main_tender_id,
            "tender_id": d.tender_id,
            "vehicle_number": d.vehicle_number,
            "dispatch_date": str(d.dispatch_date),
            "qty_mt": qty,
            "consumed_qty_qtl": consumed,
            "remaining_qty_qtl": max(qty - consumed, 0),
            "plant": d.plant.name if d.plant else None,
            "plant_id": d.plant_id,
            "material_id": d.material_id,
            "material": d.material_name,
            "status": d.status.value,
            "bill_id": d.bill_id,
            "deal_id": d.deal_id,
            "accepted_mt": float(active_receipt.accepted_mt) if active_receipt else None,
            "rejected_mt": float(active_receipt.rejected_mt) if active_receipt else None,
        })
    return result


@app.post("/api/dispatches")
def create_dispatch(
    data: DispatchCreate,
    company_id: Optional[int] = Query(None),
    main_tender_id: Optional[int] = Query(None),
    tender_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
):
    _validate_vehicle_number(data.vehicle_number)
    _assert_bill_dispatch_1to1(db, data.bill_id)

    deal_id = data.deal_id
    if not deal_id and data.bill_id:
        b = db.get(Bill, data.bill_id)
        if b and b.deal_id:
            deal_id = b.deal_id
    if not deal_id:
        raise HTTPException(400, "Deal is required to create dispatch")

    dc, dm, dt = _lineage_from_deal(db, deal_id)
    if company_id is not None and company_id != dc:
        raise HTTPException(400, "company_id does not match dispatch deal lineage")
    if main_tender_id is not None and main_tender_id != dm:
        raise HTTPException(400, "main_tender_id does not match dispatch deal lineage")
    if tender_id is not None and tender_id != dt:
        raise HTTPException(400, "tender_id does not match dispatch deal lineage")
    company_id = company_id if company_id is not None else dc
    main_tender_id = main_tender_id if main_tender_id is not None else dm
    tender_id = tender_id if tender_id is not None else dt

    d = Dispatch(
        company_id = company_id,
        main_tender_id = main_tender_id,
        tender_id = tender_id,
        bill_id = data.bill_id,
        deal_id = deal_id,
        material_id = data.material_id,
        material_name = data.material_name,
        vehicle_number = data.vehicle_number,
        dispatch_date = data.dispatch_date,
        qty_mt = data.qty_mt,
        consumed_qty_qtl = 0,
        plant_id = data.plant_id,
        driver_name = data.driver_name,
        driver_phone = data.driver_phone,
        status = DispatchStatus(data.status) if data.status else DispatchStatus.in_transit,
    )
    _enrich_dispatch_material(db, d)
    db.add(d)
    db.flush()

    auto_matched_receipt_id = _auto_match_open_receipt_for_dispatch(db, d)
    _refresh_dispatch_status(db, d.id)
    _recalculate_deal_dispatched_mt(db, [deal_id])
    db.commit(); db.refresh(d)
    return {"id": d.id, "auto_matched_receipt_id": auto_matched_receipt_id}

@app.post("/api/receipts")
def add_receipt(
    data: ReceiptCreate,
    company_id: Optional[int] = Query(None),
    main_tender_id: Optional[int] = Query(None),
    tender_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
):
    _validate_vehicle_number(data.vehicle_number)
    company_id, main_tender_id, tender_id = _resolve_context_ids(
        db, company_id=company_id, main_tender_id=main_tender_id, tender_id=tender_id
    )
    r = PlantReceipt(
        company_id=company_id,
        main_tender_id=main_tender_id,
        tender_id=tender_id,
        dispatch_id=data.dispatch_id,
        matched_dispatch_id=data.dispatch_id,
        vehicle_number=data.vehicle_number,
        plant_id=data.plant_id,
        receipt_date=data.receipt_date,
        accepted_mt=data.accepted_mt,
        rejected_mt=data.rejected_mt,
        received_qty_qtl=float(data.accepted_mt or 0) + float(data.rejected_mt or 0),
        rejection_reason=data.rejection_reason,
        source="manual",
        material_id=data.material_id,
        material_name=data.material_name,
        matched_qty_qtl=0,
        match_status="unmatched",
    )
    _enrich_receipt_material(db, r)
    db.add(r)
    db.flush()

    manual_link = data.dispatch_id is not None
    if r.dispatch_id:
        _assert_dispatch_receipt_1to1(db, r.dispatch_id, exclude_receipt_id=r.id)
        d = db.get(Dispatch, r.dispatch_id)
        if not d:
            raise HTTPException(404, "Selected dispatch not found")
        _assert_record_scope(d, company_id, main_tender_id, tender_id, label="Dispatch")
        _stamp_lineage(r, d.company_id, d.main_tender_id, d.tender_id)
        _enrich_dispatch_material(db, d)
        _validate_receipt_dispatch_keys(r, d)
        if manual_link:
            r.match_status = "manual"
            r.match_reason = "Linked manually by operator"
    else:
        _match_receipt_to_dispatch(db, r, manual=False)

    if r.dispatch_id:
        try:
            _apply_receipt_delta(db, r.dispatch_id, float(r.accepted_mt or 0), float(r.rejected_mt or 0))
            r.matched_dispatch_id = r.dispatch_id
            r.matched_qty_qtl = float(r.accepted_mt or 0) + float(r.rejected_mt or 0)
            r.match_applied_at = datetime.now()
        except HTTPException as exc:
            if manual_link:
                raise
            r.dispatch_id = None
            r.matched_dispatch_id = None
            r.matched_qty_qtl = 0
            r.match_applied_at = None
            r.match_status = "unmatched"
            r.match_reason = f"Auto-match not applied: {exc.detail}"

    # SessionLocal uses autoflush=False, so persist receipt link changes before status query.
    db.flush()
    _refresh_dispatch_status(db, r.dispatch_id)
    db.commit()
    return {"id": r.id, "match_status": r.match_status, "dispatch_id": r.dispatch_id}


@app.get("/api/receipts")
def list_receipts(plant_id: Optional[int] = None, company_id: Optional[int] = None,
                  main_tender_id: Optional[int] = None, tender_id: Optional[int] = None,
                  db: Session = Depends(get_db)):
    q = select(PlantReceipt).where(PlantReceipt.is_deleted == False).order_by(desc(PlantReceipt.created_at))
    q = _apply_scope_filters(q, PlantReceipt, company_id, main_tender_id, tender_id)
    if plant_id:
        q = q.where(PlantReceipt.plant_id == plant_id)
    receipts = db.execute(q).scalars().all()
    result = []
    for r in receipts:
        result.append({
            "id": r.id,
            "company_id": r.company_id,
            "main_tender_id": r.main_tender_id,
            "tender_id": r.tender_id,
            "vehicle_number": r.vehicle_number,
            "receipt_date": str(r.receipt_date),
            "plant": r.plant.name if r.plant else None,
            "plant_id": r.plant_id,
            "material_id": r.material_id,
            "material": r.material_name,
            "accepted_mt": float(r.accepted_mt or 0),
            "rejected_mt": float(r.rejected_mt or 0),
            "received_qty_qtl": float(r.received_qty_qtl or 0) if r.received_qty_qtl is not None else None,
            "matched_qty_qtl": float(r.matched_qty_qtl or 0),
            "match_status": r.match_status,
            "match_reason": r.match_reason,
            "matched_dispatch_id": r.matched_dispatch_id,
            "match_applied_at": str(r.match_applied_at) if r.match_applied_at else None,
            "rm_number": r.rm_number,
            "party_name": r.party_name,
            "po_number": r.po_number,
            "rejection_reason": r.rejection_reason,
            "dispatch_id": r.dispatch_id,
            "created_at": str(r.created_at),
        })
    return result


@app.get("/api/receipts/{receipt_id}/match-candidates")
def receipt_match_candidates(receipt_id: int, company_id: Optional[int] = None,
                             main_tender_id: Optional[int] = None, tender_id: Optional[int] = None,
                             db: Session = Depends(get_db)):
    r = db.get(PlantReceipt, receipt_id)
    if not r or r.is_deleted:
        raise HTTPException(404, "Receipt not found")
    _assert_record_scope(r, company_id, main_tender_id, tender_id, label="Receipt")

    candidates = _manual_match_candidates_for_receipt(db, r)
    return [{
        "id": d.id,
        "vehicle_number": d.vehicle_number,
        "dispatch_date": str(d.dispatch_date),
        "qty_mt": float(d.qty_mt or 0),
        "consumed_qty_qtl": float(d.consumed_qty_qtl or 0),
        "remaining_qty_qtl": _dispatch_remaining_qtl(d),
        "plant_id": d.plant_id,
        "plant": d.plant.name if d.plant else None,
        "material_id": d.material_id,
        "material": d.material_name,
        "status": d.status.value,
        "deal_id": d.deal_id,
        "bill_id": d.bill_id,
    } for d in candidates]


@app.post("/api/receipts/{receipt_id}/manual-match")
def manual_match_receipt(receipt_id: int, data: ReceiptManualMatch,
                         company_id: Optional[int] = None,
                         main_tender_id: Optional[int] = None,
                         tender_id: Optional[int] = None,
                         db: Session = Depends(get_db)):
    r = db.get(PlantReceipt, receipt_id)
    if not r or r.is_deleted:
        raise HTTPException(404, "Receipt not found")
    _assert_record_scope(r, company_id, main_tender_id, tender_id, label="Receipt")

    old_dispatch_id = r.dispatch_id
    old_accepted = float(r.accepted_mt or 0)
    old_rejected = float(r.rejected_mt or 0)

    d = db.get(Dispatch, data.dispatch_id)
    if not d or d.is_deleted:
        raise HTTPException(404, "Dispatch not found")
    _assert_record_scope(d, company_id, main_tender_id, tender_id, label="Dispatch")

    _enrich_receipt_material(db, r)
    _enrich_dispatch_material(db, d)
    _validate_receipt_dispatch_keys(r, d)

    _assert_dispatch_receipt_1to1(db, d.id, exclude_receipt_id=r.id)

    if old_dispatch_id and old_dispatch_id != d.id:
        _apply_receipt_delta(db, old_dispatch_id, -old_accepted, -old_rejected)

    if old_dispatch_id != d.id:
        r.dispatch_id = d.id
        _apply_receipt_delta(db, d.id, old_accepted, old_rejected)

    r.matched_dispatch_id = d.id
    r.matched_qty_qtl = old_accepted + old_rejected
    r.match_status = "manual"
    r.match_reason = "Linked manually by operator"
    r.match_applied_at = datetime.now()

    db.flush()
    _refresh_dispatch_status(db, old_dispatch_id)
    _refresh_dispatch_status(db, d.id)

    db.commit()
    return {"id": r.id, "dispatch_id": r.dispatch_id, "match_status": r.match_status}


@app.patch("/api/dispatches/{dispatch_id}")
def update_dispatch(dispatch_id: int, data: DispatchUpdate,
                    company_id: Optional[int] = None,
                    main_tender_id: Optional[int] = None,
                    tender_id: Optional[int] = None,
                    db: Session = Depends(get_db)):
    d = db.get(Dispatch, dispatch_id)
    _assert_record_scope(d, company_id, main_tender_id, tender_id, label="Dispatch")
    old_deal_id = d.deal_id
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}

    if "deal_id" in update_data and update_data["deal_id"] != old_deal_id:
        active_receipt = _active_receipt_for_dispatch(db, dispatch_id)
        if active_receipt:
            raise HTTPException(400, "Cannot change deal on a dispatch that already has a receipt")

    for k, v in update_data.items():
        setattr(d, k, v)

    if d.deal_id:
        lc, lm, lt = _lineage_from_deal(db, d.deal_id)
        _stamp_lineage(d, lc, lm, lt)

    if "vehicle_number" in update_data:
        _validate_vehicle_number(d.vehicle_number)

    if d.bill_id:
        _assert_bill_dispatch_1to1(db, d.bill_id, exclude_dispatch_id=d.id)

    _enrich_dispatch_material(db, d)

    if float(d.consumed_qty_qtl or 0) - float(d.qty_mt or 0) > 1e-9:
        raise HTTPException(400, "Dispatch quantity cannot be less than already consumed quantity")

    if 'qty_mt' in update_data or 'deal_id' in update_data:
        _recalculate_deal_dispatched_mt(db, [old_deal_id, d.deal_id])

    auto_matched_receipt_id = None
    if not _active_receipt_for_dispatch(db, d.id):
        auto_matched_receipt_id = _auto_match_open_receipt_for_dispatch(db, d)

    _refresh_dispatch_status(db, d.id)
    db.commit(); db.refresh(d)
    return {"id": d.id, "auto_matched_receipt_id": auto_matched_receipt_id}


@app.delete("/api/dispatches/{dispatch_id}")
def delete_dispatch(dispatch_id: int, company_id: Optional[int] = None,
                    main_tender_id: Optional[int] = None,
                    tender_id: Optional[int] = None,
                    db: Session = Depends(get_db)):
    d = db.get(Dispatch, dispatch_id)
    _assert_record_scope(d, company_id, main_tender_id, tender_id, label="Dispatch")
    if _active_receipt_for_dispatch(db, dispatch_id):
        raise HTTPException(400, "Cannot delete dispatch with receipt; delete receipt first")
    # soft-delete and log
    d.is_deleted = True
    d.deleted_at = datetime.now()
    _recalculate_deal_dispatched_mt(db, [d.deal_id])
    # audit
    from models import AuditLog
    al = AuditLog(entity='dispatch', entity_id=d.id, action='delete', payload={
        'bill_id': d.bill_id, 'deal_id': d.deal_id, 'qty_mt': float(d.qty_mt or 0)
    })
    db.add(al)
    db.commit(); db.refresh(d)
    return {"ok": True}


@app.patch("/api/receipts/{receipt_id}")
def update_receipt(receipt_id: int, data: ReceiptUpdate,
                   company_id: Optional[int] = None,
                   main_tender_id: Optional[int] = None,
                   tender_id: Optional[int] = None,
                   db: Session = Depends(get_db)):
    r = db.get(PlantReceipt, receipt_id)
    _assert_record_scope(r, company_id, main_tender_id, tender_id, label="Receipt")
    old_accepted = float(r.accepted_mt or 0)
    old_rejected = float(r.rejected_mt or 0)
    old_dispatch_id = r.dispatch_id

    update_data = {k: v for k, v in data.model_dump(exclude_unset=True).items()}
    for k, v in update_data.items():
        setattr(r, k, v)

    if "vehicle_number" in update_data:
        _validate_vehicle_number(r.vehicle_number)

    _enrich_receipt_material(db, r)
    r.received_qty_qtl = float(r.accepted_mt or 0) + float(r.rejected_mt or 0)

    # Revert old applied quantity effects first.
    _apply_receipt_delta(db, old_dispatch_id, -old_accepted, -old_rejected)

    manual_dispatch_set = "dispatch_id" in update_data and update_data.get("dispatch_id") is not None
    if r.dispatch_id:
        _assert_dispatch_receipt_1to1(db, r.dispatch_id, exclude_receipt_id=r.id)
        d = db.get(Dispatch, r.dispatch_id)
        if not d:
            raise HTTPException(404, "Selected dispatch not found")
        _stamp_lineage(r, d.company_id, d.main_tender_id, d.tender_id)
        _enrich_dispatch_material(db, d)
        _validate_receipt_dispatch_keys(r, d)
    else:
        _match_receipt_to_dispatch(db, r, manual=False)

    if r.dispatch_id:
        auto_linked = (not manual_dispatch_set) and (old_dispatch_id != r.dispatch_id) and (r.match_status == "auto")
        try:
            _apply_receipt_delta(db, r.dispatch_id, float(r.accepted_mt or 0), float(r.rejected_mt or 0))
            r.matched_dispatch_id = r.dispatch_id
            r.matched_qty_qtl = float(r.accepted_mt or 0) + float(r.rejected_mt or 0)
            r.match_applied_at = datetime.now()
            if manual_dispatch_set:
                r.match_status = "manual"
                r.match_reason = "Linked manually by operator"
            elif r.match_status not in ("auto", "manual"):
                r.match_status = "manual"
                r.match_reason = "Linked during receipt update"
        except HTTPException as exc:
            if not auto_linked:
                raise
            r.dispatch_id = None
            r.matched_dispatch_id = None
            r.matched_qty_qtl = 0
            r.match_applied_at = None
            r.match_status = "unmatched"
            r.match_reason = f"Auto-match not applied: {exc.detail}"
    else:
        r.matched_dispatch_id = None
        r.matched_qty_qtl = 0
        r.match_applied_at = None

    db.flush()
    _refresh_dispatch_status(db, old_dispatch_id)
    _refresh_dispatch_status(db, r.dispatch_id)

    db.commit(); db.refresh(r)
    return {"id": r.id}


@app.get('/api/audit-logs')
def list_audit_logs(limit: int = 50, db: Session = Depends(get_db)):
    q = select(AuditLog).order_by(desc(AuditLog.created_at)).limit(limit)
    logs = db.execute(q).scalars().all()
    return [{
        'id': l.id, 'entity': l.entity, 'entity_id': l.entity_id,
        'action': l.action, 'payload': l.payload, 'created_at': str(l.created_at)
    } for l in logs]


@app.delete("/api/receipts/{receipt_id}")
def delete_receipt(receipt_id: int, company_id: Optional[int] = None,
                   main_tender_id: Optional[int] = None,
                   tender_id: Optional[int] = None,
                   db: Session = Depends(get_db)):
    r = db.get(PlantReceipt, receipt_id)
    _assert_record_scope(r, company_id, main_tender_id, tender_id, label="Receipt")
    old_accepted = float(r.accepted_mt or 0)
    old_rejected = float(r.rejected_mt or 0)
    old_dispatch_id = r.dispatch_id
    _soft_delete_receipt_with_revert(db, r)
    _refresh_dispatch_status(db, old_dispatch_id)
    # soft-delete receipt and log
    from models import AuditLog
    al = AuditLog(entity='receipt', entity_id=r.id, action='delete', payload={
        'dispatch_id': old_dispatch_id,
        'accepted_mt': old_accepted,
        'rejected_mt': old_rejected,
        'material': r.material_name,
        'match_status': r.match_status,
    })
    db.add(al)
    db.commit()
    return {"ok": True}


# ── PURCHASE ORDERS ───────────────────────────────────────────────────────

@app.get("/api/purchase-orders")
def list_purchase_orders(
    status: Optional[str] = None,
    company_id: Optional[int] = None,
    main_tender_id: Optional[int] = None,
    tender_id: Optional[int] = None,
    db: Session = Depends(get_db),
):
    q = select(PurchaseOrder).order_by(desc(PurchaseOrder.created_at))
    q = _apply_scope_filters(q, PurchaseOrder, company_id, main_tender_id, tender_id)
    if status:
        try:
            q = q.where(PurchaseOrder.status == PurchaseOrderStatus(status))
        except Exception:
            raise HTTPException(400, f"Invalid status: {status}")
    rows = db.execute(q).scalars().all()
    return [_purchase_order_to_dict(r) for r in rows]


@app.get("/api/purchase-orders/{po_id}")
def get_purchase_order(
    po_id: int,
    company_id: Optional[int] = None,
    main_tender_id: Optional[int] = None,
    tender_id: Optional[int] = None,
    db: Session = Depends(get_db),
):
    po = db.get(PurchaseOrder, po_id)
    _assert_record_scope(po, company_id, main_tender_id, tender_id, label="Purchase order")
    return _purchase_order_to_dict(po)


@app.get("/api/purchase-orders/{po_id}/subtender-preview")
def get_purchase_order_subtender_preview(
    po_id: int,
    company_id: Optional[int] = None,
    main_tender_id: Optional[int] = None,
    tender_id: Optional[int] = None,
    db: Session = Depends(get_db),
):
    po = db.get(PurchaseOrder, po_id)
    _assert_record_scope(po, company_id, main_tender_id, tender_id, label="Purchase order")
    return _purchase_order_subtender_preview(db, po)


@app.post("/api/purchase-orders")
def create_purchase_order(
    data: PurchaseOrderCreate,
    company_id: Optional[int] = Query(None),
    main_tender_id: Optional[int] = Query(None),
    tender_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
):
    payload = data.model_dump()
    cid = payload.get("company_id") if payload.get("company_id") is not None else company_id
    mid = payload.get("main_tender_id") if payload.get("main_tender_id") is not None else main_tender_id
    tid = payload.get("tender_id") if payload.get("tender_id") is not None else tender_id
    cid, mid, tid = _resolve_context_ids(db, company_id=cid, main_tender_id=mid, tender_id=tid)

    plant_id, plant_name = _resolve_plant(db, payload.get("plant_name"))
    po = PurchaseOrder(
        company_id=cid,
        main_tender_id=mid,
        tender_id=tid,
        po_number=str(payload.get("po_number") or "").strip(),
        po_date=payload.get("po_date"),
        seller_name=payload.get("seller_name"),
        buyer_name=payload.get("buyer_name"),
        buyer_email=payload.get("buyer_email"),
        plant_id=plant_id,
        plant_name=plant_name,
        total_amount=payload.get("total_amount"),
        line_items=payload.get("line_items") or [],
        status=PurchaseOrderStatus.draft,
        source="manual",
        notes=payload.get("notes"),
    )
    db.add(po)
    db.commit()
    db.refresh(po)
    return _purchase_order_to_dict(po)


@app.patch("/api/purchase-orders/{po_id}")
def update_purchase_order(
    po_id: int,
    data: PurchaseOrderUpdate,
    company_id: Optional[int] = None,
    main_tender_id: Optional[int] = None,
    tender_id: Optional[int] = None,
    db: Session = Depends(get_db),
):
    po = db.get(PurchaseOrder, po_id)
    _assert_record_scope(po, company_id, main_tender_id, tender_id, label="Purchase order")

    patch = data.model_dump(exclude_unset=True)
    if "status" in patch and patch["status"] is not None:
        status_raw = str(patch.pop("status")).strip()
        try:
            po.status = PurchaseOrderStatus(status_raw)
        except Exception:
            raise HTTPException(400, f"Invalid status: {status_raw}")

    if "plant_name" in patch:
        plant_id, plant_name = _resolve_plant(db, patch.pop("plant_name"))
        po.plant_id = plant_id
        po.plant_name = plant_name

    for k, v in patch.items():
        setattr(po, k, v)

    db.commit()
    db.refresh(po)
    return _purchase_order_to_dict(po)


@app.delete("/api/purchase-orders/{po_id}")
def delete_purchase_order(
    po_id: int,
    company_id: Optional[int] = None,
    main_tender_id: Optional[int] = None,
    tender_id: Optional[int] = None,
    db: Session = Depends(get_db),
):
    po = db.get(PurchaseOrder, po_id)
    _assert_record_scope(po, company_id, main_tender_id, tender_id, label="Purchase order")
    db.delete(po)
    db.commit()
    return {"ok": True}

# ── PURCHASE & SALES BILLS ────────────────────────────────────────────────

@app.get("/api/purchase-bills")
def list_purchase_bills(status: Optional[str] = None, company_id: Optional[int] = None,
                        main_tender_id: Optional[int] = None, tender_id: Optional[int] = None,
                        db: Session = Depends(get_db)):
    q = select(PurchaseBill).order_by(desc(PurchaseBill.created_at))
    q = _apply_scope_filters(q, PurchaseBill, company_id, main_tender_id, tender_id)
    if status: q = q.where(PurchaseBill.status == PurchaseBillStatus(status))
    pbs = db.execute(q).scalars().all()
    return [{
        "id": pb.id, "pb_number": pb.pb_number,
        "company_id": pb.company_id,
        "main_tender_id": pb.main_tender_id,
        "tender_id": pb.tender_id,
        "broker": pb.broker.name, "broker_id": pb.broker_id,
        "qty_mt": float(pb.qty_mt), "rate_per_mt": float(pb.rate_per_mt),
        "total_amount": float(pb.total_amount),
        "bill_date": str(pb.bill_date), "status": pb.status.value,
        "busy_exported": pb.busy_exported,
    } for pb in pbs]

@app.get("/api/sales-bills")
def list_sales_bills(status: Optional[str] = None, company_id: Optional[int] = None,
                     main_tender_id: Optional[int] = None, db: Session = Depends(get_db)):
    q = select(SalesBill).order_by(desc(SalesBill.created_at))
    q = _apply_scope_filters(q, SalesBill, company_id, main_tender_id)
    if status: q = q.where(SalesBill.status == SalesBillStatus(status))
    sbs = db.execute(q).scalars().all()
    return [{
        "id": sb.id, "sb_number": sb.sb_number,
        "company_id": sb.company_id,
        "main_tender_id": sb.main_tender_id,
        "plant": sb.plant.name, "tender_id": sb.tender_id,
        "qty_mt": float(sb.qty_mt), "rate_per_mt": float(sb.rate_per_mt),
        "total_amount": float(sb.total_amount),
        "bill_date": str(sb.bill_date), "status": sb.status.value,
        "busy_exported": sb.busy_exported,
    } for sb in sbs]

# ── PAYMENTS ───────────────────────────────────────────────────────────────

@app.get("/api/payments")
def list_payments(status: Optional[str] = None, company_id: Optional[int] = None,
                  main_tender_id: Optional[int] = None, tender_id: Optional[int] = None,
                  db: Session = Depends(get_db)):
    q = select(Payment).order_by(desc(Payment.created_at))
    q = _apply_scope_filters(q, Payment, company_id, main_tender_id, tender_id)
    if status: q = q.where(Payment.status == PaymentStatus(status))
    pays = db.execute(q).scalars().all()
    return [{
        "id": p.id, "voucher_number": p.voucher_number,
        "company_id": p.company_id,
        "main_tender_id": p.main_tender_id,
        "tender_id": p.tender_id,
        "broker": p.broker.name, "amount": float(p.amount),
        "payment_date": str(p.payment_date) if p.payment_date else None,
        "payment_mode": p.payment_mode, "reference_no": p.reference_no,
        "status": p.status.value, "busy_exported": p.busy_exported,
    } for p in pays]

@app.post("/api/payments")
def create_payment(data: PaymentCreate, company_id: Optional[int] = None,
                   main_tender_id: Optional[int] = None, tender_id: Optional[int] = None,
                   db: Session = Depends(get_db)):
    pb = db.get(PurchaseBill, data.purchase_bill_id)
    if not pb: raise HTTPException(404, "Purchase bill not found")
    _assert_record_scope(pb, company_id, main_tender_id, tender_id, label="Purchase bill")
    count = db.execute(select(func.count(Payment.id))).scalar() or 0
    p = Payment(
        company_id       = pb.company_id,
        main_tender_id   = pb.main_tender_id,
        tender_id        = pb.tender_id,
        purchase_bill_id = data.purchase_bill_id,
        broker_id        = pb.broker_id,
        voucher_number   = f"PMT-{datetime.now().strftime('%y%m')}-{count+1:03d}",
        amount           = data.amount,
        payment_date     = data.payment_date or date.today(),
        payment_mode     = data.payment_mode,
        reference_no     = data.reference_no,
        notes            = data.notes,
    )
    db.add(p)
    pb.status = PurchaseBillStatus.paid
    db.commit(); db.refresh(p)
    return {"id": p.id, "voucher_number": p.voucher_number}

# ── BUSY EXPORT ────────────────────────────────────────────────────────────

@app.get("/api/admin/reconciliation-violations")
def list_reconciliation_violations(
    company_id: Optional[int] = None,
    main_tender_id: Optional[int] = None,
    tender_id: Optional[int] = None,
    db: Session = Depends(get_db),
):
    dq = db.query(
        Dispatch.bill_id.label("bill_id"),
        func.count(Dispatch.id).label("dispatch_count"),
    ).filter(
        Dispatch.is_deleted == False,
        Dispatch.bill_id != None,
    )
    if company_id is not None:
        dq = dq.filter(Dispatch.company_id == company_id)
    if main_tender_id is not None:
        dq = dq.filter(Dispatch.main_tender_id == main_tender_id)
    if tender_id is not None:
        dq = dq.filter(Dispatch.tender_id == tender_id)
    dispatch_rows = dq.group_by(Dispatch.bill_id).having(func.count(Dispatch.id) > 1).all()

    rq = db.query(
        PlantReceipt.dispatch_id.label("dispatch_id"),
        func.count(PlantReceipt.id).label("receipt_count"),
    ).filter(
        PlantReceipt.is_deleted == False,
        PlantReceipt.dispatch_id != None,
    )
    if company_id is not None:
        rq = rq.filter(PlantReceipt.company_id == company_id)
    if main_tender_id is not None:
        rq = rq.filter(PlantReceipt.main_tender_id == main_tender_id)
    if tender_id is not None:
        rq = rq.filter(PlantReceipt.tender_id == tender_id)
    receipt_rows = rq.group_by(PlantReceipt.dispatch_id).having(func.count(PlantReceipt.id) > 1).all()

    return {
        "dispatch_violation_count": len(dispatch_rows),
        "receipt_violation_count": len(receipt_rows),
        "dispatch_violations": [
            {"bill_id": int(r.bill_id), "dispatch_count": int(r.dispatch_count)}
            for r in dispatch_rows
        ],
        "receipt_violations": [
            {"dispatch_id": int(r.dispatch_id), "receipt_count": int(r.receipt_count)}
            for r in receipt_rows
        ],
    }


@app.get("/api/busy/party-mappings")
def list_busy_party_mappings(
    company_id: Optional[int] = None,
    q_search: Optional[str] = Query(None, alias="q"),
    db: Session = Depends(get_db),
):
    if company_id is None:
        raise HTTPException(400, "company_id is required for busy party mappings")

    q = select(BusyPartyMapping).where(BusyPartyMapping.company_id == company_id).order_by(BusyPartyMapping.source_party_name)
    rows = db.execute(q).scalars().all()

    out = [
        {
            "id": r.id,
            "company_id": r.company_id,
            "source_party_name": r.source_party_name,
            "busy_party_name": r.busy_party_name,
            "sale_purc_type_override": r.sale_purc_type_override,
            "notes": r.notes,
            "created_at": str(r.created_at),
            "updated_at": str(r.updated_at) if r.updated_at else None,
        }
        for r in rows
    ]
    if q_search:
        qs = q_search.lower().strip()
        out = [x for x in out if qs in str(x.get("source_party_name") or "").lower() or qs in str(x.get("busy_party_name") or "").lower()]
    return out


@app.get("/api/busy/master-parties")
def list_busy_master_parties(
    company_id: Optional[int] = None,
    q_search: Optional[str] = Query(None, alias="q"),
    limit: int = Query(250, ge=1, le=5000),
    db: Session = Depends(get_db),
):
    q = select(BusyPartyMaster).where(BusyPartyMaster.is_active == True)
    if company_id is not None:
        q = q.where(or_(BusyPartyMaster.company_id == company_id, BusyPartyMaster.company_id == None))
    else:
        q = q.where(BusyPartyMaster.company_id == None)

    rows = db.execute(q).scalars().all()
    if company_id is not None:
        rows.sort(key=lambda r: (0 if r.company_id == company_id else 1, _norm_party_name(r.busy_party_name)))
    else:
        rows.sort(key=lambda r: _norm_party_name(r.busy_party_name))

    if q_search:
        needle = _norm_party_name(q_search)
        rows = [
            r for r in rows
            if needle in _norm_party_name(r.busy_party_name)
            or needle in _norm_party_name(r.alias)
            or needle in _norm_party_name(r.parent_group)
            or needle in _norm_party_name(r.gstin)
        ]

    rows = rows[: max(1, int(limit))]
    return [
        {
            "id": r.id,
            "company_id": r.company_id,
            "busy_party_name": r.busy_party_name,
            "alias": r.alias,
            "parent_group": r.parent_group,
            "dealer_type": r.dealer_type,
            "gstin": r.gstin,
            "filing_frequency": r.filing_frequency,
            "state_code": r.state_code,
            "state_name": r.state_name,
            "station": r.station,
            "source_file": r.source_file,
            "updated_at": str(r.updated_at) if r.updated_at else None,
        }
        for r in rows
    ]


@app.post("/api/busy/master-parties/import")
async def import_busy_master_parties(
    file: UploadFile = File(...),
    company_id: Optional[int] = Query(None),
    replace_existing: bool = Query(False),
    db: Session = Depends(get_db),
):
    if company_id is not None and not db.get(Company, company_id):
        raise HTTPException(404, "Company not found")

    raw = await file.read()
    rows = _read_busy_party_master_rows(raw, source_file=file.filename)
    if not rows:
        raise HTTPException(400, "No party rows found in uploaded Busy file")

    scope_clause = BusyPartyMaster.company_id.is_(None) if company_id is None else (BusyPartyMaster.company_id == company_id)
    existing_rows = db.execute(select(BusyPartyMaster).where(scope_clause)).scalars().all()
    existing_by_key = {
        (str(r.name_normalized or "").strip(), _clean_gstin(r.gstin)): r
        for r in existing_rows
        if str(r.name_normalized or "").strip()
    }

    if replace_existing:
        for r in existing_rows:
            r.is_active = False

    inserted = 0
    updated = 0
    reactivated = 0

    for item in rows:
        key = (str(item.get("name_normalized") or "").strip(), _clean_gstin(item.get("gstin")))
        row = existing_by_key.get(key)
        if row:
            was_inactive = not bool(row.is_active)
            row.busy_party_name = item.get("busy_party_name")
            row.alias = item.get("alias")
            row.parent_group = item.get("parent_group")
            row.dealer_type = item.get("dealer_type")
            row.gstin = item.get("gstin")
            row.filing_frequency = item.get("filing_frequency")
            row.state_code = item.get("state_code")
            row.state_name = item.get("state_name")
            row.station = item.get("station")
            row.name_normalized = item.get("name_normalized")
            row.source_file = item.get("source_file")
            row.is_active = True
            updated += 1
            if was_inactive:
                reactivated += 1
            continue

        row = BusyPartyMaster(
            company_id=company_id,
            busy_party_name=item.get("busy_party_name"),
            alias=item.get("alias"),
            parent_group=item.get("parent_group"),
            dealer_type=item.get("dealer_type"),
            gstin=item.get("gstin"),
            filing_frequency=item.get("filing_frequency"),
            state_code=item.get("state_code"),
            state_name=item.get("state_name"),
            station=item.get("station"),
            name_normalized=item.get("name_normalized"),
            source_file=item.get("source_file"),
            is_active=True,
        )
        db.add(row)
        inserted += 1

    db.commit()
    active_count = db.execute(
        select(func.count(BusyPartyMaster.id)).where(
            scope_clause,
            BusyPartyMaster.is_active == True,
        )
    ).scalar() or 0

    return {
        "company_id": company_id,
        "replace_existing": bool(replace_existing),
        "file_name": file.filename,
        "parsed_rows": len(rows),
        "inserted": inserted,
        "updated": updated,
        "reactivated": reactivated,
        "active_total": int(active_count),
    }


@app.post("/api/busy/party-mappings/upsert")
def upsert_busy_party_mapping(
    data: BusyPartyMappingUpsert,
    company_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
):
    if company_id is None:
        raise HTTPException(400, "company_id is required for busy party mappings")
    if not db.get(Company, company_id):
        raise HTTPException(404, "Company not found")

    source_party = str(data.source_party_name or "").strip()
    busy_party = str(data.busy_party_name or "").strip()
    if not source_party:
        raise HTTPException(400, "source_party_name is required")
    if not busy_party:
        raise HTTPException(400, "busy_party_name is required")

    override = _normalize_sale_purc_override(data.sale_purc_type_override)
    row = _find_busy_party_mapping(db, company_id, source_party)
    if row and row.company_id == company_id and _norm_party_name(row.source_party_name) == _norm_party_name(source_party):
        row.source_party_name = source_party
        row.busy_party_name = busy_party
        row.sale_purc_type_override = override
        row.notes = data.notes
    else:
        row = BusyPartyMapping(
            company_id=company_id,
            source_party_name=source_party,
            busy_party_name=busy_party,
            sale_purc_type_override=override,
            notes=data.notes,
        )
        db.add(row)

    db.commit()
    db.refresh(row)
    return {
        "id": row.id,
        "company_id": row.company_id,
        "source_party_name": row.source_party_name,
        "busy_party_name": row.busy_party_name,
        "sale_purc_type_override": row.sale_purc_type_override,
        "notes": row.notes,
        "updated_at": str(row.updated_at) if row.updated_at else None,
    }


@app.get("/api/busy/purchase-bills/ready")
def list_busy_ready_purchase_bills(
    include_exported: bool = Query(True),
    include_rejected: bool = Query(False),
    company_id: Optional[int] = None,
    main_tender_id: Optional[int] = None,
    tender_id: Optional[int] = None,
    db: Session = Depends(get_db),
):
    rows, rejected = _collect_busy_ready_bill_rows(
        db,
        company_id=company_id,
        main_tender_id=main_tender_id,
        tender_id=tender_id,
        include_exported=include_exported,
    )
    if include_rejected:
        return {"rows": rows, "rejected": rejected}
    return rows


@app.post("/api/export/purchase-bills/busy-xlsx")
def export_purchase_bills_busy_xlsx(
    data: IdListPayload,
    company_id: Optional[int] = Query(None),
    main_tender_id: Optional[int] = Query(None),
    tender_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
):
    ids = []
    seen = set()
    for x in data.ids or []:
        if x in seen:
            continue
        seen.add(x)
        ids.append(int(x))

    if not ids:
        raise HTTPException(400, "ids are required")

    rows, rejected = _collect_busy_ready_bill_rows(
        db,
        company_id=company_id,
        main_tender_id=main_tender_id,
        tender_id=tender_id,
        include_exported=True,
        only_bill_ids=set(ids),
    )
    row_map = {int(r["bill_id"]): r for r in rows}
    rejected_map = {int(r["bill_id"]): r["reason"] for r in rejected}

    invalid = []
    for bill_id in ids:
        if bill_id not in row_map:
            invalid.append({
                "bill_id": bill_id,
                "reason": rejected_map.get(bill_id, "not_ready_or_out_of_scope"),
            })
    if invalid:
        raise HTTPException(400, {
            "message": "Some selected bills are not eligible for Busy export under strict 1:1 reconciliation rule",
            "invalid": invalid,
        })

    from services.busy_export import export_busy_purchase_bills_xlsx, save_export_bytes

    export_rows = []
    for bill_id in ids:
        r = row_map[bill_id]
        export_rows.append({
            "vch_series": r.get("vch_series", "Main"),
            "vch_bill_date": r.get("vch_bill_date"),
            "vch_bill_no": r.get("vch_bill_no"),
            "sale_purc_type": r.get("sale_purc_type"),
            "party_name": r.get("party_name"),
            "mc_name": r.get("mc_name", "Main Store"),
            "item_name": r.get("item_name"),
            "quantity": r.get("quantity"),
            "unit": r.get("unit", "QUINTAL"),
            "price": r.get("price_value"),
            "itc_eligibility_type": r.get("itc_eligibility_type", "Input Goods/Services"),
            "narration": r.get("narration"),
        })

    settings = get_settings()
    content = export_busy_purchase_bills_xlsx(export_rows)
    fpath = save_export_bytes(content, "purchase_bills_busy", settings, ext="xlsx")

    selected_rows = [row_map[bill_id] for bill_id in ids]
    lineage_records = [_busy_export_lineage_from_row(r) for r in selected_rows]

    pb_ids = [int(r.get("purchase_bill_id")) for r in rows if r.get("purchase_bill_id")]
    if pb_ids:
        q = select(PurchaseBill).where(PurchaseBill.id.in_(pb_ids))
        q = _apply_scope_filters(q, PurchaseBill, company_id, main_tender_id, tender_id)
        pbs = db.execute(q).scalars().all()
        for pb in pbs:
            pb.busy_exported = True
            pb.busy_export_at = datetime.now()

    company_ids = {r.get("company_id") for r in selected_rows if r.get("company_id") is not None}
    main_tender_ids = {r.get("main_tender_id") for r in selected_rows if r.get("main_tender_id") is not None}
    summary_company_id = next(iter(company_ids)) if len(company_ids) == 1 else None
    summary_main_tender_id = next(iter(main_tender_ids)) if len(main_tender_ids) == 1 else None

    exp = BusyExport(
        export_type="purchase_bill",
        record_ids=lineage_records,
        file_path=fpath,
        company_id=summary_company_id,
        main_tender_id=summary_main_tender_id,
    )
    db.add(exp)
    db.commit()

    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=busy_purchase_bills.xlsx"},
    )


@app.post("/api/export/busy-staging/busy-xlsx")
def export_busy_staging_bills_xlsx(
    data: IdListPayload,
    company_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
):
    ids = []
    seen = set()
    for x in data.ids or []:
        if x in seen:
            continue
        seen.add(x)
        ids.append(int(x))

    if not ids:
        raise HTTPException(400, "ids are required")

    rows, _rejected = _collect_busy_staging_rows(
        db,
        company_id=company_id,
        include_exported=True,
        only_ids=set(ids),
    )
    row_map = {int(r["id"]): r for r in rows}

    invalid = []
    for staging_id in ids:
        if staging_id not in row_map:
            invalid.append({
                "busy_staging_bill_id": staging_id,
                "reason": "not_found_or_out_of_scope",
            })
    if invalid:
        raise HTTPException(400, {"message": "Some selected bills are not available for export", "invalid": invalid})

    from services.busy_export import export_busy_purchase_bills_xlsx, save_export_bytes

    export_rows = []
    for staging_id in ids:
        r = row_map[staging_id]
        export_rows.append({
            "vch_series": r.get("vch_series", "Main"),
            "vch_bill_date": r.get("vch_bill_date"),
            "vch_bill_no": r.get("vch_bill_no"),
            "sale_purc_type": r.get("sale_purc_type"),
            "party_name": r.get("party_name"),
            "mc_name": r.get("mc_name", "Main Store"),
            "item_name": r.get("item_name"),
            "quantity": r.get("quantity"),
            "unit": r.get("unit", "QUINTAL"),
            "price": r.get("price_value"),
            "itc_eligibility_type": r.get("itc_eligibility_type", "Input Goods/Services"),
            "narration": r.get("narration"),
        })

    settings = get_settings()
    content = export_busy_purchase_bills_xlsx(export_rows)
    fpath = save_export_bytes(content, "busy_staging_bills", settings, ext="xlsx")

    q = select(BusyStagingBill).where(BusyStagingBill.id.in_(ids))
    q = _apply_scope_filters(q, BusyStagingBill, company_id)
    staging_rows = db.execute(q).scalars().all()
    for b in staging_rows:
        b.busy_exported = True
        b.busy_exported_at = datetime.now()

    company_ids = {r.get("company_id") for r in rows if r.get("company_id") is not None}
    summary_company_id = next(iter(company_ids)) if len(company_ids) == 1 else None

    db.commit()

    try:
        exp = BusyExport(
            export_type="busy_staging_bill",
            record_ids=[{"busy_staging_bill_id": staging_id} for staging_id in ids],
            file_path=fpath,
            company_id=summary_company_id,
        )
        db.add(exp)
        db.commit()
    except Exception as e:
        db.rollback()
        log.warning("Busy staging export logging failed: %s", e)

    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=busy_staging_bills.xlsx"},
    )

@app.post("/api/export/purchase-bills")
def export_purchase_bills_busy(ids: List[int], company_id: Optional[int] = Query(None),
                               main_tender_id: Optional[int] = Query(None),
                               db: Session = Depends(get_db)):
    from services.busy_export import export_purchase_bills, save_export
    settings = get_settings()
    q = select(PurchaseBill).where(PurchaseBill.id.in_(ids))
    q = _apply_scope_filters(q, PurchaseBill, company_id, main_tender_id)
    pbs = db.execute(q).scalars().all()
    content = export_purchase_bills(pbs, settings)
    fpath = save_export(content, "purchase_bills", settings)
    for pb in pbs:
        pb.busy_exported = True
        pb.busy_export_at = datetime.now()
    # Log export
    exp = BusyExport(export_type="purchase_bill", record_ids=ids, file_path=fpath,
                     company_id=company_id, main_tender_id=main_tender_id)
    db.add(exp); db.commit()
    return Response(content=content, media_type="text/csv",
                    headers={"Content-Disposition": f"attachment; filename=purchase_bills.csv"})


@app.get("/api/export/purchase-bills")
def export_purchase_bills_busy_get(ids: Optional[str] = Query(None), company_id: Optional[int] = Query(None),
                                   main_tender_id: Optional[int] = Query(None),
                                   db: Session = Depends(get_db)):
    if not ids:
        raise HTTPException(400, "ids query parameter required, e.g. ?ids=1,2,3")
    id_list = [int(x) for x in ids.split(',') if x.strip()]
    from services.busy_export import export_purchase_bills, save_export
    settings = get_settings()
    q = select(PurchaseBill).where(PurchaseBill.id.in_(id_list))
    q = _apply_scope_filters(q, PurchaseBill, company_id, main_tender_id)
    pbs = db.execute(q).scalars().all()
    content = export_purchase_bills(pbs, settings)
    fpath = save_export(content, "purchase_bills", settings)
    for pb in pbs:
        pb.busy_exported = True
        pb.busy_export_at = datetime.now()
    exp = BusyExport(export_type="purchase_bill", record_ids=id_list, file_path=fpath,
                     company_id=company_id, main_tender_id=main_tender_id)
    db.add(exp); db.commit()
    return Response(content=content, media_type="text/csv",
                    headers={"Content-Disposition": f"attachment; filename=purchase_bills.csv"})

@app.post("/api/export/sales-bills")
def export_sales_bills_busy(ids: List[int], company_id: Optional[int] = Query(None),
                            main_tender_id: Optional[int] = Query(None),
                            db: Session = Depends(get_db)):
    from services.busy_export import export_sales_bills, save_export
    settings = get_settings()
    q = select(SalesBill).where(SalesBill.id.in_(ids))
    q = _apply_scope_filters(q, SalesBill, company_id, main_tender_id)
    sbs = db.execute(q).scalars().all()
    content = export_sales_bills(sbs, settings)
    fpath = save_export(content, "sales_bills", settings)
    for sb in sbs:
        sb.busy_exported = True
        sb.busy_export_at = datetime.now()
    exp = BusyExport(export_type="sales_bill", record_ids=ids, file_path=fpath,
                     company_id=company_id, main_tender_id=main_tender_id)
    db.add(exp); db.commit()
    return Response(content=content, media_type="text/csv",
                    headers={"Content-Disposition": f"attachment; filename=sales_bills.csv"})


@app.get("/api/export/sales-bills")
def export_sales_bills_busy_get(ids: Optional[str] = Query(None), company_id: Optional[int] = Query(None),
                                main_tender_id: Optional[int] = Query(None),
                                db: Session = Depends(get_db)):
    if not ids:
        raise HTTPException(400, "ids query parameter required, e.g. ?ids=1,2,3")
    id_list = [int(x) for x in ids.split(',') if x.strip()]
    from services.busy_export import export_sales_bills, save_export
    settings = get_settings()
    q = select(SalesBill).where(SalesBill.id.in_(id_list))
    q = _apply_scope_filters(q, SalesBill, company_id, main_tender_id)
    sbs = db.execute(q).scalars().all()
    content = export_sales_bills(sbs, settings)
    fpath = save_export(content, "sales_bills", settings)
    for sb in sbs:
        sb.busy_exported = True
        sb.busy_export_at = datetime.now()
    exp = BusyExport(export_type="sales_bill", record_ids=id_list, file_path=fpath,
                     company_id=company_id, main_tender_id=main_tender_id)
    db.add(exp); db.commit()
    return Response(content=content, media_type="text/csv",
                    headers={"Content-Disposition": f"attachment; filename=sales_bills.csv"})

@app.post("/api/export/payments")
def export_payments_busy(ids: List[int], company_id: Optional[int] = Query(None),
                         main_tender_id: Optional[int] = Query(None),
                         db: Session = Depends(get_db)):
    from services.busy_export import export_payment_vouchers, save_export
    settings = get_settings()
    q = select(Payment).where(Payment.id.in_(ids))
    q = _apply_scope_filters(q, Payment, company_id, main_tender_id)
    pays = db.execute(q).scalars().all()
    content = export_payment_vouchers(pays, settings)
    fpath = save_export(content, "payment_vouchers", settings)
    for p in pays:
        p.busy_exported = True
        p.busy_export_at = datetime.now()
    exp = BusyExport(export_type="payment_voucher", record_ids=ids, file_path=fpath,
                     company_id=company_id, main_tender_id=main_tender_id)
    db.add(exp); db.commit()
    return Response(content=content, media_type="text/csv",
                    headers={"Content-Disposition": f"attachment; filename=payment_vouchers.csv"})


@app.get("/api/export/payments")
def export_payments_busy_get(ids: Optional[str] = Query(None), company_id: Optional[int] = Query(None),
                             main_tender_id: Optional[int] = Query(None),
                             db: Session = Depends(get_db)):
    if not ids:
        raise HTTPException(400, "ids query parameter required, e.g. ?ids=1,2,3")
    id_list = [int(x) for x in ids.split(',') if x.strip()]
    from services.busy_export import export_payment_vouchers, save_export
    settings = get_settings()
    q = select(Payment).where(Payment.id.in_(id_list))
    q = _apply_scope_filters(q, Payment, company_id, main_tender_id)
    pays = db.execute(q).scalars().all()
    content = export_payment_vouchers(pays, settings)
    fpath = save_export(content, "payment_vouchers", settings)
    for p in pays:
        p.busy_exported = True
        p.busy_export_at = datetime.now()
    exp = BusyExport(export_type="payment_voucher", record_ids=id_list, file_path=fpath,
                     company_id=company_id, main_tender_id=main_tender_id)
    db.add(exp); db.commit()
    return Response(content=content, media_type="text/csv",
                    headers={"Content-Disposition": f"attachment; filename=payment_vouchers.csv"})

# ── MARKET PRICES ──────────────────────────────────────────────────────────

@app.get("/api/market-prices")
def list_market_prices(material_id: Optional[int] = None, days: int = 30, company_id: Optional[int] = None,
                       db: Session = Depends(get_db)):
    from datetime import timedelta
    cutoff = date.today() - timedelta(days=days)
    q = select(MarketPrice).where(MarketPrice.price_date >= cutoff).order_by(MarketPrice.price_date)
    if company_id is not None:
        q = q.where(MarketPrice.company_id == company_id)
    if material_id: q = q.where(MarketPrice.material_id == material_id)
    prices = db.execute(q).scalars().all()
    return [{
        "id": p.id, "material": p.material.name, "material_id": p.material_id,
        "price_date": str(p.price_date), "price_per_mt": float(p.price_per_mt),
        "market": p.market, "source": p.source,
    } for p in prices]

@app.post("/api/market-prices")
def add_market_price(data: MarketPriceCreate, company_id: Optional[int] = Query(None), db: Session = Depends(get_db)):
    payload = data.model_dump()
    payload["company_id"] = company_id
    p = MarketPrice(**payload)
    db.add(p); db.commit(); db.refresh(p)
    return {"id": p.id}

# ── REFERENCE DATA ─────────────────────────────────────────────────────────

@app.get("/api/plants")
def get_plants(db: Session = Depends(get_db)):
    return [{"id": p.id, "name": p.name, "code": p.code}
            for p in db.execute(select(Plant).where(Plant.is_active)).scalars().all()]

@app.get("/api/materials")
def get_materials(db: Session = Depends(get_db)):
    return [{"id": m.id, "name": m.name, "code": m.code}
            for m in db.execute(select(Material).where(Material.is_active)).scalars().all()]

@app.get("/api/brokers")
def get_brokers(db: Session = Depends(get_db)):
    return [{"id": b.id, "name": b.name, "phone": b.phone, "telegram_chat_id": b.telegram_chat_id}
            for b in db.execute(select(Broker).where(Broker.is_active)).scalars().all()]

@app.post("/api/brokers")
def create_broker(data: BrokerCreate, db: Session = Depends(get_db)):
    b = Broker(**data.model_dump())
    db.add(b); db.commit(); db.refresh(b)
    return {"id": b.id, "name": b.name}

# ── REPORTS ────────────────────────────────────────────────────────────────

@app.get("/api/reports/penalty-risk")
def penalty_risk_report(cycle_id: Optional[int] = None, company_id: Optional[int] = None,
                        main_tender_id: Optional[int] = None, db: Session = Depends(get_db)):
    if not cycle_id:
        cycle = db.execute(select(SproxxCycle).where(SproxxCycle.is_active)).scalar_one_or_none()
        cycle_id = cycle.id if cycle else None
    q = select(Tender).where(Tender.cycle_id == cycle_id) if cycle_id else select(Tender)
    q = _apply_scope_filters(q, Tender, company_id, main_tender_id)
    tenders = db.execute(q).scalars().all()
    at_risk = []
    for t in tenders:
        d = _tender_to_dict(t)
        if d["summary"]["week1_pct"] < 100:
            shortfall = float(t.week1_target_mt) - d["summary"]["total_accepted_mt"]
            penalty   = shortfall * float(t.penalty_pct) / 100
            at_risk.append({
                "tender_number": t.tender_number,
                "plant": t.plant.name, "material": t.material.name,
                "week1_target_mt": float(t.week1_target_mt),
                "accepted_mt": d["summary"]["total_accepted_mt"],
                "shortfall_mt": round(max(shortfall, 0), 2),
                "week1_pct": d["summary"]["week1_pct"],
                "week1_deadline": str(t.week1_deadline),
                "estimated_penalty": round(max(penalty, 0), 2),
            })
    return sorted(at_risk, key=lambda x: x["week1_pct"])

@app.get("/api/reports/broker-performance")
def broker_performance(company_id: Optional[int] = None, main_tender_id: Optional[int] = None,
                       db: Session = Depends(get_db)):
    brokers = db.execute(select(Broker).where(Broker.is_active)).scalars().all()
    result = []
    for b in brokers:
        dq = select(Deal).where(Deal.broker_id == b.id)
        dq = _apply_scope_filters(dq, Deal, company_id, main_tender_id)
        deals = db.execute(dq).scalars().all()
        if not deals: continue
        total_deal = sum(float(d.deal_mt) for d in deals)
        total_acc  = sum(float(d.accepted_mt or 0) for d in deals)
        total_rej  = sum(float(d.rejected_mt or 0) for d in deals)
        result.append({
            "broker": b.name, "deals": len(deals),
            "deal_mt": round(total_deal, 2),
            "accepted_mt": round(total_acc, 2),
            "rejected_mt": round(total_rej, 2),
            "fulfillment_pct": round(total_acc / total_deal * 100, 1) if total_deal else 0,
            "rejection_rate": round(total_rej / (total_acc + total_rej) * 100, 1) if (total_acc + total_rej) else 0,
        })
    return sorted(result, key=lambda x: x["fulfillment_pct"], reverse=True)


# ── CRUD: TENDERS ──────────────────────────────────────────────────────────

@app.put("/api/tenders/{tender_id}")
def update_tender(tender_id: int, data: TenderCreate, company_id: Optional[int] = None,
                  main_tender_id: Optional[int] = None, db: Session = Depends(get_db)):
    t = db.get(Tender, tender_id)
    _assert_record_scope(t, company_id, main_tender_id, label="Tender")
    for k, v in data.model_dump().items():
        setattr(t, k, v)
    db.commit()
    return _tender_to_dict(t)

@app.delete("/api/tenders/{tender_id}")
def delete_tender(tender_id: int, company_id: Optional[int] = None,
                  main_tender_id: Optional[int] = None, db: Session = Depends(get_db)):
    t = db.get(Tender, tender_id)
    _assert_record_scope(t, company_id, main_tender_id, label="Tender")
    # cancel deals first
    for d in t.deals:
        d.status = DealStatus.cancelled
    t.status = TenderStatus.cancelled
    db.commit()
    return {"ok": True}

# ── CRUD: DEALS ────────────────────────────────────────────────────────────

class DealUpdate(BaseModel):
    tender_id:   Optional[int]   = None
    broker_id:   Optional[int]   = None
    material_id: Optional[int]   = None
    deal_mt:     Optional[float] = None
    rate_per_mt: Optional[float] = None
    notes:       Optional[str]   = None
    status:      Optional[str]   = None

@app.patch("/api/deals/{deal_id}")
def update_deal(deal_id: int, data: DealUpdate, company_id: Optional[int] = None,
                main_tender_id: Optional[int] = None, tender_id: Optional[int] = None,
                db: Session = Depends(get_db)):
    d = db.get(Deal, deal_id)
    _assert_record_scope(d, company_id, main_tender_id, tender_id, label="Deal")
    patch = data.model_dump(exclude_unset=True)

    if patch.get("tender_id") is not None:
        tc, tm, tid = _lineage_from_tender(db, int(patch["tender_id"]))
        if tid is None:
            raise HTTPException(404, "Tender not found")
        if company_id is not None and tc != company_id:
            raise HTTPException(400, "Selected tender is outside selected company scope")
        if main_tender_id is not None and tm != main_tender_id:
            raise HTTPException(400, "Selected tender is outside selected main tender scope")
        if tender_id is not None and tid != tender_id:
            raise HTTPException(400, "Selected tender is outside selected tender scope")

    for k, v in patch.items():
        if v is not None:
            if k == "status": setattr(d, k, DealStatus(v))
            else: setattr(d, k, v)

    lc, lm, _ = _lineage_from_tender(db, d.tender_id)
    if d.tender_id and (lc is None and lm is None):
        raise HTTPException(404, "Tender not found")
    d.company_id = lc
    d.main_tender_id = lm
    db.commit()
    return {"id": d.id, "deal_number": d.deal_number, "tender_id": d.tender_id, "status": d.status.value}

@app.delete("/api/deals/{deal_id}")
def delete_deal(deal_id: int, company_id: Optional[int] = None,
                main_tender_id: Optional[int] = None, tender_id: Optional[int] = None,
                db: Session = Depends(get_db)):
    d = db.get(Deal, deal_id)
    _assert_record_scope(d, company_id, main_tender_id, tender_id, label="Deal")
    d.status = DealStatus.cancelled
    db.commit()
    return {"ok": True}

# ── CRUD: BILLS ────────────────────────────────────────────────────────────

@app.delete("/api/bills/{bill_id}")
def delete_bill(bill_id: int, company_id: Optional[int] = None,
                main_tender_id: Optional[int] = None, tender_id: Optional[int] = None,
                db: Session = Depends(get_db)):
    b = db.get(Bill, bill_id)
    _assert_record_scope(b, company_id, main_tender_id, tender_id, label="Bill")
    # Delete image file if exists
    if b.image_path:
        settings = get_settings()
        fpath = Path(settings.upload_dir) / "bills" / b.image_path
        if fpath.exists():
            fpath.unlink()
    # Handle related dispatch: soft-delete and adjust deal quantities where possible
    dispatches = _active_dispatches_for_bill(db, b.id)
    if len(dispatches) > 1:
        raise HTTPException(
            409,
            "Cannot delete bill: multiple active dispatches found for this bill. Resolve 1:1 violation first.",
        )
    disp = dispatches[0] if dispatches else None
    affected_deal_ids = set()
    if disp:
        # cannot delete dispatch if it has a receipt; require deleting receipt first
        if disp.receipt:
            raise HTTPException(400, "Cannot delete bill: linked dispatch has a receipt. Delete receipt first.")
        affected_deal_ids.add(disp.deal_id)
        # soft-delete dispatch
        disp.is_deleted = True
        disp.deleted_at = datetime.now()

    # Remove any PurchaseBill tied to this bill
    pb = db.execute(select(PurchaseBill).where(PurchaseBill.bill_id == b.id)).scalar_one_or_none()
    if pb:
        db.delete(pb)

    # Finally delete the Bill record
    db.delete(b)
    db.flush()
    _reconcile_unloading_match_for_all_bills(db)
    _recalculate_deal_dispatched_mt(db, list(affected_deal_ids))
    db.commit()
    return {"ok": True}

# ── CRUD: BROKERS ──────────────────────────────────────────────────────────

class BrokerUpdate(BaseModel):
    name:    Optional[str] = None
    phone:   Optional[str] = None
    gstin:   Optional[str] = None
    address: Optional[str] = None
    telegram_chat_id: Optional[str] = None

@app.patch("/api/brokers/{broker_id}")
def update_broker(broker_id: int, data: BrokerUpdate, db: Session = Depends(get_db)):
    b = db.get(Broker, broker_id)
    if not b: raise HTTPException(404)
    for k, v in data.model_dump().items():
        if v is not None: setattr(b, k, v)
    db.commit()
    return {"id": b.id, "name": b.name}

@app.delete("/api/brokers/{broker_id}")
def delete_broker(broker_id: int, db: Session = Depends(get_db)):
    b = db.get(Broker, broker_id)
    if not b: raise HTTPException(404)
    b.is_active = False
    db.commit()
    return {"ok": True}

# ── CRUD: MARKET PRICES ────────────────────────────────────────────────────

@app.delete("/api/market-prices/{price_id}")
def delete_market_price(price_id: int, company_id: Optional[int] = None, db: Session = Depends(get_db)):
    p = db.get(MarketPrice, price_id)
    _assert_record_scope(p, company_id, label="Market price")
    db.delete(p)
    db.commit()
    return {"ok": True}

@app.patch("/api/market-prices/{price_id}")
def update_market_price(price_id: int, data: MarketPriceCreate, company_id: Optional[int] = None,
                        db: Session = Depends(get_db)):
    p = db.get(MarketPrice, price_id)
    _assert_record_scope(p, company_id, label="Market price")
    for k, v in data.model_dump().items():
        if v is not None: setattr(p, k, v)
    db.commit()
    return {"id": p.id}

# ── CRUD: PURCHASE BILLS ───────────────────────────────────────────────────

@app.delete("/api/purchase-bills/{pb_id}")
def delete_purchase_bill(pb_id: int, company_id: Optional[int] = None,
                         main_tender_id: Optional[int] = None, tender_id: Optional[int] = None,
                         db: Session = Depends(get_db)):
    pb = db.get(PurchaseBill, pb_id)
    _assert_record_scope(pb, company_id, main_tender_id, tender_id, label="Purchase bill")
    if pb.status == PurchaseBillStatus.paid:
        raise HTTPException(400, "Cannot delete a paid bill")
    pb.status = PurchaseBillStatus.cancelled
    db.commit()
    return {"ok": True}

# ── CRUD: PAYMENTS ─────────────────────────────────────────────────────────

@app.delete("/api/payments/{payment_id}")
def delete_payment(payment_id: int, company_id: Optional[int] = None,
                   main_tender_id: Optional[int] = None, tender_id: Optional[int] = None,
                   db: Session = Depends(get_db)):
    p = db.get(Payment, payment_id)
    _assert_record_scope(p, company_id, main_tender_id, tender_id, label="Payment")
    # Revert purchase bill status
    pb = db.get(PurchaseBill, p.purchase_bill_id)
    if pb: pb.status = PurchaseBillStatus.confirmed
    db.delete(p)
    db.commit()
    return {"ok": True}

# ── WHATSAPP INTEGRATION ROUTES ─────────────────────────────────────────────

WHATSAPP_BRIDGE_URL = os.getenv("WHATSAPP_BRIDGE_URL", "http://127.0.0.1:3001")


class WhatsAppWhitelistUpdate(BaseModel):
    jids: List[str] = []
    monitored_groups: Optional[List[dict]] = None
    auto_reply: Optional[bool] = True
    is_enabled: Optional[bool] = True


class WhatsAppContactEdit(BaseModel):
    jid: str
    name: Optional[str] = None
    is_monitored: Optional[bool] = None


class WhatsAppBatchDeleteLogs(BaseModel):
    ids: List[int] = []


class WhatsAppSendRequest(BaseModel):
    jid: str
    text: str


@app.get("/api/whatsapp/status")
async def get_whatsapp_status(db: Session = Depends(get_db)):
    """Fetch live status of the WhatsApp bridge and database configuration."""
    bridge_data = {"connected": False, "status": "disconnected", "user": None, "has_qr": False}
    try:
        async with httpx.AsyncClient(timeout=3.0) as client:
            r = await client.get(f"{WHATSAPP_BRIDGE_URL}/status")
            if r.status_code == 200:
                bridge_data = r.json()
    except Exception as e:
        log.debug("WhatsApp bridge offline or unreachable: %s", e)

    config = db.execute(select(WhatsAppConfig).order_by(WhatsAppConfig.id.desc())).scalar_one_or_none()

    return {
        "success": True,
        "bridge_online": bridge_data.get("status") != "disconnected" or bridge_data.get("connected", False),
        "status": bridge_data.get("status", "disconnected"),
        "connected": bridge_data.get("connected", False),
        "user": bridge_data.get("user"),
        "has_qr": bridge_data.get("has_qr", False),
        "config": {
            "is_enabled": config.is_enabled if config else True,
            "auto_reply": config.auto_reply if config else True,
            "whitelisted_jids": config.whitelisted_jids if config else [],
            "monitored_groups": config.monitored_groups if config else []
        }
    }


@app.get("/api/whatsapp/qr")
async def get_whatsapp_qr():
    """Proxy live QR code from the WhatsApp bridge for Web UI pairing."""
    try:
        async with httpx.AsyncClient(timeout=4.0) as client:
            r = await client.get(f"{WHATSAPP_BRIDGE_URL}/qr")
            return r.json()
    except Exception as e:
        return {"success": False, "connected": False, "error": f"WhatsApp bridge not reachable: {e}"}


@app.get("/api/whatsapp/groups")
async def get_whatsapp_groups(db: Session = Depends(get_db)):
    """Fetch synced participating groups and contacts from WhatsApp and merge with database whitelist state."""
    bridge_groups = []
    try:
        async with httpx.AsyncClient(timeout=6.0) as client:
            r = await client.get(f"{WHATSAPP_BRIDGE_URL}/groups")
            if r.status_code == 200:
                bridge_groups = r.json().get("groups", [])
    except Exception as e:
        log.warning("Could not fetch groups from WhatsApp bridge: %s", e)

    config = db.execute(select(WhatsAppConfig).order_by(WhatsAppConfig.id.desc())).scalar_one_or_none()
    whitelisted_set = set(config.whitelisted_jids or []) if config else set()
    saved_monitored = config.monitored_groups or [] if config else []

    chat_map = {}
    for g in bridge_groups:
        jid = g.get("id")
        if jid and not jid.endswith('@newsletter') and not jid.endswith('@broadcast'):
            chat_map[jid] = {
                "id": jid,
                "subject": g.get("subject") or g.get("name") or jid.split("@")[0],
                "name": g.get("name") or g.get("subject") or jid.split("@")[0],
                "is_group": g.get("is_group", jid.endswith("@g.us")),
                "is_monitored": jid in whitelisted_set
            }

    for item in saved_monitored:
        jid = item.get("id") or item.get("jid")
        if jid and not jid.endswith('@newsletter') and not jid.endswith('@broadcast'):
            subj = item.get("subject") or item.get("name") or jid.split("@")[0]
            if jid in chat_map:
                if subj and subj != jid.split("@")[0]:
                    chat_map[jid]["subject"] = subj
                    chat_map[jid]["name"] = subj
                chat_map[jid]["is_monitored"] = jid in whitelisted_set or item.get("is_monitored", True)
            else:
                chat_map[jid] = {
                    "id": jid,
                    "subject": subj,
                    "name": subj,
                    "is_group": item.get("is_group", jid.endswith("@g.us")),
                    "is_monitored": jid in whitelisted_set or item.get("is_monitored", True)
                }

    for jid in whitelisted_set:
        if jid and not jid.endswith('@newsletter') and not jid.endswith('@broadcast') and jid not in chat_map:
            chat_map[jid] = {
                "id": jid,
                "subject": jid.split("@")[0],
                "name": jid.split("@")[0],
                "is_group": jid.endswith("@g.us"),
                "is_monitored": True
            }

    return {"success": True, "groups": list(chat_map.values())}


@app.post("/api/whatsapp/whitelist")
async def update_whatsapp_whitelist(data: WhatsAppWhitelistUpdate, db: Session = Depends(get_db)):
    """Update monitored WhatsApp group/chat JIDs and preferences."""
    # Filter out any newsletter or broadcast channels
    clean_jids = [j for j in data.jids if j and not j.endswith('@newsletter') and not j.endswith('@broadcast')]
    clean_monitored = [
        m for m in (data.monitored_groups or []) 
        if m and not (m.get("id") or m.get("jid") or "").endswith('@newsletter') and not (m.get("id") or m.get("jid") or "").endswith('@broadcast')
    ]

    config = db.execute(select(WhatsAppConfig).order_by(WhatsAppConfig.id.desc())).scalar_one_or_none()
    if not config:
        config = WhatsAppConfig(
            whitelisted_jids=clean_jids,
            monitored_groups=clean_monitored,
            auto_reply=data.auto_reply if data.auto_reply is not None else True,
            is_enabled=data.is_enabled if data.is_enabled is not None else True
        )
        db.add(config)
    else:
        config.whitelisted_jids = clean_jids
        if data.monitored_groups is not None:
            config.monitored_groups = clean_monitored
        if data.auto_reply is not None:
            config.auto_reply = data.auto_reply
        if data.is_enabled is not None:
            config.is_enabled = data.is_enabled

    db.commit()

    # Sync to bridge
    try:
        async with httpx.AsyncClient(timeout=3.0) as client:
            await client.post(f"{WHATSAPP_BRIDGE_URL}/whitelist", json={"jids": clean_jids, "monitored_groups": clean_monitored})
    except Exception as e:
        log.warning("Could not sync whitelist to WhatsApp bridge: %s", e)

    return {"success": True, "whitelisted_count": len(clean_jids)}


@app.post("/api/whatsapp/contacts/edit")
async def edit_whatsapp_contact(data: WhatsAppContactEdit, db: Session = Depends(get_db)):
    """Edit friendly name/label or monitor state of a WhatsApp contact/group."""
    jid = (data.jid or "").strip()
    if not jid:
        raise HTTPException(400, "JID is required")

    config = db.execute(select(WhatsAppConfig).order_by(WhatsAppConfig.id.desc())).scalar_one_or_none()
    if not config:
        config = WhatsAppConfig(whitelisted_jids=[], monitored_groups=[])
        db.add(config)

    import copy
    groups = copy.deepcopy(config.monitored_groups or [])
    whitelisted = set(config.whitelisted_jids or [])

    found = False
    for item in groups:
        if (item.get("id") == jid or item.get("jid") == jid):
            if data.name is not None:
                item["subject"] = data.name.strip()
                item["name"] = data.name.strip()
            if data.is_monitored is not None:
                item["is_monitored"] = data.is_monitored
                if data.is_monitored:
                    whitelisted.add(jid)
                else:
                    whitelisted.discard(jid)
            found = True
            break

    if not found:
        groups.append({
            "id": jid,
            "subject": (data.name or jid.split("@")[0]).strip(),
            "name": (data.name or jid.split("@")[0]).strip(),
            "is_group": jid.endswith("@g.us"),
            "is_monitored": data.is_monitored if data.is_monitored is not None else True
        })
        if data.is_monitored is not False:
            whitelisted.add(jid)

    config.monitored_groups = groups
    config.whitelisted_jids = list(whitelisted)
    db.commit()

    # Sync to bridge
    try:
        async with httpx.AsyncClient(timeout=3.0) as client:
            await client.post(f"{WHATSAPP_BRIDGE_URL}/whitelist", json={"jids": list(whitelisted), "monitored_groups": groups})
    except Exception as e:
        log.warning("Could not sync updated contact to WhatsApp bridge: %s", e)

    return {"success": True, "jid": jid, "name": data.name, "monitored": jid in whitelisted}


@app.delete("/api/whatsapp/contacts/{jid:path}")
async def delete_whatsapp_contact(jid: str, db: Session = Depends(get_db)):
    """Remove a contact or group completely from monitored whitelist and saved contacts."""
    jid = (jid or "").strip()
    if not jid:
        raise HTTPException(400, "JID is required")

    config = db.execute(select(WhatsAppConfig).order_by(WhatsAppConfig.id.desc())).scalar_one_or_none()
    if config:
        groups = [item for item in (config.monitored_groups or []) if item.get("id") != jid and item.get("jid") != jid]
        whitelisted = [j for j in (config.whitelisted_jids or []) if j != jid]
        config.monitored_groups = groups
        config.whitelisted_jids = whitelisted
        db.commit()

    # Sync to bridge
    try:
        async with httpx.AsyncClient(timeout=3.0) as client:
            await client.delete(f"{WHATSAPP_BRIDGE_URL}/chats/{jid}")
    except Exception as e:
        log.warning("Could not delete chat from bridge: %s", e)

    return {"success": True, "message": f"Contact {jid} removed"}


@app.post("/api/whatsapp/webhook")
async def whatsapp_webhook(payload: dict, db: Session = Depends(get_db)):
    """Internal webhook called by Node.js bridge when a message is received."""
    from services.whatsapp_service import handle_whatsapp_webhook
    result = await handle_whatsapp_webhook(payload, db)
    return result


@app.post("/api/whatsapp/reconnect")
async def reconnect_whatsapp():
    """Trigger WhatsApp bridge reconnect / refresh QR."""
    try:
        async with httpx.AsyncClient(timeout=5.0) as client:
            r = await client.post(f"{WHATSAPP_BRIDGE_URL}/reconnect")
            return r.json()
    except Exception as e:
        raise HTTPException(502, f"Could not reach WhatsApp bridge: {e}")


@app.post("/api/whatsapp/disconnect")
async def disconnect_whatsapp():
    """Logout WhatsApp session and reset auth tokens."""
    try:
        async with httpx.AsyncClient(timeout=5.0) as client:
            r = await client.post(f"{WHATSAPP_BRIDGE_URL}/logout")
            return r.json()
    except Exception as e:
        raise HTTPException(502, f"Could not reach WhatsApp bridge: {e}")


@app.delete("/api/whatsapp/logs/{log_id}")
async def delete_whatsapp_log(log_id: int, db: Session = Depends(get_db)):
    """Delete a WhatsApp log record and its associated pending ingest if still unprocessed."""
    log_row = db.get(WhatsAppLog, log_id)
    if not log_row:
        raise HTTPException(404, "WhatsApp log record not found")

    # If associated with a PendingIngest that is still pending or was created from this log, delete it too
    if log_row.matched_id:
        pending = db.get(PendingIngest, log_row.matched_id)
        if pending and pending.status == IngestStatus.pending:
            db.delete(pending)
    elif log_row.message_id:
        pending = db.execute(
            select(PendingIngest).where(PendingIngest.source_message_id == log_row.message_id)
        ).scalar_one_or_none()
        if pending and pending.status == IngestStatus.pending:
            db.delete(pending)

    db.delete(log_row)
    db.commit()
    return {"success": True, "deleted_id": log_id}


@app.post("/api/whatsapp/logs/delete-batch")
async def delete_whatsapp_logs_batch(data: WhatsAppBatchDeleteLogs, db: Session = Depends(get_db)):
    """Batch delete multiple WhatsApp log records and linked pending ingests."""
    if not data.ids:
        return {"success": True, "deleted_count": 0}

    count = 0
    for log_id in data.ids:
        log_row = db.get(WhatsAppLog, log_id)
        if log_row:
            if log_row.matched_id:
                pending = db.get(PendingIngest, log_row.matched_id)
                if pending and pending.status == IngestStatus.pending:
                    db.delete(pending)
            elif log_row.message_id:
                pending = db.execute(
                    select(PendingIngest).where(PendingIngest.source_message_id == log_row.message_id)
                ).scalar_one_or_none()
                if pending and pending.status == IngestStatus.pending:
                    db.delete(pending)
            db.delete(log_row)
            count += 1

    db.commit()
    return {"success": True, "deleted_count": count}


@app.get("/api/whatsapp/logs")
def get_whatsapp_logs(
    company_id: Optional[int] = None,
    limit: int = 100,
    offset: int = 0,
    status: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Retrieve recent WhatsApp ingestion history and OCR logs."""
    q = select(WhatsAppLog)
    if company_id:
        q = q.where(WhatsAppLog.company_id == company_id)
    if status:
        q = q.where(WhatsAppLog.status == status)

    q = q.order_by(WhatsAppLog.id.desc()).offset(offset).limit(limit)
    rows = db.execute(q).scalars().all()
    total = db.execute(select(func.count(WhatsAppLog.id))).scalar() or 0

    logs_data = []
    for r in rows:
        # Ignore newsletter/broadcast channels completely from display
        if r.chat_jid and (r.chat_jid.endswith('@newsletter') or r.chat_jid.endswith('@broadcast')):
            continue

        pending = db.get(PendingIngest, r.matched_id) if r.matched_id else None
        if not pending and r.message_id:
            pending = db.execute(
                select(PendingIngest).where(PendingIngest.source_message_id == r.message_id)
            ).scalar_one_or_none()
            if pending and not r.matched_id:
                r.matched_id = pending.id
                db.commit()

        # Only auto-create pending ingest if it's a real document/bill and NOT marked ignored / irrelevant
        if not pending and (r.media_path or r.raw_text) and r.status != 'ignored' and r.doc_type not in ('irrelevant', 'spam', 'notification'):
            try:
                p_doc_type = DocumentType.purchase_bill if r.doc_type == "purchase_bill" else (DocumentType.plant_unloading if r.doc_type == "plant_unloading" else DocumentType.not_classified)
                pending = PendingIngest(
                    company_id=r.company_id,
                    source=IngestSource.whatsapp,
                    source_address=r.sender_phone or r.sender_jid,
                    source_account=f"{r.chat_name} ({r.sender_name})",
                    source_message_id=r.message_id,
                    file_name=os.path.basename(r.media_path) if r.media_path else f"wa_{r.id}.txt",
                    file_path=r.media_path or "",
                    document_type=p_doc_type,
                    extracted_payload=r.ocr_result or {},
                    status=IngestStatus.pending,
                    action_status="pending",
                    review_notes=f"WhatsApp {r.doc_type} from {r.sender_name}"
                )
                db.add(pending)
                db.commit()
                db.refresh(pending)
                r.matched_id = pending.id
                db.commit()
            except Exception as pe:
                log.warning("Could not auto-create pending ingest: %s", pe)

        raw_txt = (
            (pending.extracted_payload.get("raw_text") if pending and isinstance(pending.extracted_payload, dict) and pending.extracted_payload.get("raw_text") else None)
            or r.raw_text
            or (r.ocr_result.get("raw_text") if isinstance(r.ocr_result, dict) else "")
        )

        logs_data.append({
            "id": r.id,
            "message_id": r.message_id,
            "chat_jid": r.chat_jid,
            "chat_name": r.chat_name,
            "sender_name": r.sender_name,
            "sender_phone": r.sender_phone,
            "is_group": r.is_group,
            "doc_type": pending.document_type.value if pending and pending.document_type else r.doc_type,
            "media_path": r.media_path,
            "raw_text": raw_txt,
            "ocr_result": pending.extracted_payload if pending and pending.extracted_payload else (r.ocr_result or {}),
            "matched_id": r.matched_id,
            "pending_id": pending.id if pending else r.matched_id,
            "pending_status": pending.status.value if pending and pending.status else r.status,
            "action_status": pending.action_status if pending else "pending",
            "status": pending.status.value if pending and pending.status else r.status,
            "error_message": r.error_message,
            "reply_sent": r.reply_sent,
            "created_at": r.created_at.isoformat() if r.created_at else None
        })

    return {
        "success": True,
        "total": len(logs_data),
        "logs": logs_data
    }


@app.post("/api/whatsapp/test-send")
async def test_send_whatsapp(data: WhatsAppSendRequest):
    """Send a test WhatsApp message to verify connection."""
    from services.whatsapp_service import send_whatsapp_message
    ok = await send_whatsapp_message(data.jid, data.text)
    if not ok:
        raise HTTPException(500, "Failed to send WhatsApp message via bridge")
    return {"success": True, "message": "Sent successfully"}


# ── DEV: serve uploaded images ─────────────────────────────────────────────
@app.get("/uploads/{path:path}")
def serve_upload(path: str):
    settings = get_settings()
    fpath = Path(settings.upload_dir) / path
    if not fpath.exists(): raise HTTPException(404)
    return FileResponse(str(fpath))

if __name__ == "__main__":
    import uvicorn
    s = get_settings()
    uvicorn.run("main:app", host=s.host, port=s.port, reload=s.debug)